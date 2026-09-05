"""Banking & Statement Reconciliation Routes."""

import re
import logging
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any, Tuple
import io
from io import BytesIO
from bson import ObjectId
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Query, Response

from models.banking import (
    BankAccountIn,
    BankAccountUpdate,
    BalanceCorrectionIn,
    BankStatementLineIn,
    BankStatementLineUpdate,
    CashLedgerCreateIn,
    PeriodLockIn,
    PeriodUnlockIn,
)
from auth import require_roles

log = logging.getLogger(__name__)

banking_router = APIRouter(prefix="/api", tags=["banking"])


def _get_db(request: Request):
    return getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")


async def _get_user(request: Request) -> dict:
    getter = getattr(request.app, "get_current_user", None)
    if not getter:
        from server import get_current_user as getter
    return await getter(request)


def _oid(val: str) -> ObjectId:
    try:
        return ObjectId(str(val))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid ObjectId format")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stringify(doc: dict) -> dict:
    if not doc:
        return doc
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc.pop("_id"))
    for k, v in doc.items():
        if isinstance(v, ObjectId):
            doc[k] = str(v)
        elif isinstance(v, dict):
            doc[k] = stringify(v)
        elif isinstance(v, list):
            doc[k] = [stringify(i) if isinstance(i, dict) else (str(i) if isinstance(i, ObjectId) else i) for i in v]
    return doc


# ---------------------------------------------------------------------------
# Bank Accounts CRUD
# ---------------------------------------------------------------------------

@banking_router.get("/banking/accounts")
async def list_bank_accounts(
    request: Request,
    active: Optional[bool] = None,
    account_type: Optional[str] = None,
    search: Optional[str] = None,
):
    """List all configured bank accounts (e.g. HDFC - Online, UCO Bank - Offline)."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales", "production")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {}
    if active is not None:
        q["active"] = bool(active)
    if account_type:
        q["account_type"] = str(account_type)
    if search:
        rx = {"$regex": re.escape(str(search)), "$options": "i"}
        q["$or"] = [{"name": rx}, {"bank_name": rx}, {"account_number_last4": rx}]

    docs = await db.bank_accounts.find(q).sort("name", 1).to_list(1000)
    return [stringify(d) for d in docs]


@banking_router.post("/banking/accounts", status_code=201)
async def create_bank_account(payload: BankAccountIn, request: Request):
    """Create a new bank account."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    last4 = payload.account_number_last4.strip() if payload.account_number_last4 else ""
    if not last4 and payload.account_number:
        last4 = payload.account_number.strip()[-4:]

    doc = {
        "name": payload.name.strip(),
        "bank_name": payload.bank_name.strip(),
        "account_number_last4": last4,
        "account_number": (payload.account_number or "").strip() or None,
        "ifsc": (payload.ifsc or "").strip().upper() or None,
        "branch": (payload.branch or "").strip() or None,
        "account_type": payload.account_type,
        "opening_balance": float(payload.opening_balance),
        "opening_balance_date": payload.opening_balance_date,
        "statement_format": payload.statement_format.dict() if payload.statement_format else None,
        "active": payload.active,
        "created_at": _now_iso(),
        "created_by": u.get("email") or u.get("name", ""),
    }
    res = await db.bank_accounts.insert_one(doc)
    created = await db.bank_accounts.find_one({"_id": res.inserted_id})
    return stringify(created)


@banking_router.get("/banking/accounts/{id}")
async def get_bank_account(id: str, request: Request):
    """Retrieve a single bank account by ID."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales", "production")(u)
    db = _get_db(request)

    doc = await db.bank_accounts.find_one({"_id": _oid(id)})
    if not doc:
        raise HTTPException(404, "Bank account not found")
    return stringify(doc)


@banking_router.patch("/banking/accounts/{id}")
async def update_bank_account(id: str, payload: BankAccountUpdate, request: Request):
    """Update bank account details."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if not updates:
        doc = await db.bank_accounts.find_one({"_id": _oid(id)})
        if not doc:
            raise HTTPException(404, "Bank account not found")
        return stringify(doc)

    updates["updated_at"] = _now_iso()
    updates["updated_by"] = u.get("email") or u.get("name", "")

    res = await db.bank_accounts.update_one({"_id": _oid(id)}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Bank account not found")

    updated = await db.bank_accounts.find_one({"_id": _oid(id)})
    return stringify(updated)


@banking_router.delete("/banking/accounts/{id}")
async def delete_bank_account(id: str, request: Request):
    """Deactivate or delete a bank account."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    lines_count = await db.bank_statement_lines.count_documents({"bank_account_id": id})
    if lines_count > 0:
        # Soft-delete by setting active=False if statement lines exist
        await db.bank_accounts.update_one({"_id": _oid(id)}, {"$set": {"active": False, "updated_at": _now_iso()}})
        return {"ok": True, "message": "Bank account has existing statement lines and was deactivated."}

    res = await db.bank_accounts.delete_one({"_id": _oid(id)})
    if res.deleted_count == 0:
        raise HTTPException(404, "Bank account not found")
    return {"ok": True, "message": "Bank account deleted"}


@banking_router.post("/banking/accounts/{id}/correct-opening-balance")
@banking_router.post("/bank-accounts/{id}/correct-opening-balance")
async def correct_account_opening_balance(id: str, payload: BalanceCorrectionIn, request: Request):
    """
    Correct a bank account's opening balance with mandatory audit reason and period-lock safety check.
    """
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Correction reason is required and cannot be empty.")

    account = await db.bank_accounts.find_one({"_id": _oid(id)})
    if not account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    old_value = float(account.get("opening_balance") or 0.0)
    new_value = float(payload.new_opening_balance)
    opening_balance_date = account.get("opening_balance_date")

    # Safety check: does this account have any locked period starting on or after opening_balance_date?
    acc_id_str = str(account.get("_id", id))
    lock_query = {
        "status": "locked",
        "bank_account_id": {"$in": [acc_id_str, "all", None, ""]},
    }
    if opening_balance_date:
        lock_query["period_from"] = {"$gte": opening_balance_date}

    locked_period = await db.reconciliation_locks.find_one(lock_query, sort=[("period_from", 1)])
    if not locked_period and not opening_balance_date:
        locked_period = await db.reconciliation_locks.find_one({
            "status": "locked",
            "bank_account_id": {"$in": [acc_id_str, "all", None, ""]},
        }, sort=[("period_from", 1)])

    if locked_period:
        p_start = locked_period.get("period_from", "unknown")
        raise HTTPException(
            status_code=400,
            detail=f"This account has a locked period starting {p_start} — unlock it first if you need to correct a balance that predates it, since the locked period's figures were computed against the current balance and would become inconsistent"
        )

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "admin")

    correction_doc = {
        "bank_account_id": acc_id_str,
        "old_value": old_value,
        "new_value": new_value,
        "reason": reason,
        "corrected_by": user_email,
        "corrected_at": now,
    }

    await db.balance_corrections.insert_one(dict(correction_doc))

    await db.bank_accounts.update_one(
        {"_id": account["_id"]},
        {
            "$set": {
                "opening_balance": new_value,
                "updated_at": now,
                "updated_by": user_email,
                "last_balance_correction": {
                    "corrected_at": now,
                    "corrected_by": user_email,
                    "reason": reason,
                    "old_value": old_value,
                    "new_value": new_value,
                },
            },
            "$push": {
                "balance_corrections": correction_doc
            }
        }
    )

    log.info(f"Opening balance for account {acc_id_str} corrected from {old_value} to {new_value} by {user_email}. Reason: {reason}")

    return {
        "ok": True,
        "message": "Opening balance successfully corrected.",
        "correction": stringify(correction_doc),
        "account": stringify(await db.bank_accounts.find_one({"_id": account["_id"]})),
    }


@banking_router.get("/banking/accounts/{id}/balance-corrections")
@banking_router.get("/bank-accounts/{id}/balance-corrections")
async def list_account_balance_corrections(id: str, request: Request):
    """List all balance correction audit records for a bank account."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    docs = await db.balance_corrections.find({"bank_account_id": str(id)}).sort("corrected_at", -1).to_list(200)
    return {
        "ok": True,
        "corrections": [stringify(d) for d in docs],
    }


# ---------------------------------------------------------------------------
# Bank Statement Lines
# ---------------------------------------------------------------------------

@banking_router.get("/banking/statement-lines")
async def list_statement_lines(
    request: Request,
    bank_account_id: Optional[str] = None,
    match_status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
    skip: int = 0,
):
    """List bank statement lines with reconciliation and match status."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {}
    if bank_account_id:
        q["bank_account_id"] = str(bank_account_id)
    if match_status:
        q["match_status"] = str(match_status)
    if from_date or to_date:
        dq = {}
        if from_date:
            dq["$gte"] = str(from_date)
        if to_date:
            dq["$lte"] = str(to_date)
        q["date"] = dq
    if search:
        rx = {"$regex": re.escape(str(search)), "$options": "i"}
        q["$or"] = [{"narration": rx}, {"reference_no": rx}]

    docs = await db.bank_statement_lines.find(q).sort([("date", -1), ("_id", -1)]).skip(skip).limit(limit).to_list(limit)
    total = await db.bank_statement_lines.count_documents(q)

    # Collect cash_ledger IDs for cash_withdrawal lines to populate audit trail details
    cash_ledger_ids = [
        d["matched_to"]["ref_id"]
        for d in docs
        if isinstance(d.get("matched_to"), dict) and d["matched_to"].get("type") == "cash_withdrawal" and d["matched_to"].get("ref_id")
    ]
    cash_ledger_map = {}
    if cash_ledger_ids and hasattr(db, "cash_ledger") and db.cash_ledger is not None:
        try:
            cl_obj_ids = []
            for cid in cash_ledger_ids:
                try:
                    cl_obj_ids.append(_oid(cid))
                except Exception:
                    pass
            cl_cursor = db.cash_ledger.find({"_id": {"$in": cl_obj_ids}})
            if hasattr(cl_cursor, "to_list"):
                res = cl_cursor.to_list(len(cl_obj_ids) + 10)
                if hasattr(res, "__await__"):
                    cl_docs = await res
                elif isinstance(res, list):
                    cl_docs = res
                else:
                    cl_docs = []
                cash_ledger_map = {str(c["_id"]): c for c in cl_docs}
        except Exception:
            cash_ledger_map = {}

    wage_payments_by_cl = {}
    expenses_by_cl = {}
    if cash_ledger_ids:
        cl_str_ids = [str(c) for c in cash_ledger_ids]
        if hasattr(db, "wage_payments") and db.wage_payments is not None:
            try:
                wp_cursor = db.wage_payments.find({"cash_ledger_id": {"$in": cl_str_ids}})
                if hasattr(wp_cursor, "to_list"):
                    res = wp_cursor.to_list(1000)
                    if hasattr(res, "__await__"):
                        wp_docs = await res
                    elif isinstance(res, list):
                        wp_docs = res
                    else:
                        wp_docs = []
                    for wp in wp_docs:
                        clid = str(wp.get("cash_ledger_id"))
                        wage_payments_by_cl.setdefault(clid, []).append(wp)
            except Exception:
                wage_payments_by_cl = {}

        if hasattr(db, "expenses") and db.expenses is not None:
            try:
                exp_cursor = db.expenses.find({"cash_ledger_id": {"$in": cl_str_ids}})
                if hasattr(exp_cursor, "to_list"):
                    res = exp_cursor.to_list(1000)
                    if hasattr(res, "__await__"):
                        exp_docs = await res
                    elif isinstance(res, list):
                        exp_docs = res
                    else:
                        exp_docs = []
                    for e in exp_docs:
                        clid = str(e.get("cash_ledger_id"))
                        expenses_by_cl.setdefault(clid, []).append(e)
            except Exception:
                expenses_by_cl = {}

    items = []
    for d in docs:
        sd = stringify(d)
        if isinstance(d.get("matched_to"), dict) and d["matched_to"].get("type") == "cash_withdrawal":
            ref_id = str(d["matched_to"].get("ref_id"))
            cl_doc = cash_ledger_map.get(ref_id, {})
            wps = wage_payments_by_cl.get(ref_id, [])
            exps = expenses_by_cl.get(ref_id, [])
            allocated_wages = sum(float(wp.get("amount") or 0.0) for wp in wps)
            allocated_expenses = sum(float(e.get("amount") or 0.0) for e in exps)
            allocated = round(allocated_wages + allocated_expenses, 2)
            withdrawn = float(cl_doc.get("amount") or d.get("debit_amount") or 0.0)
            rem = float(cl_doc.get("remaining_balance") if cl_doc.get("remaining_balance") is not None else (withdrawn - allocated))

            # Combine all disbursements sorted by date
            disbursements = []
            for wp in wps:
                disbursements.append({
                    "id": str(wp.get("_id") or wp.get("id")),
                    "type": "wage_payment",
                    "type_label": "Karigar Wage",
                    "title": wp.get("worker_name") or f"Worker #{str(wp.get('worker_id'))[-6:]}",
                    "amount": float(wp.get("amount") or 0.0),
                    "date": wp.get("date"),
                    "period_from": wp.get("period_from"),
                    "period_to": wp.get("period_to"),
                    "notes": wp.get("notes") or "",
                    "override_reason": wp.get("override_reason"),
                })
            for e in exps:
                disbursements.append({
                    "id": str(e.get("_id") or e.get("id")),
                    "type": "expense",
                    "type_label": "Cash Expense",
                    "title": e.get("payee") or "Payee",
                    "category": e.get("category") or "Expense",
                    "amount": float(e.get("amount") or 0.0),
                    "date": e.get("date"),
                    "notes": e.get("notes") or "",
                })
            disbursements.sort(key=lambda x: str(x.get("date") or ""), reverse=True)

            sd["cash_ledger_info"] = {
                "cash_ledger_id": ref_id,
                "withdrawal_amount": withdrawn,
                "remaining_balance": round(rem, 2),
                "allocated_amount": allocated,
                "wage_payment_count": len(wps),
                "wage_payments": [stringify(wp) for wp in wps],
                "expense_count": len(exps),
                "expenses": [stringify(e) for e in exps],
                "disbursements": disbursements,
            }
        items.append(sd)

    return {
        "total": total,
        "items": items,
    }


@banking_router.post("/banking/statement-lines", status_code=201)
async def create_statement_lines(lines: List[BankStatementLineIn], request: Request):
    """Insert one or more bank statement lines with deduplication against existing lines."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    if not lines:
        raise HTTPException(400, "No statement lines provided")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")

    def _line_sig(doc):
        return (
            str(doc.get("date") or ""),
            round(float(doc.get("debit_amount") or 0.0), 2),
            round(float(doc.get("credit_amount") or 0.0), 2),
            str(doc.get("narration") or "").strip(),
        )

    lines_by_acc = defaultdict(list)
    for l in lines:
        row = l.dict()
        row["imported_at"] = row.get("imported_at") or now
        row["imported_by"] = row.get("imported_by") or user_email
        lines_by_acc[row.get("bank_account_id")].append(row)

    to_insert = []
    skipped_count = 0

    for acc_id, acc_lines in lines_by_acc.items():
        dates = [r["date"] for r in acc_lines if r.get("date")]
        date_q = {"date": {"$gte": min(dates), "$lte": max(dates)}} if dates else {}
        existing_cursor = db.bank_statement_lines.find({"bank_account_id": acc_id, **date_q})
        existing_docs = []
        if hasattr(existing_cursor, "to_list"):
            res = existing_cursor.to_list(100000)
            if hasattr(res, "__await__"):
                existing_docs = await res
            elif isinstance(res, list):
                existing_docs = res
        elif isinstance(existing_cursor, list):
            existing_docs = existing_cursor

        existing_counts = Counter(_line_sig(d) for d in existing_docs)
        for row in acc_lines:
            sig = _line_sig(row)
            if existing_counts[sig] > 0:
                existing_counts[sig] -= 1
                skipped_count += 1
            else:
                to_insert.append(row)

    inserted_ids = []
    if to_insert:
        res = await db.bank_statement_lines.insert_many(to_insert)
        inserted_ids = [str(i) for i in res.inserted_ids]

    summary_msg = f"{len(inserted_ids)} new, {skipped_count} skipped as duplicates." if skipped_count > 0 else f"{len(inserted_ids)} new lines inserted."
    return {
        "ok": True,
        "inserted_count": len(inserted_ids),
        "skipped_count": skipped_count,
        "message": summary_msg,
        "inserted_ids": inserted_ids,
    }


# ---------------------------------------------------------------------------
# Bank Statement Import & Config
# ---------------------------------------------------------------------------

from models.sku_map import SheetLocator, HeaderLocator
from models.banking import (
    StatementImportConfigIn,
    StatementImportConfigUpdate,
    STATEMENT_CANONICAL_FIELDS,
)
from routes.online_orders import _parse_tabular_bytes, _resolve_column, _detect_tabular_file_format, _parse_xlrd_workbook, _is_html_content, _parse_html_table


def _extract_statement_metadata(content: bytes, filename: str) -> Dict[str, str]:
    """
    Extract account metadata (account_number, ifsc, branch) from statement header rows.
    """
    fmt = _detect_tabular_file_format(content, filename)
    rows_raw = []
    if fmt == "html_xls":
        try:
            rows_raw = _parse_html_table(content)
        except Exception:
            pass
    elif fmt == "xls":
        try:
            rows_raw = _parse_xlrd_workbook(content, SheetLocator(type="first_sheet"))
        except Exception:
            # Try HTML fallback
            if _is_html_content(content):
                try:
                    rows_raw = _parse_html_table(content)
                except Exception:
                    pass
    elif fmt == "xlsx":
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
            sheet = wb.worksheets[0]
            for r in sheet.iter_rows(values_only=True):
                rows_raw.append([str(c) if c is not None else "" for c in r])
        except Exception:
            if _is_html_content(content):
                try:
                    rows_raw = _parse_html_table(content)
                except Exception:
                    pass
    elif fmt == "csv":
        try:
            import csv, io
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                import csv, io
                text = content.decode("latin-1")
            except Exception:
                text = ""
        if text:
            reader = csv.reader(io.StringIO(text))
            rows_raw = list(reader)

    meta: Dict[str, str] = {}
    for r in rows_raw:
        row_cells = []
        for c in r:
            s = str(c).strip()
            if s and (not row_cells or s != row_cells[-1]):
                row_cells.append(s)

        for i, cell in enumerate(row_cells):
            cell_lc = cell.lower()
            if not meta.get("account_number"):
                if re.search(r"^(?:account\s*no\.?|acct\s*no\.?|a/c\s*no\.?|number\s*:?)$", cell_lc) and i + 1 < len(row_cells):
                    next_val = re.sub(r"[^\d]", "", row_cells[i + 1])
                    if 9 <= len(next_val) <= 20:
                        meta["account_number"] = next_val
                else:
                    m_acc = re.search(r"(?:account\s*no\.?|acct\s*no\.?|a/c\s*no\.?|number\s*:)\s*(\d{9,20})", cell, re.IGNORECASE)
                    if m_acc:
                        meta["account_number"] = m_acc.group(1)

            if not meta.get("ifsc"):
                if re.search(r"^ifsc\s*(?:code)?\s*:?$", cell_lc) and i + 1 < len(row_cells):
                    next_val = row_cells[i + 1].strip().upper()
                    if re.match(r"^[A-Z]{4}0[A-Z0-9]{6}$", next_val):
                        meta["ifsc"] = next_val
                else:
                    m_ifsc = re.search(r"\b([A-Z]{4}0[A-Z0-9]{6})\b", cell)
                    if m_ifsc:
                        meta["ifsc"] = m_ifsc.group(1).upper()

            if not meta.get("branch"):
                if re.search(r"^branch\s*:?$", cell_lc) and i + 1 < len(row_cells):
                    next_val = row_cells[i + 1].strip()
                    if next_val and next_val.lower() != "branch":
                        meta["branch"] = next_val
                else:
                    m_br = re.search(r"branch\s*:\s*([A-Za-z0-9\s]+?)(?:\s{2,}|$)", cell, re.IGNORECASE)
                    if m_br:
                        val = m_br.group(1).strip()
                        if val:
                            meta["branch"] = val

    if meta.get("account_number") and not meta.get("account_number_last4"):
        meta["account_number_last4"] = meta["account_number"][-4:]

    return meta


def _parse_date_to_iso(date_str: str, custom_format: Optional[str] = None) -> str:
    """Normalize various bank statement date formats to standard ISO YYYY-MM-DD."""
    s = str(date_str or "").strip()
    if not s:
        return ""
    if custom_format:
        try:
            return datetime.strptime(s, custom_format).strftime("%Y-%m-%d")
        except Exception:
            pass
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]

    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
        "%Y/%m/%d", "%Y-%m-%d",
        "%d %b %Y", "%d-%b-%Y", "%d/%b/%Y",
        "%d %B %Y", "%d-%B-%Y", "%d/%B/%Y",
        "%b %d, %Y", "%B %d, %Y",
        "%d-%b-%y", "%d/%b/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s.split(" ")[0] if " " in s and len(s) > 11 else s, fmt).strftime("%Y-%m-%d")
        except Exception:
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
    return s[:10]


def _parse_amount(val: Any) -> float:
    """Safely convert dirty amount strings with currency symbols and commas to float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ["-", "--", "NA", "N/A", "null", "None"]:
        return 0.0
    s_clean = re.sub(r"[₹RsINR,\s]", "", s, flags=re.IGNORECASE)
    s_clean = re.sub(r"(dr|cr)", "", s_clean, flags=re.IGNORECASE).strip()
    if s_clean.startswith("(") and s_clean.endswith(")"):
        s_clean = f"-{s_clean[1:-1]}"
    try:
        return float(s_clean)
    except Exception:
        return 0.0


DEFAULT_BANK_FORMAT_TEMPLATES = {
    "uco": {
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {
            "type": "scan_for_columns",
            "must_contain_any": [
                "Tran. Date", "Transaction Date", "Value Date",
                "Withdrawl", "Withdrawal", "Deposit", "Balance",
                "Narration", "Description"
            ]
        },
        "skip_rows_after_header": 0,
        "column_map": {
            "date": "Tran. Date",
            "narration": "Narration",
            "reference": "Chq. No.",
            "debit_amount": "Withdrawl",
            "credit_amount": "Deposit",
            "balance": "Balance",
        },
        "date_format": "%d/%m/%Y",
        "notes": "Standard UCO Bank Corporate / Retail Netbanking XLS/CSV template (date from Tran. Date)"
    },
    "hdfc": {
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {
            "type": "scan_for_columns",
            "must_contain_any": [
                "Date", "Narration", "Chq./Ref.No.",
                "Withdrawal Amt.", "Deposit Amt.", "Closing Balance"
            ]
        },
        "skip_rows_after_header": 0,
        "column_map": {
            "date": "Date",
            "narration": "Narration",
            "reference": "Chq./Ref.No.",
            "debit_amount": "Withdrawal Amt.",
            "credit_amount": "Deposit Amt.",
            "balance": "Closing Balance",
        },
        "date_format": "%d/%m/%Y",
        "notes": "Standard HDFC Bank Statement template"
    }
}


@banking_router.get("/banking/accounts/{id}/statement-format")
@banking_router.get("/bank-accounts/{id}/statement-format")
async def get_statement_format(id: str, request: Request):
    """Get the saved statement column-mapping format for a bank account."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    doc = await db.bank_accounts.find_one({"_id": _oid(id)})
    if not doc:
        raise HTTPException(404, "Bank account not found")
    return {
        "bank_account_id": id,
        "bank_name": doc.get("bank_name"),
        "canonical_fields": STATEMENT_CANONICAL_FIELDS,
        "statement_format": doc.get("statement_format"),
        "default_templates": DEFAULT_BANK_FORMAT_TEMPLATES,
    }


@banking_router.put("/banking/accounts/{id}/statement-format")
@banking_router.put("/bank-accounts/{id}/statement-format")
@banking_router.post("/banking/accounts/{id}/statement-format")
@banking_router.post("/bank-accounts/{id}/statement-format")
async def save_statement_format(id: str, payload: StatementImportConfigIn, request: Request):
    """Configure or update the statement column-mapping format for a bank account."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    fmt_data = payload.dict()
    res = await db.bank_accounts.update_one(
        {"_id": _oid(id)},
        {"$set": {"statement_format": fmt_data, "updated_at": _now_iso()}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Bank account not found")

    updated = await db.bank_accounts.find_one({"_id": _oid(id)})
    return stringify(updated)


async def _handle_statement_import(
    id: str,
    file: UploadFile,
    dry_run: bool,
    confirm_account_update: bool,
    request: Request,
) -> dict:
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    acc = await db.bank_accounts.find_one({"_id": _oid(id)})
    if not acc:
        raise HTTPException(404, "Bank account not found")

    cfg = acc.get("statement_format")
    if not cfg or not cfg.get("column_map"):
        bank_name_lower = (acc.get("bank_name") or acc.get("name") or "").lower()
        if "uco" in bank_name_lower:
            cfg = DEFAULT_BANK_FORMAT_TEMPLATES.get("uco")
        elif "hdfc" in bank_name_lower:
            cfg = DEFAULT_BANK_FORMAT_TEMPLATES.get("hdfc")
        else:
            cfg = DEFAULT_BANK_FORMAT_TEMPLATES.get("uco")

    if not cfg or not cfg.get("column_map"):
        raise HTTPException(
            400,
            f"Bank account '{acc.get('name')}' does not have a saved statement column-map format. "
            "Please configure the statement format before importing."
        )

    sheet_loc = SheetLocator(**(cfg.get("sheet_locator") or {"type": "first_sheet"}))
    header_loc = HeaderLocator(**(cfg.get("header_locator") or {"type": "fixed_row", "row": 0}))
    skip_rows = int(cfg.get("skip_rows_after_header") or 0)
    col_map = cfg.get("column_map") or {}
    custom_date_fmt = cfg.get("date_format")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Uploaded file is empty")

    detected_fmt = _detect_tabular_file_format(content, file.filename)
    log.info("Statement import: file=%s, size=%d bytes, detected_format=%s", file.filename, len(content), detected_fmt)

    headers, raw_rows = _parse_tabular_bytes(content, file.filename, sheet_loc, header_loc, skip_rows)
    log.info("Statement import: parsed headers=%s, data_rows=%d", headers, len(raw_rows) if raw_rows else 0)
    if not headers or not raw_rows:
        raise HTTPException(400, "Could not extract tabular header and data rows from file")

    date_col = _resolve_column(col_map.get("date", ""), headers)
    narr_col = _resolve_column(col_map.get("narration", ""), headers)
    ref_col = _resolve_column(col_map.get("reference", ""), headers) if col_map.get("reference") else None
    debit_col = _resolve_column(col_map.get("debit_amount", ""), headers) if col_map.get("debit_amount") else None
    credit_col = _resolve_column(col_map.get("credit_amount", ""), headers) if col_map.get("credit_amount") else None
    bal_col = _resolve_column(col_map.get("balance", ""), headers) if col_map.get("balance") else None

    if not date_col:
        raise HTTPException(400, f"Configured date column '{col_map.get('date')}' not found in file headers: {headers}")
    if not narr_col:
        raise HTTPException(400, f"Configured narration column '{col_map.get('narration')}' not found in file headers: {headers}")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")

    statement_lines = []
    for r in raw_rows:
        date_raw = r.get(date_col, "")
        narr_raw = r.get(narr_col, "")
        if not date_raw and not narr_raw:
            continue

        ref_raw = r.get(ref_col, "") if ref_col else ""
        debit_val = _parse_amount(r.get(debit_col)) if debit_col else 0.0
        credit_val = _parse_amount(r.get(credit_col)) if credit_col else 0.0
        bal_val = _parse_amount(r.get(bal_col)) if bal_col and r.get(bal_col) not in ["", None] else None

        line_doc = {
            "bank_account_id": id,
            "date": _parse_date_to_iso(date_raw, custom_date_fmt),
            "narration": str(narr_raw).strip(),
            "reference_no": str(ref_raw).strip() if ref_raw else "",
            "debit_amount": abs(float(debit_val)),
            "credit_amount": abs(float(credit_val)),
            "running_balance": float(bal_val) if bal_val is not None else None,
            "match_status": "unmatched",
            "matched_to": None,
            "remarks": "",
            "imported_at": now,
            "imported_by": user_email,
        }
        statement_lines.append(line_doc)

    if not statement_lines:
        raise HTTPException(400, "No valid statement transaction rows extracted from the file")

    # Extract metadata from statement header for auto-populate suggestion (requires explicit user confirmation)
    extracted_meta = _extract_statement_metadata(content, file.filename)
    suggested_update: Dict[str, str] = {}
    if extracted_meta.get("account_number") and not acc.get("account_number"):
        suggested_update["account_number"] = extracted_meta["account_number"]
    if extracted_meta.get("ifsc") and not acc.get("ifsc"):
        suggested_update["ifsc"] = extracted_meta["ifsc"]
    if extracted_meta.get("branch") and not acc.get("branch"):
        suggested_update["branch"] = extracted_meta["branch"]
    if extracted_meta.get("account_number_last4") and not acc.get("account_number_last4"):
        suggested_update["account_number_last4"] = extracted_meta["account_number_last4"]

    applied_account_update = None
    is_confirmed = isinstance(confirm_account_update, bool) and confirm_account_update
    if is_confirmed and suggested_update and not dry_run:
        upd = dict(suggested_update)
        upd["updated_at"] = _now_iso()
        upd["updated_by"] = user_email
        await db.bank_accounts.update_one({"_id": _oid(id)}, {"$set": upd})
        applied_account_update = suggested_update
        suggested_update = {}

    # Deduplicate against existing lines on the same bank_account_id
    # matching signature: (date, debit_amount, credit_amount, narration)
    dates = [l["date"] for l in statement_lines if l.get("date")]
    date_q = {"date": {"$gte": min(dates), "$lte": max(dates)}} if dates else {}

    existing_query = {"bank_account_id": id, **date_q}
    existing_cursor = db.bank_statement_lines.find(existing_query)
    existing_docs = []
    if hasattr(existing_cursor, "to_list"):
        res = existing_cursor.to_list(100000)
        if hasattr(res, "__await__"):
            existing_docs = await res
        elif isinstance(res, list):
            existing_docs = res
    elif isinstance(existing_cursor, list):
        existing_docs = existing_cursor

    def _line_sig(doc):
        return (
            str(doc.get("date") or ""),
            round(float(doc.get("debit_amount") or 0.0), 2),
            round(float(doc.get("credit_amount") or 0.0), 2),
            str(doc.get("narration") or "").strip(),
        )

    existing_counts = Counter(_line_sig(d) for d in existing_docs)

    new_lines_to_insert = []
    skipped_count = 0
    for l in statement_lines:
        sig = _line_sig(l)
        if existing_counts[sig] > 0:
            existing_counts[sig] -= 1
            skipped_count += 1
        else:
            new_lines_to_insert.append(l)

    if dry_run:
        res_dict = {
            "ok": True,
            "dry_run": True,
            "bank_account_id": id,
            "bank_account_name": acc.get("name"),
            "total_file_rows": len(raw_rows),
            "parsed_count": len(statement_lines),
            "new_count": len(new_lines_to_insert),
            "skipped_count": skipped_count,
            "message": f"Parsed {len(statement_lines)} rows ({len(new_lines_to_insert)} new, {skipped_count} skipped as duplicates)." if skipped_count > 0 else f"Parsed {len(statement_lines)} rows ({len(new_lines_to_insert)} new).",
            "sample": [stringify(s) for s in (new_lines_to_insert[:5] if new_lines_to_insert else statement_lines[:5])],
        }
        if suggested_update:
            res_dict["suggested_account_update"] = suggested_update
            res_dict["requires_account_confirmation"] = True
            res_dict["account_update_prompt"] = (
                "Statement header contains bank details: "
                + ", ".join(f"{k.replace('_', ' ').title()}: {v}" for k, v in suggested_update.items())
                + ". Confirm update to save these to the bank account."
            )
        return res_dict

    inserted_ids = []
    if new_lines_to_insert:
        res = await db.bank_statement_lines.insert_many(new_lines_to_insert)
        inserted_ids = [str(i) for i in res.inserted_ids]

    inserted_count = len(inserted_ids)
    if skipped_count > 0:
        summary_msg = f"{inserted_count} new, {skipped_count} skipped as duplicates."
    else:
        summary_msg = f"{inserted_count} new lines inserted."

    res_dict = {
        "ok": True,
        "dry_run": False,
        "bank_account_id": id,
        "bank_account_name": acc.get("name"),
        "total_file_rows": len(raw_rows),
        "parsed_count": len(statement_lines),
        "inserted_count": inserted_count,
        "skipped_count": skipped_count,
        "message": summary_msg,
        "sample": [stringify(s) for s in new_lines_to_insert[:5]],
    }
    if suggested_update:
        res_dict["suggested_account_update"] = suggested_update
        res_dict["requires_account_confirmation"] = True
        res_dict["account_update_prompt"] = (
            "Statement header contains bank details: "
            + ", ".join(f"{k.replace('_', ' ').title()}: {v}" for k, v in suggested_update.items())
            + ". Confirm update to save these to the bank account."
        )
    if applied_account_update:
        res_dict["applied_account_update"] = applied_account_update
    return res_dict


@banking_router.post("/bank-accounts/{id}/statement/import")
@banking_router.post("/banking/accounts/{id}/statement/import")
async def import_bank_statement(
    id: str,
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    confirm_account_update: bool = Query(False),
):
    """Import statement file (CSV/XLS/XLSX) using the bank account's saved column-map configuration."""
    return await _handle_statement_import(id, file, dry_run, confirm_account_update, request)


async def _check_statement_line_dependents(db, line_doc: dict, target_action: str = "reclassify"):
    """
    Check if a statement line has dependent records (e.g. cash_ledger entry with wage payments or expenses).
    Blocks the reclassification/re-matching if dependent records exist.
    If an unused cash_ledger entry exists with NO dependents, cleanly removes the unused cash_ledger entry.
    """
    if not line_doc:
        return

    line_id_str = str(line_doc.get("_id") or line_doc.get("id"))
    matched_to = line_doc.get("matched_to") or {}
    matched_type = matched_to.get("type") if isinstance(matched_to, dict) else None
    ref_id = matched_to.get("ref_id") if isinstance(matched_to, dict) else None

    # Check if this line is linked to a cash_ledger entry
    cash_entry = None
    if hasattr(db, "cash_ledger") and db.cash_ledger is not None:
        if matched_type == "cash_withdrawal" and ref_id:
            try:
                cash_entry = await db.cash_ledger.find_one({"_id": _oid(ref_id)})
            except Exception:
                cash_entry = None
        elif matched_type in [None, "cash_withdrawal"]:
            try:
                cash_entry = await db.cash_ledger.find_one({"source_statement_line_id": line_id_str})
            except Exception:
                cash_entry = None

    if cash_entry:
        clid = str(cash_entry["_id"])
        wage_payments_count = 0
        expenses_count = 0

        if hasattr(db, "wage_payments") and db.wage_payments is not None:
            try:
                wp_cursor = db.wage_payments.find({"cash_ledger_id": clid})
                if hasattr(wp_cursor, "to_list"):
                    wps = wp_cursor.to_list(1000)
                    if hasattr(wps, "__await__"):
                        wps = await wps
                    wage_payments_count = len(wps) if isinstance(wps, list) else 0
                elif hasattr(db.wage_payments, "count_documents"):
                    res = db.wage_payments.count_documents({"cash_ledger_id": clid})
                    if hasattr(res, "__await__"):
                        wage_payments_count = await res
                    else:
                        wage_payments_count = int(res or 0)
            except Exception:
                wage_payments_count = 0

        if hasattr(db, "expenses") and db.expenses is not None:
            try:
                exp_cursor = db.expenses.find({"cash_ledger_id": clid})
                if hasattr(exp_cursor, "to_list"):
                    exps = exp_cursor.to_list(1000)
                    if hasattr(exps, "__await__"):
                        exps = await exps
                    expenses_count = len(exps) if isinstance(exps, list) else 0
                elif hasattr(db.expenses, "count_documents"):
                    res = db.expenses.count_documents({"cash_ledger_id": clid})
                    if hasattr(res, "__await__"):
                        expenses_count = await res
                    else:
                        expenses_count = int(res or 0)
            except Exception:
                expenses_count = 0

        if wage_payments_count > 0 or expenses_count > 0:
            details = []
            if wage_payments_count > 0:
                details.append(f"{wage_payments_count} wage payment(s)")
            if expenses_count > 0:
                details.append(f"{expenses_count} cash expense(s)")
            dep_str = " and ".join(details)
            raise HTTPException(
                400,
                f"Cannot {target_action} statement line '{line_id_str}': it is linked to cash ledger entry '{clid}' "
                f"which has active dependent records ({dep_str}). "
                f"Please delete or reassign the dependent wage payments/expenses before re-matching or reclassifying this line.",
            )

        # If cash ledger entry has 0 dependents, clean it up so no orphaned cash pool remains
        try:
            await db.cash_ledger.delete_one({"_id": cash_entry["_id"]})
        except Exception:
            pass


async def _check_period_locked(db, bank_account_id: Optional[str], line_date: Optional[str], action: str = "modify"):
    """
    Checks if a transaction date for a given bank account falls within an active locked reconciliation period.
    If locked, raises HTTPException(400) blocking the action until an admin unlocks the period.
    """
    if not line_date or not hasattr(db, "reconciliation_locks") or db.reconciliation_locks is None:
        return

    line_date_str = str(line_date).strip()[:10]
    acc_id_str = str(bank_account_id) if bank_account_id else ""

    q = {
        "status": "locked",
        "period_from": {"$lte": line_date_str},
        "period_to": {"$gte": line_date_str},
    }
    if acc_id_str:
        q["bank_account_id"] = {"$in": [acc_id_str, "all", None, ""]}
    
    lock = None
    try:
        lock = await db.reconciliation_locks.find_one(q)
    except Exception:
        lock = None

    if lock:
        acc_label = "all bank accounts" if lock.get("bank_account_id") in [None, "all", ""] else f"bank account '{acc_id_str}'"
        locked_by = lock.get("locked_by") or "Administrator"
        locked_at = str(lock.get("locked_at") or "")[:10]
        raise HTTPException(
            400,
            f"Cannot {action} transaction on {line_date_str}: Reconciliation period ({lock.get('period_from')} to {lock.get('period_to')}) "
            f"for {acc_label} is finalized and locked by {locked_by} on {locked_at}. "
            f"An administrator must explicitly unlock this period before making any changes.",
        )


@banking_router.patch("/banking/statement-lines/{id}/match")
async def match_statement_line(id: str, payload: BankStatementLineUpdate, request: Request):
    """Update match status or link a statement line to an internal payment/settlement/expense."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    doc = await db.bank_statement_lines.find_one({"_id": _oid(id)})
    if not doc:
        raise HTTPException(404, "Statement line not found")

    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if not updates:
        return stringify(doc)

    # Check if line date falls in a locked period
    await _check_period_locked(db, doc.get("bank_account_id"), doc.get("date"), "edit or rematch")

    # Check if re-matching / reclassification touches a line with dependent records
    if "match_status" in updates or "matched_to" in updates:
        new_status = updates.get("match_status", doc.get("match_status"))
        new_matched = updates.get("matched_to", doc.get("matched_to"))
        if new_status != doc.get("match_status") or new_matched != doc.get("matched_to"):
            await _check_statement_line_dependents(db, doc, "re-match")

    updates["updated_at"] = _now_iso()
    updates["updated_by"] = u.get("email") or u.get("name", "")

    res = await db.bank_statement_lines.update_one({"_id": _oid(id)}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Statement line not found")

    updated = await db.bank_statement_lines.find_one({"_id": _oid(id)})
    return stringify(updated)


# ---------------------------------------------------------------------------
def _parse_dt(d_val: Any) -> Optional[datetime]:
    if not d_val:
        return None
    s = str(d_val).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None


def _compute_match_confidence(
    stmt_amount: float,
    doc_amount: float,
    stmt_date_str: Optional[str],
    doc_date_str: Optional[str],
    narration: Optional[str] = "",
    doc_name_or_ref: Optional[str] = "",
) -> Tuple[float, List[str]]:
    """
    Computes a match confidence score between 0.0 and 1.0 (0% - 100%) and returns explanation reasons.
    - Exact amount (diff <= 0.001) + Exact date (0 days diff): 0.95 to 1.0 (95%+ high confidence)
    - Exact amount + 1 day diff: ~0.85 (85% confidence, pending individual review if threshold is 95%)
    - Exact amount + 2-3 days diff: ~0.70 to 0.80
    - Non-zero amount diff (within tolerance) + exact date: ~0.85 - 0.90
    """
    amt_diff = abs(round(stmt_amount, 2) - round(doc_amount, 2))
    reasons = []

    # 1. Amount scoring (up to 50 points)
    if amt_diff <= 0.001:
        amt_score = 50.0
        reasons.append("Exact amount match")
    elif amt_diff <= 1.0:
        amt_score = 50.0 - (amt_diff * 15.0)
        reasons.append(f"Amount difference of ₹{amt_diff:.2f}")
    else:
        amt_score = max(20.0, 50.0 - (amt_diff * 5.0))
        reasons.append(f"Amount difference of ₹{amt_diff:.2f}")

    # 2. Date scoring (up to 40 points)
    stmt_dt = _parse_dt(stmt_date_str)
    doc_dt = _parse_dt(doc_date_str)
    day_diff = abs((stmt_dt - doc_dt).days) if (stmt_dt and doc_dt) else None

    if day_diff == 0:
        date_score = 40.0
        reasons.append("Exact date match")
    elif day_diff == 1:
        date_score = 30.0
        reasons.append("1 day date difference")
    elif day_diff == 2:
        date_score = 20.0
        reasons.append("2 days date difference")
    elif day_diff == 3:
        date_score = 15.0
        reasons.append("3 days date difference")
    else:
        date_score = 10.0
        if day_diff is not None:
            reasons.append(f"{day_diff} days date difference")

    # 3. Text / Narration heuristic (up to 10 points)
    text_score = 5.0
    if narration and doc_name_or_ref:
        narr_tokens = set(re.findall(r'[a-zA-Z0-9]+', str(narration).lower()))
        ref_tokens = set(re.findall(r'[a-zA-Z0-9]+', str(doc_name_or_ref).lower()))
        stop_words = {"pvt", "ltd", "payment", "bank", "neft", "rtgs", "upi", "imps", "dr", "cr", "to", "by", "for", "the", "a", "an", "and", "in", "of"}
        overlap = (narr_tokens & ref_tokens) - stop_words
        if overlap:
            text_score = 10.0
            reasons.append(f"Narration match ({', '.join(list(overlap)[:2])})")

    total_score = amt_score + date_score + text_score
    confidence = min(1.0, max(0.0, round(total_score / 100.0, 2)))

    if amt_diff <= 0.001 and day_diff == 0:
        confidence = max(0.95, confidence)

    return confidence, reasons


async def _handle_reconcile_account(
    id: str,
    date_window_days: int,
    amount_tolerance: float,
    dry_run: bool,
    min_confidence: float,
    request: Request,
) -> dict:
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    acc = await db.bank_accounts.find_one({"_id": _oid(id)})
    if not acc:
        raise HTTPException(404, "Bank account not found")

    account_id_str = str(acc["_id"])
    account_type = acc.get("account_type", "b2b_client")

    raw_conf = min_confidence.default if hasattr(min_confidence, "default") else min_confidence
    conf_float = float(raw_conf) if raw_conf is not None else 0.95
    norm_min_confidence = conf_float / 100.0 if conf_float > 1.0 else conf_float

    raw_window = date_window_days.default if hasattr(date_window_days, "default") else date_window_days
    date_window_days = int(raw_window) if raw_window is not None else 3

    raw_tol = amount_tolerance.default if hasattr(amount_tolerance, "default") else amount_tolerance
    amount_tolerance = float(raw_tol) if raw_tol is not None else 1.0

    raw_dry = dry_run.default if hasattr(dry_run, "default") else dry_run
    dry_run = bool(raw_dry) if raw_dry is not None else False

    unmatched_lines = await db.bank_statement_lines.find({
        "bank_account_id": account_id_str,
        "match_status": "unmatched",
    }).sort("date", 1).to_list(5000)

    if not unmatched_lines:
        return {
            "ok": True,
            "dry_run": dry_run,
            "bank_account_id": account_id_str,
            "bank_account_name": acc.get("name"),
            "total_unmatched_evaluated": 0,
            "auto_matched_count": 0,
            "pending_review_count": 0,
            "ambiguous_count": 0,
            "no_match_count": 0,
            "min_confidence": norm_min_confidence,
            "min_confidence_percent": int(round(norm_min_confidence * 100)),
            "matched_details": [],
            "pending_review_details": [],
        }

    claimed_ids = set()
    matched_results = []
    pending_review_results = []
    ambiguous_count = 0
    no_match_count = 0

    now = _now_iso()
    account_filter = {"$in": [None, "", account_id_str]}

    online_settlements_raw = await db.online_settlements.find({
        "bank_account_id": account_filter,
    }).to_list(10000)

    client_payments_raw = await db.payments.find({
        "type": {"$ne": "vendor_payment"},
        "vendor_id": {"$in": [None, ""]},
        "bank_account_id": account_filter,
    }).to_list(10000)

    vendor_payments_raw = await db.payments.find({
        "$or": [
            {"type": "vendor_payment"},
            {"vendor_id": {"$nin": [None, ""]}},
        ],
        "bank_account_id": account_filter,
    }).to_list(10000)

    expenses_raw = await db.expenses.find({
        "bank_account_id": account_filter,
    }).to_list(10000)

    for line in unmatched_lines:
        line_id = line["_id"]
        stmt_dt = _parse_dt(line.get("date"))
        credit = float(line.get("credit_amount") or 0.0)
        debit = float(line.get("debit_amount") or 0.0)

        # ── CREDIT Matching (Money In)
        if credit > 0:
            candidates = []
            if account_type == "online_channel":
                for s in online_settlements_raw:
                    s_id = str(s["_id"])
                    if s_id in claimed_ids:
                        continue
                    s_amount = float(s.get("net_payout") or s.get("settlement_amount") or 0.0)
                    s_dt = _parse_dt(s.get("settlement_date") or s.get("date") or s.get("created_at"))
                    if abs(s_amount - credit) <= amount_tolerance:
                        if stmt_dt and s_dt:
                            if abs((s_dt - stmt_dt).days) <= date_window_days:
                                candidates.append(("settlement", s, s_amount, s.get("settlement_date") or s.get("date") or s.get("created_at"), s.get("channel") or s.get("reference_number") or ""))
                        else:
                            candidates.append(("settlement", s, s_amount, s.get("settlement_date") or s.get("date") or s.get("created_at"), s.get("channel") or s.get("reference_number") or ""))
            else:
                for p in client_payments_raw:
                    p_id = str(p["_id"])
                    if p_id in claimed_ids:
                        continue
                    p_amount = float(p.get("amount") or 0.0)
                    p_dt = _parse_dt(p.get("payment_date") or p.get("created_at"))
                    if abs(p_amount - credit) <= amount_tolerance:
                        if stmt_dt and p_dt:
                            if abs((p_dt - stmt_dt).days) <= date_window_days:
                                candidates.append(("payment", p, p_amount, p.get("payment_date") or p.get("created_at"), p.get("customer_name") or p.get("reference") or ""))
                        else:
                            candidates.append(("payment", p, p_amount, p.get("payment_date") or p.get("created_at"), p.get("customer_name") or p.get("reference") or ""))

            if len(candidates) == 1:
                target_type, match_doc, cand_amount, cand_date, cand_ref = candidates[0]
                ref_id = str(match_doc["_id"])
                confidence, reasons = _compute_match_confidence(
                    credit,
                    cand_amount,
                    line.get("date"),
                    cand_date,
                    line.get("narration"),
                    cand_ref,
                )

                matched_item = {
                    "statement_line_id": str(line_id),
                    "statement_date": line.get("date"),
                    "narration": line.get("narration"),
                    "amount": credit,
                    "side": "credit",
                    "matched_type": target_type,
                    "ref_id": ref_id,
                    "candidate_title": cand_ref or target_type.replace("_", " ").title(),
                    "confidence_score": confidence,
                    "confidence_percent": int(round(confidence * 100)),
                    "confidence_reasons": reasons,
                }

                if confidence >= norm_min_confidence:
                    claimed_ids.add(ref_id)
                    matched_results.append(matched_item)
                    if not dry_run:
                        await db.bank_statement_lines.update_one(
                            {"_id": line_id},
                            {"$set": {
                                "match_status": "matched",
                                "matched_to": {"type": target_type, "ref_id": ref_id},
                                "updated_at": now,
                            }}
                        )
                        if target_type == "settlement":
                            await db.online_settlements.update_one(
                                {"_id": match_doc["_id"]},
                                {"$set": {"bank_account_id": account_id_str, "updated_at": now}}
                            )
                        elif target_type == "payment":
                            await db.payments.update_one(
                                {"_id": match_doc["_id"]},
                                {"$set": {"bank_account_id": account_id_str, "updated_at": now}}
                            )
                else:
                    # Low confidence match left for individual review
                    pending_review_results.append(matched_item)
            elif len(candidates) > 1:
                ambiguous_count += 1
            else:
                no_match_count += 1

        # ── DEBIT Matching (Money Out)
        elif debit > 0:
            candidates = []
            for vp in vendor_payments_raw:
                vp_id = str(vp["_id"])
                if vp_id in claimed_ids:
                    continue
                vp_amount = float(vp.get("amount") or 0.0)
                vp_dt = _parse_dt(vp.get("payment_date") or vp.get("created_at"))
                if abs(vp_amount - debit) <= amount_tolerance:
                    if stmt_dt and vp_dt:
                        if abs((vp_dt - stmt_dt).days) <= date_window_days:
                            candidates.append(("vendor_payment", vp, vp_amount, vp.get("payment_date") or vp.get("created_at"), vp.get("vendor_name") or vp.get("payee") or ""))
                    else:
                        candidates.append(("vendor_payment", vp, vp_amount, vp.get("payment_date") or vp.get("created_at"), vp.get("vendor_name") or vp.get("payee") or ""))

            for exp in expenses_raw:
                exp_id = str(exp["_id"])
                if exp_id in claimed_ids:
                    continue
                exp_amount = float(exp.get("amount") or 0.0)
                exp_dt = _parse_dt(exp.get("date") or exp.get("created_at"))
                if abs(exp_amount - debit) <= amount_tolerance:
                    if stmt_dt and exp_dt:
                        if abs((exp_dt - stmt_dt).days) <= date_window_days:
                            candidates.append(("expense", exp, exp_amount, exp.get("date") or exp.get("created_at"), exp.get("payee") or exp.get("category") or ""))
                    else:
                        candidates.append(("expense", exp, exp_amount, exp.get("date") or exp.get("created_at"), exp.get("payee") or exp.get("category") or ""))

            if len(candidates) == 1:
                target_type, match_doc, cand_amount, cand_date, cand_ref = candidates[0]
                ref_id = str(match_doc["_id"])
                confidence, reasons = _compute_match_confidence(
                    debit,
                    cand_amount,
                    line.get("date"),
                    cand_date,
                    line.get("narration"),
                    cand_ref,
                )

                matched_item = {
                    "statement_line_id": str(line_id),
                    "statement_date": line.get("date"),
                    "narration": line.get("narration"),
                    "amount": debit,
                    "side": "debit",
                    "matched_type": target_type,
                    "ref_id": ref_id,
                    "candidate_title": cand_ref or target_type.replace("_", " ").title(),
                    "confidence_score": confidence,
                    "confidence_percent": int(round(confidence * 100)),
                    "confidence_reasons": reasons,
                }

                if confidence >= norm_min_confidence:
                    claimed_ids.add(ref_id)
                    matched_results.append(matched_item)
                    if not dry_run:
                        await db.bank_statement_lines.update_one(
                            {"_id": line_id},
                            {"$set": {
                                "match_status": "matched",
                                "matched_to": {"type": target_type, "ref_id": ref_id},
                                "updated_at": now,
                            }}
                        )
                        if target_type == "vendor_payment":
                            await db.payments.update_one(
                                {"_id": match_doc["_id"]},
                                {"$set": {"bank_account_id": account_id_str, "updated_at": now}}
                            )
                        elif target_type == "expense":
                            await db.expenses.update_one(
                                {"_id": match_doc["_id"]},
                                {"$set": {"bank_account_id": account_id_str, "updated_at": now}}
                            )
                else:
                    # Low confidence match left for individual review
                    pending_review_results.append(matched_item)
            elif len(candidates) > 1:
                ambiguous_count += 1
            else:
                no_match_count += 1

    return {
        "ok": True,
        "dry_run": dry_run,
        "bank_account_id": account_id_str,
        "bank_account_name": acc.get("name"),
        "total_unmatched_evaluated": len(unmatched_lines),
        "auto_matched_count": len(matched_results),
        "pending_review_count": len(pending_review_results),
        "ambiguous_count": ambiguous_count,
        "no_match_count": no_match_count,
        "min_confidence": norm_min_confidence,
        "min_confidence_percent": int(round(norm_min_confidence * 100)),
        "date_window_days": date_window_days,
        "amount_tolerance": amount_tolerance,
        "matched_details": matched_results,
        "pending_review_details": pending_review_results,
    }


@banking_router.post("/banking/accounts/{id}/reconcile")
@banking_router.post("/bank-accounts/{id}/reconcile")
@banking_router.post("/banking/accounts/{id}/bulk-confirm-matches")
@banking_router.post("/bank-accounts/{id}/bulk-confirm-matches")
async def reconcile_bank_account(
    id: str,
    request: Request,
    date_window_days: int = Query(3, ge=0, le=30),
    amount_tolerance: float = Query(1.0, ge=0.0, le=100.0),
    min_confidence: float = Query(0.95, ge=0.0, le=100.0),
    dry_run: bool = Query(False),
):
    """
    Run auto-reconciliation on unmatched statement lines against ERP records.
    Only auto-confirms matches with confidence >= min_confidence (default 95%+ exact amount & date match).
    Low confidence matches remain unmatched for individual review.
    Transfer pairs are NEVER bulk-confirmed and always require explicit individual review.
    """
    return await _handle_reconcile_account(id, date_window_days, amount_tolerance, dry_run, min_confidence, request)


# ---------------------------------------------------------------------------
# Transfer Pairs & Reconciliation Summary (Stage 4 & Stage 5)
# ---------------------------------------------------------------------------

from models.banking import TransferConfirmIn, CashWithdrawalConfirmIn

# Common cash withdrawal narration indicators: ATM, CASH, SELF, CWDR, NFS, EAW, etc.
_CASH_WITHDRAWAL_REGEX = re.compile(
    r'(?:^|[\s\-_/.,:])(ATM|CASH|SELF|CWDR|NFS|EAW|CSH|SELF\s*CHQ|SELF\s*CHEQUE|CASH\s*WDL|CASH\s*WITHDRAWAL)(?:$|[\s\-_/.,:])',
    re.IGNORECASE,
)


def _is_cash_withdrawal_candidate(narration: Optional[str]) -> bool:
    if not narration:
        return False
    return bool(_CASH_WITHDRAWAL_REGEX.search(narration))


@banking_router.get("/banking/cash-withdrawals/suggested")
@banking_router.get("/bank-accounts/cash-withdrawals/suggested")
async def get_suggested_cash_withdrawals(
    request: Request,
    bank_account_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    date_window_days: int = Query(14, ge=0, le=60),
):
    """Scan unmatched debit statement lines for potential cash withdrawal patterns (ATM, CASH, SELF, etc.) and matching manual cash ledger entries."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {
        "match_status": "unmatched",
        "debit_amount": {"$gt": 0},
    }
    if bank_account_id and bank_account_id != "all":
        q["bank_account_id"] = str(bank_account_id)
    if from_date or to_date:
        dq = {}
        if from_date:
            dq["$gte"] = str(from_date)
        if to_date:
            dq["$lte"] = str(to_date)
        q["date"] = dq

    docs = await db.bank_statement_lines.find(q).sort("date", -1).to_list(5000)

    # Fetch accounts map
    accounts = await db.bank_accounts.find({}).to_list(1000)
    acc_map = {str(a["_id"]): a.get("name", "Unknown Account") for a in accounts}

    # Fetch unlinked manual cash ledger entries (source_statement_line_id is None / empty)
    unlinked_cl_q: Dict[str, Any] = {"source_statement_line_id": {"$in": [None, ""]}}
    if bank_account_id and bank_account_id != "all":
        unlinked_cl_q["bank_account_id"] = str(bank_account_id)
    unlinked_cls = []
    if hasattr(db, "cash_ledger") and db.cash_ledger is not None:
        try:
            unlinked_cls = await db.cash_ledger.find(unlinked_cl_q).sort("date", -1).to_list(1000)
        except Exception:
            unlinked_cls = []

    window_days = int(getattr(date_window_days, "default", date_window_days) or 14)

    candidates = []
    matched_cl_ids = set()
    for d in docs:
        narration = d.get("narration") or ""
        stmt_amt = float(d.get("debit_amount") or 0.0)
        stmt_dt = _parse_dt(d.get("date"))
        line_acc_id = str(d.get("bank_account_id"))

        # Look for matching unlinked manual cash entry:
        # Cash is usually recorded on or slightly before statement import (cl_date <= stmt_date + 1 day leeway)
        matching_cl = None
        for cl in unlinked_cls:
            cl_id = str(cl["_id"])
            if cl_id in matched_cl_ids:
                continue
            cl_acc_id = str(cl.get("bank_account_id"))
            if cl_acc_id == line_acc_id or not line_acc_id:
                cl_amt = float(cl.get("amount") or 0.0)
                cl_dt = _parse_dt(cl.get("date"))
                if abs(cl_amt - stmt_amt) <= 1.0:
                    if stmt_dt and cl_dt:
                        diff_days = (stmt_dt - cl_dt).days
                        # Manual entry was created on or before statement line (within date_window_days)
                        if -1 <= diff_days <= window_days:
                            matching_cl = cl
                            break
                    else:
                        matching_cl = cl
                        break

        is_pattern_match = _is_cash_withdrawal_candidate(narration)

        if matching_cl or is_pattern_match:
            c_doc = stringify(d)
            c_doc["bank_account_name"] = acc_map.get(line_acc_id, "Unknown Account")
            c_doc["amount"] = stmt_amt
            if matching_cl:
                matched_cl_ids.add(str(matching_cl["_id"]))
                c_doc["existing_cash_ledger_id"] = str(matching_cl["_id"])
                c_doc["existing_cash_ledger_date"] = matching_cl.get("date")
                c_doc["existing_cash_ledger_amount"] = float(matching_cl.get("amount") or 0.0)
                c_doc["existing_cash_ledger_remaining"] = float(matching_cl.get("remaining_balance") or 0.0)
                c_doc["existing_cash_ledger_notes"] = matching_cl.get("notes") or ""
                c_doc["is_existing_manual_entry"] = True
                c_doc["suggestion_reason"] = f"Matches existing manual cash withdrawal (₹{matching_cl.get('amount'):.2f} on {matching_cl.get('date')})"
            else:
                c_doc["is_existing_manual_entry"] = False
                c_doc["suggestion_reason"] = "Narration matches cash withdrawal pattern (ATM/CASH/SELF)"
            candidates.append(c_doc)

    return {
        "ok": True,
        "total_suggestions": len(candidates),
        "candidates": candidates,
    }


@banking_router.post("/banking/cash-withdrawals/confirm")
@banking_router.post("/bank-accounts/cash-withdrawals/confirm")
async def confirm_cash_withdrawal(payload: CashWithdrawalConfirmIn, request: Request):
    """Explicitly confirm a statement line as cash withdrawal, link to existing cash_ledger or create new, and mark line matched."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    line = await db.bank_statement_lines.find_one({"_id": _oid(payload.statement_line_id)})
    if not line:
        raise HTTPException(404, f"Statement line '{payload.statement_line_id}' not found")

    amount = float(line.get("debit_amount") or 0.0)
    if amount <= 0:
        raise HTTPException(400, "Cash withdrawal line must have a debit amount > 0")

    # Guard against orphaning dependent records if reclassifying
    if line.get("match_status") in ["matched", "transfer"]:
        await _check_statement_line_dependents(db, line, "reclassify as cash withdrawal")

    # Guard against modifying locked period
    await _check_period_locked(db, line.get("bank_account_id"), line.get("date"), "confirm cash withdrawal")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")
    line_acc_id = str(line.get("bank_account_id"))

    cash_ledger_id = None
    cash_ledger_doc = None
    is_linked_existing = False

    # 1. If explicit existing_cash_ledger_id provided, link to it
    if payload.existing_cash_ledger_id:
        cash_ledger_doc = await db.cash_ledger.find_one({"_id": _oid(payload.existing_cash_ledger_id)})
        if not cash_ledger_doc:
            raise HTTPException(404, f"Existing cash ledger entry '{payload.existing_cash_ledger_id}' not found")
        if cash_ledger_doc.get("source_statement_line_id") and str(cash_ledger_doc["source_statement_line_id"]) != str(line["_id"]):
            raise HTTPException(400, f"Cash ledger entry is already linked to statement line '{cash_ledger_doc['source_statement_line_id']}'")
        cash_ledger_id = str(cash_ledger_doc["_id"])
        is_linked_existing = True
        await db.cash_ledger.update_one(
            {"_id": cash_ledger_doc["_id"]},
            {"$set": {
                "source_statement_line_id": str(line["_id"]),
                "bank_account_id": line_acc_id,
                "updated_at": now,
                "updated_by": user_email,
            }}
        )
    else:
        # Check if there is an exact matching unlinked manual cash ledger entry for this bank account, amount, and date
        unlinked_match = await db.cash_ledger.find_one({
            "bank_account_id": line_acc_id,
            "amount": amount,
            "date": line.get("date"),
            "source_statement_line_id": {"$in": [None, ""]},
        })
        if unlinked_match:
            cash_ledger_doc = unlinked_match
            cash_ledger_id = str(unlinked_match["_id"])
            is_linked_existing = True
            await db.cash_ledger.update_one(
                {"_id": unlinked_match["_id"]},
                {"$set": {
                    "source_statement_line_id": str(line["_id"]),
                    "updated_at": now,
                    "updated_by": user_email,
                }}
            )
        else:
            # Create document in cash_ledger
            cash_ledger_doc = {
                "bank_account_id": line_acc_id,
                "source_statement_line_id": str(line["_id"]),
                "date": line.get("date"),
                "amount": amount,
                "remaining_balance": amount,
                "notes": payload.notes or line.get("narration") or "Cash withdrawal from bank",
                "created_at": now,
                "created_by": user_email,
            }
            res = await db.cash_ledger.insert_one(cash_ledger_doc)
            cash_ledger_id = str(res.inserted_id)

    # Update statement line to matched with matched_to: cash_withdrawal
    await db.bank_statement_lines.update_one(
        {"_id": line["_id"]},
        {"$set": {
            "match_status": "matched",
            "matched_to": {
                "type": "cash_withdrawal",
                "ref_id": cash_ledger_id,
            },
            "updated_at": now,
            "updated_by": user_email,
        }}
    )

    remaining_balance = float(cash_ledger_doc.get("remaining_balance") if cash_ledger_doc and "remaining_balance" in cash_ledger_doc else amount)

    return {
        "ok": True,
        "message": "Cash withdrawal successfully linked to Cash Ledger." if is_linked_existing else "Cash withdrawal successfully confirmed and added to Cash Ledger.",
        "statement_line_id": str(line["_id"]),
        "cash_ledger_id": cash_ledger_id,
        "amount": amount,
        "remaining_balance": remaining_balance,
        "is_linked_to_existing": is_linked_existing,
    }


@banking_router.post("/banking/cash-ledger", status_code=201)
@banking_router.post("/bank-accounts/cash-ledger", status_code=201)
async def create_cash_ledger_entry(payload: CashLedgerCreateIn, request: Request):
    """Directly create a cash_ledger entry (bank_account, amount, date) independent of bank statement import."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    acc = await db.bank_accounts.find_one({"_id": _oid(payload.bank_account_id)})
    if not acc:
        raise HTTPException(404, f"Bank account '{payload.bank_account_id}' not found")

    amount = float(payload.amount or 0.0)
    if amount <= 0:
        raise HTTPException(400, "Cash withdrawal amount must be > 0")

    await _check_period_locked(db, str(payload.bank_account_id), payload.date, "create manual cash withdrawal")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")

    cash_ledger_doc = {
        "bank_account_id": str(payload.bank_account_id),
        "source_statement_line_id": None,
        "date": payload.date,
        "amount": amount,
        "remaining_balance": amount,
        "notes": payload.notes or f"Manual cash withdrawal from {acc.get('name', 'bank')}",
        "created_at": now,
        "created_by": user_email,
    }
    res = await db.cash_ledger.insert_one(cash_ledger_doc)
    cash_ledger_doc["_id"] = res.inserted_id

    return {
        "ok": True,
        "message": "Cash withdrawal entry recorded successfully.",
        "cash_ledger": stringify(cash_ledger_doc),
        "id": str(res.inserted_id),
        "amount": amount,
        "remaining_balance": amount,
    }


@banking_router.delete("/banking/cash-ledger/{cash_ledger_id}")
@banking_router.delete("/bank-accounts/cash-ledger/{cash_ledger_id}")
async def delete_cash_ledger_entry(cash_ledger_id: str, request: Request):
    """Delete an unlinked manual cash_ledger entry (if it has no linked statement line and no wage payments/expenses)."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    doc = await db.cash_ledger.find_one({"_id": _oid(cash_ledger_id)})
    if not doc:
        raise HTTPException(404, f"Cash ledger entry '{cash_ledger_id}' not found")

    if doc.get("source_statement_line_id"):
        raise HTTPException(
            400,
            f"Cannot delete cash ledger entry '{cash_ledger_id}': it is linked to bank statement line '{doc['source_statement_line_id']}'. "
            f"Please unmatch or reclassify the statement line first."
        )

    wage_count = 0
    if hasattr(db, "wage_payments") and db.wage_payments is not None:
        try:
            wage_count = await db.wage_payments.count_documents({"cash_ledger_id": str(cash_ledger_id)})
        except Exception:
            wage_count = 0
    expense_count = 0
    if hasattr(db, "expenses") and db.expenses is not None:
        try:
            expense_count = await db.expenses.count_documents({"cash_ledger_id": str(cash_ledger_id)})
        except Exception:
            expense_count = 0
    advance_count = 0
    if hasattr(db, "advances") and db.advances is not None:
        try:
            advance_count = await db.advances.count_documents({"cash_ledger_id": str(cash_ledger_id)})
        except Exception:
            advance_count = 0

    if wage_count > 0 or expense_count > 0 or advance_count > 0:
        details = []
        if wage_count > 0:
            details.append(f"{wage_count} wage payment(s)")
        if expense_count > 0:
            details.append(f"{expense_count} cash expense(s)")
        if advance_count > 0:
            details.append(f"{advance_count} karigar advance/payment(s)")
        raise HTTPException(
            400,
            f"Cannot delete cash ledger entry '{cash_ledger_id}': it has active dependent records ({' and '.join(details)}). "
            f"Please delete or reassign the wage payments/expenses before deleting this entry."
        )

    await _check_period_locked(db, doc.get("bank_account_id"), doc.get("date"), "delete cash ledger entry")
    await db.cash_ledger.delete_one({"_id": doc["_id"]})

    return {"ok": True, "message": f"Cash ledger entry '{cash_ledger_id}' deleted successfully."}


@banking_router.get("/banking/cash-ledger")
@banking_router.get("/bank-accounts/cash-ledger")
async def list_cash_ledger(
    request: Request,
    bank_account_id: Optional[str] = None,
    limit: int = 200,
):
    """List cash ledger entries and total cash-in-hand remaining balance with linked wage disbursements."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {}
    if bank_account_id and bank_account_id != "all":
        q["bank_account_id"] = str(bank_account_id)

    docs = await db.cash_ledger.find(q).sort("date", -1).to_list(limit)
    total_withdrawn = sum(float(d.get("amount") or 0.0) for d in docs)
    total_remaining_balance = sum(float(d.get("remaining_balance") or 0.0) for d in docs)

    # Populate linked wage payments and cash expenses per cash_ledger entry
    cash_ids = [str(d["_id"]) for d in docs]
    wage_payments_by_cl = {}
    expenses_by_cl = {}
    if cash_ids:
        if hasattr(db, "wage_payments") and db.wage_payments is not None:
            try:
                wp_cursor = db.wage_payments.find({"cash_ledger_id": {"$in": cash_ids}})
                if hasattr(wp_cursor, "to_list"):
                    res = wp_cursor.to_list(1000)
                    if hasattr(res, "__await__"):
                        wp_docs = await res
                    elif isinstance(res, list):
                        wp_docs = res
                    else:
                        wp_docs = []
                    for wp in wp_docs:
                        clid = str(wp.get("cash_ledger_id"))
                        wage_payments_by_cl.setdefault(clid, []).append(wp)
            except Exception:
                wage_payments_by_cl = {}

        if hasattr(db, "expenses") and db.expenses is not None:
            try:
                exp_cursor = db.expenses.find({"cash_ledger_id": {"$in": cash_ids}})
                if hasattr(exp_cursor, "to_list"):
                    res = exp_cursor.to_list(1000)
                    if hasattr(res, "__await__"):
                        exp_docs = await res
                    elif isinstance(res, list):
                        exp_docs = res
                    else:
                        exp_docs = []
                    for e in exp_docs:
                        clid = str(e.get("cash_ledger_id"))
                        expenses_by_cl.setdefault(clid, []).append(e)
            except Exception:
                expenses_by_cl = {}

        advances_by_cl = {}
        if hasattr(db, "advances") and db.advances is not None:
            try:
                adv_cursor = db.advances.find({"cash_ledger_id": {"$in": cash_ids}})
                if hasattr(adv_cursor, "to_list"):
                    res = adv_cursor.to_list(1000)
                    if hasattr(res, "__await__"):
                        adv_docs = await res
                    elif isinstance(res, list):
                        adv_docs = res
                    else:
                        adv_docs = []
                    for a in adv_docs:
                        clid = str(a.get("cash_ledger_id"))
                        advances_by_cl.setdefault(clid, []).append(a)
            except Exception:
                advances_by_cl = {}

    items = []
    for d in docs:
        sd = stringify(d)
        cid = str(d["_id"])
        wps = wage_payments_by_cl.get(cid, [])
        exps = expenses_by_cl.get(cid, [])
        advs = advances_by_cl.get(cid, []) if cash_ids else []
        allocated_wages = sum(float(wp.get("amount") or 0.0) for wp in wps)
        allocated_expenses = sum(float(e.get("amount") or 0.0) for e in exps)
        allocated_advances = sum(float(a.get("amount") or 0.0) for a in advs)
        allocated = round(allocated_wages + allocated_expenses + allocated_advances, 2)
        withdrawn = float(d.get("amount") or 0.0)
        rem = float(d.get("remaining_balance") if d.get("remaining_balance") is not None else (withdrawn - allocated))

        disbursements = []
        for wp in wps:
            disbursements.append({
                "id": str(wp.get("_id") or wp.get("id")),
                "type": "wage_payment",
                "type_label": "Karigar Wage",
                "title": wp.get("worker_name") or f"Worker #{str(wp.get('worker_id'))[-6:]}",
                "amount": float(wp.get("amount") or 0.0),
                "date": wp.get("date"),
                "period_from": wp.get("period_from"),
                "period_to": wp.get("period_to"),
                "notes": wp.get("notes") or "",
                "override_reason": wp.get("override_reason"),
            })
        for a in advs:
            t_label = "Karigar Advance" if a.get("txn_type") == "advance" else "Karigar Payment"
            disbursements.append({
                "id": str(a.get("_id") or a.get("id")),
                "type": a.get("txn_type") or "advance",
                "type_label": t_label,
                "title": a.get("worker_name") or f"Karigar #{str(a.get('worker_id'))[-6:]}",
                "amount": float(a.get("amount") or 0.0),
                "date": a.get("date"),
                "notes": a.get("notes") or "",
            })
        for e in exps:
            disbursements.append({
                "id": str(e.get("_id") or e.get("id")),
                "type": "expense",
                "type_label": "Cash Expense",
                "title": e.get("payee") or "Payee",
                "category": e.get("category") or "Expense",
                "amount": float(e.get("amount") or 0.0),
                "date": e.get("date"),
                "notes": e.get("notes") or "",
            })
        disbursements.sort(key=lambda x: str(x.get("date") or ""), reverse=True)

        sd["allocated_amount"] = allocated
        sd["remaining_balance"] = round(rem, 2)
        sd["statement_linked"] = bool(d.get("source_statement_line_id"))
        sd["source_statement_line_id"] = d.get("source_statement_line_id")
        sd["is_manual_entry"] = not bool(d.get("source_statement_line_id"))
        sd["wage_payment_count"] = len(wps)
        sd["wage_payments"] = [stringify(wp) for wp in wps]
        sd["expense_count"] = len(exps)
        sd["expenses"] = [stringify(e) for e in exps]
        sd["disbursements"] = disbursements
        items.append(sd)

    return {
        "ok": True,
        "total_count": len(docs),
        "total_withdrawn": round(total_withdrawn, 2),
        "total_remaining_balance": round(total_remaining_balance, 2),
        "items": items,
    }


@banking_router.get("/banking/cash-ledger/{cash_ledger_id}")
@banking_router.get("/bank-accounts/cash-ledger/{cash_ledger_id}")
async def get_cash_ledger_detail(cash_ledger_id: str, request: Request):
    """Get cash ledger entry details with all linked wage payments and cash expenses funded by this withdrawal."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales", "production")(u)
    db = _get_db(request)

    doc = await db.cash_ledger.find_one({"_id": _oid(cash_ledger_id)})
    if not doc:
        raise HTTPException(404, f"Cash ledger entry '{cash_ledger_id}' not found")

    wage_payments = []
    if hasattr(db, "wage_payments") and db.wage_payments is not None:
        try:
            wp_cursor = db.wage_payments.find({"cash_ledger_id": str(cash_ledger_id)})
            if hasattr(wp_cursor, "to_list"):
                res = wp_cursor.to_list(500)
                if hasattr(res, "__await__"):
                    wage_payments = await res
                elif isinstance(res, list):
                    wage_payments = res
        except Exception:
            wage_payments = []

    expenses = []
    if hasattr(db, "expenses") and db.expenses is not None:
        try:
            exp_cursor = db.expenses.find({"cash_ledger_id": str(cash_ledger_id)})
            if hasattr(exp_cursor, "to_list"):
                res = exp_cursor.to_list(500)
                if hasattr(res, "__await__"):
                    expenses = await res
                elif isinstance(res, list):
                    expenses = res
        except Exception:
            expenses = []

    wage_payments_list = [stringify(wp) for wp in wage_payments]
    expenses_list = [stringify(e) for e in expenses]

    allocated_wages = sum(float(wp.get("amount") or 0.0) for wp in wage_payments)
    allocated_expenses = sum(float(e.get("amount") or 0.0) for e in expenses)
    allocated_amount = round(allocated_wages + allocated_expenses, 2)
    withdrawal_amount = float(doc.get("amount") or 0.0)
    remaining_balance = float(doc.get("remaining_balance") if doc.get("remaining_balance") is not None else (withdrawal_amount - allocated_amount))

    disbursements = []
    for wp in wage_payments:
        disbursements.append({
            "id": str(wp.get("_id") or wp.get("id")),
            "type": "wage_payment",
            "type_label": "Karigar Wage",
            "title": wp.get("worker_name") or f"Worker #{str(wp.get('worker_id'))[-6:]}",
            "amount": float(wp.get("amount") or 0.0),
            "date": wp.get("date"),
            "period_from": wp.get("period_from"),
            "period_to": wp.get("period_to"),
            "notes": wp.get("notes") or "",
            "override_reason": wp.get("override_reason"),
        })
    for e in expenses:
        disbursements.append({
            "id": str(e.get("_id") or e.get("id")),
            "type": "expense",
            "type_label": "Cash Expense",
            "title": e.get("payee") or "Payee",
            "category": e.get("category") or "Expense",
            "amount": float(e.get("amount") or 0.0),
            "date": e.get("date"),
            "notes": e.get("notes") or "",
        })
    disbursements.sort(key=lambda x: str(x.get("date") or ""), reverse=True)

    cash_ledger_dict = stringify(doc)
    cash_ledger_dict["statement_linked"] = bool(doc.get("source_statement_line_id"))
    cash_ledger_dict["source_statement_line_id"] = doc.get("source_statement_line_id")
    cash_ledger_dict["is_manual_entry"] = not bool(doc.get("source_statement_line_id"))

    return {
        "ok": True,
        "cash_ledger": cash_ledger_dict,
        "withdrawal_amount": withdrawal_amount,
        "allocated_amount": allocated_amount,
        "remaining_balance": round(remaining_balance, 2),
        "statement_linked": bool(doc.get("source_statement_line_id")),
        "source_statement_line_id": doc.get("source_statement_line_id"),
        "is_manual_entry": not bool(doc.get("source_statement_line_id")),
        "wage_payments": wage_payments_list,
        "wage_payment_count": len(wage_payments_list),
        "expenses": expenses_list,
        "expense_count": len(expenses_list),
        "disbursements": disbursements,
        "disbursement_count": len(disbursements),
    }


@banking_router.get("/banking/transfers/suggested")
@banking_router.get("/bank-accounts/transfers/suggested")
async def get_suggested_transfers(
    request: Request,
    date_window_days: int = Query(3, ge=0, le=30),
    amount_tolerance: float = Query(1.0, ge=0.0, le=100.0),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    """Scan across bank accounts for potential transfer pairs (debit in account A, credit in account B)."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {"match_status": "unmatched"}
    if from_date or to_date:
        dq = {}
        if from_date:
            dq["$gte"] = str(from_date)
        if to_date:
            dq["$lte"] = str(to_date)
        q["date"] = dq

    all_unmatched = await db.bank_statement_lines.find(q).sort("date", 1).to_list(10000)

    # Fetch accounts map
    accounts = await db.bank_accounts.find({}).to_list(1000)
    acc_map = {str(a["_id"]): a.get("name", "Unknown Account") for a in accounts}

    debits = [l for l in all_unmatched if float(l.get("debit_amount") or 0.0) > 0]
    credits = [l for l in all_unmatched if float(l.get("credit_amount") or 0.0) > 0]

    suggested_pairs = []
    used_credit_ids = set()

    for d in debits:
        d_id = str(d["_id"])
        d_acc = str(d.get("bank_account_id"))
        d_amt = float(d.get("debit_amount") or 0.0)
        d_dt = _parse_dt(d.get("date"))

        for c in credits:
            c_id = str(c["_id"])
            c_acc = str(c.get("bank_account_id"))

            # Transfers must be between DIFFERENT accounts
            if d_acc == c_acc:
                continue

            if c_id in used_credit_ids:
                continue

            c_amt = float(c.get("credit_amount") or 0.0)
            if abs(d_amt - c_amt) <= amount_tolerance:
                c_dt = _parse_dt(c.get("date"))
                day_diff = abs((c_dt - d_dt).days) if (c_dt and d_dt) else 0

                if day_diff <= date_window_days:
                    pair_doc = {
                        "pair_id": f"{d_id}:{c_id}",
                        "from_line": {
                            "id": d_id,
                            "bank_account_id": d_acc,
                            "bank_account_name": acc_map.get(d_acc, d_acc),
                            "date": d.get("date"),
                            "narration": d.get("narration"),
                            "reference_no": d.get("reference_no", ""),
                            "amount": d_amt,
                            "side": "debit",
                        },
                        "to_line": {
                            "id": c_id,
                            "bank_account_id": c_acc,
                            "bank_account_name": acc_map.get(c_acc, c_acc),
                            "date": c.get("date"),
                            "narration": c.get("narration"),
                            "reference_no": c.get("reference_no", ""),
                            "amount": c_amt,
                            "side": "credit",
                        },
                        "amount_diff": round(abs(d_amt - c_amt), 2),
                        "day_diff": day_diff,
                    }
                    suggested_pairs.append(pair_doc)

    return {
        "ok": True,
        "total_suggestions": len(suggested_pairs),
        "date_window_days": date_window_days,
        "amount_tolerance": amount_tolerance,
        "pairs": suggested_pairs,
    }


@banking_router.post("/banking/transfers/confirm")
@banking_router.post("/bank-accounts/transfers/confirm")
async def confirm_transfer_pair(payload: TransferConfirmIn, request: Request):
    """Explicitly confirm a transfer pair between two bank accounts."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    from_line = await db.bank_statement_lines.find_one({"_id": _oid(payload.from_line_id)})
    if not from_line:
        raise HTTPException(404, f"Sending statement line '{payload.from_line_id}' not found")

    to_line = await db.bank_statement_lines.find_one({"_id": _oid(payload.to_line_id)})
    if not to_line:
        raise HTTPException(404, f"Receiving statement line '{payload.to_line_id}' not found")

    if str(from_line.get("bank_account_id")) == str(to_line.get("bank_account_id")):
        raise HTTPException(400, "Transfer must be between two different bank accounts")

    if float(from_line.get("debit_amount") or 0.0) <= 0:
        raise HTTPException(400, "From-line must be a debit (withdrawal)")
    if float(to_line.get("credit_amount") or 0.0) <= 0:
        raise HTTPException(400, "To-line must be a credit (deposit)")

    # Guard against orphaning dependent records (e.g. wage payments / expenses drawn from cash_withdrawal)
    await _check_statement_line_dependents(db, from_line, "reclassify as transfer")
    await _check_statement_line_dependents(db, to_line, "reclassify as transfer")

    # Guard against modifying locked reconciliation period
    await _check_period_locked(db, from_line.get("bank_account_id"), from_line.get("date"), "transfer")
    await _check_period_locked(db, to_line.get("bank_account_id"), to_line.get("date"), "transfer")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")

    await db.bank_statement_lines.update_one(
        {"_id": from_line["_id"]},
        {"$set": {
            "match_status": "transfer",
            "matched_to": {"type": "transfer", "ref_id": str(to_line["_id"])},
            "transfer_notes": payload.notes or "",
            "confirmed_by": user_email,
            "updated_at": now,
        }}
    )

    await db.bank_statement_lines.update_one(
        {"_id": to_line["_id"]},
        {"$set": {
            "match_status": "transfer",
            "matched_to": {"type": "transfer", "ref_id": str(from_line["_id"])},
            "transfer_notes": payload.notes or "",
            "confirmed_by": user_email,
            "updated_at": now,
        }}
    )

    return {
        "ok": True,
        "message": "Transfer pair successfully confirmed and marked.",
        "from_line_id": str(from_line["_id"]),
        "to_line_id": str(to_line["_id"]),
    }


@banking_router.get("/banking/reconciliation/summary")
@banking_router.get("/bank-accounts/reconciliation/summary")
async def get_reconciliation_summary(
    request: Request,
    bank_account_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    """
    Get financial reconciliation summary across bank statements.
    EXCLUDES transfer pairs from income/expense totals as internal liquidity moves.
    """
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {}
    if bank_account_id:
        q["bank_account_id"] = str(bank_account_id)
    if from_date or to_date:
        dq = {}
        if from_date:
            dq["$gte"] = str(from_date)
        if to_date:
            dq["$lte"] = str(to_date)
        q["date"] = dq

    docs = await db.bank_statement_lines.find(q).to_list(20000)
    accounts = await db.bank_accounts.find({}).to_list(1000)
    acc_map = {str(a["_id"]): a for a in accounts}

    total_income = 0.0
    matched_income = 0.0
    unmatched_income = 0.0

    total_expenses = 0.0
    matched_expenses = 0.0
    unmatched_expenses = 0.0

    total_transfers = 0.0
    transfer_count = 0
    ignored_count = 0

    per_account_stats = {}
    for acc_id_key, acc_doc in acc_map.items():
        per_account_stats[acc_id_key] = {
            "bank_account_id": acc_id_key,
            "name": acc_doc.get("name"),
            "bank_name": acc_doc.get("bank_name"),
            "account_type": acc_doc.get("account_type"),
            "opening_balance": acc_doc.get("opening_balance", 0.0),
            "opening_balance_date": acc_doc.get("opening_balance_date"),
            "last_balance_correction": acc_doc.get("last_balance_correction"),
            "balance_corrections_count": len(acc_doc.get("balance_corrections") or []),
            "income": 0.0,
            "matched_income": 0.0,
            "unmatched_income": 0.0,
            "expenses": 0.0,
            "matched_expenses": 0.0,
            "unmatched_expenses": 0.0,
            "transfers_in": 0.0,
            "transfers_out": 0.0,
            "total_lines": 0,
        }

    for d in docs:
        acc_id = str(d.get("bank_account_id"))
        status = d.get("match_status", "unmatched")
        credit = float(d.get("credit_amount") or 0.0)
        debit = float(d.get("debit_amount") or 0.0)

        acc_stat = per_account_stats.get(acc_id)
        if acc_stat:
            acc_stat["total_lines"] += 1

        if status == "transfer":
            transfer_count += 1
            if debit > 0:
                total_transfers += debit
                if acc_stat:
                    acc_stat["transfers_out"] += debit
            elif credit > 0:
                if acc_stat:
                    acc_stat["transfers_in"] += credit
            # Strictly EXCLUDED from income/expenses
            continue

        if status == "ignored":
            ignored_count += 1
            continue

        # Revenue / Income (Credits)
        if credit > 0:
            total_income += credit
            if acc_stat:
                acc_stat["income"] += credit
            if status == "matched":
                matched_income += credit
                if acc_stat:
                    acc_stat["matched_income"] += credit
            else:
                unmatched_income += credit
                if acc_stat:
                    acc_stat["unmatched_income"] += credit

        # Expenses / Outflows (Debits)
        if debit > 0:
            total_expenses += debit
            if acc_stat:
                acc_stat["expenses"] += debit
            if status == "matched":
                matched_expenses += debit
                if acc_stat:
                    acc_stat["matched_expenses"] += debit
            else:
                unmatched_expenses += debit
    for acc_id_key, acc_stat in per_account_stats.items():
        acc_stat["total_reconciled_credits"] = round(acc_stat["matched_income"], 2)
        acc_stat["total_reconciled_debits"] = round(acc_stat["matched_expenses"], 2)
        acc_stat["net_statement_flow"] = round(acc_stat["income"] - acc_stat["expenses"], 2)

    # Cash in hand remaining balance from cash_ledger
    cash_docs = []
    if hasattr(db, "cash_ledger") and db.cash_ledger is not None:
        try:
            cash_q: Dict[str, Any] = {}
            if bank_account_id and bank_account_id != "all":
                cash_q["bank_account_id"] = str(bank_account_id)
            cursor = db.cash_ledger.find(cash_q)
            if hasattr(cursor, "to_list"):
                res = cursor.to_list(5000)
                if hasattr(res, "__await__"):
                    cash_docs = await res
                elif isinstance(res, list):
                    cash_docs = res
        except Exception:
            cash_docs = []

    total_cash_in_hand = round(sum(float(cd.get("remaining_balance") or 0.0) for cd in cash_docs), 2)
    total_cash_withdrawn = round(sum(float(cd.get("amount") or 0.0) for cd in cash_docs), 2)

    net_operating_cashflow = round(total_income - total_expenses, 2)

    return {
        "ok": True,
        "filter": {
            "bank_account_id": bank_account_id,
            "from_date": from_date,
            "to_date": to_date,
        },
        "summary": {
            "total_income": round(total_income, 2),
            "matched_income": round(matched_income, 2),
            "unmatched_income": round(unmatched_income, 2),
            "total_expenses": round(total_expenses, 2),
            "matched_expenses": round(matched_expenses, 2),
            "unmatched_expenses": round(unmatched_expenses, 2),
            "net_operating_cashflow": net_operating_cashflow,
            "total_transfers_volume": round(total_transfers, 2),
            "transfer_lines_count": transfer_count,
            "ignored_lines_count": ignored_count,
            "total_statement_lines": len(docs),
            "total_cash_in_hand": total_cash_in_hand,
            "total_cash_withdrawn": total_cash_withdrawn,
        },
        "accounts": list(per_account_stats.values()),
    }


@banking_router.get("/banking/unmatched-erp-candidates")
@banking_router.get("/bank-accounts/unmatched-erp-candidates")
async def get_unmatched_erp_candidates(
    request: Request,
    bank_account_id: Optional[str] = None,
    side: Optional[str] = Query("all"),
    search: Optional[str] = None,
    limit: int = 200,
):
    """
    List ERP-side transactions (settlements, client payments, vendor payments, expenses)
    that are not yet reconciled / linked to a bank account.
    """
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    account_type = "b2b_client"
    account_filter = None
    if bank_account_id and bank_account_id != "all":
        acc = await db.bank_accounts.find_one({"_id": _oid(bank_account_id)})
        if acc:
            account_type = acc.get("account_type", "b2b_client")
            account_filter = {"$in": [None, "", str(acc["_id"])]}

    candidates = []

    # 1. Credits (Settlements / Client Payments)
    if side in ["credit", "all"]:
        if account_type == "online_channel" or not bank_account_id or bank_account_id == "all":
            s_q = {}
            if account_filter:
                s_q["bank_account_id"] = account_filter
            if search:
                s_rx = {"$regex": re.escape(str(search)), "$options": "i"}
                s_q["$or"] = [{"seller_order_id": s_rx}, {"order_release_id": s_rx}, {"platform": s_rx}]
            settlements = await db.online_settlements.find(s_q).sort("settlement_date", -1).limit(limit).to_list(limit)
            for s in settlements:
                candidates.append({
                    "type": "settlement",
                    "id": str(s["_id"]),
                    "date": s.get("settlement_date") or s.get("created_at", "")[:10],
                    "amount": float(s.get("net_payout") or s.get("settlement_amount") or 0.0),
                    "description": f"{s.get('platform', 'Online').upper()} Settlement - Order {s.get('seller_order_id') or s.get('order_release_id') or str(s.get('_id', ''))}",
                    "side": "credit",
                    "party": s.get("platform", "Online"),
                    "reference": s.get("payment_id") or s.get("neft_ref") or "",
                })

        if account_type == "b2b_client" or not bank_account_id or bank_account_id == "all":
            p_q = {
                "type": {"$ne": "vendor_payment"},
                "vendor_id": {"$in": [None, ""]},
            }
            if account_filter:
                p_q["bank_account_id"] = account_filter
            if search:
                p_rx = {"$regex": re.escape(str(search)), "$options": "i"}
                p_q["$or"] = [{"client_name": p_rx}, {"reference": p_rx}, {"payment_no": p_rx}]
            payments = await db.payments.find(p_q).sort("payment_date", -1).limit(limit).to_list(limit)
            for p in payments:
                candidates.append({
                    "type": "payment",
                    "id": str(p["_id"]),
                    "date": p.get("payment_date", "")[:10],
                    "amount": float(p.get("amount") or 0.0),
                    "description": f"Client Payment {p.get('payment_no', '')} - {p.get('client_name', 'Client')}",
                    "side": "credit",
                    "party": p.get("client_name", "Client"),
                    "reference": p.get("reference") or p.get("payment_no") or "",
                })

    # 2. Debits (Vendor Payments / Expenses)
    if side in ["debit", "all"]:
        vp_q = {
            "$or": [{"type": "vendor_payment"}, {"vendor_id": {"$nin": [None, ""]}}],
        }
        if account_filter:
            vp_q["bank_account_id"] = account_filter
        if search:
            vp_rx = {"$regex": re.escape(str(search)), "$options": "i"}
            vp_q["$or"] = [{"vendor_name": vp_rx}, {"reference": vp_rx}, {"payment_no": vp_rx}]
        vendor_payments = await db.payments.find(vp_q).sort("payment_date", -1).limit(limit).to_list(limit)
        for vp in vendor_payments:
            candidates.append({
                "type": "vendor_payment",
                "id": str(vp["_id"]),
                "date": vp.get("payment_date", "")[:10],
                "amount": float(vp.get("amount") or 0.0),
                "description": f"Vendor Payment {vp.get('payment_no', '')} - {vp.get('vendor_name', 'Vendor')}",
                "side": "debit",
                "party": vp.get("vendor_name", "Vendor"),
                "reference": vp.get("reference") or vp.get("payment_no") or "",
            })

        e_q = {
            "statement_line_id": {"$in": [None, ""]},
            "paid_via": {"$ne": "cash"},
        }
        if account_filter:
            e_q["bank_account_id"] = account_filter
        if search:
            e_rx = {"$regex": re.escape(str(search)), "$options": "i"}
            e_q["$or"] = [{"payee": e_rx}, {"category": e_rx}, {"notes": e_rx}]
        expenses = await db.expenses.find(e_q).sort("date", -1).limit(limit).to_list(limit)
        for exp in expenses:
            candidates.append({
                "type": "expense",
                "id": str(exp["_id"]),
                "date": exp.get("date", "")[:10],
                "amount": float(exp.get("amount") or 0.0),
                "description": f"{exp.get('category', 'Expense')} - {exp.get('payee', 'Payee')}",
                "side": "debit",
                "party": exp.get("payee", ""),
                "reference": exp.get("category", ""),
            })

    return {
        "ok": True,
        "total": len(candidates),
        "candidates": candidates[:limit],
    }


# ---------------------------------------------------------------------------
# Period Locking & Finalization (Stage 5)
# ---------------------------------------------------------------------------

@banking_router.post("/banking/periods/lock")
@banking_router.post("/bank-accounts/periods/lock")
async def lock_reconciliation_period(payload: PeriodLockIn, request: Request):
    """Finalize and lock a reconciliation period against any edits or unmatching."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = _get_db(request)

    if payload.period_from > payload.period_to:
        raise HTTPException(400, "period_from cannot be after period_to")

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")
    acc_id_str = str(payload.bank_account_id) if payload.bank_account_id and payload.bank_account_id != "all" else "all"

    # Check if already locked
    existing = await db.reconciliation_locks.find_one({
        "status": "locked",
        "bank_account_id": acc_id_str,
        "period_from": payload.period_from,
        "period_to": payload.period_to,
    })
    if existing:
        return {
            "ok": True,
            "message": f"Period {payload.period_from} to {payload.period_to} is already locked.",
            "lock": stringify(existing),
        }

    lock_doc = {
        "bank_account_id": acc_id_str,
        "period_from": payload.period_from,
        "period_to": payload.period_to,
        "status": "locked",
        "locked_at": now,
        "locked_by": user_email,
        "lock_reason": payload.reason or "Monthly reconciliation finalized",
        "history": [{
            "action": "locked",
            "timestamp": now,
            "user": user_email,
            "reason": payload.reason or "Monthly reconciliation finalized",
        }],
    }
    res = await db.reconciliation_locks.insert_one(lock_doc)
    lock_doc["_id"] = res.inserted_id

    log.info(f"Reconciliation period {payload.period_from} to {payload.period_to} locked for account {acc_id_str} by {user_email}")

    return {
        "ok": True,
        "message": f"Reconciliation period {payload.period_from} to {payload.period_to} successfully locked.",
        "lock": stringify(lock_doc),
    }


@banking_router.post("/banking/periods/unlock")
@banking_router.post("/bank-accounts/periods/unlock")
async def unlock_reconciliation_period(payload: PeriodUnlockIn, request: Request):
    """Unlock a previously locked reconciliation period (Admin only, audit logged)."""
    u = await _get_user(request)
    require_roles("admin")(u)  # ADMIN ONLY!
    db = _get_db(request)

    now = _now_iso()
    user_email = u.get("email") or u.get("name", "")
    acc_id_str = str(payload.bank_account_id) if payload.bank_account_id and payload.bank_account_id != "all" else "all"

    # Find active lock
    lock = await db.reconciliation_locks.find_one({
        "status": "locked",
        "bank_account_id": {"$in": [acc_id_str, "all", None, ""]},
        "period_from": {"$lte": payload.period_from},
        "period_to": {"$gte": payload.period_to},
    })
    if not lock:
        lock = await db.reconciliation_locks.find_one({
            "status": "locked",
            "period_from": payload.period_from,
            "period_to": payload.period_to,
        })

    if not lock:
        raise HTTPException(404, f"No active lock found for period {payload.period_from} to {payload.period_to}")

    unlock_history_entry = {
        "action": "unlocked",
        "timestamp": now,
        "user": user_email,
        "reason": payload.reason or "Admin unlocked for adjustment",
    }

    await db.reconciliation_locks.update_one(
        {"_id": lock["_id"]},
        {
            "$set": {
                "status": "unlocked",
                "unlocked_at": now,
                "unlocked_by": user_email,
                "unlock_reason": payload.reason or "Admin unlocked for adjustment",
            },
            "$push": {"history": unlock_history_entry},
        }
    )

    log.info(f"Reconciliation period {payload.period_from} to {payload.period_to} unlocked by admin {user_email}. Reason: {payload.reason}")

    return {
        "ok": True,
        "message": f"Reconciliation period {payload.period_from} to {payload.period_to} successfully unlocked.",
        "unlocked_by": user_email,
        "unlocked_at": now,
        "reason": payload.reason,
    }


@banking_router.get("/banking/periods/locks")
@banking_router.get("/bank-accounts/periods/locks")
async def list_period_locks(request: Request, bank_account_id: Optional[str] = None):
    """List active and past reconciliation period locks."""
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    q: Dict[str, Any] = {}
    if bank_account_id and bank_account_id != "all":
        q["bank_account_id"] = {"$in": [str(bank_account_id), "all", None, ""]}

    docs = await db.reconciliation_locks.find(q).sort("period_from", -1).to_list(100)
    return {
        "ok": True,
        "locks": [stringify(d) for d in docs],
    }


# ---------------------------------------------------------------------------
# Month-End Reconciliation Statement & Accountant Export (Stage 6)
# ---------------------------------------------------------------------------

def _generate_reconciliation_excel(
    acc_doc: Optional[dict],
    from_date: Optional[str],
    to_date: Optional[str],
    summary_data: dict,
    categorized_data: dict,
    lock_doc: Optional[dict],
    user_email: str,
) -> bytes:
    """
    Generates a CA-ready, audit-grade Multi-Tab Excel Workbook formatted for Month-End Bank Reconciliation.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Define color scheme & typography
    font_family = "Calibri"
    title_font = Font(name=font_family, size=15, bold=True, color="1E3A8A")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="475569")
    h2_font = Font(name=font_family, size=12, bold=True, color="0F172A")
    tbl_hdr_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    tbl_hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    subhdr_fill = PatternFill("solid", fgColor="334155")
    card_hdr_fill = PatternFill("solid", fgColor="F1F5F9")
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    total_fill = PatternFill("solid", fgColor="E2E8F0")
    success_fill = PatternFill("solid", fgColor="DCFCE7")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")

    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    green_bold = Font(name=font_family, size=10, bold=True, color="166534")
    red_bold = Font(name=font_family, size=10, bold=True, color="991B1B")
    amber_bold = Font(name=font_family, size=10, bold=True, color="92400E")

    thin_border_side = Side(style="thin", color="CBD5E1")
    thick_bottom_side = Side(style="medium", color="1E3A8A")
    double_bottom_side = Side(style="double", color="1E3A8A")

    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(top=thin_border_side, bottom=double_bottom_side, left=thin_border_side, right=thin_border_side)

    CURR_FMT = "#,##0.00"

    def auto_fit_columns(ws, max_widths=None):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if cell.number_format and "0.00" in cell.number_format and isinstance(cell.value, (int, float)):
                    val = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val))
            applied_w = max(max_len + 3, 12)
            if max_widths and col_letter in max_widths:
                applied_w = min(applied_w, max_widths[col_letter])
            ws.column_dimensions[col_letter].width = applied_w

    # =========================================================================
    # TAB 1: Executive Summary & Reconciliation Certificate
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title & Metadata
    ws1["A1"] = "SSK FOOTCARE - BANK RECONCILIATION STATEMENT"
    ws1["A1"].font = title_font
    ws1["A2"] = "Official Month-End Financial Audit & Reconciliation Certificate"
    ws1["A2"].font = subtitle_font

    acc_name = acc_doc.get("name", "All Bank Accounts") if acc_doc else "All Bank Accounts"
    bank_name = acc_doc.get("bank_name", "Consolidated") if acc_doc else "Consolidated"
    acc_num = acc_doc.get("account_number") or acc_doc.get("account_number_last4") or "N/A" if acc_doc else "All"
    ifsc = acc_doc.get("ifsc", "N/A") if acc_doc else "N/A"
    branch = acc_doc.get("branch", "N/A") if acc_doc else "N/A"
    period_label = f"{from_date or 'Start'} to {to_date or 'Present'}"
    is_locked = lock_doc and lock_doc.get("status") == "locked"
    lock_status_label = f"FINALIZED & LOCKED ({lock_doc.get('locked_at', '')[:10]} by {lock_doc.get('locked_by', 'Admin')})" if is_locked else "OPEN / DRAFT"

    meta_entries = [
        ("Bank Account:", acc_name, "Reconciliation Period:", period_label),
        ("Bank & Branch:", f"{bank_name} ({branch})", "Lock / Audit Status:", lock_status_label),
        ("Account Number / IFSC:", f"{acc_num} / {ifsc}", "Report Generated On:", f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} by {user_email}"),
    ]

    curr_row = 4
    for left_k, left_v, right_k, right_v in meta_entries:
        ws1.cell(row=curr_row, column=1, value=left_k).font = bold_font
        ws1.cell(row=curr_row, column=2, value=left_v).font = regular_font
        ws1.cell(row=curr_row, column=3, value=right_k).font = bold_font
        ws1.cell(row=curr_row, column=4, value=right_v).font = bold_font if "Status" in right_k and is_locked else regular_font
        if "Status" in right_k and is_locked:
            ws1.cell(row=curr_row, column=4).fill = warn_fill
        curr_row += 1

    curr_row += 1

    # Balance Reconciliation Table
    ws1.cell(row=curr_row, column=1, value="1. RECONCILIATION SUMMARY (BALANCE PROOF)").font = h2_font
    curr_row += 1

    rec_headers = ["Line Item / Balance Component", "Amount (INR)", "Reconciliation Status / Note"]
    for c_idx, h_text in enumerate(rec_headers, start=1):
        c = ws1.cell(row=curr_row, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx != 2 else "right")
    curr_row += 1

    op_bal = summary_data.get("opening_balance", 0.0)
    matched_inc = summary_data.get("matched_income", 0.0)
    unmatched_inc = summary_data.get("unmatched_income", 0.0)
    total_inc = summary_data.get("total_income", 0.0)
    matched_exp = summary_data.get("matched_expenses", 0.0)
    unmatched_exp = summary_data.get("unmatched_expenses", 0.0)
    total_exp = summary_data.get("total_expenses", 0.0)
    stmt_close_bal = summary_data.get("statement_closing_balance", op_bal + total_inc - total_exp)
    reconciled_erp_bal = summary_data.get("reconciled_erp_balance", op_bal + matched_inc - matched_exp)
    variance = round(stmt_close_bal - reconciled_erp_bal, 2)

    rec_rows = [
        ("Bank Statement Opening Balance", op_bal, "Opening ledger position as of period start", regular_font),
        ("(+) Matched Inflows / Revenue & Receipts", matched_inc, "Verified and linked against ERP invoices & settlements", regular_font),
        ("(-) Matched Outflows / Operating Expenses & Payments", matched_exp, "Verified and linked against ERP expenses & vendor payments", regular_font),
        ("(=) Reconciled ERP Closing Position", reconciled_erp_bal, "Fully accounted & verified ERP cash position", bold_font),
        ("(+) Unmatched Statement Credits (Pending Inflows)", unmatched_inc, "Receipts on bank statement awaiting ERP booking", regular_font),
        ("(-) Unmatched Statement Debits (Pending Outflows)", unmatched_exp, "Debits on bank statement awaiting ERP booking", regular_font),
        ("(=) Bank Statement Closing Balance", stmt_close_bal, "Recorded bank statement closing figure", bold_font),
        ("Reconciliation Variance (Discrepancy)", variance, "Zero variance indicates 100% reconciled period" if variance == 0 else "Pending items require accountant allocation", bold_font),
    ]

    for item, amt, note, fnt in rec_rows:
        c1 = ws1.cell(row=curr_row, column=1, value=item)
        c2 = ws1.cell(row=curr_row, column=2, value=amt)
        c3 = ws1.cell(row=curr_row, column=3, value=note)

        c1.font = fnt
        c2.font = fnt
        c3.font = regular_font

        c1.border = cell_border
        c2.border = cell_border
        c3.border = cell_border

        c2.number_format = CURR_FMT
        c2.alignment = Alignment(horizontal="right")

        if item.startswith("(=)"):
            c1.fill = total_fill
            c2.fill = total_fill
            c3.fill = total_fill
        elif "Variance" in item:
            fill_to_use = success_fill if variance == 0 else warn_fill
            font_to_use = green_bold if variance == 0 else red_bold
            c1.fill = fill_to_use
            c2.fill = fill_to_use
            c3.fill = fill_to_use
            c1.font = font_to_use
            c2.font = font_to_use
            c3.font = font_to_use

        curr_row += 1

    curr_row += 2

    # Categorized Volume Breakdown Table
    ws1.cell(row=curr_row, column=1, value="2. TRANSACTION TYPE & CATEGORY AUDIT BREAKDOWN").font = h2_font
    curr_row += 1

    vol_headers = ["Transaction Category", "Flow Direction", "Transaction Count", "Total Volume (INR)", "Audit Description"]
    for c_idx, h_text in enumerate(vol_headers, start=1):
        c = ws1.cell(row=curr_row, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = subhdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx not in [3, 4] else "right")
    curr_row += 1

    rev_list = categorized_data.get("revenue", [])
    exp_list = categorized_data.get("expenses", [])
    vp_list = categorized_data.get("vendor_payments", [])
    trf_list = categorized_data.get("transfers", [])
    cash_list = categorized_data.get("cash_withdrawals", [])
    unmatched_list = categorized_data.get("unmatched", [])

    rev_vol = sum(float(r.get("credit_amount") or 0.0) for r in rev_list)
    exp_vol = sum(float(r.get("debit_amount") or 0.0) for r in exp_list)
    vp_vol = sum(float(r.get("debit_amount") or 0.0) for r in vp_list)
    trf_vol = sum(float(r.get("amount") or 0.0) for r in trf_list)
    cash_vol = sum(float(r.get("debit_amount") or 0.0) for r in cash_list)
    unmatched_vol = sum(float(r.get("debit_amount") or 0.0) + float(r.get("credit_amount") or 0.0) for r in unmatched_list)

    breakdown_rows = [
        ("1. Revenue & Client Receipts", "Credit (Inflow)", len(rev_list), rev_vol, "B2B client invoices & online marketplace payouts"),
        ("2. Operating Overheads & Expenses", "Debit (Outflow)", len(exp_list), exp_vol, "Factory rent, utilities, consumables, courier, GST"),
        ("3. Vendor Raw Material Payments", "Debit (Outflow)", len(vp_list), vp_vol, "Leather, sole, upper, insole supplier NEFT/RTGS"),
        ("4. Cash Withdrawals (Karigar Wages / Petty Cash)", "Debit (Outflow)", len(cash_list), cash_vol, "ATM/Bank cash drawn for karigar wages & cash expenses"),
        ("5. Inter-Account Transfers", "Internal Transfer", len(trf_list), trf_vol, "Excluded from P&L as neutral internal liquidity moves"),
        ("6. Unmatched / Pending Transactions", "Debit & Credit", len(unmatched_list), unmatched_vol, "Unclassified statement entries awaiting audit review"),
    ]

    for cat_title, flow, count, vol, desc in breakdown_rows:
        c1 = ws1.cell(row=curr_row, column=1, value=cat_title)
        c2 = ws1.cell(row=curr_row, column=2, value=flow)
        c3 = ws1.cell(row=curr_row, column=3, value=count)
        c4 = ws1.cell(row=curr_row, column=4, value=vol)
        c5 = ws1.cell(row=curr_row, column=5, value=desc)

        for cell in (c1, c2, c3, c4, c5):
            cell.font = regular_font
            cell.border = cell_border

        c3.alignment = Alignment(horizontal="right")
        c4.alignment = Alignment(horizontal="right")
        c4.number_format = CURR_FMT
        curr_row += 1

    auto_fit_columns(ws1, {"A": 45, "B": 25, "C": 35, "D": 30, "E": 45})

    # =========================================================================
    # TAB 2: Revenue & Client Receipts
    # =========================================================================
    ws2 = wb.create_sheet("1. Revenue & Receipts")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "1. REVENUE & CLIENT RECEIPTS (MATCHED CREDITS)"
    ws2["A1"].font = h2_font

    r_headers = ["Date", "Statement Narration", "Ref / UTR", "Client / Source", "ERP Voucher #", "Statement Credit (INR)", "Reconciled Amount (INR)", "Status"]
    for c_idx, h_text in enumerate(r_headers, start=1):
        c = ws2.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx not in [6, 7] else "right")

    r_row = 4
    for item in rev_list:
        ws2.cell(row=r_row, column=1, value=item.get("date")).font = regular_font
        ws2.cell(row=r_row, column=2, value=item.get("narration")).font = regular_font
        ws2.cell(row=r_row, column=3, value=item.get("reference_no") or "-").font = regular_font
        ws2.cell(row=r_row, column=4, value=item.get("party_name") or "-").font = bold_font
        ws2.cell(row=r_row, column=5, value=item.get("erp_voucher") or "-").font = regular_font
        
        c6 = ws2.cell(row=r_row, column=6, value=float(item.get("credit_amount") or 0.0))
        c7 = ws2.cell(row=r_row, column=7, value=float(item.get("reconciled_amount") or item.get("credit_amount") or 0.0))
        c8 = ws2.cell(row=r_row, column=8, value="MATCHED")

        c6.number_format = CURR_FMT
        c7.number_format = CURR_FMT
        c6.alignment = Alignment(horizontal="right")
        c7.alignment = Alignment(horizontal="right")
        c8.font = green_bold

        for col_i in range(1, 9):
            ws2.cell(row=r_row, column=col_i).border = cell_border
            if r_row % 2 == 0:
                ws2.cell(row=r_row, column=col_i).fill = zebra_fill
        r_row += 1

    # Total Row
    if rev_list:
        ws2.cell(row=r_row, column=1, value="TOTAL REVENUE & RECEIPTS").font = bold_font
        ws2.cell(row=r_row, column=6, value=rev_vol).font = bold_font
        ws2.cell(row=r_row, column=6).number_format = CURR_FMT
        ws2.cell(row=r_row, column=7, value=rev_vol).font = bold_font
        ws2.cell(row=r_row, column=7).number_format = CURR_FMT
        for col_i in range(1, 9):
            ws2.cell(row=r_row, column=col_i).border = total_border
            ws2.cell(row=r_row, column=col_i).fill = total_fill

    auto_fit_columns(ws2)

    # =========================================================================
    # TAB 3: Operating Overheads & Expenses
    # =========================================================================
    ws3 = wb.create_sheet("2. Operating Expenses")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "2. OPERATING OVERHEADS & EXPENSES (MATCHED DEBITS)"
    ws3["A1"].font = h2_font

    e_headers = ["Date", "Statement Narration", "Ref / UTR", "Payee / Beneficiary", "Expense Category", "Statement Debit (INR)", "Reconciled Amount (INR)", "Status"]
    for c_idx, h_text in enumerate(e_headers, start=1):
        c = ws3.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx not in [6, 7] else "right")

    e_row = 4
    for item in exp_list:
        ws3.cell(row=e_row, column=1, value=item.get("date")).font = regular_font
        ws3.cell(row=e_row, column=2, value=item.get("narration")).font = regular_font
        ws3.cell(row=e_row, column=3, value=item.get("reference_no") or "-").font = regular_font
        ws3.cell(row=e_row, column=4, value=item.get("party_name") or "-").font = bold_font
        ws3.cell(row=e_row, column=5, value=item.get("category") or "General").font = regular_font
        
        c6 = ws3.cell(row=e_row, column=6, value=float(item.get("debit_amount") or 0.0))
        c7 = ws3.cell(row=e_row, column=7, value=float(item.get("reconciled_amount") or item.get("debit_amount") or 0.0))
        c8 = ws3.cell(row=e_row, column=8, value="MATCHED")

        c6.number_format = CURR_FMT
        c7.number_format = CURR_FMT
        c6.alignment = Alignment(horizontal="right")
        c7.alignment = Alignment(horizontal="right")
        c8.font = green_bold

        for col_i in range(1, 9):
            ws3.cell(row=e_row, column=col_i).border = cell_border
            if e_row % 2 == 0:
                ws3.cell(row=e_row, column=col_i).fill = zebra_fill
        e_row += 1

    if exp_list:
        ws3.cell(row=e_row, column=1, value="TOTAL OPERATING EXPENSES").font = bold_font
        ws3.cell(row=e_row, column=6, value=exp_vol).font = bold_font
        ws3.cell(row=e_row, column=6).number_format = CURR_FMT
        ws3.cell(row=e_row, column=7, value=exp_vol).font = bold_font
        ws3.cell(row=e_row, column=7).number_format = CURR_FMT
        for col_i in range(1, 9):
            ws3.cell(row=e_row, column=col_i).border = total_border
            ws3.cell(row=e_row, column=col_i).fill = total_fill

    auto_fit_columns(ws3)

    # =========================================================================
    # TAB 4: Vendor Payments
    # =========================================================================
    ws4 = wb.create_sheet("3. Vendor Payments")
    ws4.views.sheetView[0].showGridLines = True
    ws4["A1"] = "3. VENDOR RAW MATERIAL PAYMENTS (MATCHED DEBITS)"
    ws4["A1"].font = h2_font

    v_headers = ["Date", "Statement Narration", "Ref / UTR", "Vendor / Supplier", "Payment Voucher #", "Statement Debit (INR)", "Reconciled Amount (INR)", "Status"]
    for c_idx, h_text in enumerate(v_headers, start=1):
        c = ws4.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx not in [6, 7] else "right")

    v_row = 4
    for item in vp_list:
        ws4.cell(row=v_row, column=1, value=item.get("date")).font = regular_font
        ws4.cell(row=v_row, column=2, value=item.get("narration")).font = regular_font
        ws4.cell(row=v_row, column=3, value=item.get("reference_no") or "-").font = regular_font
        ws4.cell(row=v_row, column=4, value=item.get("party_name") or "-").font = bold_font
        ws4.cell(row=v_row, column=5, value=item.get("erp_voucher") or "-").font = regular_font
        
        c6 = ws4.cell(row=v_row, column=6, value=float(item.get("debit_amount") or 0.0))
        c7 = ws4.cell(row=v_row, column=7, value=float(item.get("reconciled_amount") or item.get("debit_amount") or 0.0))
        c8 = ws4.cell(row=v_row, column=8, value="MATCHED")

        c6.number_format = CURR_FMT
        c7.number_format = CURR_FMT
        c6.alignment = Alignment(horizontal="right")
        c7.alignment = Alignment(horizontal="right")
        c8.font = green_bold

        for col_i in range(1, 9):
            ws4.cell(row=v_row, column=col_i).border = cell_border
            if v_row % 2 == 0:
                ws4.cell(row=v_row, column=col_i).fill = zebra_fill
        v_row += 1

    if vp_list:
        ws4.cell(row=v_row, column=1, value="TOTAL VENDOR PAYMENTS").font = bold_font
        ws4.cell(row=v_row, column=6, value=vp_vol).font = bold_font
        ws4.cell(row=v_row, column=6).number_format = CURR_FMT
        ws4.cell(row=v_row, column=7, value=vp_vol).font = bold_font
        ws4.cell(row=v_row, column=7).number_format = CURR_FMT
        for col_i in range(1, 9):
            ws4.cell(row=v_row, column=col_i).border = total_border
            ws4.cell(row=v_row, column=col_i).fill = total_fill

    auto_fit_columns(ws4)

    # =========================================================================
    # TAB 5: Cash Withdrawals & Karigar Wages / Expense Funding
    # =========================================================================
    ws5 = wb.create_sheet("4. Cash & Karigar Wages")
    ws5.views.sheetView[0].showGridLines = True
    ws5["A1"] = "4. CASH WITHDRAWALS & KARIGAR WAGES / CASH DISBURSEMENTS"
    ws5["A1"].font = h2_font

    c_headers = [
        "Withdrawal Date",
        "Statement Narration",
        "Withdrawn (INR)",
        "Disbursed to Wages & Expenses (INR)",
        "Unallocated Cash in Hand (INR)",
        "Cash Pool Ref",
        "Disbursement Details (What did this cash fund?)",
    ]
    for c_idx, h_text in enumerate(c_headers, start=1):
        c = ws5.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx not in [3, 4, 5] else "right")

    c_row = 4
    total_disbursed_all = 0.0
    total_unalloc_all = 0.0

    for item in cash_list:
        w_amt = float(item.get("debit_amount") or item.get("amount") or 0.0)
        alloc_amt = float(item.get("allocated_amount") or 0.0)
        unalloc_amt = float(item.get("remaining_balance") or (w_amt - alloc_amt))
        total_disbursed_all += alloc_amt
        total_unalloc_all += unalloc_amt

        ws5.cell(row=c_row, column=1, value=item.get("date")).font = regular_font
        ws5.cell(row=c_row, column=2, value=item.get("narration")).font = regular_font
        
        c3 = ws5.cell(row=c_row, column=3, value=w_amt)
        c4 = ws5.cell(row=c_row, column=4, value=alloc_amt)
        c5 = ws5.cell(row=c_row, column=5, value=unalloc_amt)
        c6 = ws5.cell(row=c_row, column=6, value=item.get("cash_ledger_id") or "-")
        c7 = ws5.cell(row=c_row, column=7, value=item.get("disbursements_summary") or "Unallocated Cash in Hand")

        c3.number_format = CURR_FMT
        c4.number_format = CURR_FMT
        c5.number_format = CURR_FMT
        c3.alignment = Alignment(horizontal="right")
        c4.alignment = Alignment(horizontal="right")
        c5.alignment = Alignment(horizontal="right")

        c3.font = bold_font
        c4.font = regular_font
        c5.font = amber_bold if unalloc_amt > 0 else regular_font
        c6.font = regular_font
        c7.font = regular_font

        for col_i in range(1, 8):
            ws5.cell(row=c_row, column=col_i).border = cell_border
            if c_row % 2 == 0:
                ws5.cell(row=c_row, column=col_i).fill = zebra_fill
        c_row += 1

    if cash_list:
        ws5.cell(row=c_row, column=1, value="TOTAL CASH WITHDRAWALS").font = bold_font
        ws5.cell(row=c_row, column=3, value=cash_vol).font = bold_font
        ws5.cell(row=c_row, column=3).number_format = CURR_FMT
        ws5.cell(row=c_row, column=4, value=total_disbursed_all).font = bold_font
        ws5.cell(row=c_row, column=4).number_format = CURR_FMT
        ws5.cell(row=c_row, column=5, value=total_unalloc_all).font = bold_font
        ws5.cell(row=c_row, column=5).number_format = CURR_FMT
        for col_i in range(1, 8):
            ws5.cell(row=c_row, column=col_i).border = total_border
            ws5.cell(row=c_row, column=col_i).fill = total_fill

    auto_fit_columns(ws5, {"G": 60})

    # =========================================================================
    # TAB 6: Inter-Account Transfers
    # =========================================================================
    ws6 = wb.create_sheet("5. Inter-Account Transfers")
    ws6.views.sheetView[0].showGridLines = True
    ws6["A1"] = "5. INTER-ACCOUNT TRANSFERS (INTERNAL LIQUIDITY MOVES - P&L EXCLUDED)"
    ws6["A1"].font = h2_font

    t_headers = ["Date", "Sending Account (Debit)", "Receiving Account (Credit)", "Statement Narration", "Amount (INR)", "Confirmed By", "Notes"]
    for c_idx, h_text in enumerate(t_headers, start=1):
        c = ws6.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = tbl_hdr_fill
        c.alignment = Alignment(horizontal="left" if c_idx != 5 else "right")

    t_row = 4
    for item in trf_list:
        ws6.cell(row=t_row, column=1, value=item.get("date")).font = regular_font
        ws6.cell(row=t_row, column=2, value=item.get("from_account_name") or "-").font = bold_font
        ws6.cell(row=t_row, column=3, value=item.get("to_account_name") or "-").font = bold_font
        ws6.cell(row=t_row, column=4, value=item.get("narration") or "-").font = regular_font
        
        c5 = ws6.cell(row=t_row, column=5, value=float(item.get("amount") or 0.0))
        c6 = ws6.cell(row=t_row, column=6, value=item.get("confirmed_by") or "-")
        c7 = ws6.cell(row=t_row, column=7, value=item.get("notes") or "-")

        c5.number_format = CURR_FMT
        c5.alignment = Alignment(horizontal="right")
        c5.font = bold_font

        for col_i in range(1, 8):
            ws6.cell(row=t_row, column=col_i).border = cell_border
            if t_row % 2 == 0:
                ws6.cell(row=t_row, column=col_i).fill = zebra_fill
        t_row += 1

    if trf_list:
        ws6.cell(row=t_row, column=1, value="TOTAL INTERNAL TRANSFERS").font = bold_font
        ws6.cell(row=t_row, column=5, value=trf_vol).font = bold_font
        ws6.cell(row=t_row, column=5).number_format = CURR_FMT
        for col_i in range(1, 8):
            ws6.cell(row=t_row, column=col_i).border = total_border
            ws6.cell(row=t_row, column=col_i).fill = total_fill

    auto_fit_columns(ws6)

    # =========================================================================
    # TAB 7: Pending & Unmatched Transactions
    # =========================================================================
    ws7 = wb.create_sheet("6. Pending & Unmatched")
    ws7.views.sheetView[0].showGridLines = True
    ws7["A1"] = "6. PENDING & UNMATCHED STATEMENT ENTRIES (REQUIRING AUDIT ATTENTION)"
    ws7["A1"].font = h2_font

    u_headers = ["Date", "Bank Account", "Statement Narration", "Reference No", "Debit (INR)", "Credit (INR)", "Status", "Auditor / Accountant Remarks"]
    for c_idx, h_text in enumerate(u_headers, start=1):
        c = ws7.cell(row=3, column=c_idx, value=h_text)
        c.font = tbl_hdr_font
        c.fill = PatternFill("solid", fgColor="B91C1C" if unmatched_list else "1E3A8A")
        c.alignment = Alignment(horizontal="left" if c_idx not in [5, 6] else "right")

    u_row = 4
    for item in unmatched_list:
        ws7.cell(row=u_row, column=1, value=item.get("date")).font = regular_font
        ws7.cell(row=u_row, column=2, value=item.get("bank_account_name") or "-").font = bold_font
        ws7.cell(row=u_row, column=3, value=item.get("narration")).font = regular_font
        ws7.cell(row=u_row, column=4, value=item.get("reference_no") or "-").font = regular_font
        
        c5 = ws7.cell(row=u_row, column=5, value=float(item.get("debit_amount") or 0.0) if float(item.get("debit_amount") or 0.0) > 0 else "")
        c6 = ws7.cell(row=u_row, column=6, value=float(item.get("credit_amount") or 0.0) if float(item.get("credit_amount") or 0.0) > 0 else "")
        c7 = ws7.cell(row=u_row, column=7, value=str(item.get("match_status") or "unmatched").upper())
        c8 = ws7.cell(row=u_row, column=8, value=item.get("remarks") or "-")

        if isinstance(c5.value, (int, float)):
            c5.number_format = CURR_FMT
            c5.alignment = Alignment(horizontal="right")
            c5.font = red_bold
        if isinstance(c6.value, (int, float)):
            c6.number_format = CURR_FMT
            c6.alignment = Alignment(horizontal="right")
            c6.font = green_bold
        c7.font = amber_bold

        for col_i in range(1, 9):
            ws7.cell(row=u_row, column=col_i).border = cell_border
            if u_row % 2 == 0:
                ws7.cell(row=u_row, column=col_i).fill = zebra_fill
        u_row += 1

    if not unmatched_list:
        ws7.cell(row=4, column=1, value="No pending or unmatched lines found for this period. Reconciliation is 100% complete.").font = green_bold

    auto_fit_columns(ws7)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@banking_router.get("/banking/reconciliation/export")
@banking_router.get("/bank-accounts/reconciliation/export")
async def export_reconciliation_report(
    request: Request,
    bank_account_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    format: str = Query("excel"),
):
    """
    Generate and download an official CA / Accountant-Ready Bank Reconciliation Report (Excel format).
    Includes statement balance, ERP reconciled balance, variance proof, grouped transactions (Revenue, Expenses, Vendor, Transfers, Cash), and unmatched items.
    """
    u = await _get_user(request)
    require_roles("admin", "manager", "sales")(u)
    db = _get_db(request)

    acc_id_str = str(bank_account_id) if bank_account_id and bank_account_id != "all" else None
    acc_doc = None
    if acc_id_str:
        acc_doc = await db.bank_accounts.find_one({"_id": _oid(acc_id_str)})

    # Fetch statement lines
    q: Dict[str, Any] = {}
    if acc_id_str:
        q["bank_account_id"] = acc_id_str
    if from_date or to_date:
        dq = {}
        if from_date:
            dq["$gte"] = str(from_date)
        if to_date:
            dq["$lte"] = str(to_date)
        q["date"] = dq

    docs = await db.bank_statement_lines.find(q).sort("date", 1).to_list(50000)

    # Fetch accounts map
    accounts = await db.bank_accounts.find({}).to_list(1000)
    acc_map = {str(a["_id"]): a for a in accounts}

    # Fetch active period lock if any
    lock_q: Dict[str, Any] = {"status": "locked"}
    if acc_id_str:
        lock_q["bank_account_id"] = {"$in": [acc_id_str, "all", None, ""]}
    if from_date:
        lock_q["period_to"] = {"$gte": from_date}
    if to_date:
        lock_q["period_from"] = {"$lte": to_date}

    lock_doc = await db.reconciliation_locks.find_one(lock_q) if hasattr(db, "reconciliation_locks") and db.reconciliation_locks is not None else None

    # Fetch ERP records to enrich details
    client_payments_map = {}
    vendor_payments_map = {}
    expenses_map = {}
    settlements_map = {}
    cash_ledger_map = {}

    if hasattr(db, "payments") and db.payments is not None:
        try:
            p_list = await db.payments.find({}).to_list(20000)
            for p in p_list:
                pid = str(p["_id"])
                if p.get("type") == "vendor_payment" or p.get("vendor_id") or p.get("vendor_name"):
                    vendor_payments_map[pid] = p
                else:
                    client_payments_map[pid] = p
        except Exception:
            pass

    if hasattr(db, "expenses") and db.expenses is not None:
        try:
            e_list = await db.expenses.find({}).to_list(20000)
            for e in e_list:
                expenses_map[str(e["_id"])] = e
        except Exception:
            pass

    if hasattr(db, "online_settlements") and db.online_settlements is not None:
        try:
            s_list = await db.online_settlements.find({}).to_list(20000)
            for s in s_list:
                settlements_map[str(s["_id"])] = s
        except Exception:
            pass

    if hasattr(db, "cash_ledger") and db.cash_ledger is not None:
        try:
            c_list = await db.cash_ledger.find({}).to_list(5000)
            for c in c_list:
                cash_ledger_map[str(c["_id"])] = c
                if c.get("source_statement_line_id"):
                    cash_ledger_map[str(c.get("source_statement_line_id"))] = c
        except Exception:
            pass

    # Fetch wage payments & cash expenses for cash disbursements
    wage_payments_by_cash_id = defaultdict(list)
    cash_expenses_by_cash_id = defaultdict(list)

    if hasattr(db, "wage_payments") and db.wage_payments is not None:
        try:
            wp_list = await db.wage_payments.find({"paid_via": "cash"}).to_list(20000)
            for wp in wp_list:
                clid = str(wp.get("cash_ledger_id") or "")
                if clid:
                    wage_payments_by_cash_id[clid].append(wp)
        except Exception:
            pass

    if hasattr(db, "expenses") and db.expenses is not None:
        try:
            ce_list = await db.expenses.find({"paid_via": "cash"}).to_list(20000)
            for ce in ce_list:
                clid = str(ce.get("cash_ledger_id") or "")
                if clid:
                    cash_expenses_by_cash_id[clid].append(ce)
        except Exception:
            pass

    # Build categorized lists
    rev_items = []
    exp_items = []
    vp_items = []
    trf_items = []
    cash_items = []
    unmatched_items = []

    total_income = 0.0
    matched_income = 0.0
    unmatched_income = 0.0
    total_expenses = 0.0
    matched_expenses = 0.0
    unmatched_expenses = 0.0

    lines_by_id = {str(d["_id"]): d for d in docs}

    for d in docs:
        lid = str(d["_id"])
        status = d.get("match_status", "unmatched")
        matched_to = d.get("matched_to") or {}
        m_type = matched_to.get("type") if isinstance(matched_to, dict) else None
        ref_id = str(matched_to.get("ref_id") or "") if isinstance(matched_to, dict) else ""

        credit = float(d.get("credit_amount") or 0.0)
        debit = float(d.get("debit_amount") or 0.0)
        acc_info = acc_map.get(str(d.get("bank_account_id")), {})

        if status == "transfer":
            peer_line = lines_by_id.get(ref_id)
            if not peer_line:
                try:
                    peer_line = await db.bank_statement_lines.find_one({"_id": _oid(ref_id)})
                except Exception:
                    peer_line = None
            peer_acc = acc_map.get(str(peer_line.get("bank_account_id") if peer_line else ""), {})

            if debit > 0:
                trf_items.append({
                    "date": d.get("date"),
                    "from_account_name": acc_info.get("name", "Current Account"),
                    "to_account_name": peer_acc.get("name", "Peer Account"),
                    "narration": d.get("narration"),
                    "amount": debit,
                    "confirmed_by": d.get("confirmed_by") or d.get("imported_by") or "-",
                    "notes": d.get("transfer_notes") or "",
                })
            continue

        if status == "ignored":
            continue

        if credit > 0:
            total_income += credit
            if status == "matched":
                matched_income += credit
                if m_type == "settlement" and ref_id in settlements_map:
                    s_doc = settlements_map[ref_id]
                    rev_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": f"Marketplace Settlement ({s_doc.get('order_release_id', '')})",
                        "erp_voucher": s_doc.get("order_release_id") or "Settlement",
                        "credit_amount": credit,
                        "reconciled_amount": float(s_doc.get("net_payout") or credit),
                    })
                elif m_type == "payment" and ref_id in client_payments_map:
                    p_doc = client_payments_map[ref_id]
                    rev_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": p_doc.get("client_name") or "Client Payment",
                        "erp_voucher": p_doc.get("payment_no") or p_doc.get("reference") or "Payment",
                        "credit_amount": credit,
                        "reconciled_amount": float(p_doc.get("amount") or credit),
                    })
                else:
                    rev_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": "Direct Client Receipt",
                        "erp_voucher": ref_id or "Matched",
                        "credit_amount": credit,
                        "reconciled_amount": credit,
                    })
            else:
                unmatched_income += credit
                unmatched_items.append({
                    "date": d.get("date"),
                    "bank_account_name": acc_info.get("name", "Account"),
                    "narration": d.get("narration"),
                    "reference_no": d.get("reference_no"),
                    "debit_amount": 0.0,
                    "credit_amount": credit,
                    "match_status": status,
                    "remarks": d.get("remarks"),
                })

        if debit > 0:
            total_expenses += debit
            if status == "matched":
                matched_expenses += debit
                if m_type == "cash_withdrawal":
                    cash_doc = cash_ledger_map.get(ref_id) or cash_ledger_map.get(lid)
                    clid = str(cash_doc["_id"]) if cash_doc else ref_id
                    wps = wage_payments_by_cash_id.get(clid, [])
                    exps = cash_expenses_by_cash_id.get(clid, [])
                    alloc = sum(float(x.get("amount") or 0.0) for x in wps) + sum(float(x.get("amount") or 0.0) for x in exps)
                    rem = float(cash_doc.get("remaining_balance") if cash_doc else (debit - alloc))

                    disb_details = []
                    if wps:
                        wp_strs = [f"Karigar {wp.get('worker_name', 'Worker')} (₹{float(wp.get('amount') or 0.0):,.2f})" for wp in wps[:5]]
                        if len(wps) > 5:
                            wp_strs.append(f"+{len(wps)-5} more karigars")
                        disb_details.append(f"Wages: {', '.join(wp_strs)}")
                    if exps:
                        exp_strs = [f"{e.get('category', 'Expense')} (₹{float(e.get('amount') or 0.0):,.2f})" for e in exps[:5]]
                        if len(exps) > 5:
                            exp_strs.append(f"+{len(exps)-5} more expenses")
                        disb_details.append(f"Expenses: {', '.join(exp_strs)}")

                    disb_summary = "; ".join(disb_details) if disb_details else "Unallocated Cash in Hand"

                    cash_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "debit_amount": debit,
                        "allocated_amount": round(alloc, 2),
                        "remaining_balance": round(rem, 2),
                        "cash_ledger_id": clid,
                        "disbursements_summary": disb_summary,
                    })
                elif m_type in ["vendor_payment", "payment"] and (ref_id in vendor_payments_map or (ref_id in client_payments_map and client_payments_map[ref_id].get("vendor_name"))):
                    v_doc = vendor_payments_map.get(ref_id) or client_payments_map.get(ref_id, {})
                    vp_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": v_doc.get("vendor_name") or v_doc.get("client_name") or "Vendor Payment",
                        "erp_voucher": v_doc.get("payment_no") or v_doc.get("reference") or "Payment",
                        "debit_amount": debit,
                        "reconciled_amount": float(v_doc.get("amount") or debit),
                    })
                elif m_type == "expense" and ref_id in expenses_map:
                    e_doc = expenses_map[ref_id]
                    exp_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": e_doc.get("payee") or "Expense Payee",
                        "category": e_doc.get("category") or "Direct Expense",
                        "debit_amount": debit,
                        "reconciled_amount": float(e_doc.get("amount") or debit),
                    })
                else:
                    exp_items.append({
                        "date": d.get("date"),
                        "narration": d.get("narration"),
                        "reference_no": d.get("reference_no"),
                        "party_name": "Direct Expense",
                        "category": "Direct Expense",
                        "debit_amount": debit,
                        "reconciled_amount": debit,
                    })
            else:
                unmatched_expenses += debit
                unmatched_items.append({
                    "date": d.get("date"),
                    "bank_account_name": acc_info.get("name", "Account"),
                    "narration": d.get("narration"),
                    "reference_no": d.get("reference_no"),
                    "debit_amount": debit,
                    "credit_amount": 0.0,
                    "match_status": status,
                    "remarks": d.get("remarks"),
                })

    op_balance = float(acc_doc.get("opening_balance") or 0.0) if acc_doc else sum(float(a.get("opening_balance") or 0.0) for a in accounts)
    stmt_closing = round(op_balance + total_income - total_expenses, 2)
    erp_closing = round(op_balance + matched_income - matched_expenses, 2)

    summary_data = {
        "opening_balance": op_balance,
        "total_income": round(total_income, 2),
        "matched_income": round(matched_income, 2),
        "unmatched_income": round(unmatched_income, 2),
        "total_expenses": round(total_expenses, 2),
        "matched_expenses": round(matched_expenses, 2),
        "unmatched_expenses": round(unmatched_expenses, 2),
        "statement_closing_balance": stmt_closing,
        "reconciled_erp_balance": erp_closing,
    }

    categorized_data = {
        "revenue": rev_items,
        "expenses": exp_items,
        "vendor_payments": vp_items,
        "transfers": trf_items,
        "cash_withdrawals": cash_items,
        "unmatched": unmatched_items,
    }

    user_email = u.get("email") or u.get("name", "Auditor")
    excel_bytes = _generate_reconciliation_excel(
        acc_doc=acc_doc,
        from_date=from_date,
        to_date=to_date,
        summary_data=summary_data,
        categorized_data=categorized_data,
        lock_doc=lock_doc,
        user_email=user_email,
    )

    acc_slug = acc_doc.get("name", "Consolidated").replace(" ", "_") if acc_doc else "Consolidated"
    date_slug = f"{from_date or 'Start'}_to_{to_date or 'Present'}"
    filename = f"Bank_Reconciliation_Statement_{acc_slug}_{date_slug}.xlsx"

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )






