"""SKU Map, Listing Import & Marketplace Resolver Routes."""

import re
import io
import csv
import uuid
import difflib
import logging
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any, Tuple
from bson import ObjectId
from fastapi import APIRouter, HTTPException, Request, Depends, UploadFile, File
from fastapi.responses import Response
from pymongo.errors import DuplicateKeyError

from models.sku_map import (
    SkuMapIn,
    SkuMapUpdate,
    ParserTemplateIn,
    StyleColorMappingIn,
    SkuResolveIn,
    UnresolvedMapIn,
    ListingImportCommitIn,
)
from auth import require_roles
from rate_limiter import upload_rate_limiter

log = logging.getLogger(__name__)

sku_map_router = APIRouter(prefix="/api", tags=["SKU Mapping & Marketplace Resolver"])


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stringify(doc: dict) -> dict:
    if not doc:
        return {}
    d = dict(doc)
    if "_id" in d:
        d["id"] = str(d.pop("_id"))
    for k, v in list(d.items()):
        if isinstance(v, ObjectId):
            d[k] = str(v)
    return d


def oid(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(400, "Invalid object ID format")


async def _get_user(request: Request):
    import server
    if getattr(server, "get_current_user", None) is not None:
        return await server.get_current_user(request)
    from auth import get_current_user_factory
    db = getattr(request.app, "mongodb", None) or server.db
    fn = await get_current_user_factory(db)
    return await fn(request)


async def log_activity_db(db, action: str, category: str, details: str, email: str):
    try:
        await db.audit_logs.insert_one({
            "action": action,
            "category": category,
            "details": details,
            "by": email,
            "created_at": now_iso()
        })
    except Exception:
        pass


def _norm_key(value: Optional[str]) -> str:
    """Canonical key for marketplace/order identities: trimmed, case-folded."""
    return (value or "").strip().casefold()


def _norm_marketplace(value: Optional[str]) -> str:
    """Canonical marketplace/source name used by unique-key guards."""
    return _norm_key(value)


def normalize_image_url(raw: str) -> str:
    """Rewrite common share-link formats (Dropbox / OneDrive / Google Drive) to direct-download URLs."""
    from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
    if not raw or not isinstance(raw, str):
        return raw
    val = raw.strip()
    if not val:
        return val
    try:
        parts = urlsplit(val)
    except Exception:
        return val
    host = (parts.hostname or "").lower()

    if host.endswith("dropbox.com") and host != "dl.dropboxusercontent.com":
        qs = [(k, v) for (k, v) in parse_qsl(parts.query, keep_blank_values=True) if k.lower() != "dl"]
        return urlunsplit((parts.scheme, "dl.dropboxusercontent.com", parts.path, urlencode(qs), ""))

    if "api.onedrive.com/v1.0/shares/u!" in val:
        m = re.search(r"/shares/u!([^/]+)", val)
        if m:
            b64_str = m.group(1)
            padding = 4 - (len(b64_str) % 4)
            if padding != 4:
                b64_str += "=" * padding
            try:
                import base64
                decoded = base64.urlsafe_b64decode(b64_str).decode("utf-8")
                return f"https://api.onedrive.com/v1.0/shares/u!{m.group(1)}/root/content"
            except Exception:
                pass

    if "drive.google.com" in host:
        m = re.search(r"/d/([a-zA-Z0-9_-]+)", val) or re.search(r"id=([a-zA-Z0-9_-]+)", val)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/uc?export=view&id={file_id}"

    return val


# ── split_leaf_sku helper ────────────────────────────────────────────
DEFAULT_SIZE_LABELS = ["XS", "S", "M", "L", "XL", "XXL", "Free Size", "FreeSize", "FS"]
NUMERIC_SIZE_RE     = re.compile(r"^\d{1,2}$")


def split_leaf_sku(
    leaf_sku: str,
    size_labels: Optional[List[str]] = None,
) -> Tuple[str, Optional[str], List[str]]:
    """Split leaf_sku into (group_id, size_token, flags)."""
    raw = (leaf_sku or "").strip()
    if not raw:
        return "", None, ["empty_leaf_sku"]

    labels = size_labels or DEFAULT_SIZE_LABELS
    labels_lc = {s.lower() for s in labels}

    last_dash  = raw.rfind("-")
    last_under = raw.rfind("_")
    cut = max(last_dash, last_under)
    if cut <= 0 or cut == len(raw) - 1:
        return raw, None, ["group_id_derivation_failed"]

    head = raw[:cut]
    tail = raw[cut + 1:]

    is_numeric_size = bool(NUMERIC_SIZE_RE.match(tail))
    is_label_size   = tail.lower() in labels_lc

    if is_numeric_size or is_label_size:
        return head, tail, []

    return raw, None, ["group_id_derivation_failed"]


async def _update_unmatched_jobs_for_sku_mapping(mapping_id: str, mapping_doc: dict, db=None):
    """Scan and resolve all 'unmatched' production jobs that can now be resolved by SKU mapping."""
    if db is None:
        import server
        db = server.db
    style_id = mapping_doc.get("style_id")
    style_code = mapping_doc.get("style_code")
    source_name = mapping_doc.get("source_name", "").strip()
    external_sku = mapping_doc.get("external_sku", "").strip()

    if not style_id or not style_code or not source_name or not external_sku:
        return

    jobs = await db.production_jobs.find({
        "style_match_status": "unmatched",
        "client_name": {"$regex": f"^{re.escape(source_name)}$", "$options": "i"},
        "style_code": {"$regex": f"^{re.escape(external_sku)}$", "$options": "i"},
    }).to_list(2000)

    if not jobs:
        return

    color_map = mapping_doc.get("color_map") or {}
    size_map = mapping_doc.get("size_map") or {}

    def translate(m: dict, val: str) -> str:
        if not val:
            return val
        if val in m:
            return m[val]
        val_lower = val.lower()
        for k, v in m.items():
            if k.lower() == val_lower:
                return v
        return val

    for j in jobs:
        ext_color = j.get("color") or ""
        ext_size = str(j.get("size") or "")

        translated_color = translate(color_map, ext_color)
        translated_size = translate(size_map, ext_size)

        now = now_iso()
        history_entry = {
            "event": "sku_mapped",
            "at": now,
            "by": "system_sku_mapper",
            "style_id": str(style_id),
            "style_code": style_code,
            "mapping_id": str(mapping_id),
            "color_translated": translated_color != ext_color,
            "size_translated": translated_size != ext_size,
        }

        upd = {
            "style_id": str(style_id),
            "style_code": style_code,
            "style_match_status": "matched",
            "mapped_from_sku": external_sku,
            "color": translated_color,
            "size": translated_size,
            "updated_at": now,
        }

        await db.production_jobs.update_one(
            {"_id": j["_id"]},
            {"$set": upd, "$push": {"history": history_entry}}
        )


async def resolve_style(
    source_type: str,
    source_name: str,
    external_sku: str,
    external_color: Optional[str] = None,
    external_size: Optional[str] = None,
    db=None,
) -> dict:
    """Canonical resolver: external SKU → internal style + translated color/size."""
    if db is None:
        import server
        db = server.db
    ext_sku    = (external_sku   or "").strip()
    ext_color  = (external_color or "").strip()
    ext_size   = (external_size  or "").strip()
    src_name   = (source_name    or "").strip()
    src_type   = (source_type    or "").strip()

    mapping = await db.sku_map.find_one({
        "source_type": src_type,
        "source_name_key": _norm_marketplace(src_name),
        "external_sku_key": _norm_key(ext_sku),
    })
    if not mapping:
        mapping = await db.sku_map.find_one({
            "source_type": src_type,
            "source_name": {"$regex": f"^{re.escape(src_name)}$", "$options": "i"},
            "external_sku": {"$regex": f"^{re.escape(ext_sku)}$",  "$options": "i"},
        })

    if mapping:
        style = await db.styles.find_one({"_id": ObjectId(mapping["style_id"])})
        if style:
            color_map: dict = mapping.get("color_map") or {}
            size_map:  dict = mapping.get("size_map")  or {}

            def translate(m: dict, val: str) -> tuple:
                val = (val or "").strip()
                if not m:
                    return val, True
                if val in m:
                    return m[val], True
                val_lower = val.lower()
                for k, v in m.items():
                    if k.lower() == val_lower:
                        return v, True
                return val, False

            resolved_color, color_exact = translate(color_map, ext_color)
            resolved_size,  size_exact  = translate(size_map,  ext_size)

            return {
                "style_id":            str(style["_id"]),
                "style_code":          style.get("code", ""),
                "color":               resolved_color,
                "size":                resolved_size,
                "matched":             True,
                "matched_exact":       color_exact and size_exact,
                "color_matched_exact": color_exact,
                "size_matched_exact":  size_exact,
                "unmapped_color":      ext_color if not color_exact else None,
                "unmapped_size":       ext_size  if not size_exact  else None,
                "match_via":           "sku_map",
                "mapping_id":          str(mapping.get("_id") or mapping.get("id") or ""),
                "mapped_from_sku":     ext_sku,
            }

    # Fallback: direct style code lookup
    style = await db.styles.find_one({
        "code": {"$regex": f"^{re.escape(ext_sku)}$", "$options": "i"}
    })
    if style:
        return {
            "style_id":            str(style["_id"]),
            "style_code":          style.get("code", ""),
            "color":               ext_color,
            "size":                ext_size,
            "matched":             True,
            "matched_exact":       True,
            "color_matched_exact": True,
            "size_matched_exact":  True,
            "unmapped_color":      None,
            "unmapped_size":       None,
            "match_via":           "style_code",
            "mapping_id":          None,
            "mapped_from_sku":     None,
        }

    return {
        "style_id":            None,
        "style_code":          None,
        "color":               ext_color,
        "size":                ext_size,
        "matched":             False,
        "matched_exact":       False,
        "color_matched_exact": False,
        "size_matched_exact":  False,
        "unmapped_color":      ext_color or None,
        "unmapped_size":       ext_size  or None,
        "match_via":           None,
        "mapping_id":          None,
        "mapped_from_sku":     None,
    }


# ── Marketplace SKU Resolver Engine ─────────────────────────────────────────

SUPPORTED_MARKETPLACES = ["myntra", "ajio", "flipkart", "nykaa", "amazon", "website", "unicommerce"]

DEFAULT_PARSER_PATTERNS = {
    "myntra":      r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "ajio":        r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "flipkart":    r"^(?P<style>.+?)[_-](?P<color>[A-Za-z]{1,4})[_-](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "nykaa":       r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "amazon":      r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "website":     r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
    "unicommerce": r"^(?P<style>.+?)[-_](?P<color>[A-Za-z]{1,4})[-_](?P<size>[0-9]{1,4}(?:\.[0-9]{1,2})?)$",
}


async def _seed_parser_templates(db=None):
    if db is None:
        import server
        db = server.db
    inserted = 0
    for mp in SUPPORTED_MARKETPLACES:
        exists = await db.sku_parser_templates.find_one({"marketplace": mp})
        if not exists:
            await db.sku_parser_templates.insert_one({
                "marketplace":  mp,
                "template":     "STYLE-COLOR-SIZE",
                "pattern":      DEFAULT_PARSER_PATTERNS[mp],
                "separator":    "-" if mp not in ("flipkart",) else "_",
                "active":       True,
                "example":      {"myntra": "CC-058-BR-38", "flipkart": "FL_AK_002_GO-4"}.get(mp, "STYLE-COLOR-SIZE"),
                "created_at":   now_iso(),
                "updated_at":   now_iso(),
            })
            inserted += 1
    return inserted


async def _get_parser_template(marketplace: str, db=None):
    if db is None:
        import server
        db = server.db
    tmpl = await db.sku_parser_templates.find_one({"marketplace": marketplace, "active": True})
    if tmpl:
        return tmpl
    return {
        "marketplace": marketplace,
        "pattern":     DEFAULT_PARSER_PATTERNS.get(marketplace, DEFAULT_PARSER_PATTERNS["myntra"]),
        "template":    "STYLE-COLOR-SIZE (default)",
    }


def _parse_sku(sku: str, pattern: str):
    if not sku:
        return {"ok": False, "error": "empty sku"}
    try:
        m = re.match(pattern, sku.strip())
    except re.error as e:
        return {"ok": False, "error": f"bad regex: {e}"}
    if not m:
        return {"ok": False, "error": f"pattern did not match: {pattern}"}
    d = m.groupdict()
    style = (d.get("style") or "").strip()
    color = (d.get("color") or "").strip()
    size  = (d.get("size")  or "").strip()
    if not (style and color and size):
        return {"ok": False, "error": "missing style/color/size group"}
    return {"ok": True, "style": style, "color": color, "size": size}


async def _resolve_marketplace_sku(marketplace: str, sku: str, db=None):
    if db is None:
        import server
        db = server.db
    tmpl = await _get_parser_template(marketplace, db=db)
    parsed = _parse_sku(sku, tmpl["pattern"])
    out = {
        "marketplace":     marketplace,
        "raw_sku":         sku,
        "template":        tmpl.get("template"),
        "pattern":         tmpl.get("pattern"),
        "parsed_ok":       parsed["ok"],
        "parsed_style":    parsed.get("style"),
        "parsed_color":    parsed.get("color"),
        "parsed_size":     parsed.get("size"),
        "parse_error":     parsed.get("error"),
        "erp_style_id":    None,
        "erp_style_code":  None,
        "erp_color_code":  None,
        "erp_size":        parsed.get("size"),
        "mapping_id":      None,
        "resolved":        False,
        "resolution_status": "parse_failed" if not parsed["ok"] else "unmapped",
    }
    if not parsed["ok"]:
        return out

    mapping = await db.marketplace_style_color_mapping.find_one({
        "marketplace_key":            _norm_marketplace(marketplace),
        "marketplace_style_code_key": _norm_key(parsed["style"]),
        "marketplace_color_code_key": _norm_key(parsed["color"]),
        "active":                    {"$ne": False},
    })
    if not mapping:
        mapping = await db.marketplace_style_color_mapping.find_one({
            "marketplace":            marketplace,
            "marketplace_style_code": {"$regex": f"^{re.escape(parsed['style'])}$", "$options": "i"},
            "marketplace_color_code": {"$regex": f"^{re.escape(parsed['color'])}$", "$options": "i"},
            "active":                 {"$ne": False},
        })
    if not mapping:
        out["resolution_status"] = "unmapped"
        return out

    style = await db.styles.find_one({
        "code": {"$regex": f"^{re.escape(mapping['erp_style_code'])}$", "$options": "i"},
    })
    if not style:
        out["resolution_status"] = "erp_style_missing"
        out["erp_style_code"] = mapping["erp_style_code"]
        out["erp_color_code"] = mapping["erp_color_code"]
        return out

    out["erp_style_id"]   = str(style["_id"])
    out["erp_style_code"] = style["code"]
    out["erp_color_code"] = mapping["erp_color_code"]
    out["mapping_id"]     = str(mapping["_id"])
    out["resolved"]       = True
    out["resolution_status"] = "resolved"
    return out


async def _log_unresolved_sku(marketplace: str, raw_sku: str, resolution: dict,
                              source: str = "import", order_id: Optional[str] = None, db=None):
    if db is None:
        import server
        db = server.db
    key = {
        "marketplace":            marketplace,
        "raw_sku":                raw_sku,
        "marketplace_style_code": resolution.get("parsed_style"),
        "marketplace_color_code": resolution.get("parsed_color"),
    }
    now = now_iso()
    upd = {
        "$set": {
            **key,
            "parsed_size":         resolution.get("parsed_size"),
            "resolution_status":   resolution.get("resolution_status"),
            "parse_error":         resolution.get("parse_error"),
            "last_source":         source,
            "last_seen_at":        now,
        },
        "$setOnInsert": {"created_at": now, "occurrences": 0, "status": "open"},
        "$inc":         {"occurrences": 1},
    }
    if order_id:
        upd["$addToSet"] = {"seen_order_ids": order_id}
    await db.unresolved_sku_queue.update_one(key, upd, upsert=True)


# ---------- SKU MAP ENDPOINTS ----------

@sku_map_router.get("/sku-map")
async def list_sku_map(
    request: Request,
    style_id: Optional[str] = None,
    source_type: Optional[str] = None,
    source_name: Optional[str] = None,
    search: Optional[str] = None,
    needs_style_code: Optional[bool] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    query: dict = {}
    if style_id:
        query["style_id"] = style_id
    if source_type:
        query["source_type"] = source_type
    if source_name:
        query["source_name"] = {"$regex": re.escape(source_name), "$options": "i"}
    if needs_style_code is True:
        query["needs_style_code"] = True
    elif needs_style_code is False:
        query["needs_style_code"] = {"$ne": True}
    if search:
        query["$or"] = [
            {"external_sku": {"$regex": re.escape(search), "$options": "i"}},
            {"external_style_name": {"$regex": re.escape(search), "$options": "i"}},
            {"source_name": {"$regex": re.escape(search), "$options": "i"}},
            {"style_code": {"$regex": re.escape(search), "$options": "i"}},
        ]
    docs = await db.sku_map.find(query).sort("created_at", -1).to_list(2000)
    return [stringify(d) for d in docs]


@sku_map_router.get("/sku-map/resolve")
async def resolve_sku_endpoint(
    source_type: str,
    source_name: str,
    external_sku: str,
    external_color: Optional[str] = None,
    external_size: Optional[str] = None,
    request: Request = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    return await resolve_style(
        source_type=source_type,
        source_name=source_name,
        external_sku=external_sku,
        external_color=external_color,
        external_size=external_size,
        db=db,
    )


@sku_map_router.post("/sku-map")
async def create_sku_map(payload: SkuMapIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    style = await db.styles.find_one({"_id": oid(payload.style_id)})
    if not style:
        raise HTTPException(404, f"Style '{payload.style_id}' not found")
    src_type = (payload.source_type or "").strip()
    src_name = (payload.source_name or "").strip()
    ext_sku = (payload.external_sku or "").strip()
    existing = await db.sku_map.find_one({
        "source_type": src_type,
        "source_name_key": _norm_marketplace(src_name),
        "external_sku_key": _norm_key(ext_sku),
    })
    if existing:
        raise HTTPException(409, f"A mapping for source '{src_name}' / SKU '{ext_sku}' already exists")
    doc = payload.model_dump()
    doc["source_type"] = src_type
    doc["source_name"] = src_name
    doc["external_sku"] = ext_sku
    doc["source_name_key"] = _norm_marketplace(src_name)
    doc["external_sku_key"] = _norm_key(ext_sku)
    doc["style_code"] = style["code"]
    doc["image_url"] = normalize_image_url(payload.image_url or "")
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    doc["created_by"] = u.get("email") or u.get("name", "")
    try:
        res = await db.sku_map.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(409, f"A mapping for source '{payload.source_name}' / SKU '{payload.external_sku}' already exists")
    await log_activity_db(db, "CREATE", "sku_map", f"Mapped {payload.external_sku} ({payload.source_name}) → {style['code']}", u.get("email") or u.get("name", ""))
    ret = dict(doc)
    ret.pop("_id", None)
    ret["id"] = str(res.inserted_id)
    await _update_unmatched_jobs_for_sku_mapping(res.inserted_id, doc, db=db)
    return ret


@sku_map_router.put("/sku-map/{mid}")
@sku_map_router.patch("/sku-map/{mid}")
async def update_sku_map(mid: str, payload: SkuMapUpdate, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    existing = await db.sku_map.find_one({"_id": oid(mid)})
    if not existing:
        raise HTTPException(404, "Mapping not found")
    update: dict = {"updated_at": now_iso()}
    if payload.external_style_name is not None:
        update["external_style_name"] = payload.external_style_name
    if payload.external_style_id is not None:
        update["external_style_id"] = payload.external_style_id
    if payload.color_map is not None:
        update["color_map"] = payload.color_map
    if payload.size_map is not None:
        update["size_map"] = payload.size_map
    if payload.image_url is not None:
        update["image_url"] = normalize_image_url(payload.image_url)
    await db.sku_map.update_one({"_id": oid(mid)}, {"$set": update})
    await log_activity_db(db, "UPDATE", "sku_map", f"Updated mapping {mid}", u.get("email") or u.get("name", ""))
    updated_doc = await db.sku_map.find_one({"_id": oid(mid)})
    if updated_doc:
        await _update_unmatched_jobs_for_sku_mapping(mid, updated_doc, db=db)
    return stringify(await db.sku_map.find_one({"_id": oid(mid)}))


@sku_map_router.delete("/sku-map/{mid}")
async def delete_sku_map(mid: str, request: Request):
    u = await _get_user(request)
    require_roles("admin")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    existing = await db.sku_map.find_one({"_id": oid(mid)})
    if not existing:
        raise HTTPException(404, "Mapping not found")
    await db.sku_map.delete_one({"_id": oid(mid)})
    await log_activity_db(db, "DELETE", "sku_map", f"Deleted mapping {mid} ({existing.get('source_name')} / {existing.get('external_sku')})", u.get("email") or u.get("name", ""))
    return {"ok": True}


@sku_map_router.get("/sku-map/template")
async def download_sku_map_template(format: str = "xlsx", request: Request = None):
    await _get_user(request)
    headers = [
        "style_code", "color", "size", "external_sku",
        "source_type", "source_name", "external_style_name", "image_url"
    ]
    sample_rows = [
        ["SSK-OXF-01", "Tan", "8 UK", "MYN-OXF-TAN-8", "online_channel", "myntra", "Classic Oxford Formal Shoes", "https://www.dropbox.com/s/sample/shoe.jpg?dl=0"],
        ["SSK-OXF-01", "Tan", "9 UK", "MYN-OXF-TAN-9", "online_channel", "myntra", "Classic Oxford Formal Shoes", ""],
        ["SSK-OXF-01", "Black", "8 UK", "MYN-OXF-BLK-8", "online_channel", "myntra", "Classic Oxford Formal Shoes", ""],
        ["SSK-MOC-02", "Navy", "7 UK", "BAT-MOC-NAV-7", "b2b_client", "Bata India Ltd", "Navy Suede Moccasin", ""],
    ]
    if format.lower() == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in sample_rows:
            w.writerow(r)
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="sku_mapping_template.csv"'}
        )
    else:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SKU Mapping Template"
        ws.append(headers)
        for r in sample_rows:
            ws.append(r)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="sku_mapping_template.xlsx"'}
        )


@sku_map_router.post("/sku-map/bulk", dependencies=[Depends(upload_rate_limiter)])
async def bulk_create_sku_map(
    file: UploadFile = File(...),
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")

    content = await file.read()
    filename = (file.filename or "").lower()
    raw_rows: list[dict] = []

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = None
            for row_cells in ws.iter_rows(values_only=True):
                if headers is None:
                    if any(c is not None for c in row_cells):
                        headers = [str(c or "").strip().lower().replace(" ", "_") for c in row_cells]
                    continue
                if all(c is None for c in row_cells):
                    continue
                raw_rows.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row_cells])))
        except Exception as exc:
            raise HTTPException(400, f"Could not read .xlsx file: {exc}")
    else:
        try:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                raw_rows.append({k.strip().lower().replace(" ", "_"): (v or "").strip()
                                 for k, v in row.items()})
        except Exception as exc:
            raise HTTPException(400, f"Could not read CSV file: {exc}")

    if not raw_rows:
        raise HTTPException(400, "File is empty or has no data rows")

    ALIAS_MAP = {
        "style": "style_code", "stylecode": "style_code", "internal_style": "style_code",
        "colour": "color", "color_name": "color", "uk_size": "size", "size_uk": "size",
        "sku": "external_sku", "client": "source_name", "marketplace": "source_name",
        "channel": "source_name", "type": "source_type", "photo": "image_url",
    }

    def clean_row(r: dict) -> dict:
        out = {}
        for k, v in r.items():
            norm_k = str(k or "").strip().lower().replace(" ", "_").replace("-", "_")
            canonical_k = ALIAS_MAP.get(norm_k, norm_k)
            val = str(v).strip() if v is not None else ""
            if canonical_k not in out or not out[canonical_k]:
                out[canonical_k] = val
        return out

    cleaned_rows = [clean_row(r) for r in raw_rows]
    all_styles = await db.styles.find({}, {"code": 1}).to_list(10000)
    style_map: dict[str, dict] = {s["code"].strip().upper(): {"id": str(s["_id"]), "code": s["code"]}
                                   for s in all_styles}

    REQUIRED = ("style_code", "color", "size", "external_sku", "source_type", "source_name")
    errors: list[dict] = []
    warnings: list[dict] = []
    groups: dict[tuple, dict] = {}

    for idx, row in enumerate(cleaned_rows, start=2):
        if not row.get("source_type") and row.get("source_name"):
            src_nm = row["source_name"].lower()
            if any(ch in src_nm for ch in ["myntra", "flipkart", "nykaa", "ajio", "amazon", "website"]):
                row["source_type"] = "online_channel"
            else:
                row["source_type"] = "b2b_client"

        missing = [f for f in REQUIRED if not row.get(f, "").strip()]
        if missing:
            errors.append({"row": idx, "reason": f"Missing required field(s): {', '.join(missing)}"})
            continue

        s_code    = row["style_code"].strip()
        color     = row["color"].strip()
        size      = row["size"].strip()
        ext_sku   = row["external_sku"].strip()
        src_type  = row["source_type"].strip()
        src_name  = row["source_name"].strip()
        ext_style = row.get("external_style_name", "").strip()
        img_url   = row.get("image_url", "").strip()

        style_entry = style_map.get(s_code.upper())
        if not style_entry:
            errors.append({"row": idx, "reason": f"style_code '{s_code}' not found in Style Master"})
            continue

        color_key = color.strip().casefold()
        gk = (s_code.upper(), src_type.casefold(), _norm_marketplace(src_name), color_key)

        if gk not in groups:
            groups[gk] = {
                "style_id":           style_entry["id"],
                "style_code":         style_entry["code"],
                "source_type":        src_type,
                "source_name":        src_name,
                "source_name_key":    _norm_marketplace(src_name),
                "color":              color,
                "color_key":          color_key,
                "external_style_name": ext_style,
                "image_url_raw":      img_url,
                "size_rows":          [],
                "image_url_rows":     [],
            }
        else:
            existing_img = groups[gk]["image_url_raw"]
            if img_url and existing_img and img_url != existing_img:
                warnings.append({
                    "row": idx,
                    "reason": f"Row has image_url '{img_url}' but group already has a different image_url '{existing_img}'",
                })
            elif img_url and not existing_img:
                groups[gk]["image_url_raw"] = img_url

        if img_url:
            groups[gk]["image_url_rows"].append((idx, img_url))
        groups[gk]["size_rows"].append((idx, size, ext_sku))

    created = 0
    updated = 0

    for gk, g in groups.items():
        new_size_map: dict[str, str] = {}
        for (_idx, sz, ext_sku) in g["size_rows"]:
            new_size_map[sz] = ext_sku

        raw_img = g["image_url_raw"] or ""
        norm_img = normalize_image_url(raw_img)

        existing = await db.sku_map.find_one({
            "source_type":     g["source_type"],
            "source_name_key": g["source_name_key"],
            "style_id":        g["style_id"],
            "color_key":       g["color_key"],
        })

        if existing:
            merged_size_map = {**(existing.get("size_map") or {}), **new_size_map}
            upd: dict = {
                "size_map":   merged_size_map,
                "updated_at": now_iso(),
                "updated_by": u.get("email") or u.get("name", ""),
            }
            if g.get("external_style_name"):
                upd["external_style_name"] = g["external_style_name"]
            if norm_img:
                upd["image_url"] = norm_img
            await db.sku_map.update_one({"_id": existing["_id"]}, {"$set": upd})
            updated_doc = {**existing, **upd}
            await _update_unmatched_jobs_for_sku_mapping(existing["_id"], updated_doc, db=db)
            updated += 1
        else:
            doc = {
                "style_id":            g["style_id"],
                "style_code":          g["style_code"],
                "source_type":         g["source_type"],
                "source_name":         g["source_name"],
                "source_name_key":     g["source_name_key"],
                "color":               g["color"],
                "color_key":           g["color_key"],
                "external_style_name": g["external_style_name"],
                "external_sku":        next(iter(new_size_map.values()), g["style_code"]),
                "size_map":            new_size_map,
                "color_map":           {},
                "image_url":           norm_img,
                "created_at":          now_iso(),
                "updated_at":          now_iso(),
                "created_by":          u.get("email") or u.get("name", ""),
            }
            try:
                res = await db.sku_map.insert_one(doc)
                await _update_unmatched_jobs_for_sku_mapping(res.inserted_id, doc, db=db)
                created += 1
            except DuplicateKeyError:
                existing2 = await db.sku_map.find_one({
                    "source_type":     g["source_type"],
                    "source_name_key": g["source_name_key"],
                    "style_id":        g["style_id"],
                    "color_key":       g["color_key"],
                })
                if existing2:
                    merged = {**(existing2.get("size_map") or {}), **new_size_map}
                    await db.sku_map.update_one(
                        {"_id": existing2["_id"]},
                        {"$set": {"size_map": merged, "updated_at": now_iso()}},
                    )
                    updated += 1

    await log_activity_db(
        db,
        "BULK_CREATE", "sku_map",
        f"Bulk import: {created} created, {updated} updated, {len(errors)} errors, {len(warnings)} warnings",
        u.get("email") or u.get("name", "")
    )
    return {
        "created":  created,
        "updated":  updated,
        "errors":   errors,
        "warnings": warnings,
    }


@sku_map_router.get("/sku-map/unmapped")
async def sku_map_unmapped(request: Request):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    jobs = await db.production_jobs.find({
        "archived":           {"$ne": True},
        "stage":              {"$ne": "dispatched"},
        "style_match_status": "unmatched",
    }).to_list(5000)

    groups: dict[tuple, dict] = {}
    for j in jobs:
        src_type = "b2b_client"
        src_name = j.get("client_name") or "(unknown)"
        key = (src_type, src_name)
        if key not in groups:
            groups[key] = {
                "source_type":  src_type,
                "source_name":  src_name,
                "job_count":    0,
                "external_skus": [],
                "jobs":         [],
            }
        g = groups[key]
        g["job_count"] += 1
        ext_sku = j.get("style_code") or "(blank)"
        if ext_sku not in g["external_skus"]:
            g["external_skus"].append(ext_sku)
        g["jobs"].append({
            "id":                  str(j["_id"]),
            "po_number":           j.get("po_number"),
            "style_code":          j.get("style_code"),
            "color":               j.get("color"),
            "size":                j.get("size"),
            "quantity":            j.get("quantity"),
            "stage":               j.get("stage"),
            "style_match_status":  j.get("style_match_status"),
            "created_at":          j.get("created_at"),
        })

    result = list(groups.values())
    result.sort(key=lambda g: -g["job_count"])
    return result


# ---------- LISTING IMPORT (Stage 1 parse → Stage 2 link) ----------

@sku_map_router.post("/sku-map/listing-import/parse", dependencies=[Depends(upload_rate_limiter)])
async def listing_import_parse(
    file: UploadFile = File(...),
    platform: str = "other",
    source_type: str = "online_channel",
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")

    content = await file.read()
    filename = (file.filename or "").lower()
    raw_rows: list[dict] = []

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = None
            for row_cells in ws.iter_rows(values_only=True):
                if headers is None:
                    if any(c is not None for c in row_cells):
                        headers = [str(c or "").strip().lower().replace(" ", "_") for c in row_cells]
                    continue
                if all(c is None for c in row_cells):
                    continue
                raw_rows.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row_cells])))
        except Exception as exc:
            raise HTTPException(400, f"Could not read .xlsx file: {exc}")
    else:
        try:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                raw_rows.append({k.strip().lower().replace(" ", "_"): (v or "").strip()
                                 for k, v in row.items()})
        except Exception as exc:
            raise HTTPException(400, f"Could not read CSV file: {exc}")

    if not raw_rows:
        raise HTTPException(400, "File is empty or has no data rows")

    ALIAS_MAP = {
        "style_id": "external_style_id", "myntra_style_id": "external_style_id",
        "style_name": "external_style_name", "product_name": "external_style_name",
        "color": "color", "colour": "color", "size": "size", "sku": "external_sku",
        "sellerskucode": "external_sku", "seller_sku_code": "external_sku",
        "image_url": "image_url", "image": "image_url",
    }

    def clean_row(r: dict) -> dict:
        out = {}
        for k, v in r.items():
            raw_key = str(k or "").strip().lower()
            norm_k = raw_key.replace(" ", "_").replace("-", "_")
            compact_k = raw_key.replace(" ", "").replace("_", "").replace("-", "")
            canonical_k = ALIAS_MAP.get(norm_k) or ALIAS_MAP.get(compact_k) or norm_k
            val = str(v).strip() if v is not None else ""
            if canonical_k not in out or not out[canonical_k]:
                out[canonical_k] = val
        return out

    cleaned_rows = [clean_row(r) for r in raw_rows]
    groups: dict[str, dict] = {}

    for row in cleaned_rows:
        ext_style_name = row.get("external_style_name", "").strip()
        ext_style_id   = row.get("external_style_id", "").strip()
        color_label    = row.get("color", "").strip()
        ext_sku        = row.get("external_sku", "").strip()
        size_label     = row.get("size", "").strip()
        image_url      = row.get("image_url", "").strip()

        if not size_label and ext_sku:
            _, derived_size, flags = split_leaf_sku(ext_sku)
            if derived_size and not flags:
                size_label = derived_size

        effective_size = size_label if size_label else ("One Size" if ext_sku else "")
        base_key = ext_style_id if ext_style_id else ext_style_name
        if not base_key:
            continue

        color_norm = color_label.strip().casefold()
        gk = f"{base_key}/{color_norm}"

        if gk not in groups:
            groups[gk] = {
                "group_key":            gk,
                "external_style_name":  ext_style_name,
                "external_style_id":    ext_style_id,
                "base_key":             base_key,
                "color_label":          color_label or color_norm,
                "platform":             platform,
                "source_type":          source_type,
                "source_name":          platform,
                "sku_count":            0,
                "sample_skus":          [],
                "size_sku_map":         {},
                "image_url":            image_url,
            }
        g = groups[gk]
        g["sku_count"] += 1
        if effective_size and ext_sku:
            g["size_sku_map"][effective_size] = ext_sku
        if ext_sku and ext_sku not in g["sample_skus"]:
            g["sample_skus"].append(ext_sku)
        if not g["image_url"] and image_url:
            g["image_url"] = image_url

    if not groups:
        raise HTTPException(400, "Could not identify any SKU groups from this file — check that it contains columns for style name/ID and colour.")

    for g in groups.values():
        g["sample_skus"] = g["sample_skus"][:5]

    from collections import defaultdict
    base_to_groups: dict[str, list] = defaultdict(list)
    name_to_groups: dict[str, list] = defaultdict(list)
    for gk, g in groups.items():
        base_to_groups[g["base_key"]].append(gk)
        if g.get("external_style_name"):
            name_to_groups[g["external_style_name"].strip().casefold()].append(gk)

    for gk, g in groups.items():
        sibs = set(base_to_groups[g["base_key"]])
        if g.get("external_style_name"):
            sibs.update(name_to_groups[g["external_style_name"].strip().casefold()])
        sibs.discard(gk)
        g["sibling_group_keys"] = sorted(list(sibs))

    all_styles = await db.styles.find({}, {"code": 1, "name": 1}).to_list(10000)
    style_codes_upper = {s["code"].upper(): {"id": str(s["_id"]), "code": s["code"], "name": s.get("name", "")}
                         for s in all_styles}
    all_codes_list = list(style_codes_upper.keys())

    def suggest_styles(base: str) -> list[str]:
        if not base:
            return []
        clean = re.sub(r'[^A-Za-z0-9]', '', base).upper()
        matches = difflib.get_close_matches(clean, all_codes_list, n=3, cutoff=0.4)
        return [style_codes_upper[m]["id"] for m in matches]

    groups_list = []
    for gk, g in groups.items():
        suggested = suggest_styles(g["base_key"])
        suggested_base_match = None
        if suggested:
            top_id = suggested[0]
            for s in all_styles:
                if str(s["_id"]) == top_id:
                    suggested_base_match = s["code"]
                    break
        g["suggested_style_ids"] = suggested
        g["suggested_base_match"] = suggested_base_match
        groups_list.append(g)

    session_id = str(uuid.uuid4())
    session_doc = {
        "session_id":   session_id,
        "filename":     file.filename,
        "platform":     platform,
        "source_type":  source_type,
        "status":       "pending_link",
        "groups":       groups_list,
        "group_count":  len(groups_list),
        "sku_count":    sum(g["sku_count"] for g in groups_list),
        "created_at":   now_iso(),
        "created_by":   u.get("email") or u.get("name", ""),
        "committed_at": None,
    }
    await db.listing_import_sessions.insert_one(session_doc)

    return {
        "session_id":  session_id,
        "filename":    file.filename,
        "platform":    platform,
        "group_count": len(groups_list),
        "sku_count":   session_doc["sku_count"],
        "groups":      groups_list,
    }


@sku_map_router.get("/sku-map/listing-import/sessions")
async def listing_import_list_sessions(request: Request):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    docs = await db.listing_import_sessions.find(
        {}, {"groups": 0}
    ).sort("created_at", -1).to_list(200)
    return [stringify(d) for d in docs]


@sku_map_router.get("/sku-map/listing-import/sessions/{session_id}")
async def listing_import_get_session(session_id: str, request: Request):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    doc = await db.listing_import_sessions.find_one({"session_id": session_id})
    if not doc:
        raise HTTPException(404, f"No listing import session '{session_id}'")
    return stringify(doc)


@sku_map_router.post("/sku-map/listing-import/sessions/{session_id}/commit")
async def listing_import_commit(
    session_id: str,
    payload: ListingImportCommitIn,
    request: Request,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")

    session_doc = await db.listing_import_sessions.find_one({"session_id": session_id})
    if not session_doc:
        raise HTTPException(404, f"No listing import session '{session_id}'")
    if session_doc.get("status") == "committed":
        raise HTTPException(409, "This import session has already been committed.")

    groups_by_key: dict[str, dict] = {g["group_key"]: g for g in (session_doc.get("groups") or [])}
    linked_count = 0
    unlinked_count = 0
    errors: list[dict] = []

    for decision in payload.decisions:
        gk = decision.group_key
        group = groups_by_key.get(gk)
        if not group:
            errors.append({"group_key": gk, "reason": "Group key not found in session"})
            continue

        style_id = (decision.style_id or "").strip() or None
        style_code = None
        style_doc = None

        if style_id:
            try:
                style_doc = await db.styles.find_one({"_id": oid(style_id)})
            except Exception:
                style_doc = None
            if not style_doc:
                errors.append({"group_key": gk, "reason": f"style_id '{style_id}' not found"})
                continue
            style_code = style_doc.get("code", "")

        size_sku_map = group.get("size_sku_map", {})
        color_label  = group.get("color_label", "")
        color_key    = color_label.strip().casefold()
        source_type  = group.get("source_type", "online_channel")
        source_name  = group.get("source_name", group.get("platform", ""))
        src_name_key = _norm_marketplace(source_name)

        sample_skus = group.get("sample_skus", [])
        group_ext_sku = next(iter(size_sku_map.values()), None) or (sample_skus[0] if sample_skus else "")

        if style_id:
            existing = await db.sku_map.find_one({
                "source_type":     source_type,
                "source_name_key": src_name_key,
                "style_id":        style_id,
                "color_key":       color_key,
            })
        else:
            existing = await db.sku_map.find_one({
                "source_type":              source_type,
                "source_name_key":          src_name_key,
                "needs_style_code":         True,
                "external_style_name_key":  _norm_key(group.get("external_style_name", group.get("base_key", gk))),
                "color_key":                color_key,
            })

        doc_base = {
            "source_type":        source_type,
            "source_name":        source_name,
            "source_name_key":    src_name_key,
            "color":              color_label,
            "color_key":          color_key,
            "external_style_name": group.get("external_style_name", ""),
            "external_style_id":  group.get("external_style_id", ""),
            "external_style_name_key": _norm_key(group.get("external_style_name", group.get("base_key", gk))),
            "external_sku":       group_ext_sku,
            "external_sku_key":   _norm_key(group_ext_sku),
            "size_map":           size_sku_map,
            "color_map":          {},
            "image_url":          normalize_image_url(group.get("image_url", "")),
            "needs_style_code":   style_id is None,
            "listing_session_id": session_id,
            "updated_at":         now_iso(),
            "updated_by":         u.get("email") or u.get("name", ""),
        }
        if style_id:
            doc_base["style_id"]   = style_id
            doc_base["style_code"] = style_code

        if existing:
            merged_size_map = dict(existing.get("size_map") or {})
            merged_size_map.update(size_sku_map)
            doc_base["size_map"] = merged_size_map
            await db.sku_map.update_one(
                {"_id": existing["_id"]},
                {"$set": doc_base},
            )
        else:
            doc_base["created_at"] = now_iso()
            doc_base["created_by"] = u.get("email") or u.get("name", ""),
            await db.sku_map.insert_one(doc_base)

        if style_id:
            linked_count += 1
        else:
            unlinked_count += 1

        label = f"{source_name} / {group.get('external_style_name', gk)} / {color_label}"
        if style_id:
            await log_activity_db(db, "CREATE", "sku_map", f"Listing import: linked {label} → {style_code}", u.get("email") or u.get("name", ""))
        else:
            await log_activity_db(db, "CREATE", "sku_map", f"Listing import: unlinked {label} (needs_style_code)", u.get("email") or u.get("name", ""))

    await db.listing_import_sessions.update_one(
        {"session_id": session_id},
        {"$set": {"status": "committed", "committed_at": now_iso(), "committed_by": u.get("email") or u.get("name", "")}},
    )

    return {
        "linked":   linked_count,
        "unlinked": unlinked_count,
        "errors":   errors,
    }


# ───────────── Parser Templates Endpoints ─────────────

@sku_map_router.get("/marketplace/parser-templates")
async def list_parser_templates(request: Request):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    docs = await db.sku_parser_templates.find({}).sort("marketplace", 1).to_list(50)
    return [stringify(d) for d in docs]


@sku_map_router.post("/marketplace/parser-templates")
async def upsert_parser_template(request: Request, payload: ParserTemplateIn):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    try:
        cp = re.compile(payload.pattern)
    except re.error as e:
        raise HTTPException(400, f"Invalid regex: {e}")
    if not all(g in cp.groupindex for g in ("style", "color", "size")):
        raise HTTPException(400, "Pattern must contain named groups (?P<style>), (?P<color>), (?P<size>)")
    now = now_iso()
    doc = {
        "marketplace": payload.marketplace,
        "template":    payload.template,
        "pattern":     payload.pattern,
        "separator":   payload.separator,
        "active":      payload.active,
        "example":     payload.example,
        "updated_at":  now,
    }
    await db.sku_parser_templates.update_one(
        {"marketplace": payload.marketplace},
        {"$set": doc, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )
    ret = await db.sku_parser_templates.find_one({"marketplace": payload.marketplace})
    await log_activity_db(db, "UPDATE", "sku_parser_templates", f"Set parser for {payload.marketplace}", u.get("email") or u.get("name", ""))
    return stringify(ret)


@sku_map_router.delete("/marketplace/parser-templates/{tid}")
async def delete_parser_template(request: Request, tid: str):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    try:
        await db.sku_parser_templates.delete_one({"_id": ObjectId(tid)})
    except Exception:
        raise HTTPException(404, "Template not found")
    return {"ok": True}


# ───────────── Style-Color Mapping Endpoints ─────────────

@sku_map_router.get("/marketplace/style-color-mapping")
async def list_marketplace_mappings(
    request: Request,
    marketplace: Optional[str] = None,
    search:      Optional[str] = None,
    active:      Optional[bool] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    q = {}
    if marketplace: q["marketplace"] = str(marketplace)
    if active is not None: q["active"] = bool(active)
    if search:
        rx = {"$regex": re.escape(str(search)), "$options": "i"}
        q["$or"] = [
            {"marketplace_style_code": rx},
            {"marketplace_color_code": rx},
            {"erp_style_code":         rx},
            {"erp_color_code":         rx},
        ]
    docs = await db.marketplace_style_color_mapping.find(q).sort("marketplace", 1).to_list(2000)
    return [stringify(d) for d in docs]


@sku_map_router.post("/marketplace/style-color-mapping")
async def upsert_marketplace_mapping(request: Request, payload: StyleColorMappingIn):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    style = await db.styles.find_one({
        "code": {"$regex": f"^{re.escape(payload.erp_style_code)}$", "$options": "i"},
    })
    if not style:
        raise HTTPException(400, f"ERP style '{payload.erp_style_code}' not found. Create it first.")

    marketplace = (payload.marketplace or "").strip().lower()
    mp_style = (payload.marketplace_style_code or "").strip()
    mp_color = (payload.marketplace_color_code or "").strip()
    key = {
        "marketplace_key":            _norm_marketplace(marketplace),
        "marketplace_style_code_key": _norm_key(mp_style),
        "marketplace_color_code_key": _norm_key(mp_color),
    }
    now = now_iso()
    doc = {
        **key,
        "erp_style_code": style["code"],
        "erp_style_id":   str(style["_id"]),
        "erp_color_code": payload.erp_color_code,
        "active":         payload.active,
        "updated_at":     now,
        "updated_by":     u.get("email") or u.get("name", ""),
    }
    await db.marketplace_style_color_mapping.update_one(
        key, {"$set": doc, "$setOnInsert": {"created_at": now}}, upsert=True,
    )
    ret = await db.marketplace_style_color_mapping.find_one(key)
    await log_activity_db(db, "UPDATE", "marketplace_style_color_mapping",
                        f"{payload.marketplace}: {payload.marketplace_style_code}/{payload.marketplace_color_code} → {style['code']}/{payload.erp_color_code}",
                        u.get("email") or u.get("name", ""))
    return stringify(ret)


@sku_map_router.delete("/marketplace/style-color-mapping/{mid}")
async def delete_marketplace_mapping(request: Request, mid: str):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    try:
        res = await db.marketplace_style_color_mapping.delete_one({"_id": ObjectId(mid)})
    except Exception:
        raise HTTPException(404, "Mapping not found")
    if not res.deleted_count:
        raise HTTPException(404, "Mapping not found")
    return {"ok": True}


# ───────────── Parse / Resolve Endpoints ─────────────

@sku_map_router.post("/marketplace/parse-sku")
async def parse_sku_endpoint(request: Request, payload: SkuResolveIn):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    return await _resolve_marketplace_sku(payload.marketplace, payload.sku, db=db)


@sku_map_router.post("/marketplace/parse-batch")
async def parse_sku_batch(request: Request, payload: dict):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    marketplace = payload.get("marketplace")
    skus = payload.get("skus") or []
    if not marketplace or not isinstance(skus, list):
        raise HTTPException(400, "Provide {marketplace, skus: [str]}")
    out = []
    for sku in skus[:1000]:
        out.append(await _resolve_marketplace_sku(marketplace, sku, db=db))
    return {"count": len(out), "results": out}


# ───────────── Unresolved SKU Queue ─────────────

@sku_map_router.get("/marketplace/unresolved")
async def list_unresolved(
    request: Request,
    marketplace: Optional[str] = None,
    status: Optional[str] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    q = {}
    if marketplace: q["marketplace"] = marketplace
    q["status"] = status or "open"
    docs = await db.unresolved_sku_queue.find(q).sort([("occurrences", -1), ("last_seen_at", -1)]).to_list(1000)
    return [stringify(d) for d in docs]


@sku_map_router.post("/marketplace/unresolved/map")
async def map_and_replay_unresolved(request: Request, payload: UnresolvedMapIn):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")

    style = await db.styles.find_one({
        "code": {"$regex": f"^{re.escape(payload.erp_style_code)}$", "$options": "i"},
    })
    if not style:
        raise HTTPException(400, f"ERP style '{payload.erp_style_code}' not found. Create it first.")

    marketplace = (payload.marketplace or "").strip().lower()
    mp_style = (payload.marketplace_style_code or "").strip()
    mp_color = (payload.marketplace_color_code or "").strip()
    key = {
        "marketplace_key":            _norm_marketplace(marketplace),
        "marketplace_style_code_key": _norm_key(mp_style),
        "marketplace_color_code_key": _norm_key(mp_color),
    }
    now = now_iso()
    await db.marketplace_style_color_mapping.update_one(
        key,
        {"$set": {
            **key,
            "erp_style_code": style["code"],
            "erp_style_id":   str(style["_id"]),
            "erp_color_code": payload.erp_color_code,
            "active":         True,
            "updated_at":     now,
            "updated_by":     u.get("email") or u.get("name", ""),
        }, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )

    close_res = await db.unresolved_sku_queue.update_many(
        {**key, "status": "open"},
        {"$set": {"status": "mapped", "mapped_at": now, "mapped_by": u.get("email") or u.get("name", "")}},
    )

    await log_activity_db(db, "UPDATE", "unresolved_sku_queue",
                        f"Mapped {payload.marketplace}:{payload.marketplace_style_code}/{payload.marketplace_color_code} → {style['code']}/{payload.erp_color_code} (closed {close_res.modified_count} queue rows)",
                        u.get("email") or u.get("name", ""))
    return {
        "ok":            True,
        "mapping_key":   key,
        "closed_queue_rows": close_res.modified_count,
    }


@sku_map_router.delete("/marketplace/unresolved/{qid}")
async def dismiss_unresolved(request: Request, qid: str):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    try:
        res = await db.unresolved_sku_queue.update_one(
            {"_id": ObjectId(qid)},
            {"$set": {"status": "dismissed", "dismissed_at": now_iso(), "dismissed_by": u.get("email") or u.get("name", "")}},
        )
    except Exception:
        raise HTTPException(404, "Queue item not found")
    if not res.matched_count:
        raise HTTPException(404, "Queue item not found")
    return {"ok": True}
