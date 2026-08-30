"""Style Master, BOM Engineering, Costing, Online Lifecycle, and Color Master Routes."""

import io
import os
import re
import json
import logging
import inspect
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any, Tuple
from collections import defaultdict

from fastapi import APIRouter, HTTPException, Request, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pydantic import BaseModel, Field, field_validator
from pydantic_core import PydanticCustomError
from pymongo.errors import DuplicateKeyError
from pymongo import ReturnDocument

from auth import get_current_user_factory, require_roles
from models.styles import (
    StyleIn,
    PlannedComponent,
    StyleLifecycleUpsert,
    OnlineStatusPatchIn,
    ColorMasterIn,
    ColorMasterUpdate,
    ONLINE_STATUS_SEQUENCE,
    ONLINE_STATUS_SIDE_BRANCHES,
    PLANNED_COMPONENTS,
)
from models.plm import DEFAULT_PLM_FOLDERS
from rate_limiter import upload_rate_limiter

log = logging.getLogger("styles_routes")

styles_router = APIRouter(prefix="/api", tags=["Style Master & BOM Engineering"])


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stringify(doc: dict) -> dict:
    if not doc:
        return {}
    d = dict(doc)
    if "_id" in d:
        d["id"] = str(d.pop("_id"))
    for k, v in list(d.items()):
        if isinstance(v, ObjectId):
            d[k] = str(v)
    return d


def oid(val: Any) -> ObjectId:
    if isinstance(val, ObjectId):
        return val
    try:
        return ObjectId(str(val))
    except Exception:
        raise HTTPException(400, f"Invalid ObjectId '{val}'")


def get_db():
    import server
    return getattr(server, "db", None)


async def _get_user(request: Optional[Request] = None):
    if request is not None:
        user = getattr(request.state, "user", None)
        if user:
            return user
    import server
    if getattr(server, "get_current_user", None) is not None:
        return await server.get_current_user(request)
    from auth import get_current_user_factory
    fn = await get_current_user_factory(get_db())
    return await fn(request)


async def log_activity(action: str, category: str, details: str, email: str = "system"):
    try:
        db = get_db()
        if db is not None and hasattr(db, "audit_logs"):
            res = db.audit_logs.insert_one({
                "action": action,
                "category": category,
                "details": details,
                "by": email,
                "created_at": now_iso(),
            })
            if inspect.isawaitable(res):
                await res
    except Exception as e:
        log.warning(f"Failed to write audit log: {e}")


def normalize_image_url(raw: str) -> str:
    """Rewrite common share-link formats (Dropbox / OneDrive / Google Drive) to direct-download URLs."""
    from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
    if not raw or not isinstance(raw, str):
        return raw
    val = raw.strip()
    if not val:
        return val
    try:
        parts = urlsplit(val)
    except Exception:
        return val
    host = (parts.hostname or "").lower()

    if host.endswith("dropbox.com") and host != "dl.dropboxusercontent.com":
        qs = [(k, v) for (k, v) in parse_qsl(parts.query, keep_blank_values=True) if k.lower() != "dl"]
        return urlunsplit((parts.scheme, "dl.dropboxusercontent.com", parts.path, urlencode(qs), ""))

    if "api.onedrive.com/v1.0/shares/u!" in val:
        m = re.search(r"/shares/u!([^/]+)", val)
        if m:
            b64_str = m.group(1)
            padding = 4 - (len(b64_str) % 4)
            if padding != 4:
                b64_str += "=" * padding
            try:
                import base64
                decoded = base64.urlsafe_b64decode(b64_str).decode("utf-8")
                return f"https://api.onedrive.com/v1.0/shares/u!{m.group(1)}/root/content"
            except Exception:
                pass

    if "drive.google.com" in host:
        m = re.search(r"/d/([a-zA-Z0-9_-]+)", val) or re.search(r"id=([a-zA-Z0-9_-]+)", val)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/uc?export=view&id={file_id}"

    return val


# ── Style Code Generation (Phase A0) ─────────────────────────────────────────
STYLE_CODE_PREFIX = "SSK_"
STYLE_CODE_PAD = 5
STYLE_CODE_RE = re.compile(rf"^{re.escape(STYLE_CODE_PREFIX)}\d{{{STYLE_CODE_PAD},}}$")


async def _next_style_code() -> str:
    """Atomically increment the style_code counter and return the next code."""
    db = get_db()
    doc = await db.counters.find_one_and_update(
        {"_id": "style_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = int(doc.get("seq", 1))
    return f"{STYLE_CODE_PREFIX}{seq:0{STYLE_CODE_PAD}d}"


def build_catalogue_sku(style_code: str, color_code: str, size: Optional[str] = None) -> str:
    """Canonical catalogue-SKU builder — MUST be used everywhere a catalogue SKU needs to be produced."""
    sc = (style_code or "").strip()
    cc = (color_code or "").strip().upper()
    if not sc or not cc:
        return ""
    if size is None or str(size).strip() == "":
        return f"{sc}-{cc}"
    return f"{sc}-{cc}-{str(size).strip()}"


# ── Footwear GST Rate Threshold Configuration ────────────────────────────────
FOOTWEAR_GST_CONFIG = {
    "threshold": 2500.0,
    "rate_below_or_equal": 5.0,
    "rate_above": 18.0,
}


def suggest_gst_pct(price: Optional[float]) -> float:
    """Return suggested GST % based on price threshold config."""
    if price is None:
        return FOOTWEAR_GST_CONFIG["rate_below_or_equal"]
    try:
        p = float(price)
        if p <= 0:
            return FOOTWEAR_GST_CONFIG["rate_below_or_equal"]
        return FOOTWEAR_GST_CONFIG["rate_above"] if p > FOOTWEAR_GST_CONFIG["threshold"] else FOOTWEAR_GST_CONFIG["rate_below_or_equal"]
    except (ValueError, TypeError):
        return FOOTWEAR_GST_CONFIG["rate_below_or_equal"]


@styles_router.get("/config/gst")
async def get_gst_config():
    """Return current footwear GST rate threshold configuration."""
    return FOOTWEAR_GST_CONFIG


# ── Style Costing Engine ─────────────────────────────────────────────────────
def compute_style_costing(style: dict) -> dict:
    materials_cost = 0.0
    for b in style.get("bom", []):
        rate = float(b.get("rate", 0))
        qty = float(b.get("quantity", 0))
        raw_yld = b.get("yield_per_unit")
        def_yld = b.get("default_yield_per_unit")
        if raw_yld is not None and float(raw_yld) > 0:
            yld = float(raw_yld)
        elif def_yld is not None and float(def_yld) > 0:
            yld = float(def_yld)
        else:
            yld = 1.0
        waste = float(b.get("waste_pct", 0) or 0)
        materials_cost += (rate * qty / yld) * (1 + waste / 100)
    labor_items = style.get("labor", [])
    labor_cost = sum(float(l.get("rate", 0)) for l in labor_items)
    labor_is_set = len(labor_items) > 0
    base_cost = materials_cost + labor_cost
    overhead_cost = base_cost * (style.get("overhead_pct", 0) / 100)
    packing = style.get("packing_cost", 0)
    total_cost = base_cost + overhead_cost + packing
    margin_pct = style.get("margin_pct", 0)
    suggested_margin_amount = total_cost * (margin_pct / 100)
    suggested_target_price = total_cost + suggested_margin_amount
    raw_gst_pct = style.get("gst_pct")
    gst_pct = float(raw_gst_pct) if raw_gst_pct is not None else suggest_gst_pct(suggested_target_price)
    gst_amount = suggested_target_price * (gst_pct / 100)
    suggested_target_price_with_gst = suggested_target_price + gst_amount
    return {
        "materials_cost": round(materials_cost, 2),
        "labor_cost": round(labor_cost, 2),
        "labor_is_set": labor_is_set,
        "overhead_cost": round(overhead_cost, 2),
        "packing_cost": round(packing, 2),
        "total_cost": round(total_cost, 2),
        "suggested_margin_amount": round(suggested_margin_amount, 2),
        "suggested_target_price": round(suggested_target_price, 2),
        "gst_amount": round(gst_amount, 2),
        "suggested_target_price_with_gst": round(suggested_target_price_with_gst, 2),
        "selling_price": round(suggested_target_price, 2),
        "final_price": round(suggested_target_price_with_gst, 2),
        "margin_amount": round(suggested_margin_amount, 2),
    }


def compute_style_costing_from_jobs(style: dict, jobs: list) -> dict:
    """Computes style costing given a list of pre-fetched production jobs."""
    c = compute_style_costing(style)
    c["planned_labor_cost"] = c["labor_cost"]
    c["labor_source"] = "estimated"
    c["is_assigned"] = False
    c["assigned_roles"] = []

    try:
        assigned_roles_map = {}
        job_rates = []

        for job in jobs:
            assignments = job.get("assignments") or {}
            if isinstance(assignments, dict) and assignments:
                total_job_rate = 0.0
                has_valid = False
                for role, asgn in assignments.items():
                    if isinstance(asgn, dict):
                        rate = float(asgn.get("rate_per_pair") or 0)
                        if rate > 0:
                            total_job_rate += rate
                            has_valid = True
                            if role not in assigned_roles_map:
                                assigned_roles_map[role] = {
                                    "role": role,
                                    "worker_id": asgn.get("worker_id", ""),
                                    "worker_name": asgn.get("worker_name", ""),
                                    "rate_per_pair": round(rate, 2),
                                }
                if has_valid:
                    job_rates.append(total_job_rate)

        if job_rates:
            actual_labor_cost = round(sum(job_rates) / len(job_rates), 2)
            c["actual_labor_cost"] = actual_labor_cost
            c["labor_cost"] = actual_labor_cost
            c["labor_source"] = "actual"
            c["is_assigned"] = True
            c["assigned_roles"] = list(assigned_roles_map.values())

            materials_cost = c["materials_cost"]
            base_cost = materials_cost + actual_labor_cost
            overhead_pct = float(style.get("overhead_pct", 0) or 0)
            overhead_cost = base_cost * (overhead_pct / 100)
            packing = c["packing_cost"]
            total_cost = base_cost + overhead_cost + packing

            margin_pct = float(style.get("margin_pct", 0) or 0)
            suggested_margin_amount = total_cost * (margin_pct / 100)
            suggested_target_price = total_cost + suggested_margin_amount
            gst_pct = float(style.get("gst_pct", 0) or 0)
            gst_amount = suggested_target_price * (gst_pct / 100)
            suggested_target_price_with_gst = suggested_target_price + gst_amount

            c["overhead_cost"] = round(overhead_cost, 2)
            c["total_cost"] = round(total_cost, 2)
            c["suggested_margin_amount"] = round(suggested_margin_amount, 2)
            c["suggested_target_price"] = round(suggested_target_price, 2)
            c["gst_amount"] = round(gst_amount, 2)
            c["suggested_target_price_with_gst"] = round(suggested_target_price_with_gst, 2)
            c["selling_price"] = round(suggested_target_price, 2)
            c["final_price"] = round(suggested_target_price_with_gst, 2)
            c["margin_amount"] = round(suggested_margin_amount, 2)
    except Exception:
        pass

    return c


async def compute_style_costing_async(style: dict, db=None) -> dict:
    """Computes style costing, automatically incorporating real worker assignment rates."""
    if db is None:
        db = get_db()
    try:
        style_id = str(style.get("_id", style.get("id", "")))
        style_code = style.get("code", "")

        if not style_id and not style_code:
            return compute_style_costing_from_jobs(style, [])

        query_conditions = []
        if style_id:
            query_conditions.append({"style_id": style_id})
        if style_code:
            query_conditions.append({"style_code": style_code})

        fetched_jobs = await db.production_jobs.find(
            {"$or": query_conditions} if len(query_conditions) > 1 else query_conditions[0]
        ).to_list(500)
        return compute_style_costing_from_jobs(style, fetched_jobs)
    except Exception:
        return compute_style_costing_from_jobs(style, [])


async def compute_po_profitability(po_line: dict, style_obj: dict, db=None) -> dict:
    """Compute real profit for a PO line: unit_price (negotiated) minus actual costs."""
    if db is None:
        db = get_db()
    c = await compute_style_costing_async(style_obj, db)
    bom_cost = float(c.get("materials_cost", 0))
    overhead = float(c.get("overhead_cost", 0))
    packing = float(c.get("packing_cost", 0) or style_obj.get("packing_cost", 0) or 0)
    labor_cost = float(c.get("labor_cost", 0))
    labor_source = c.get("labor_source", "estimated")

    unit_price = float(po_line.get("unit_price", 0))
    total_cost = round(bom_cost + labor_cost + packing, 2)
    profit = round(unit_price - total_cost, 2) if unit_price > 0 else None
    profit_pct = round(profit / unit_price * 100, 1) if (profit is not None and unit_price > 0) else None

    return {
        "style_code": style_obj.get("code", "") or po_line.get("style_code", ""),
        "unit_price": round(unit_price, 2),
        "bom_cost": round(bom_cost, 2),
        "labor_cost": round(labor_cost, 2),
        "labor_source": labor_source,
        "is_estimated": labor_source == "estimated",
        "overhead_cost": round(overhead, 2),
        "packing_cost": round(packing, 2),
        "total_cost": total_cost,
        "profit": profit,
        "profit_pct": profit_pct,
    }


# ── Style Lifecycle Helpers ──────────────────────────────────────────────────
def _default_lifecycle(style_id: str, style_code: str) -> dict:
    now = now_iso()
    return {
        "style_id":                style_id,
        "style_code":              style_code,
        "online_status":           "draft",
        "online_status_history":   [{
            "status":     "draft",
            "changed_at": now,
            "by":         "system",
            "notes":      "Auto-initialised on first read",
        }],
        "sale_channels":           [],
        "mrp":                     None,
        "online_selling_price":    None,
        "platform_commission_pct": {},
        "planned_min_stock":       25,
        "planned_components":      [{"component": c, "planned_qty": 0} for c in PLANNED_COMPONENTS],
        "planned_colors":          [],
        "planned_sizes":           [],
        "sole_mould_name":         "",
        "sole_shape":              "",
        "pattern_number":          "",
        "photoshoot_link":         "",
        "catalogue_link":          "",
        "back_track_number":       "",
        "went_live_at":            None,
        "created_at":              now,
        "updated_at":              now,
    }


async def _get_or_create_lifecycle(style_id: str) -> dict:
    """Look up the lifecycle doc for a style. If missing, insert a default (draft) one."""
    db = get_db()
    style = await db.styles.find_one({"_id": oid(style_id)})
    if not style:
        raise HTTPException(404, f"Style '{style_id}' not found")
    doc = await db.style_lifecycle.find_one({"style_id": str(style["_id"])})
    if doc:
        return doc
    doc = _default_lifecycle(str(style["_id"]), style["code"])
    try:
        res = await db.style_lifecycle.insert_one(doc)
        doc["_id"] = res.inserted_id
    except DuplicateKeyError:
        doc = await db.style_lifecycle.find_one({"style_id": str(style["_id"])})
    return doc


async def _generate_back_track_number(style_code: str) -> str:
    """Return '{style_code}-{YYYYMMDD}-{seq}' where seq is the next per-(code,date) counter."""
    db = get_db()
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    prefix = f"{style_code}-{today}-"
    existing = await db.style_lifecycle.count_documents({
        "back_track_number": {"$regex": f"^{re.escape(prefix)}"}
    })
    return f"{prefix}{existing + 1:03d}"


def _validate_online_status_transition(current: str, to_status: str):
    """Raises 400 if the transition is not allowed."""
    if to_status not in (ONLINE_STATUS_SEQUENCE + list(ONLINE_STATUS_SIDE_BRANCHES)):
        raise HTTPException(400, f"Unknown online_status '{to_status}'")
    if to_status in ONLINE_STATUS_SIDE_BRANCHES:
        return
    if current == to_status:
        return
    if current in ONLINE_STATUS_SIDE_BRANCHES:
        raise HTTPException(400,
            f"Cannot transition from side-branch '{current}' back into the pipeline. "
            f"Un-archive is not supported.")
    try:
        cur_idx = ONLINE_STATUS_SEQUENCE.index(current)
        new_idx = ONLINE_STATUS_SEQUENCE.index(to_status)
    except ValueError:
        raise HTTPException(400, f"Invalid transition from '{current}' to '{to_status}'")
    if new_idx != cur_idx + 1:
        raise HTTPException(400,
            f"Invalid transition: {current} → {to_status}. "
            f"Only forward, one step at a time (next allowed: "
            f"{ONLINE_STATUS_SEQUENCE[cur_idx + 1] if cur_idx + 1 < len(ONLINE_STATUS_SEQUENCE) else '—'}).")


# ── Color Master Seed & Resolution ───────────────────────────────────────────
DEFAULT_COLOR_MASTER = [
    ("Tan",         "TN"),
    ("Beige",       "BG"),
    ("Gold",        "GD"),
    ("Silver",      "SL"),
    ("Blue",        "BL"),
    ("Navy",        "NV"),
    ("Brown",       "BR"),
    ("Gunmetal",    "GN"),
    ("Maroon",      "MR"),
    ("Pink",        "PK"),
    ("Black",       "BK"),
    ("White",       "WH"),
    ("Cream",       "CR"),
    ("Deep Peach",  "DP"),
    ("Grey",        "GY"),
    ("Red",         "RD"),
    ("Green",       "GR"),
    ("Yellow",      "YL"),
    ("Orange",      "OR"),
    ("Purple",      "PR"),
    ("Rose Gold",   "RG"),
    ("Copper",      "CP"),
    ("Bronze",      "BZ"),
    ("Nude",        "ND"),
    ("Olive",       "OV"),
]


async def _seed_color_master() -> int:
    """Idempotently seed the color_master collection with the default palette."""
    db = get_db()
    try:
        await db.color_master.create_index("color_code", unique=True, name="color_master_code_unique")
    except Exception as e:
        log.warning(f"Could not create color_master unique index: {e}")
    try:
        await db.color_master.create_index("color_name_lc", name="color_master_name_lc")
    except Exception as e:
        log.warning(f"Could not create color_master name index: {e}")

    inserted = 0
    for name, code in DEFAULT_COLOR_MASTER:
        existing = await db.color_master.find_one({"color_code": code})
        if existing:
            continue
        by_name = await db.color_master.find_one({"color_name_lc": name.lower()})
        if by_name:
            continue
        await db.color_master.insert_one({
            "color_name":    name,
            "color_name_lc": name.lower(),
            "color_code":    code,
            "active":        True,
            "created_at":    now_iso(),
            "updated_at":    now_iso(),
        })
        inserted += 1
    return inserted


async def resolve_color_code(color_name: str) -> Optional[str]:
    """Look up a colour's short code by name (case-insensitive)."""
    if not color_name:
        return None
    db = get_db()
    doc = await db.color_master.find_one({
        "color_name_lc": color_name.strip().lower(),
        "active": {"$ne": False},
    })
    return doc["color_code"] if doc else None


# ── Listing Format Configs & Catalogue Export Models ─────────────────────────
Platform = Literal["myntra", "flipkart", "ajio", "nykaa", "website", "other"]

CANONICAL_FIELDS = [
    "group_id", "leaf_sku", "size",
    "color_primary", "color_family",
    "style_description", "mrp", "selling_price",
    "brand", "listing_status",
]

EXPORT_SOURCE_TYPES = [
    "group_sku", "leaf_sku", "style_code", "size",
    "color_name", "color_code", "style", "lifecycle",
    "constant", "blank",
]


class ExportColumn(BaseModel):
    name: str
    source: Literal[
        "group_sku", "leaf_sku", "style_code", "size",
        "color_name", "color_code",
        "style", "lifecycle", "constant", "blank",
    ]
    key: Optional[str] = None
    value: Optional[Any] = None
    notes: Optional[str] = None
    required: bool = False


class ExportTemplate(BaseModel):
    sheet_name: str = "Sheet1"
    header_row_index: int = 0
    pre_header_rows: Optional[List[List[Any]]] = None
    post_header_rows: Optional[List[List[Any]]] = None
    columns: List[ExportColumn]

    @field_validator("columns")
    @classmethod
    def _cols_non_empty(cls, v):
        if not v or len(v) == 0:
            raise PydanticCustomError(
                "export_columns_empty",
                "export_template.columns must contain at least one column"
            )
        if not any(c.source == "leaf_sku" for c in v):
            raise PydanticCustomError(
                "export_columns_leaf_sku",
                "export_template.columns must include exactly one column with source='leaf_sku'"
            )
        return v


class SheetLocator(BaseModel):
    type: Literal["fixed_name", "name_contains", "first_sheet"]
    name: Optional[str] = None
    substring: Optional[str] = None

    @field_validator("name")
    @classmethod
    def _clean_name(cls, v):
        return v.strip() if isinstance(v, str) else v

    @field_validator("substring")
    @classmethod
    def _clean_sub(cls, v):
        return v.strip() if isinstance(v, str) else v


class HeaderLocator(BaseModel):
    type: Literal["fixed_row", "scan_for_columns"]
    row: Optional[int] = None
    must_contain_any: Optional[List[str]] = None


class ListingFormatConfigIn(BaseModel):
    platform: Platform
    sheet_locator: SheetLocator
    header_locator: HeaderLocator
    skip_rows_after_header: int = 0
    column_map: Dict[str, Optional[str]]
    has_native_group_id: bool = False
    active: bool = True
    notes: Optional[str] = ""
    export_template: Optional[ExportTemplate] = None

    @field_validator("column_map")
    @classmethod
    def _validate_column_map(cls, v):
        if not isinstance(v, dict):
            raise PydanticCustomError("column_map_type", "column_map must be an object")
        if not v.get("leaf_sku"):
            raise PydanticCustomError(
                "column_map_leaf_sku",
                "column_map.leaf_sku is required"
            )
        return v


class ListingFormatConfigUpdate(BaseModel):
    sheet_locator: Optional[SheetLocator] = None
    header_locator: Optional[HeaderLocator] = None
    skip_rows_after_header: Optional[int] = None
    column_map: Optional[Dict[str, Optional[str]]] = None
    has_native_group_id: Optional[bool] = None
    active: Optional[bool] = None
    notes: Optional[str] = None
    export_template: Optional[ExportTemplate] = None

    @field_validator("column_map")
    @classmethod
    def _validate_column_map_opt(cls, v):
        if v is None:
            return v
        if not isinstance(v, dict):
            raise PydanticCustomError("column_map_type", "column_map must be an object")
        if not v.get("leaf_sku"):
            raise PydanticCustomError(
                "column_map_leaf_sku",
                "column_map.leaf_sku is required"
            )
        return v


DEFAULT_LISTING_FORMAT_CONFIGS = [
    {
        "platform": "myntra",
        "sheet_locator": {"type": "fixed_name", "name": "styledashboard"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "group_id":          "Style Id",
            "leaf_sku":          "SellerSkuCode",
            "size":              None,
            "color_primary":     "Colour",
            "color_family":      None,
            "style_description": "Style Name",
            "mrp":               "MRP",
            "selling_price":     "Selling Price",
            "brand":             "Brand",
            "listing_status":    "Listing Status",
        },
        "has_native_group_id": True,
        "active": True,
        "notes": "Myntra Style Dashboard export. Sheet name is fixed. Size is embedded in SellerSkuCode.",
        "export_template": {
            "sheet_name": "styledashboard",
            "header_row_index": 0,
            "columns": [
                {"name": "Style Id",       "source": "blank",     "notes": "Myntra assigns after ingest"},
                {"name": "SellerSkuCode",  "source": "leaf_sku",  "required": True},
                {"name": "Style Name",     "source": "style",     "key": "name"},
                {"name": "Brand",          "source": "lifecycle", "key": "brand"},
                {"name": "Category",       "source": "style",     "key": "category"},
                {"name": "Colour",         "source": "color_name"},
                {"name": "Size",           "source": "size"},
                {"name": "MRP",            "source": "lifecycle", "key": "mrp"},
                {"name": "Selling Price",  "source": "lifecycle", "key": "online_selling_price"},
                {"name": "Description",    "source": "style",     "key": "description"},
                {"name": "Image URL",      "source": "style",     "key": "image_url"},
                {"name": "Listing Status", "source": "constant",  "value": "Active"},
            ],
        },
    },
    {
        "platform": "ajio",
        "sheet_locator": {"type": "name_contains", "substring": "_Styles_"},
        "header_locator": {"type": "scan_for_columns", "must_contain_any": ["*Style Code", "*Item SKU", "*Size", "*Primary Color"]},
        "skip_rows_after_header": 0,
        "column_map": {
            "group_id":          "*Style Code",
            "leaf_sku":          "*Item SKU",
            "size":              "*Size",
            "color_primary":     "*Primary Color",
            "color_family":      "*Color Family",
            "style_description": "*Product Description",
            "mrp":               "*MRP",
            "selling_price":     "*Selling Price",
            "brand":             "*Brand",
            "listing_status":    "*Listing Status",
        },
        "has_native_group_id": True,
        "active": True,
        "notes": "Ajio catalogue export. Sheet name pattern *_Styles_*. Header row may shift — scanner handles rows 0-10.",
        "export_template": {
            "sheet_name": "SSK_Styles_Export",
            "header_row_index": 2,
            "pre_header_rows": [
                ["SSK Footcare — new listing upload"],
                ["Generated automatically. Do not edit header row."],
            ],
            "columns": [
                {"name": "*Style Code",           "source": "group_sku",  "required": True},
                {"name": "*Style Description",   "source": "style",      "key": "description"},
                {"name": "*Item SKU",            "source": "leaf_sku",   "required": True},
                {"name": "*Brand",               "source": "lifecycle",  "key": "brand"},
                {"name": "*MRP",                 "source": "lifecycle",  "key": "mrp"},
                {"name": "*Size",                "source": "size",       "required": True},
                {"name": "*Primary Color",       "source": "color_name", "required": True},
                {"name": "*Color Family",        "source": "lifecycle",  "key": "color_family"},
                {"name": "*Upper Material",      "source": "lifecycle",  "key": "upper_material"},
                {"name": "*Sole Material",       "source": "lifecycle",  "key": "sole_material"},
                {"name": "*Product Description", "source": "style",      "key": "description"},
                {"name": "*Selling Price",       "source": "lifecycle",  "key": "online_selling_price"},
                {"name": "*Listing Status",      "source": "constant",   "value": "Active"},
            ],
        },
    },
    {
        "platform": "flipkart",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 1,
        "column_map": {
            "group_id":          None,
            "leaf_sku":          "Seller SKU Id",
            "size":              None,
            "color_primary":     "Color",
            "color_family":      None,
            "style_description": "Product Title",
            "mrp":               "MRP",
            "selling_price":     "Your Selling Price",
            "brand":             "Brand",
            "listing_status":    "Listing Status",
        },
        "has_native_group_id": False,
        "active": True,
        "notes": "Flipkart Listings export. Row 1 (after header) is a description/hint row and must be skipped.",
        "export_template": {
            "sheet_name": "Listings",
            "header_row_index": 0,
            "post_header_rows": [
                [
                    "Seller SKU Id (must be unique per size)",
                    "Full product title (max 200 chars)",
                    "Brand name", "MRP (INR)", "Your selling price (INR)",
                    "Colour name", "Size (numeric)", "Product description",
                    "Public image URL", "Active",
                ],
            ],
            "columns": [
                {"name": "Seller SKU Id",     "source": "leaf_sku",   "required": True},
                {"name": "Product Title",    "source": "style",      "key": "name"},
                {"name": "Brand",            "source": "lifecycle",  "key": "brand"},
                {"name": "MRP",              "source": "lifecycle",  "key": "mrp"},
                {"name": "Your Selling Price","source": "lifecycle", "key": "online_selling_price"},
                {"name": "Color",            "source": "color_name"},
                {"name": "Size",             "source": "size"},
                {"name": "Description",      "source": "style",      "key": "description"},
                {"name": "Image URL",        "source": "style",      "key": "image_url"},
                {"name": "Listing Status",   "source": "constant",   "value": "Active"},
            ],
        },
    },
]


async def _seed_listing_format_configs() -> int:
    """Idempotently seed listing_format_configs with default platform formats."""
    db = get_db()
    try:
        await db.listing_format_configs.create_index("platform", unique=True, name="lfc_platform_unique")
    except Exception as e:
        log.warning(f"Could not create listing_format_configs index: {e}")
    inserted = 0
    patched = 0
    for cfg in DEFAULT_LISTING_FORMAT_CONFIGS:
        existing = await db.listing_format_configs.find_one({"platform": cfg["platform"]})
        if existing:
            if existing.get("seeded", False) and not existing.get("export_template") and cfg.get("export_template"):
                await db.listing_format_configs.update_one(
                    {"platform": cfg["platform"]},
                    {"$set": {
                        "export_template": cfg["export_template"],
                        "updated_at":      now_iso(),
                    }},
                )
                patched += 1
            continue
        doc = dict(cfg)
        doc["created_at"] = now_iso()
        doc["updated_at"] = now_iso()
        doc["seeded"]     = True
        await db.listing_format_configs.insert_one(doc)
        inserted += 1
    return inserted


class CatalogueExportRequest(BaseModel):
    style_id: str
    platform: Platform
    colors: Optional[List[str]] = None
    sizes:  Optional[List[str]] = None


def _resolve_export_source(
    col: dict,
    *,
    style: dict,
    lifecycle: Optional[dict],
    style_code: str,
    color_name: str,
    color_code: str,
    size: str,
) -> Any:
    src = col.get("source")
    if src == "blank":
        return ""
    if src == "constant":
        return col.get("value", "")
    if src == "group_sku":
        return build_catalogue_sku(style_code, color_code)
    if src == "leaf_sku":
        return build_catalogue_sku(style_code, color_code, size)
    if src == "style_code":
        return style_code
    if src == "size":
        return size
    if src == "color_name":
        return color_name
    if src == "color_code":
        return color_code
    if src == "style":
        key = col.get("key")
        if not key:
            return ""
        return style.get(key, "") if isinstance(style, dict) else ""
    if src == "lifecycle":
        key = col.get("key")
        if not key or not lifecycle:
            return ""
        return lifecycle.get(key, "")
    return ""


async def _upsert_provisional_sku_map(
    *,
    style_id: str,
    platform: str,
    group_sku: str,
    color_name: str,
    sizes_covered: List[str],
    user_email: str,
) -> str:
    db = get_db()
    filter_ = {
        "source_type":  "online_channel",
        "source_name":  platform,
        "external_sku": group_sku,
    }
    existing = await db.sku_map.find_one(filter_)
    color_map = {color_name: color_name}
    size_map  = {s: str(s) for s in sizes_covered}
    if existing:
        update: Dict[str, Any] = {}
        if existing.get("style_id") != style_id:
            update["style_id"] = style_id
        merged_cm = {**(existing.get("color_map") or {}), **color_map}
        merged_sm = {**(existing.get("size_map")  or {}), **size_map}
        if merged_cm != (existing.get("color_map") or {}):
            update["color_map"] = merged_cm
        if merged_sm != (existing.get("size_map") or {}):
            update["size_map"] = merged_sm
        if existing.get("status") not in ("confirmed", "auto_confirmed"):
            update["status"] = "pending_platform_confirmation"
        if not update:
            return "unchanged"
        update["updated_at"] = now_iso()
        await db.sku_map.update_one({"_id": existing["_id"]}, {"$set": update})
        return "updated"
    doc = {
        "style_id":            style_id,
        "source_type":         "online_channel",
        "source_name":         platform,
        "external_sku":        group_sku,
        "external_style_name": "",
        "color_map":           color_map,
        "size_map":            size_map,
        "status":              "pending_platform_confirmation",
        "created_via":         "catalogue_export",
        "created_at":          now_iso(),
        "updated_at":          now_iso(),
    }
    try:
        await db.sku_map.insert_one(doc)
        return "created"
    except DuplicateKeyError:
        return "updated"


# ═══════════════════════════════════════════════════════════════════════
# ══ ENDPOINTS: STYLE MASTER, BOM, LIFECYCLE & COLOR MASTER ═════════════
# ═══════════════════════════════════════════════════════════════════════

@styles_router.get("/styles/summary")
async def list_styles_summary(
    request: Request,
    status: Optional[str] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
):
    """Lightweight summary list of styles: code, name, category, status, image_thumbnail_url, cost_summary.
    Uses batch job fetching to avoid N+1 queries.
    """
    await _get_user(request)
    db = get_db()
    query: dict = {}
    if not include_deleted:
        query["$or"] = [
            {"active": {"$ne": False}},
            {"deleted_at": {"$exists": False}},
        ]
    if status:
        query["status"] = str(status)
    if search:
        search_regex = {"$regex": re.escape(str(search)), "$options": "i"}
        query["$or"] = [
            {"code": search_regex},
            {"name": search_regex},
            {"description": search_regex}
        ]
    docs = await db.styles.find(query).sort("created_at", -1).to_list(2000)
    pipeline_ids = {
        d["style_id"]
        for d in await db.style_lifecycle.find({}, {"style_id": 1}).to_list(20000)
    }

    # ── Batch-fetch production jobs for all styles in a single query ──────────
    style_ids = [str(d.get("_id") or d.get("id")) for d in docs if (d.get("_id") or d.get("id"))]
    style_codes = [d.get("code") for d in docs if d.get("code")]

    jobs_by_style_id = defaultdict(list)
    jobs_by_style_code = defaultdict(list)

    if style_ids or style_codes:
        job_or = []
        if style_ids:
            job_or.append({"style_id": {"$in": style_ids}})
        if style_codes:
            job_or.append({"style_code": {"$in": style_codes}})
        job_query = {"$or": job_or} if len(job_or) > 1 else job_or[0]
        all_jobs = await db.production_jobs.find(job_query).to_list(50000)
        for job in all_jobs:
            jid = job.get("style_id")
            if jid:
                jobs_by_style_id[str(jid)].append(job)
            jcode = job.get("style_code")
            if jcode:
                jobs_by_style_code[jcode].append(job)

    out = []
    for d in docs:
        d = stringify(d)
        sid = str(d.get("id") or d.get("_id") or "")
        scode = d.get("code", "")

        matched_jobs = []
        seen_job_ids = set()
        for job in jobs_by_style_id.get(sid, []):
            job_id_key = str(job.get("_id", id(job)))
            if job_id_key not in seen_job_ids:
                seen_job_ids.add(job_id_key)
                matched_jobs.append(job)
        for job in jobs_by_style_code.get(scode, []):
            job_id_key = str(job.get("_id", id(job)))
            if job_id_key not in seen_job_ids:
                seen_job_ids.add(job_id_key)
                matched_jobs.append(job)

        costing_full = compute_style_costing_from_jobs(d, matched_jobs)
        cost_summary = {
            "materials_cost": costing_full.get("materials_cost", 0.0),
            "labor_cost": costing_full.get("labor_cost", 0.0),
            "total_cost": costing_full.get("total_cost", 0.0),
            "selling_price": costing_full.get("selling_price", 0.0),
            "suggested_target_price": costing_full.get("suggested_target_price", 0.0),
            "is_assigned": costing_full.get("is_assigned", False),
        }
        out.append({
            "id": d.get("id"),
            "code": d.get("code"),
            "name": d.get("name"),
            "category": d.get("category", "Footwear"),
            "status": d.get("status", "active"),
            "image_thumbnail_url": d.get("image_thumbnail_url") or d.get("image_url") or "",
            "image_url": d.get("image_url") or "",
            "image_display_url": d.get("image_display_url") or "",
            "description": d.get("description", ""),
            "gst_pct": d.get("gst_pct", 5),
            "margin_pct": d.get("margin_pct", 25),
            "insole_mould_name": d.get("insole_mould_name", ""),
            "sole_mould_name": d.get("sole_mould_name", ""),
            "in_online_pipeline": d.get("id") in pipeline_ids,
            "cost_summary": cost_summary,
            "costing": cost_summary,
        })
    return out


@styles_router.get("/styles")
async def list_styles(
    request: Request,
    status: Optional[str] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
    full: bool = True,
):
    if not full:
        return await list_styles_summary(request, status=status, search=search, include_deleted=include_deleted)
    await _get_user(request)
    db = get_db()
    query: dict = {}
    if not include_deleted:
        query["$or"] = [
            {"active": {"$ne": False}},
            {"deleted_at": {"$exists": False}},
        ]
    if status:
        query["status"] = str(status)
    if search:
        search_regex = {"$regex": re.escape(str(search)), "$options": "i"}
        query["$or"] = [
            {"code": search_regex},
            {"name": search_regex},
            {"description": search_regex}
        ]
    docs = await db.styles.find(query).sort("created_at", -1).to_list(1000)
    pipeline_ids = {
        d["style_id"]
        for d in await db.style_lifecycle.find({}, {"style_id": 1}).to_list(20000)
    }

    # ── Batch-fetch production jobs for all styles in a single query ──────────
    style_ids = [str(d.get("_id") or d.get("id")) for d in docs if (d.get("_id") or d.get("id"))]
    style_codes = [d.get("code") for d in docs if d.get("code")]

    jobs_by_style_id = defaultdict(list)
    jobs_by_style_code = defaultdict(list)

    if style_ids or style_codes:
        job_or = []
        if style_ids:
            job_or.append({"style_id": {"$in": style_ids}})
        if style_codes:
            job_or.append({"style_code": {"$in": style_codes}})
        job_query = {"$or": job_or} if len(job_or) > 1 else job_or[0]
        all_jobs = await db.production_jobs.find(job_query).to_list(50000)
        for job in all_jobs:
            jid = job.get("style_id")
            if jid:
                jobs_by_style_id[str(jid)].append(job)
            jcode = job.get("style_code")
            if jcode:
                jobs_by_style_code[jcode].append(job)

    out = []
    for d in docs:
        d = stringify(d)
        sid = str(d.get("id") or d.get("_id") or "")
        scode = d.get("code", "")

        matched_jobs = []
        seen_job_ids = set()
        for job in jobs_by_style_id.get(sid, []):
            job_id_key = str(job.get("_id", id(job)))
            if job_id_key not in seen_job_ids:
                seen_job_ids.add(job_id_key)
                matched_jobs.append(job)
        for job in jobs_by_style_code.get(scode, []):
            job_id_key = str(job.get("_id", id(job)))
            if job_id_key not in seen_job_ids:
                seen_job_ids.add(job_id_key)
                matched_jobs.append(job)

        d["costing"] = compute_style_costing_from_jobs(d, matched_jobs)
        d["in_online_pipeline"] = d["id"] in pipeline_ids
        out.append(d)
    return out


@styles_router.get("/styles/bulk/template")
async def get_styles_template():
    import pandas as pd
    columns = [
        "Name", "Category", "Description", "Base Size",
        "Insole Mould Name", "Sole Mould Name", "Default Pairs Per Carton",
        "Overhead %", "Packing Cost", "Margin %", "GST %", "Image URL",
        "Labor: Cutting", "Labor: Fitting", "Labor: Pasting", "Labor: Finishing", "Labor: Packing"
    ]
    sample_data = [
        {
            "Name": "Classic Oxford", "Category": "Footwear", "Description": "Men's leather shoe",
            "Base Size": 7, "Insole Mould Name": "INSOLE-OX-01", "Sole Mould Name": "SOLE-OX-01",
            "Default Pairs Per Carton": 12,
            "Overhead %": 10, "Packing Cost": 15, "Margin %": 25, "GST %": 5, "Image URL": "",
            "Labor: Cutting": 12, "Labor: Fitting": 18, "Labor: Pasting": 10, "Labor: Finishing": 8, "Labor: Packing": 5
        }
    ]
    df = pd.DataFrame(sample_data, columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="style_master_template.xlsx"'}
    )


@styles_router.post("/styles/bulk/preview", dependencies=[Depends(upload_rate_limiter)])
async def bulk_upload_preview(file: UploadFile = File(...), request: Request = None):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    import pandas as pd
    content = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel/CSV file: {str(e)}")

    col_map = {}
    labor_cols = []
    for col in df.columns:
        c_str = str(col).strip()
        c_lower = c_str.lower()
        norm = re.sub(r'[\s_]+', ' ', c_lower)
        if c_lower.startswith("labor:"):
            labor_cols.append(col)
        elif norm in ("name", "style name", "stylename"):
            col_map[col] = "name"
        elif norm in ("category", "cat"):
            col_map[col] = "category"
        elif norm in ("description", "desc"):
            col_map[col] = "description"
        elif norm in ("base size", "basesize", "base_size", "size"):
            col_map[col] = "base_size"
        elif norm in ("insole mould name", "insole mould", "insole mold name", "insole mold", "insole_mould_name", "insole_mould"):
            col_map[col] = "insole_mould_name"
        elif norm in ("sole mould name", "sole mould", "sole mold name", "sole mold", "sole_mould_name", "sole_mould"):
            col_map[col] = "sole_mould_name"
        elif norm in ("default pairs per carton", "pairs per carton", "default_pairs_per_carton", "pairs_per_carton", "carton pairs", "default pairs"):
            col_map[col] = "default_pairs_per_carton"
        elif norm in ("overhead %", "overhead pct", "overhead percentage", "overhead", "overhead_pct"):
            col_map[col] = "overhead_pct"
        elif norm in ("packing cost", "packing_cost", "packing"):
            col_map[col] = "packing_cost"
        elif norm in ("margin %", "margin pct", "margin percentage", "margin", "margin_pct"):
            col_map[col] = "margin_pct"
        elif norm in ("gst %", "gst pct", "gst percentage", "gst", "gst_pct"):
            col_map[col] = "gst_pct"
        elif norm in ("image url", "image_url", "image", "image link", "photo url", "photo_url"):
            col_map[col] = "image_url"

    if "name" not in col_map.values():
        raise HTTPException(400, "Missing required column: Name")

    df = df.rename(columns=col_map)

    def _parse_float(val, default: float) -> float:
        if pd.isna(val): return default
        try:
            s = str(val).strip().rstrip("%")
            if not s or s.lower() == "nan": return default
            return float(s)
        except Exception: return default

    def _parse_str(val, default: str = "") -> str:
        if pd.isna(val): return default
        s = str(val).strip()
        if not s or s.lower() == "nan": return default
        if s.endswith(".0"):
            try:
                f = float(s)
                if f.is_integer(): return str(int(f))
            except Exception: pass
        return s

    def _parse_pairs_per_carton(val) -> Optional[Dict[str, Any]]:
        if pd.isna(val): return None
        if isinstance(val, (int, float)):
            if float(val) > 0:
                return {"default": int(val) if float(val).is_integer() else float(val)}
            return None
        s = str(val).strip()
        if not s or s.lower() == "nan": return None
        if s.startswith("{") and s.endswith("}"):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, dict): return parsed
            except Exception: pass
        try:
            f = float(s)
            if f > 0: return {"default": int(f) if f.is_integer() else f}
        except Exception: pass
        return None

    preview = []
    errors = []
    total_rows = len(df)

    for idx, row in df.iterrows():
        row_num = idx + 2
        name = _parse_str(row.get("name"), "")
        if not name:
            errors.append(f"Row {row_num}: Missing Name")
            continue

        category = _parse_str(row.get("category"), "Footwear")
        description = _parse_str(row.get("description"), "")
        base_size = _parse_str(row.get("base_size"), "7")
        insole_mould = _parse_str(row.get("insole_mould_name"), "") or None
        sole_mould = _parse_str(row.get("sole_mould_name"), "") or None
        default_pairs = _parse_pairs_per_carton(row.get("default_pairs_per_carton"))

        overhead_pct = _parse_float(row.get("overhead_pct"), 0.0)
        packing_cost = _parse_float(row.get("packing_cost"), 0.0)
        margin_pct = _parse_float(row.get("margin_pct"), 25.0)
        gst_pct = _parse_float(row.get("gst_pct"), 5.0)

        raw_img = _parse_str(row.get("image_url"), "")
        norm_img = normalize_image_url(raw_img) if raw_img else ""

        labor = []
        for lc in labor_cols:
            op_name = str(lc).split(":", 1)[1].strip()
            val = row.get(lc)
            if pd.notna(val) and str(val).strip() != "" and str(val).strip().lower() != "nan":
                try:
                    rate = float(str(val).strip().rstrip("%"))
                    labor.append({"name": op_name, "rate": rate})
                except Exception:
                    pass

        preview.append({
            "row_number": row_num,
            "name": name,
            "category": category or "Footwear",
            "description": description,
            "base_size": base_size or "7",
            "insole_mould_name": insole_mould,
            "sole_mould_name": sole_mould,
            "default_pairs_per_carton": default_pairs,
            "overhead_pct": overhead_pct,
            "packing_cost": packing_cost,
            "margin_pct": margin_pct,
            "gst_pct": gst_pct,
            "image_url": norm_img,
            "image_display_url": norm_img,
            "image_thumbnail_url": norm_img,
            "labor": labor,
            "bom": [],
        })

    return {
        "preview": preview,
        "total_rows": total_rows,
        "valid_rows": len(preview),
        "errors": errors,
    }


@styles_router.post("/styles/bulk")
async def bulk_upload_styles(payload: dict, request: Request = None):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    user_email = u.get("email", "") if isinstance(u, dict) else getattr(u, "email", "")

    styles_list = payload.get("styles", [])
    if not styles_list:
        raise HTTPException(400, "No styles provided")

    success = 0
    errors = []
    created = []

    for idx, row in enumerate(styles_list):
        row_num = row.get("row_number") or row.get("row") or (idx + 2)
        try:
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Row {row_num}: Missing Name")
                continue

            category = str(row.get("category") or "Footwear").strip() or "Footwear"
            description = str(row.get("description") or "").strip()
            base_size = str(row.get("base_size") or "7").strip() or "7"
            insole_mould_name = row.get("insole_mould_name") or None
            sole_mould_name = row.get("sole_mould_name") or None

            try: overhead_pct = float(row.get("overhead_pct", 0) or 0)
            except Exception: overhead_pct = 0.0
            try: packing_cost = float(row.get("packing_cost", 0) or 0)
            except Exception: packing_cost = 0.0
            try: margin_pct = float(row.get("margin_pct", 25) if row.get("margin_pct") is not None else 25)
            except Exception: margin_pct = 25.0
            try: gst_pct = float(row.get("gst_pct", 5) if row.get("gst_pct") is not None else 5)
            except Exception: gst_pct = 5.0

            raw_img = str(row.get("image_url") or "").strip()
            norm_img = normalize_image_url(raw_img) if raw_img else ""
            raw_display = str(row.get("image_display_url") or "").strip()
            raw_thumb = str(row.get("image_thumbnail_url") or "").strip()
            image_display = normalize_image_url(raw_display) if raw_display else norm_img
            image_thumb = normalize_image_url(raw_thumb) if raw_thumb else norm_img

            default_pairs_per_carton = row.get("default_pairs_per_carton")
            labor = row.get("labor") or []

            generated_code = None
            for _ in range(5):
                candidate = await _next_style_code()
                if not await db.styles.find_one({"code": candidate}):
                    generated_code = candidate
                    break
            if not generated_code:
                errors.append(f"Row {row_num}: Failed to generate a unique style code")
                continue

            doc = {
                "code": generated_code,
                "name": name,
                "category": category,
                "description": description,
                "base_size": base_size,
                "insole_mould_name": insole_mould_name,
                "sole_mould_name": sole_mould_name,
                "overhead_pct": overhead_pct,
                "packing_cost": packing_cost,
                "margin_pct": margin_pct,
                "gst_pct": gst_pct,
                "image_url": norm_img,
                "image_display_url": image_display,
                "image_thumbnail_url": image_thumb,
                "default_pairs_per_carton": default_pairs_per_carton,
                "bom": [],
                "labor": labor,
                "status": "inactive",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }

            try:
                res = await db.styles.insert_one(doc)
            except DuplicateKeyError:
                errors.append(f"Row {row_num}: Style code '{generated_code}' collision")
                continue

            try:
                await db.style_folders.insert_one({
                    "style_id": str(res.inserted_id),
                    "style_code": generated_code,
                    "folders": DEFAULT_PLM_FOLDERS,
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                })
            except Exception:
                pass

            doc.pop("_id", None)
            doc["id"] = str(res.inserted_id)
            costing = await compute_style_costing_async(doc, db)
            doc["costing"] = costing

            await log_activity(
                "style.create",
                "styles",
                f"Bulk created style {generated_code} — {name}",
                user_email,
            )

            created.append({
                "row": row_num,
                "code": generated_code,
                "name": name,
                "costing": costing,
            })
            success += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    await log_activity(
        "BULK_CREATE",
        "styles",
        f"Bulk import: {success} created, {len(errors)} errors",
        user_email,
    )

    return {
        "ok": True,
        "success_count": success,
        "errors": errors,
        "created": created,
    }


@styles_router.get("/styles/{sid}")
async def get_style(sid: str, request: Request):
    await _get_user(request)
    db = get_db()
    d = await db.styles.find_one({"_id": oid(sid)})
    if not d:
        raise HTTPException(404, "Not found")
    d = stringify(d)
    d["costing"] = await compute_style_costing_async(d, db)
    return d


@styles_router.post("/styles")
async def create_style(payload: StyleIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    doc = payload.model_dump()
    _u = normalize_image_url(doc.get("image_url", "") or "")
    if _u:
        doc["image_url"] = _u
        if not doc.get("image_display_url"):   doc["image_display_url"]   = _u
        if not doc.get("image_thumbnail_url"): doc["image_thumbnail_url"] = _u

    generated_code = None
    for _ in range(5):
        candidate = await _next_style_code()
        if not await db.styles.find_one({"code": candidate}):
            generated_code = candidate
            break
    if not generated_code:
        raise HTTPException(500, "Failed to generate a unique style code — please retry")
    doc["code"] = generated_code
    doc["status"] = "active" if len(doc.get("bom", [])) > 0 else "inactive"
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    try:
        res = await db.styles.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail=f"Style code '{generated_code}' collision — please retry")
    doc.pop("_id", None)
    doc["id"] = str(res.inserted_id)
    doc["costing"] = await compute_style_costing_async(doc, db)

    try:
        await db.style_folders.insert_one({
            "style_id": str(res.inserted_id),
            "style_code": generated_code,
            "folders": DEFAULT_PLM_FOLDERS,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    except Exception:
        pass
    await log_activity("style.create", "styles",
                       f"Created style {generated_code} — {doc.get('name','')}", u["email"])
    return doc


@styles_router.patch("/styles/{sid}")
async def update_style(sid: str, payload: StyleIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    existing = await db.styles.find_one({"_id": oid(sid)})
    if not existing:
        raise HTTPException(404, "Style not found")
    update = payload.model_dump()
    _u = normalize_image_url(update.get("image_url", "") or "")
    if _u:
        update["image_url"] = _u
        if not update.get("image_display_url"):   update["image_display_url"]   = _u
        if not update.get("image_thumbnail_url"): update["image_thumbnail_url"] = _u

    supplied_code = (update.pop("code", "") or "").strip()
    if supplied_code and supplied_code != existing.get("code", ""):
        raise HTTPException(
            400,
            f"Style code is immutable — attempted to change '{existing.get('code','')}' to '{supplied_code}'"
        )
    update["status"] = "active" if len(update.get("bom", [])) > 0 else "inactive"
    update["updated_at"] = now_iso()
    await db.styles.update_one({"_id": oid(sid)}, {"$set": update})
    d = stringify(await db.styles.find_one({"_id": oid(sid)}))
    d["costing"] = await compute_style_costing_async(d, db)
    return d


@styles_router.delete("/styles/{sid}")
async def delete_style(sid: str, request: Request):
    """Cascade-aware soft-delete for a Style Master record."""
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()

    style = await db.styles.find_one({"_id": oid(sid)})
    if not style:
        raise HTTPException(404, "Style not found")

    style_code = style.get("code", sid)
    now = now_iso()
    blockers: list = []

    po_refs = await db.purchase_orders.find_one({
        "line_items.style_code": style_code,
        "status": {"$nin": ["cancelled", "closed"]},
    })
    if po_refs:
        blockers.append(
            f"Active PO '{po_refs.get('po_number', po_refs['_id'])}' references style {style_code}"
        )

    active_job = await db.production_jobs.find_one({
        "style_id": ObjectId(sid),
        "stage": {"$nin": ["dispatched", "cancelled"]},
    })
    if active_job:
        blockers.append(
            f"Active production job '{active_job.get('job_number', active_job['_id'])}' "
            f"is still in progress for style {style_code}"
        )

    nonzero_inv = await db.fg_inventory.find_one({
        "style_id": ObjectId(sid),
        "$or": [
            {"ready_stock_qty": {"$gt": 0}},
            {"reserved_qty":    {"$gt": 0}},
        ],
    })
    if nonzero_inv:
        blockers.append(
            f"FG inventory has non-zero stock (color={nonzero_inv.get('color')}, "
            f"size={nonzero_inv.get('size')}) for style {style_code}. Zero out stock first."
        )

    if blockers:
        raise HTTPException(
            409,
            {
                "message": f"Cannot delete style '{style_code}' — {len(blockers)} blocker(s) must be resolved first.",
                "blockers": blockers,
            }
        )

    cascade_summary: dict = {}
    lc_del = await db.style_lifecycle.delete_one({"style_id": sid})
    cascade_summary["lifecycle_deleted"] = lc_del.deleted_count

    sku_del = await db.sku_map.delete_many({"style_id": sid})
    cascade_summary["sku_map_deleted"] = sku_del.deleted_count

    bom_res = await db.style_component_mapping.update_many(
        {"style_id": sid},
        {"$set": {"active": False, "updated_at": now}},
    )
    cascade_summary["bom_links_deactivated"] = bom_res.modified_count

    zero_inv_del = await db.fg_inventory.delete_many({
        "style_id": ObjectId(sid),
        "ready_stock_qty": {"$lte": 0},
        "reserved_qty":    {"$lte": 0},
        "in_transit_qty":  {"$lte": 0},
        "return_qty":      {"$lte": 0},
        "damaged_qty":     {"$lte": 0},
        "liquidation_qty": {"$lte": 0},
    })
    nonzero_inv_flag = await db.fg_inventory.update_many(
        {"style_id": ObjectId(sid)},
        {"$set": {"active": False, "updated_at": now}},
    )
    cascade_summary["fg_inventory_zeroed_deleted"] = zero_inv_del.deleted_count
    cascade_summary["fg_inventory_nonzero_flagged"] = nonzero_inv_flag.modified_count

    await db.styles.update_one(
        {"_id": oid(sid)},
        {"$set": {
            "active":     False,
            "deleted_at": now,
            "updated_at": now,
        }}
    )

    await log_activity(
        "DELETE", "styles",
        f"Style '{style_code}' soft-deleted. Cascade: {cascade_summary}",
        u["email"],
    )

    return {
        "ok":             True,
        "id":             sid,
        "style_code":     style_code,
        "cascade":        cascade_summary,
    }


# ── Style Lifecycle & Online Pipeline Endpoints ──────────────────────────────
@styles_router.get("/style-lifecycle/{style_id}")
async def get_style_lifecycle(style_id: str, request: Request):
    await _get_user(request)
    doc = await _get_or_create_lifecycle(style_id)
    return stringify(doc)


@styles_router.post("/styles/{sid}/pipeline")
async def add_style_to_online_pipeline(sid: str, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    try:
        s_oid = ObjectId(sid)
    except Exception:
        raise HTTPException(400, "Invalid style id")
    style = await db.styles.find_one({"_id": s_oid})
    if not style:
        raise HTTPException(404, "Style not found")

    existing = await db.style_lifecycle.find_one({"style_id": sid})
    if existing:
        return {"ok": True, "already_in_pipeline": True, "lifecycle": stringify(existing)}

    doc = _default_lifecycle(sid, style["code"])
    res = await db.style_lifecycle.insert_one(doc)
    doc = await db.style_lifecycle.find_one({"_id": res.inserted_id})
    await log_activity(
        "ADD_TO_PIPELINE", "style_lifecycle",
        f"Added style {style['code']} to Online Style Pipeline.",
        u["email"],
    )
    return {"ok": True, "already_in_pipeline": False, "lifecycle": stringify(doc)}


@styles_router.delete("/styles/{sid}/pipeline")
async def remove_style_from_online_pipeline(sid: str, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    lc = await db.style_lifecycle.find_one({"style_id": sid})
    if not lc:
        return {"ok": True, "was_in_pipeline": False}
    style = await db.styles.find_one({"_id": ObjectId(sid)})
    style_code = (style or {}).get("code", sid)
    await db.style_lifecycle.delete_one({"style_id": sid})
    await log_activity(
        "REMOVE_FROM_PIPELINE", "style_lifecycle",
        f"Removed style {style_code} from Online Style Pipeline.",
        u["email"],
    )
    return {"ok": True, "was_in_pipeline": True}


@styles_router.get("/styles/not-in-pipeline")
async def list_styles_not_in_pipeline(request: Request, search: Optional[str] = None):
    await _get_user(request)
    db = get_db()
    q: dict = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        q["$or"] = [{"code": rx}, {"name": rx}]
    styles = await db.styles.find(q).sort("code", 1).to_list(5000)
    ids_in_pipeline = {
        d["style_id"]
        for d in await db.style_lifecycle.find({}, {"style_id": 1}).to_list(20000)
    }
    out = []
    for s in styles:
        sid = str(s["_id"])
        if sid in ids_in_pipeline:
            continue
        out.append({
            "id":                   sid,
            "code":                 s.get("code"),
            "name":                 s.get("name", ""),
            "image_url":            s.get("image_url", ""),
            "image_display_url":    s.get("image_display_url", ""),
            "image_thumbnail_url":  s.get("image_thumbnail_url", ""),
        })
    return out


@styles_router.put("/style-lifecycle/{style_id}")
async def upsert_style_lifecycle(style_id: str, payload: StyleLifecycleUpsert, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    existing = await _get_or_create_lifecycle(style_id)
    update: dict = {"updated_at": now_iso()}
    payload_dict = payload.model_dump(exclude_none=True)

    if "planned_components" in payload_dict:
        pcs = payload_dict["planned_components"]
        by_name = {c["component"]: int(c.get("planned_qty") or 0) for c in pcs}
        payload_dict["planned_components"] = [
            {"component": name, "planned_qty": by_name.get(name, 0)}
            for name in PLANNED_COMPONENTS
        ]

    payload_dict.pop("online_status", None)

    for k, v in payload_dict.items():
        update[k] = v

    await db.style_lifecycle.update_one({"style_id": str(existing.get("style_id"))}, {"$set": update})
    await log_activity(
        "UPDATE", "style_lifecycle",
        f"Updated lifecycle for {existing.get('style_code')}: {', '.join(payload_dict.keys())}",
        u["email"],
    )
    return stringify(await db.style_lifecycle.find_one({"style_id": str(existing.get("style_id"))}))


@styles_router.patch("/styles/{sid}/online-status")
async def patch_style_online_status(sid: str, payload: OnlineStatusPatchIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    lifecycle = await _get_or_create_lifecycle(sid)

    current = lifecycle.get("online_status", "draft")
    to_status = payload.to_status
    _validate_online_status_transition(current, to_status)

    now = now_iso()
    history_entry = {
        "status":     to_status,
        "changed_at": now,
        "by":         u["email"],
        "notes":      (payload.notes or "").strip(),
        "from":       current,
    }
    update = {
        "$set":  {"online_status": to_status, "updated_at": now},
        "$push": {"online_status_history": history_entry},
    }

    seed_result = None
    if to_status == "live" and current != "live":
        style = await db.styles.find_one({"_id": oid(sid)})
        style_code = style["code"] if style else lifecycle.get("style_code", "")
        if not lifecycle.get("back_track_number"):
            back_track = await _generate_back_track_number(style_code)
            update["$set"]["back_track_number"] = back_track
        update["$set"]["went_live_at"] = now

    await db.style_lifecycle.update_one({"style_id": str(lifecycle["style_id"])}, update)
    updated_doc = await db.style_lifecycle.find_one({"style_id": str(lifecycle["style_id"])})

    if to_status == "live" and current != "live":
        from routes.inventory import _seed_fg_inventory_for_lifecycle
        seed_result = await _seed_fg_inventory_for_lifecycle(updated_doc, u["email"], db=db)

    await log_activity(
        "STATUS", "style_lifecycle",
        f"{updated_doc.get('style_code')}: {current} → {to_status}"
        + (f" [seeded: {seed_result['created']} FG rows]" if seed_result else ""),
        u["email"],
    )
    resp = stringify(updated_doc)
    if seed_result:
        resp["seed_result"] = seed_result
    return resp


@styles_router.get("/styles/online")
async def list_online_styles(
    request: Request,
    online_status:  Optional[str] = None,
    sale_channel:   Optional[str] = None,
    search:         Optional[str] = None,
):
    await _get_user(request)
    db = get_db()
    style_query: dict = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        style_query["$or"] = [{"code": rx}, {"name": rx}]

    styles = await db.styles.find(style_query).sort("code", 1).to_list(5000)
    style_ids_str = [str(s["_id"]) for s in styles]

    lifecycles = await db.style_lifecycle.find({"style_id": {"$in": style_ids_str}}).to_list(5000)
    lc_by_id = {l["style_id"]: l for l in lifecycles}
    style_ids_in_pipeline = set(lc_by_id.keys())

    mappings = await db.sku_map.find({"style_id": {"$in": list(style_ids_in_pipeline)}}).to_list(20000)
    maps_by_style: dict = {}
    for m in mappings:
        maps_by_style.setdefault(m["style_id"], []).append({
            "id":                  str(m["_id"]),
            "source_type":         m.get("source_type"),
            "source_name":         m.get("source_name"),
            "external_sku":        m.get("external_sku"),
            "external_style_name": m.get("external_style_name", ""),
        })

    out = []
    for s in styles:
        sid = str(s["_id"])
        if sid not in style_ids_in_pipeline:
            continue
        lc = lc_by_id[sid]
        if online_status and lc.get("online_status") != online_status:
            continue
        if sale_channel and sale_channel not in (lc.get("sale_channels") or []):
            continue

        out.append({
            "style_id":               sid,
            "style_code":             s.get("code"),
            "style_name":             s.get("name", ""),
            "image_url":              s.get("image_url", ""),
            "image_display_url":      s.get("image_display_url", ""),
            "image_thumbnail_url":    s.get("image_thumbnail_url", ""),
            "online_status":          lc.get("online_status", "draft"),
            "online_status_history":  lc.get("online_status_history", []),
            "sale_channels":          lc.get("sale_channels", []),
            "mrp":                    lc.get("mrp"),
            "online_selling_price":   lc.get("online_selling_price"),
            "gst_pct":                s.get("gst_pct", 5),
            "platform_commission_pct": lc.get("platform_commission_pct", {}),
            "planned_min_stock":      lc.get("planned_min_stock", 25),
            "planned_components":     lc.get("planned_components", []),
            "planned_colors":         lc.get("planned_colors", []),
            "planned_sizes":          lc.get("planned_sizes", []),
            "sole_mould_name":        lc.get("sole_mould_name", ""),
            "sole_shape":             lc.get("sole_shape", ""),
            "pattern_number":         lc.get("pattern_number", ""),
            "photoshoot_link":        lc.get("photoshoot_link", ""),
            "catalogue_link":         lc.get("catalogue_link", ""),
            "back_track_number":      lc.get("back_track_number", ""),
            "went_live_at":           lc.get("went_live_at"),
            "channel_skus":           maps_by_style.get(sid, []),
        })

    def sort_key(row):
        st = row["online_status"]
        try:
            idx = ONLINE_STATUS_SEQUENCE.index(st)
        except ValueError:
            idx = 99 if st == "archived" else 98
        return (idx, row["style_code"] or "")
    out.sort(key=sort_key)
    return out


# ── Catalogue Codes & Color Master Endpoints ─────────────────────────────────
@styles_router.get("/styles/{sid}/catalogue-codes")
async def get_style_catalogue_codes(sid: str, request: Request):
    await _get_user(request)
    db = get_db()
    style = await db.styles.find_one({"_id": oid(sid)})
    if not style:
        raise HTTPException(404, "Style not found")
    style_code = style.get("code", "")

    lifecycle = await db.style_lifecycle.find_one({"style_id": str(style["_id"])})
    colors: List[str] = []
    sizes: List[str] = []
    if lifecycle:
        colors = [c for c in (lifecycle.get("planned_colors") or []) if c]
        sizes = [str(s) for s in (lifecycle.get("planned_sizes") or []) if s]

    if not colors or not sizes:
        cursor = db.fg_inventory.find({"style_id": str(style["_id"])})
        cset, sset = set(), set()
        async for r in cursor:
            if r.get("color"): cset.add(r["color"])
            if r.get("size"):  sset.add(str(r["size"]))
        if not colors: colors = sorted(cset)
        if not sizes:
            def _ssort(x):
                try:    return (0, int(x))
                except: return (1, x)
            sizes = sorted(sset, key=_ssort)

    color_rows = []
    for cname in colors:
        code = await resolve_color_code(cname)
        group_sku = build_catalogue_sku(style_code, code) if code else ""
        color_rows.append({
            "color_name":  cname,
            "color_code":  code or "",
            "mapped":      bool(code),
            "group_sku":   group_sku,
            "size_skus":   [
                {
                    "size":     sz,
                    "leaf_sku": build_catalogue_sku(style_code, code, sz) if code else "",
                }
                for sz in sizes
            ],
        })

    unmapped_colors = [r["color_name"] for r in color_rows if not r["mapped"]]

    return {
        "style_id":         str(style["_id"]),
        "style_code":       style_code,
        "style_name":       style.get("name", ""),
        "colors":           colors,
        "sizes":            sizes,
        "rows":             color_rows,
        "unmapped_colors":  unmapped_colors,
    }


@styles_router.get("/color-master")
async def list_color_master(request: Request, active: Optional[bool] = None, search: Optional[str] = None):
    await _get_user(request)
    db = get_db()
    q: Dict[str, Any] = {}
    if active is not None:
        q["active"] = active
    if search:
        rgx = {"$regex": re.escape(search), "$options": "i"}
        q["$or"] = [{"color_name": rgx}, {"color_code": rgx}]
    docs = await db.color_master.find(q).sort("color_name", 1).to_list(1000)
    return [stringify(d) for d in docs]


@styles_router.post("/color-master")
async def create_color(payload: ColorMasterIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    if await db.color_master.find_one({"color_code": payload.color_code}):
        raise HTTPException(409, f"color_code '{payload.color_code}' already exists")
    if await db.color_master.find_one({"color_name_lc": payload.color_name.lower()}):
        raise HTTPException(409, f"color_name '{payload.color_name}' already exists")
    doc = {
        "color_name":    payload.color_name,
        "color_name_lc": payload.color_name.lower(),
        "color_code":    payload.color_code,
        "active":        payload.active,
        "created_at":    now_iso(),
        "updated_at":    now_iso(),
    }
    try:
        res = await db.color_master.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(409, f"color_code '{payload.color_code}' already exists")
    doc.pop("_id", None)
    doc["id"] = str(res.inserted_id)
    await log_activity("color.create", "color_master",
                       f"Added colour {payload.color_name} ({payload.color_code})", u["email"])
    return doc


@styles_router.put("/color-master/{cid}")
async def update_color(cid: str, payload: ColorMasterUpdate, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    existing = await db.color_master.find_one({"_id": oid(cid)})
    if not existing:
        raise HTTPException(404, "Colour not found")
    update: Dict[str, Any] = {}
    if payload.color_name is not None:
        by_name = await db.color_master.find_one({
            "color_name_lc": payload.color_name.lower(),
            "_id": {"$ne": oid(cid)},
        })
        if by_name:
            raise HTTPException(409, f"color_name '{payload.color_name}' already exists")
        update["color_name"]    = payload.color_name
        update["color_name_lc"] = payload.color_name.lower()
    if payload.color_code is not None:
        by_code = await db.color_master.find_one({
            "color_code": payload.color_code,
            "_id": {"$ne": oid(cid)},
        })
        if by_code:
            raise HTTPException(409, f"color_code '{payload.color_code}' already exists")
        update["color_code"] = payload.color_code
    if payload.active is not None:
        update["active"] = payload.active
    if not update:
        return stringify(existing)
    update["updated_at"] = now_iso()
    await db.color_master.update_one({"_id": oid(cid)}, {"$set": update})
    fresh = await db.color_master.find_one({"_id": oid(cid)})
    await log_activity("color.update", "color_master",
                       f"Updated colour {fresh.get('color_name')} ({fresh.get('color_code')})", u["email"])
    return stringify(fresh)


# ── Listing Format Configs & Catalogue Export Endpoints ──────────────────────
@styles_router.get("/listing-format-configs")
async def list_listing_format_configs(request: Request, active: Optional[bool] = None):
    await _get_user(request)
    db = get_db()
    q: Dict[str, Any] = {}
    if active is not None:
        q["active"] = active
    docs = await db.listing_format_configs.find(q).sort("platform", 1).to_list(1000)
    return [stringify(d) for d in docs]


@styles_router.get("/listing-format-configs/{platform}")
async def get_listing_format_config(platform: str, request: Request):
    await _get_user(request)
    db = get_db()
    doc = await db.listing_format_configs.find_one({"platform": platform.lower()})
    if not doc:
        raise HTTPException(404, f"No listing-format config for platform '{platform}'")
    return stringify(doc)


@styles_router.post("/listing-format-configs")
async def create_listing_format_config(payload: ListingFormatConfigIn, request: Request):
    u = await _get_user(request)
    require_roles("admin")(u)
    db = get_db()
    if await db.listing_format_configs.find_one({"platform": payload.platform}):
        raise HTTPException(409, f"Config for platform '{payload.platform}' already exists — use PUT to update")
    doc = payload.model_dump()
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    doc["seeded"]     = False
    try:
        res = await db.listing_format_configs.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(409, f"Config for platform '{payload.platform}' already exists")
    doc.pop("_id", None)
    doc["id"] = str(res.inserted_id)
    await log_activity("listing_format.create", "listing_format_configs",
                       f"Added listing format config for {payload.platform}", u["email"])
    return doc


@styles_router.put("/listing-format-configs/{platform}")
async def update_listing_format_config(platform: str, payload: ListingFormatConfigUpdate, request: Request):
    u = await _get_user(request)
    require_roles("admin")(u)
    db = get_db()
    platform_lc = platform.lower()
    existing = await db.listing_format_configs.find_one({"platform": platform_lc})
    if not existing:
        raise HTTPException(404, f"No listing-format config for platform '{platform_lc}'")
    update: Dict[str, Any] = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        return stringify(existing)
    update["updated_at"] = now_iso()
    await db.listing_format_configs.update_one({"platform": platform_lc}, {"$set": update})
    fresh = await db.listing_format_configs.find_one({"platform": platform_lc})
    await log_activity("listing_format.update", "listing_format_configs",
                       f"Updated listing format config for {platform_lc}: {', '.join(update.keys())}", u["email"])
    return stringify(fresh)


@styles_router.get("/listing-format-configs/_meta/canonical-fields")
async def get_canonical_fields(request: Request):
    await _get_user(request)
    return {"canonical_fields": CANONICAL_FIELDS}


@styles_router.post("/catalogue-export")
async def catalogue_export(payload: CatalogueExportRequest, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()

    style = await db.styles.find_one({"_id": oid(payload.style_id)})
    if not style:
        raise HTTPException(404, "Style not found")
    style_code = style.get("code", "")
    if not style_code:
        raise HTTPException(400, "Style has no style_code — cannot export a catalogue file")

    cfg = await db.listing_format_configs.find_one({"platform": payload.platform})
    if not cfg:
        raise HTTPException(404, f"No listing-format config for platform '{payload.platform}' — create one first")
    export_template = cfg.get("export_template")
    if not export_template:
        raise HTTPException(
            400,
            f"Platform '{payload.platform}' has no export_template configured. "
            "Add one via PUT /api/listing-format-configs/{platform} before generating catalogue files."
        )
    if not cfg.get("active", True):
        raise HTTPException(400, f"Platform '{payload.platform}' config is inactive")

    lifecycle = await db.style_lifecycle.find_one({"style_id": str(style["_id"])})

    colors = [c for c in (payload.colors or []) if c and c.strip()]
    sizes  = [str(s) for s in (payload.sizes  or []) if s and str(s).strip()]
    if not colors and lifecycle:
        colors = [c for c in (lifecycle.get("planned_colors") or []) if c]
    if not sizes and lifecycle:
        sizes = [str(s) for s in (lifecycle.get("planned_sizes") or []) if s]
    if not colors:
        raise HTTPException(400, "No colours to export. Pass `colors` in the body or set lifecycle.planned_colors.")
    if not sizes:
        raise HTTPException(400, "No sizes to export. Pass `sizes` in the body or set lifecycle.planned_sizes.")

    color_rows: List[Dict[str, str]] = []
    unmapped: List[str] = []
    for cname in colors:
        cc = await resolve_color_code(cname)
        if not cc:
            unmapped.append(cname)
        else:
            color_rows.append({"color_name": cname, "color_code": cc})
    if unmapped:
        raise HTTPException(
            400,
            f"These colours are not in the color master: {', '.join(unmapped)}. "
            "Add them via POST /api/color-master before exporting."
        )

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = (export_template.get("sheet_name") or "Sheet1")[:31]

    columns: List[dict] = export_template.get("columns") or []
    if not columns:
        raise HTTPException(500, "export_template.columns is empty — refusing to write an empty file")
    header_row_index = int(export_template.get("header_row_index") or 0)
    pre_header_rows  = export_template.get("pre_header_rows")  or []
    post_header_rows = export_template.get("post_header_rows") or []

    row_cursor = 1
    for filler in pre_header_rows:
        for i, cell in enumerate(filler):
            ws.cell(row=row_cursor, column=i + 1, value=cell)
        row_cursor += 1
    target_header_row = header_row_index + 1
    while row_cursor < target_header_row:
        row_cursor += 1

    for i, col in enumerate(columns):
        ws.cell(row=row_cursor, column=i + 1, value=col.get("name", ""))
    row_cursor += 1

    for filler in post_header_rows:
        for i, cell in enumerate(filler):
            ws.cell(row=row_cursor, column=i + 1, value=cell)
        row_cursor += 1

    rows_written = 0
    sku_map_summary = {"created": 0, "updated": 0, "unchanged": 0}
    style_dict = stringify(style)
    lifecycle_dict = stringify(lifecycle) if lifecycle else None

    for cr in color_rows:
        group_sku = build_catalogue_sku(style_code, cr["color_code"])
        for sz in sizes:
            for i, col in enumerate(columns):
                val = _resolve_export_source(
                    col,
                    style=style_dict,
                    lifecycle=lifecycle_dict,
                    style_code=style_code,
                    color_name=cr["color_name"],
                    color_code=cr["color_code"],
                    size=sz,
                )
                ws.cell(row=row_cursor, column=i + 1, value=val)
            row_cursor += 1
            rows_written += 1

        outcome = await _upsert_provisional_sku_map(
            style_id=str(style["_id"]),
            platform=payload.platform,
            group_sku=group_sku,
            color_name=cr["color_name"],
            sizes_covered=sizes,
            user_email=u["email"],
        )
        sku_map_summary[outcome] = sku_map_summary.get(outcome, 0) + 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await log_activity(
        "catalogue.export",
        "catalogue_export",
        f"Generated {payload.platform} listing file for style {style_code} "
        f"({len(color_rows)} colours × {len(sizes)} sizes = {rows_written} rows; "
        f"sku_map: {sku_map_summary})",
        u["email"],
    )

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{style_code}_{payload.platform}_listing_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Style-Code":         style_code,
            "X-Rows-Written":       str(rows_written),
            "X-Colors":             str(len(color_rows)),
            "X-Sizes":              str(len(sizes)),
            "X-SkuMap-Created":     str(sku_map_summary.get("created", 0)),
            "X-SkuMap-Updated":     str(sku_map_summary.get("updated", 0)),
            "X-SkuMap-Unchanged":   str(sku_map_summary.get("unchanged", 0)),
            "Access-Control-Expose-Headers":
                "Content-Disposition, X-Style-Code, X-Rows-Written, X-Colors, "
                "X-Sizes, X-SkuMap-Created, X-SkuMap-Updated, X-SkuMap-Unchanged",
        },
    )


@styles_router.post("/catalogue-export/preview")
async def catalogue_export_preview(payload: CatalogueExportRequest, request: Request):
    await _get_user(request)
    db = get_db()
    style = await db.styles.find_one({"_id": oid(payload.style_id)})
    if not style:
        raise HTTPException(404, "Style not found")
    style_code = style.get("code", "")
    cfg = await db.listing_format_configs.find_one({"platform": payload.platform})
    if not cfg:
        raise HTTPException(404, f"No listing-format config for platform '{payload.platform}'")
    export_template = cfg.get("export_template")
    if not export_template:
        raise HTTPException(400, f"Platform '{payload.platform}' has no export_template configured")

    lifecycle = await db.style_lifecycle.find_one({"style_id": str(style["_id"])})
    colors = [c for c in (payload.colors or []) if c and c.strip()]
    sizes  = [str(s) for s in (payload.sizes  or []) if s and str(s).strip()]
    if not colors and lifecycle: colors = [c for c in (lifecycle.get("planned_colors") or []) if c]
    if not sizes  and lifecycle: sizes  = [str(s) for s in (lifecycle.get("planned_sizes")  or []) if s]

    color_rows, unmapped = [], []
    for cname in colors:
        cc = await resolve_color_code(cname)
        (unmapped if not cc else color_rows).append({"color_name": cname, "color_code": cc} if cc else cname)

    style_dict = stringify(style)
    lifecycle_dict = stringify(lifecycle) if lifecycle else None
    columns = export_template.get("columns") or []
    header  = [c.get("name", "") for c in columns]
    rows    = []
    for cr in color_rows:
        for sz in sizes:
            row = []
            for col in columns:
                row.append(_resolve_export_source(
                    col,
                    style=style_dict, lifecycle=lifecycle_dict,
                    style_code=style_code,
                    color_name=cr["color_name"], color_code=cr["color_code"],
                    size=sz,
                ))
            rows.append(row)

    return {
        "style_code":       style_code,
        "platform":         payload.platform,
        "sheet_name":       export_template.get("sheet_name"),
        "header_row_index": export_template.get("header_row_index", 0),
        "header":           header,
        "rows":             rows,
        "row_count":        len(rows),
        "colors":           [cr["color_name"] for cr in color_rows],
        "sizes":            sizes,
        "unmapped_colors":  unmapped,
    }


@styles_router.post("/costing/preview")
async def costing_preview(payload: StyleIn, request: Request):
    await _get_user(request)
    return compute_style_costing(payload.model_dump())
