"""Marketplace Online Orders, Configured Imports, Dispatch Imports, Monthly Reports, and Settlement Routes."""

import io
import os
import re
import csv
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any, Tuple
from io import BytesIO
import inspect

from fastapi import APIRouter, HTTPException, Request, Depends, UploadFile, File, Query
from bson import ObjectId
from pydantic import BaseModel, Field, field_validator
from pydantic_core import PydanticCustomError
from pymongo.errors import DuplicateKeyError
from pymongo import ReturnDocument

from auth import get_current_user_factory, require_roles
from models.orders import POIn, POLineItem, ProductionStageUpdate
from models.sku_map import Platform, SheetLocator, HeaderLocator
from rate_limiter import upload_rate_limiter, bulk_import_rate_limiter

log = logging.getLogger("online_orders_routes")

online_orders_router = APIRouter(prefix="/api", tags=["Online Orders & Marketplace Imports"])


def get_db():
    import server
    return server.db


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def oid(v) -> ObjectId:
    try:
        return ObjectId(v)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")


def stringify(doc: dict) -> dict:
    if doc is None:
        return doc
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc.pop("_id"))
    for key, value in doc.items():
        if isinstance(value, ObjectId):
            doc[key] = str(value)
        elif isinstance(value, dict):
            doc[key] = stringify(value)
        elif isinstance(value, list):
            doc[key] = [stringify(item) if isinstance(item, dict) else (str(item) if isinstance(item, ObjectId) else item) for item in value]
    return doc


async def _get_user(request: Optional[Request] = None):
    if request is not None:
        user = getattr(request.state, "user", None)
        if user:
            return user
    import server
    if getattr(server, "get_current_user", None) is not None:
        return await server.get_current_user(request)
    from auth import get_current_user_factory
    fn = await get_current_user_factory(get_db())
    return await fn(request)


async def _log_activity(action: str, category: str, details: str, email: str, db=None):
    db = db if db is not None else get_db()
    try:
        await db.audit_logs.insert_one({
            "action": action,
            "category": category,
            "details": details,
            "user": email,
            "timestamp": now_iso(),
        })
    except Exception as e:
        log.warning(f"Failed to log activity: {e}")


# ═══════════════════════════════════════════════════════════════════════
# Canonical Fields & Schemas
# ═══════════════════════════════════════════════════════════════════════

ORDER_CANONICAL_FIELDS = [
    "order_id", "order_item_id", "shipment_id",
    "order_date", "dispatch_by_date",
    "leaf_sku", "myntra_sku_code", "color", "size",
    "product_title", "qty",
    "selling_price", "invoice_amount",
    "order_state", "tracking_id",
    "buyer_name", "city", "state", "pincode",
    "bin_barcode",
]

DISPATCH_CANONICAL_FIELDS = [
    "order_id", "order_release_id",
    "leaf_sku", "channel_sku",
    "packed_on", "status",
    "mrp", "selling_value",
    "cgst", "sgst", "igst",
    "tracking_id",
    "destination_city", "destination_state", "destination_pincode",
    "store_packet_id",
    "product_title", "qty",
]

MONTHLY_REPORT_CANONICAL_FIELDS = [
    "order_id", "order_release_id",
    "leaf_sku", "size", "product_title",
    "order_status",
    "packed_on", "delivered_on", "cancelled_on",
    "rto_creation_date", "return_creation_date",
    "final_amount", "total_mrp", "discount", "seller_price",
]

SETTLEMENT_CANONICAL_FIELDS = [
    "order_ref", "leaf_sku",
    "gross_amount", "commission", "shipping_fee", "rto_charge", "gst_on_fees", "fixed_fee", "fees_total", "net_payout",
    "settlement_date", "payment_id",
]

ConfigRole = Literal["order", "dispatch", "monthly_report", "settlement"]


class OrderImportFormatConfigIn(BaseModel):
    platform: Platform
    role: ConfigRole = "order"
    sheet_locator: SheetLocator
    header_locator: HeaderLocator
    skip_rows_after_header: int = 0
    column_map: Dict[str, Optional[str]]
    known_sku_prefixes_to_strip: List[str] = Field(default_factory=list)
    known_sku_prefix_replacements: Dict[str, str] = Field(default_factory=dict)
    is_picklist: bool = False
    active: bool = True
    notes: Optional[str] = ""

    @field_validator("column_map")
    @classmethod
    def _order_column_map_leaf_sku(cls, v):
        if not isinstance(v, dict):
            raise PydanticCustomError("column_map_type", "column_map must be an object")
        if not v.get("leaf_sku"):
            raise PydanticCustomError(
                "column_map_leaf_sku",
                "column_map.leaf_sku is required — every order/picklist file must expose our internal SKU column"
            )
        return v


class OrderImportFormatConfigUpdate(BaseModel):
    sheet_locator: Optional[SheetLocator] = None
    header_locator: Optional[HeaderLocator] = None
    skip_rows_after_header: Optional[int] = None
    column_map: Optional[Dict[str, Optional[str]]] = None
    known_sku_prefixes_to_strip: Optional[List[str]] = None
    known_sku_prefix_replacements: Optional[Dict[str, str]] = None
    is_picklist: Optional[bool] = None
    active: Optional[bool] = None
    notes: Optional[str] = None

    @field_validator("column_map")
    @classmethod
    def _order_column_map_opt(cls, v):
        if v is None: return v
        if not isinstance(v, dict):
            raise PydanticCustomError("column_map_type", "column_map must be an object")
        if not v.get("leaf_sku"):
            raise PydanticCustomError(
                "column_map_leaf_sku",
                "column_map.leaf_sku is required"
            )
        return v


DEFAULT_ORDER_IMPORT_CONFIGS = [
    {
        "platform": "flipkart",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_id":         "Order Id",
            "order_item_id":    "ORDER ITEM ID",
            "shipment_id":      "Shipment ID",
            "order_date":       "Ordered On",
            "leaf_sku":         "SKU",
            "product_title":    "Product",
            "qty":              "Quantity",
            "selling_price":    "Selling Price Per Item",
            "invoice_amount":   "Invoice Amount",
            "order_state":      "Order State",
            "tracking_id":      "Tracking ID",
            "dispatch_by_date": "Dispatch by date",
            "buyer_name":       "Buyer name",
            "city":             "City",
            "state":            "State",
            "pincode":          "PIN Code",
        },
        "known_sku_prefixes_to_strip": ["TH"],
        "known_sku_prefix_replacements": {},
        "is_picklist": False,
        "active": True,
        "notes": "Flipkart order-CSV export.",
    },
    {
        "platform": "myntra",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_id":         None,
            "myntra_sku_code":  "myntraSkuCode",
            "leaf_sku":         "sellerSkuCode",
            "product_title":    "productDescription",
            "qty":              "quantity",
            "bin_barcode":      "binBarcode",
        },
        "known_sku_prefixes_to_strip": [],
        "known_sku_prefix_replacements": {"FLL": "FL"},
        "is_picklist": True,
        "active": True,
        "notes": "Myntra picklist (OP-xxxxx.csv).",
    },
    {
        "platform": "myntra",
        "role": "dispatch",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_id":            "Order id",
            "order_release_id":    "Order_release_id",
            "leaf_sku":            "Seller_sku_code",
            "channel_sku":         "Myntra SKU code",
            "packed_on":           "Packed On",
            "status":              "Status",
            "mrp":                 "MRP",
            "selling_value":       "Selling value",
            "cgst":                "CGST",
            "sgst":                "SGST",
            "igst":                "IGST",
            "tracking_id":         "Tracking_id",
            "destination_city":    "Destination City",
            "destination_state":   "Destination state",
            "destination_pincode": "Destination pincode",
            "store_packet_id":     "Store Packet ID",
        },
        "known_sku_prefixes_to_strip": [],
        "known_sku_prefix_replacements": {"FLL": "FL"},
        "is_picklist": False,
        "active": True,
        "notes": "Myntra daily dispatch file (Packed_order_data.csv).",
    },
    {
        "platform": "myntra",
        "role": "monthly_report",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_id":             "order id fk",
            "order_release_id":     "order release id",
            "leaf_sku":             "seller sku code",
            "size":                 "size",
            "order_status":         "order status",
            "packed_on":            "packed on",
            "delivered_on":         "delivered on",
            "cancelled_on":         "cancelled on",
            "rto_creation_date":    "rto creation date",
            "return_creation_date": "return creation date",
            "final_amount":         "final amount",
            "total_mrp":            "total mrp",
            "discount":             "discount",
            "seller_price":         "seller price",
        },
        "known_sku_prefixes_to_strip": [],
        "known_sku_prefix_replacements": {"FLL": "FL"},
        "is_picklist": False,
        "active": True,
        "notes": "Myntra Monthly_order_report.csv.",
    },
    {
        "platform": "myntra",
        "role": "settlement",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_ref":       "order release id",
            "leaf_sku":        "seller sku code",
            "gross_amount":    "gross amount",
            "commission":      "commission",
            "shipping_fee":    "shipping fee",
            "rto_charge":      "rto charge",
            "net_payout":      "net payout",
            "settlement_date": "settlement date",
            "payment_id":      "payment id",
        },
        "known_sku_prefixes_to_strip": [],
        "known_sku_prefix_replacements": {"FLL": "FL"},
        "is_picklist": False,
        "active": True,
        "notes": "Myntra Settlement Advice CSV.",
    },
    {
        "platform": "flipkart",
        "role": "settlement",
        "sheet_locator": {"type": "first_sheet"},
        "header_locator": {"type": "fixed_row", "row": 0},
        "skip_rows_after_header": 0,
        "column_map": {
            "order_ref":       "order_id",
            "leaf_sku":        "sku",
            "gross_amount":    "sale_amount",
            "commission":      "commission",
            "shipping_fee":    "shipping_fee",
            "rto_charge":      "reverse_shipping_fee",
            "net_payout":      "bank_settlement_value",
            "settlement_date": "settlement_date",
            "payment_id":      "neft_id",
        },
        "known_sku_prefixes_to_strip": ["TH"],
        "known_sku_prefix_replacements": {},
        "is_picklist": False,
        "active": True,
        "notes": "Flipkart Settlement Report.",
    },
]


async def _seed_order_import_format_configs(db=None) -> int:
    db = db if db is not None else get_db()
    inserted = 0
    try:
        await db.order_import_format_configs.update_many(
            {"role": {"$exists": False}},
            {"$set": {"role": "order"}}
        )
    except Exception as e:
        log.warning(f"Could not backfill role on order_import_format_configs: {e}")

    try:
        idx_info = await db.order_import_format_configs.index_information()
        if "oifc_platform_unique" in idx_info:
            await db.order_import_format_configs.drop_index("oifc_platform_unique")
    except Exception as e:
        log.warning(f"Could not inspect/drop old index: {e}")

    try:
        await db.order_import_format_configs.create_index(
            [("platform", 1), ("role", 1)],
            unique=True,
            name="oifc_platform_role_unique"
        )
    except Exception as e:
        log.warning(f"Could not create order_import_format_configs composite index: {e}")

    for cfg in DEFAULT_ORDER_IMPORT_CONFIGS:
        role = cfg.get("role", "order")
        existing = await db.order_import_format_configs.find_one(
            {"platform": cfg["platform"], "role": role}
        )
        if not existing:
            existing = await db.order_import_format_configs.find_one(
                {"platform": cfg["platform"], "role": {"$exists": False}}
            )

        if existing:
            update_fields = {}
            if "role" not in existing:
                update_fields["role"] = role
            if "known_sku_prefixes_to_strip" not in existing and "known_sku_prefixes_to_strip" in cfg:
                update_fields["known_sku_prefixes_to_strip"] = cfg["known_sku_prefixes_to_strip"]
            if "known_sku_prefix_replacements" not in existing and "known_sku_prefix_replacements" in cfg:
                update_fields["known_sku_prefix_replacements"] = cfg["known_sku_prefix_replacements"]
            if "is_picklist" not in existing and "is_picklist" in cfg:
                update_fields["is_picklist"] = cfg["is_picklist"]
            if update_fields:
                try:
                    await db.order_import_format_configs.update_one(
                        {"_id": existing["_id"]},
                        {"$set": update_fields}
                    )
                except Exception as e:
                    log.warning(f"Could not update order_import_format_config {cfg['platform']}: {e}")
            continue

        doc = dict(cfg)
        doc["role"] = role
        doc["created_at"] = now_iso()
        doc["updated_at"] = now_iso()
        doc["created_by"] = "system"
        try:
            await db.order_import_format_configs.insert_one(doc)
            inserted += 1
        except DuplicateKeyError:
            pass
    return inserted


# ═══════════════════════════════════════════════════════════════════════
# Helpers: File Parsers & Column Matching
# ═══════════════════════════════════════════════════════════════════════

def _norm_token(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def strip_known_prefixes(leaf_sku: str, prefixes: List[str]) -> str:
    s = (leaf_sku or "").strip()
    for pfx in prefixes or []:
        pfx_clean = str(pfx or "").strip()
        if not pfx_clean:
            continue
        for delim in ["-", "_", ""]:
            full = f"{pfx_clean}{delim}"
            if s.upper().startswith(full.upper()):
                s = s[len(full):].strip()
                break
    return s


COMMON_COLUMN_ALIAS_GROUPS = [
    {"withdrawl", "withdrawal", "withdrawals", "withdrawalamt", "withdrawalamount", "debit", "debitamount", "dramount", "dr"},
    {"deposit", "deposits", "depositamt", "depositamount", "credit", "creditamount", "cramount", "cr"},
    {"trandate", "transactiondate", "txndate", "date", "txdate", "postingdate"},
    {"valuedate"},
    {"narration", "description", "particulars", "remarks", "transactionremarks", "transactionparticulars"},
    {"balance", "closingbalance", "runningbalance", "netbalance"},
    {"chqno", "chqrefno", "chequeno", "refno", "referenceno", "reference", "utr", "utrno", "chqrefno", "chequereferenceno"},
]


def _resolve_column(target: str, actual_headers: List[str]) -> Optional[str]:
    if not target or not actual_headers:
        return None
    for h in actual_headers:
        if h == target:
            return h
    t_lower = target.strip().lower()
    for h in actual_headers:
        if h.strip().lower() == t_lower:
            return h
    t_norm = _norm_token(target)
    for h in actual_headers:
        if _norm_token(h) == t_norm:
            return h

    # Check alias groups (e.g. withdrawal <-> withdrawl)
    for group in COMMON_COLUMN_ALIAS_GROUPS:
        if t_norm in group:
            for h in actual_headers:
                if _norm_token(h) in group:
                    return h
    return None


def _detect_tabular_file_format(content: bytes, filename: str = "") -> str:
    """
    Detect whether file is legacy Excel (.xls / OLE2), modern Excel (.xlsx / OOXML zip), or CSV.
    Uses magic bytes inspection first, falling back to file extension or content decoding.
    Raises HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv") if unsupported.
    """
    if not content:
        raise HTTPException(400, "Uploaded file is empty")

    fname_lc = (filename or "").strip().lower()

    # 1. Magic bytes check (highest priority)
    # Legacy OLE2 compound document signature for .xls: D0 CF 11 E0
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls"

    # OOXML / Zip archive signature for .xlsx: PK\x03\x04, PK\x05\x06, PK\x07\x08
    if content.startswith(b"PK\x03\x04") or content.startswith(b"PK\x05\x06") or content.startswith(b"PK\x07\x08"):
        return "xlsx"

    # 2. Filename extension checks
    if fname_lc.endswith(".xls"):
        return "xls"
    if fname_lc.endswith(".xlsx") or fname_lc.endswith(".xlsm") or fname_lc.endswith(".xltx"):
        return "xlsx"
    if fname_lc.endswith(".csv") or fname_lc.endswith(".tsv") or fname_lc.endswith(".txt"):
        try:
            content.decode("utf-8-sig")
            return "csv"
        except UnicodeDecodeError:
            try:
                content.decode("latin-1")
                return "csv"
            except Exception:
                raise HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv")

    # 3. Plain text / CSV sniffing (no null bytes in initial chunk)
    if b"\x00" not in content[:1024]:
        try:
            content.decode("utf-8-sig")
            return "csv"
        except UnicodeDecodeError:
            try:
                content.decode("latin-1")
                return "csv"
            except Exception:
                pass

    raise HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv")


def _parse_xlrd_workbook(content: bytes, sheet_locator: SheetLocator) -> List[List[str]]:
    try:
        import xlrd
    except ImportError:
        raise HTTPException(500, "xlrd is required to parse legacy .xls Excel files")
    try:
        wb = xlrd.open_workbook(file_contents=content, formatting_info=True)
    except Exception:
        try:
            wb = xlrd.open_workbook(file_contents=content)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse .xls file: {str(e)}")

    sheet = None
    if sheet_locator.type == "first_sheet":
        sheet = wb.sheet_by_index(0)
    elif sheet_locator.type == "fixed_name":
        if sheet_locator.name not in wb.sheet_names():
            raise HTTPException(400, f"Sheet '{sheet_locator.name}' not found in workbook")
        sheet = wb.sheet_by_name(sheet_locator.name)
    elif sheet_locator.type == "name_contains":
        sub = (sheet_locator.substring or "").lower()
        for sname in wb.sheet_names():
            if sub in sname.lower():
                sheet = wb.sheet_by_name(sname)
                break
        if not sheet:
            sheet = wb.sheet_by_index(0)
    else:
        sheet = wb.sheet_by_index(0)

    rows_raw = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            ctype = sheet.cell_type(r, c)
            val = sheet.cell_value(r, c)
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(val, wb.datemode)
                    row.append(dt.strftime("%Y-%m-%d %H:%M:%S") if (dt.hour or dt.minute or dt.second) else dt.strftime("%Y-%m-%d"))
                except Exception:
                    row.append(str(val))
            elif ctype == xlrd.XL_CELL_NUMBER:
                if isinstance(val, float) and val.is_integer():
                    row.append(str(int(val)))
                else:
                    row.append(str(val))
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                row.append("TRUE" if val else "FALSE")
            elif ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
                row.append("")
            else:
                row.append(str(val) if val is not None else "")
        rows_raw.append(row)

    for (rlo, rhi, clo, chi) in getattr(sheet, "merged_cells", []):
        top_left_val = rows_raw[rlo][clo] if rlo < len(rows_raw) and clo < len(rows_raw[rlo]) else ""
        if top_left_val:
            for r in range(rlo, min(rhi, len(rows_raw))):
                for c in range(clo, min(chi, len(rows_raw[r]))):
                    if not rows_raw[r][c]:
                        rows_raw[r][c] = top_left_val

    return rows_raw


def _parse_tabular_bytes(
    content: bytes,
    filename: str,
    sheet_locator: SheetLocator,
    header_locator: HeaderLocator,
    skip_rows_after_header: int = 0,
) -> Tuple[List[str], List[Dict[str, str]]]:
    fmt = _detect_tabular_file_format(content, filename)

    if fmt == "xls":
        rows_raw = _parse_xlrd_workbook(content, sheet_locator)

    elif fmt == "xlsx":
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "openpyxl is required to parse Excel files")
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
            sheet = None
            if sheet_locator.type == "first_sheet":
                sheet = wb.worksheets[0]
            elif sheet_locator.type == "fixed_name":
                if sheet_locator.name not in wb.sheetnames:
                    raise HTTPException(400, f"Sheet '{sheet_locator.name}' not found in workbook")
                sheet = wb[sheet_locator.name]
            elif sheet_locator.type == "name_contains":
                sub = (sheet_locator.substring or "").lower()
                for sname in wb.sheetnames:
                    if sub in sname.lower():
                        sheet = wb[sname]
                        break
                if not sheet:
                    sheet = wb.worksheets[0]
            else:
                sheet = wb.worksheets[0]

            rows_raw = []
            for r in sheet.iter_rows(values_only=True):
                rows_raw.append([str(c) if c is not None else "" for c in r])
        except Exception as e:
            # If openpyxl fails because file was actually legacy .xls (OLE2 format)
            if "xlrd" in str(e).lower() or content.startswith(b"\xd0\xcf\x11\xe0"):
                try:
                    rows_raw = _parse_xlrd_workbook(content, sheet_locator)
                except Exception:
                    raise HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv")
            else:
                raise HTTPException(400, f"Failed to parse Excel file: {str(e)}")

    elif fmt == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except Exception:
                raise HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv")
        reader = csv.reader(io.StringIO(text))
        rows_raw = list(reader)
    else:
        raise HTTPException(400, "Unsupported file format — please upload .xls, .xlsx, or .csv")

    if not rows_raw:
        return [], []

    header_row_idx = 0
    if header_locator.type == "fixed_row":
        header_row_idx = max(0, header_locator.row or 0)
    elif header_locator.type == "scan_for_columns":
        raw_kws = header_locator.must_contain_any or []
        if not raw_kws:
            raw_kws = [
                "Tran. Date", "Transaction Date", "Txn Date", "Value Date", "Date",
                "Withdrawl", "Withdrawal", "Withdrawals", "Withdrawal Amt.", "Debit", "Debit Amount", "Dr Amount",
                "Deposit", "Deposits", "Deposit Amt.", "Credit", "Credit Amount", "Cr Amount",
                "Balance", "Closing Balance", "Running Balance",
                "Narration", "Description", "Particulars", "Remarks",
                "Chq. No.", "Chq./Ref.No.", "Cheque No", "Ref No", "Reference No"
            ]
        must_contain = [_norm_token(kw) for kw in raw_kws if kw]
        best_idx = 0
        best_score = 0
        for idx, r in enumerate(rows_raw):
            row_tokens = [_norm_token(c) for c in r if c and str(c).strip()]
            if not row_tokens:
                continue
            matched_kws = set()
            for kw in must_contain:
                for cell in row_tokens:
                    if kw == cell or (len(kw) >= 4 and kw in cell):
                        matched_kws.add(kw)
                        break
            score = len(matched_kws)
            if score > best_score:
                best_score = score
                best_idx = idx

        if best_score > 0:
            header_row_idx = best_idx
        else:
            header_row_idx = 0

    if header_row_idx >= len(rows_raw):
        return [], []

    raw_headers = rows_raw[header_row_idx]
    headers = [str(h).strip() for h in raw_headers]

    data_start = header_row_idx + 1 + max(0, skip_rows_after_header)
    parsed_rows = []
    for r in rows_raw[data_start:]:
        if not any(str(c).strip() for c in r):
            continue
        row_dict = {}
        for h_idx, h_name in enumerate(headers):
            if not h_name:
                continue
            val = r[h_idx].strip() if h_idx < len(r) else ""
            row_dict[h_name] = val
        parsed_rows.append(row_dict)

    return headers, parsed_rows


# ═══════════════════════════════════════════════════════════════════════
# Endpoints: Order Import Format Configs
# ═══════════════════════════════════════════════════════════════════════

@online_orders_router.get("/order-import-format-configs")
async def list_order_import_format_configs(
    request: Request,
    active: Optional[bool] = None,
    role: Optional[ConfigRole] = None,
):
    await _get_user(request)
    db = get_db()
    q = {}
    if active is not None:
        q["active"] = active
    if role:
        q["role"] = role
    docs = await db.order_import_format_configs.find(q).sort("platform", 1).to_list(1000)
    return [stringify(d) for d in docs]


@online_orders_router.get("/order-import-format-configs/{platform}")
async def get_order_import_format_config(
    platform: str,
    request: Request,
    role: ConfigRole = "order",
):
    await _get_user(request)
    db = get_db()
    platform_lc = platform.lower()
    doc = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": role}
    )
    if not doc and role == "order":
        doc = await db.order_import_format_configs.find_one(
            {"platform": platform_lc, "role": {"$exists": False}}
        )
    if not doc:
        raise HTTPException(404, f"Order import format config for platform '{platform}' with role '{role}' not found")
    return stringify(doc)


@online_orders_router.post("/order-import-format-configs")
async def create_order_import_format_config(payload: OrderImportFormatConfigIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = payload.platform.lower()
    role = payload.role or "order"
    if await db.order_import_format_configs.find_one({"platform": platform_lc, "role": role}):
        raise HTTPException(409, f"Config for platform '{platform_lc}' with role '{role}' already exists")
    doc = payload.model_dump()
    doc["platform"] = platform_lc
    doc["role"] = role
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    doc["created_by"] = u.get("email", "unknown")
    res = await db.order_import_format_configs.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    doc.pop("_id", None)
    await _log_activity(
        "order_import_format.create", "order_import_format_configs",
        f"Created order import format config for platform={platform_lc}, role={role}",
        u["email"], db=db
    )
    return doc


@online_orders_router.patch("/order-import-format-configs/{platform}")
async def update_order_import_format_config(
    platform: str,
    payload: OrderImportFormatConfigUpdate,
    request: Request,
    role: ConfigRole = "order",
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.lower()
    existing = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": role}
    )
    if not existing and role == "order":
        existing = await db.order_import_format_configs.find_one(
            {"platform": platform_lc, "role": {"$exists": False}}
        )
    if not existing:
        raise HTTPException(404, f"Config for platform '{platform_lc}' with role '{role}' not found")
    update = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    update["updated_at"] = now_iso()
    await db.order_import_format_configs.update_one({"_id": existing["_id"]}, {"$set": update})
    fresh = await db.order_import_format_configs.find_one({"_id": existing["_id"]})
    await _log_activity(
        "order_import_format.update", "order_import_format_configs",
        f"Updated order import format config for platform={platform_lc}, role={role}",
        u["email"], db=db
    )
    return stringify(fresh)


@online_orders_router.delete("/order-import-format-configs/{platform}")
async def delete_order_import_format_config(
    platform: str,
    request: Request,
    role: ConfigRole = "order",
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.lower()
    res = await db.order_import_format_configs.delete_one({"platform": platform_lc, "role": role})
    if res.deleted_count == 0 and role == "order":
        res = await db.order_import_format_configs.delete_one({"platform": platform_lc, "role": {"$exists": False}})
    if res.deleted_count == 0:
        raise HTTPException(404, f"Config for platform '{platform_lc}' with role '{role}' not found")
    await _log_activity(
        "order_import_format.delete", "order_import_format_configs",
        f"Deleted order import format config for platform={platform_lc}, role={role}",
        u["email"], db=db
    )
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════
# Endpoints: Order Imports & Marketplace Pipelines
# ═══════════════════════════════════════════════════════════════════════

class OnlineOrderImportResult(BaseModel):
    channel: str
    imported: int
    unresolved: int
    errors: List[dict]


@online_orders_router.post("/online-orders/import", dependencies=[Depends(bulk_import_rate_limiter)])
async def import_online_orders(
    file: UploadFile = File(...),
    channel: str = "myntra",
    order_date: Optional[str] = None,
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()

    channel = channel.strip().lower()
    if channel not in ["myntra", "ajio", "flipkart", "nykaa", "amazon", "website", "unicommerce"]:
        raise HTTPException(400, f"Unknown channel '{channel}'. Must be one of: myntra, ajio, flipkart, nykaa, amazon, website, unicommerce")

    today = (order_date or now_iso()[:10])

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))

    def norm(row: dict) -> dict:
        return {k.strip().lower().replace(" ", "_"): (v or "").strip() for k, v in row.items()}

    imported = 0
    unresolved = 0
    errors = []
    fulfilled_from_stock = 0
    picklist_lines_by_order: Dict[str, List[dict]] = {}
    in_flight_covered: Dict[tuple, int] = {}

    from routes.pos import _get_stage_durations, _compute_deadline
    from routes.sku_map import resolve_style
    from routes.wms import _generate_picklist_for_order

    durations = await _get_stage_durations(db=db)
    jobs_to_insert = []
    entered = now_iso()
    deadline = _compute_deadline(entered, durations.get("procurement", 24))

    all_styles = await db.styles.find({}, {"code": 1, "_id": 1}).to_list(10000)
    styles_by_code = {s["code"].strip().upper(): str(s["_id"]) for s in all_styles}

    for row_idx, raw_row in enumerate(reader, start=2):
        r = norm(raw_row)
        order_id = r.get("order_id") or r.get("order_number") or r.get("order_no") or r.get("shipment_id") or ""
        raw_sku = r.get("style_sku") or r.get("sku") or r.get("seller_sku") or r.get("sku_code") or ""
        qty_str = r.get("quantity") or r.get("qty") or "1"
        color = r.get("color") or ""
        size = r.get("size") or ""
        description = r.get("description") or r.get("product_name") or ""
        unit_price = float(r.get("unit_price") or r.get("price") or r.get("mrp") or 0.0)
        delivery_date = r.get("delivery_date") or r.get("expected_delivery") or ""

        if not order_id:
            errors.append({"row": row_idx, "order_id": "", "style_sku": raw_sku, "reason": "Missing order_id"})
            continue
        if not raw_sku:
            errors.append({"row": row_idx, "order_id": order_id, "style_sku": "", "reason": "Missing style_sku"})
            continue
        try:
            quantity = int(qty_str)
            if quantity <= 0:
                raise ValueError()
        except ValueError:
            errors.append({"row": row_idx, "order_id": order_id, "style_sku": raw_sku, "reason": f"Invalid quantity '{qty_str}'"})
            continue

        result = await resolve_style(
            source_type="online_channel",
            source_name=channel,
            external_sku=raw_sku,
            external_color=color or None,
            external_size=size or None,
            db=db,
        )

        if not result["matched"]:
            unresolved += 1
            errors.append({
                "row": row_idx, "order_id": order_id, "style_sku": raw_sku,
                "reason": f"Style SKU '{raw_sku}' not found in Style Master or SKU Mappings for channel '{channel}'",
            })
            continue

        match_status = "mapped" if result["match_via"] in ("sku_map", "marketplace_resolver") else "matched"
        style_doc_id = result.get("style_id") or styles_by_code.get(result["style_code"].upper())
        resolved_color = result.get("color") or color or ""
        resolved_size = str(result.get("size") or size or "")

        covered_qty = 0
        remaining_qty = quantity

        if style_doc_id:
            loc_docs = await db.fg_location_inventory.find({
                "style_id": str(style_doc_id),
                "color": {"$regex": f"^{re.escape(resolved_color)}$", "$options": "i"} if resolved_color else {"$in": ["", None]},
                "size": resolved_size,
                "qty": {"$gt": 0},
            }).to_list(100)

            total_avail = max(0, sum(loc.get("qty", 0) - loc.get("reserved_qty", 0) for loc in loc_docs))
            inflight_key = (str(style_doc_id), resolved_color.lower(), resolved_size)
            prior_claimed = in_flight_covered.get(inflight_key, 0)
            effective_avail = max(0, total_avail - prior_claimed)

            if effective_avail > 0:
                covered_qty = min(quantity, effective_avail)
                remaining_qty = quantity - covered_qty
                in_flight_covered[inflight_key] = prior_claimed + covered_qty

                if order_id not in picklist_lines_by_order:
                    picklist_lines_by_order[order_id] = []
                picklist_lines_by_order[order_id].append({
                    "style_id": str(style_doc_id),
                    "style_code": result["style_code"],
                    "color": resolved_color,
                    "size": resolved_size,
                    "ordered_qty": quantity,
                    "qty": quantity,
                })

        if remaining_qty <= 0:
            imported += 1
            fulfilled_from_stock += covered_qty
            continue

        job = {
            "po_id": None,
            "po_number": order_id,
            "client_name": channel,
            "channel": channel,
            "source_type": "online_channel",
            "order_date": today,
            "style_code": result["style_code"],
            "style_id": result["style_id"],
            "style_match_status": match_status,
            **({"mapped_from_sku": result["mapped_from_sku"], "sku_mapping_id": result["mapping_id"]} if result["match_via"] in ("sku_map", "marketplace_resolver") else {}),
            "description": description,
            "color": result["color"],
            "size": result["size"],
            "quantity": remaining_qty,
            "original_order_qty": quantity,
            "fulfilled_from_stock_qty": covered_qty,
            "unit_price": unit_price,
            "amount": round(unit_price * remaining_qty, 2),
            "completed_qty": 0,
            "rejected_qty": 0,
            "delivery_date": delivery_date,
            "stage": "procurement",
            "stage_entered_at": entered,
            "stage_deadline": deadline,
            "split_from_job_id": None,
            "split_history": None,
            "created_at": now_iso(),
            "updated_at": now_iso(),
            "history": [{"stage": "procurement", "at": now_iso(), "by": u["email"],
                         "notes": f"Auto-created from {channel} CSV import"
                                  + (f" (partial: {covered_qty} pairs shipped from ready stock)" if covered_qty else "")}],
        }
        jobs_to_insert.append(job)
        imported += 1
        fulfilled_from_stock += covered_qty

    if jobs_to_insert:
        await db.production_jobs.insert_many(jobs_to_insert)

    picklists_created = []
    for oid_key, lines in picklist_lines_by_order.items():
        try:
            pl_doc, covered_map, uncovered_map = await _generate_picklist_for_order(
                oid_key, channel, lines, u["email"], db=db)
            if pl_doc.get("_id"):
                picklists_created.append({
                    "picklist_no": pl_doc.get("picklist_no"),
                    "order_id": oid_key,
                    "items": pl_doc.get("total_items", 0),
                    "qty": pl_doc.get("total_qty", 0),
                })
        except Exception as pe:
            log.warning(f"Picklist generation failed for order {oid_key}: {pe}")

    await _log_activity(
        "IMPORT", "online_orders",
        f"{channel.capitalize()} CSV import: {imported} orders, {fulfilled_from_stock} pairs from stock, "
        f"{len(picklists_created)} picklists, {unresolved} unresolved, {len(errors)-unresolved} errors",
        u["email"], db=db
    )
    return {
        "channel": channel,
        "imported": imported,
        "unresolved": unresolved,
        "fulfilled_from_stock": fulfilled_from_stock,
        "picklists_created": picklists_created,
        "errors": errors,
    }


@online_orders_router.get("/online-orders")
async def list_online_orders(
    request: Request,
    channel: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    style_match_status: Optional[str] = None,
):
    await _get_user(request)
    db = get_db()
    query: dict = {"source_type": "online_channel"}
    if channel:
        query["channel"] = channel.lower()
    if style_match_status:
        query["style_match_status"] = style_match_status
    if from_date or to_date:
        date_q: dict = {}
        if from_date:
            date_q["$gte"] = from_date
        if to_date:
            date_q["$lte"] = to_date + "T23:59:59.999Z"
        query["created_at"] = date_q
    docs = await db.production_jobs.find(query).sort("created_at", -1).to_list(5000)
    return [stringify(j) for j in docs]


# ═══════════════════════════════════════════════════════════════════════
# Endpoints: Configured Order Import, Dispatch, Monthly & Settlements
# ═══════════════════════════════════════════════════════════════════════

def _sanitize_sheet_loc(data: dict) -> dict:
    d = dict(data or {})
    if d.get("type") == "first":
        d["type"] = "first_sheet"
    return d


def _sanitize_header_loc(data: dict) -> dict:
    d = dict(data or {})
    if d.get("type") == "row":
        d["type"] = "fixed_row"
        if "row_1_based" in d:
            d["row"] = max(0, int(d.pop("row_1_based")) - 1)
    return d


@online_orders_router.post("/online-orders/import-configured", dependencies=[Depends(upload_rate_limiter)])
async def import_configured_online_orders(
    file: UploadFile = File(...),
    platform: str = Query(..., description="Platform identifier matching order_import_format_configs"),
    order_date: Optional[str] = None,
    dry_run: bool = False,
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.strip().lower()

    cfg_doc = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": "order", "active": True}
    )
    if not cfg_doc:
        cfg_doc = await db.order_import_format_configs.find_one(
            {"platform": platform_lc, "active": True}
        )
    if not cfg_doc:
        raise HTTPException(400, f"No active order import format config found for platform '{platform_lc}'")

    content = await file.read()
    sheet_loc = SheetLocator(**_sanitize_sheet_loc(cfg_doc.get("sheet_locator", {"type": "first_sheet"})))
    header_loc = HeaderLocator(**_sanitize_header_loc(cfg_doc.get("header_locator", {"type": "fixed_row", "row": 0})))
    skip_rows = int(cfg_doc.get("skip_rows_after_header", 0) or 0)

    headers, rows = _parse_tabular_bytes(
        content=content,
        filename=file.filename or "",
        sheet_locator=sheet_loc,
        header_locator=header_loc,
        skip_rows_after_header=skip_rows,
    )

    if not rows:
        raise HTTPException(400, "No data rows found in uploaded file")

    col_map = cfg_doc.get("column_map") or {}
    resolved_cols = {}
    for canon_field, target_name in col_map.items():
        if target_name:
            actual = _resolve_column(target_name, headers)
            if actual:
                resolved_cols[canon_field] = actual

    leaf_sku_col = resolved_cols.get("leaf_sku")
    if not leaf_sku_col:
        raise HTTPException(400, f"Configured leaf_sku column '{col_map.get('leaf_sku')}' not found in file headers")

    from routes.sku_map import resolve_style, split_leaf_sku
    from routes.pos import _get_stage_durations, _compute_deadline

    prefixes_to_strip = cfg_doc.get("known_sku_prefixes_to_strip", [])
    prefix_replacements = cfg_doc.get("known_sku_prefix_replacements", {})
    is_picklist = bool(cfg_doc.get("is_picklist", False))

    batch_name = os.path.splitext(file.filename or "import")[0]
    today = (order_date or now_iso()[:10])
    durations = await _get_stage_durations(db=db)
    entered = now_iso()
    deadline = _compute_deadline(entered, durations.get("procurement", 24))

    matched_rows = []
    unresolved_rows = []

    for r_idx, r in enumerate(rows, start=1):
        raw_leaf = r.get(leaf_sku_col, "").strip()
        if not raw_leaf:
            continue

        # Prefix replacement & stripping
        cleaned_leaf = raw_leaf
        for wrong, right in prefix_replacements.items():
            if cleaned_leaf.startswith(wrong):
                cleaned_leaf = right + cleaned_leaf[len(wrong):]
        cleaned_leaf = strip_known_prefixes(cleaned_leaf, prefixes_to_strip)

        qty_str = r.get(resolved_cols.get("qty", ""), "1") if resolved_cols.get("qty") else "1"
        try:
            qty = int(qty_str)
            if qty <= 0: qty = 1
        except Exception:
            qty = 1

        color_val = r.get(resolved_cols.get("color", ""), "") if resolved_cols.get("color") else ""
        size_val = r.get(resolved_cols.get("size", ""), "") if resolved_cols.get("size") else ""

        if not color_val or not size_val:
            group_id, size_token, _ = split_leaf_sku(cleaned_leaf)
            if not size_val and size_token:
                size_val = size_token

        result = await resolve_style(
            source_type="online_channel",
            source_name=platform_lc,
            external_sku=cleaned_leaf,
            external_color=color_val or None,
            external_size=size_val or None,
            db=db,
        )

        order_id = r.get(resolved_cols.get("order_id", ""), "") if resolved_cols.get("order_id") else ""
        if not order_id and is_picklist:
            order_id = batch_name

        is_row_matched = bool(result.get("matched")) and (result.get("matched_exact") is not False) and (not result.get("unmapped_size")) and (not result.get("unmapped_color"))

        if is_row_matched:
            matched_rows.append({
                "row_index": r_idx,
                "order_id": order_id or batch_name,
                "style_code": result["style_code"],
                "style_id": result["style_id"],
                "color": result["color"],
                "size": result["size"],
                "quantity": qty,
                "unit_price": float(r.get(resolved_cols.get("selling_price", ""), 0.0) or 0.0) if resolved_cols.get("selling_price") else 0.0,
            })
        else:
            unresolved_rows.append({
                "row_index": r_idx,
                "order_id": order_id or batch_name,
                "raw_sku": raw_leaf,
                "cleaned_sku": cleaned_leaf,
                "color": color_val,
                "size": size_val,
                "quantity": qty,
            })

    if not dry_run and len(matched_rows) == 0:
        raise HTTPException(status_code=400, detail="Nothing to commit — no rows matched.")

    if not dry_run:
        jobs = []
        for m in matched_rows:
            jobs.append({
                "po_id": None,
                "po_number": m["order_id"],
                "client_name": platform_lc,
                "channel": platform_lc,
                "source_type": "online_channel",
                "order_date": today,
                "style_code": m["style_code"],
                "style_id": m["style_id"],
                "style_match_status": "matched",
                "color": m["color"],
                "size": m["size"],
                "quantity": m["quantity"],
                "unit_price": m["unit_price"],
                "amount": round(m["unit_price"] * m["quantity"], 2),
                "completed_qty": 0,
                "rejected_qty": 0,
                "stage": "procurement",
                "stage_entered_at": entered,
                "stage_deadline": deadline,
                "created_at": now_iso(),
                "updated_at": now_iso(),
                "history": [{"stage": "procurement", "at": now_iso(), "by": u["email"],
                             "notes": f"Configured import from {platform_lc}"}],
            })
        if jobs and hasattr(db, "production_jobs"):
            try:
                p_res = db.production_jobs.insert_many(jobs)
                if inspect.isawaitable(p_res):
                    await p_res
            except Exception:
                pass

        if unresolved_rows and hasattr(db, "online_order_exceptions"):
            exc_docs = [{
                "order_id": u["order_id"],
                "raw_sku": u.get("raw_sku", ""),
                "reason": "Unmapped color/size: not found in SKU map",
                "created_at": now_iso(),
            } for u in unresolved_rows]
            try:
                e_res = db.online_order_exceptions.insert_many(exc_docs)
                if inspect.isawaitable(e_res):
                    await e_res
            except Exception:
                pass

    return {
        "platform": platform_lc,
        "total_rows": len(rows),
        "matched_count": len(matched_rows),
        "unresolved_count": len(unresolved_rows),
        "dry_run": dry_run,
        "matched": matched_rows[:100],
        "unresolved": unresolved_rows[:100],
        "stats": {
            "total_rows_read": len(rows),
            "matched": len(matched_rows),
            "unmatched": len(unresolved_rows),
        },
        "committed": {
            "jobs_created": len(matched_rows) if not dry_run else 0,
            "exceptions_queued": len(unresolved_rows) if not dry_run else 0,
        },
    }


@online_orders_router.post("/online-orders/dispatch-import", dependencies=[Depends(bulk_import_rate_limiter)])
async def import_dispatch_orders(
    file: UploadFile = File(...),
    platform: str = Query(..., description="Platform identifier matching order_import_format_configs"),
    dry_run: bool = False,
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.strip().lower()

    cfg_doc = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": "dispatch", "active": True}
    )
    if not cfg_doc:
        raise HTTPException(400, f"No active dispatch import format config found for platform '{platform_lc}'")

    content = await file.read()
    sheet_loc = SheetLocator(**cfg_doc.get("sheet_locator", {"type": "first_sheet"}))
    header_loc = HeaderLocator(**cfg_doc.get("header_locator", {"type": "fixed_row", "row": 0}))
    skip_rows = int(cfg_doc.get("skip_rows_after_header", 0) or 0)

    headers, rows = _parse_tabular_bytes(
        content=content,
        filename=file.filename or "",
        sheet_locator=sheet_loc,
        header_locator=header_loc,
        skip_rows_after_header=skip_rows,
    )

    if not rows:
        raise HTTPException(400, "No data rows found in dispatch file")

    col_map = cfg_doc.get("column_map") or {}
    resolved_cols = {}
    for canon_field, target_name in col_map.items():
        if target_name:
            actual = _resolve_column(target_name, headers)
            if actual:
                resolved_cols[canon_field] = actual

    leaf_sku_col = resolved_cols.get("leaf_sku")
    if not leaf_sku_col:
        raise HTTPException(400, f"Configured leaf_sku column '{col_map.get('leaf_sku')}' not found in headers")

    from routes.sku_map import resolve_style, split_leaf_sku

    prefixes_to_strip = cfg_doc.get("known_sku_prefixes_to_strip", [])
    prefix_replacements = cfg_doc.get("known_sku_prefix_replacements", {})

    matched_rows = []
    unresolved_rows = []

    for r_idx, r in enumerate(rows, start=1):
        raw_leaf = r.get(leaf_sku_col, "").strip()
        if not raw_leaf:
            continue

        cleaned_leaf = raw_leaf
        for wrong, right in prefix_replacements.items():
            if cleaned_leaf.startswith(wrong):
                cleaned_leaf = right + cleaned_leaf[len(wrong):]
        cleaned_leaf = strip_known_prefixes(cleaned_leaf, prefixes_to_strip)

        group_id, size_token, _ = split_leaf_sku(cleaned_leaf)
        color_val = ""
        size_val = size_token or ""

        result = await resolve_style(
            source_type="online_channel",
            source_name=platform_lc,
            external_sku=cleaned_leaf,
            external_color=color_val or None,
            external_size=size_val or None,
            db=db,
        )

        order_id = r.get(resolved_cols.get("order_id", ""), "") if resolved_cols.get("order_id") else ""
        order_rel = r.get(resolved_cols.get("order_release_id", ""), "") if resolved_cols.get("order_release_id") else ""

        if result["matched"]:
            matched_rows.append({
                "row_index": r_idx,
                "order_id": order_id,
                "order_release_id": order_rel,
                "style_code": result["style_code"],
                "style_id": result["style_id"],
                "color": result["color"],
                "size": result["size"],
                "packed_on": r.get(resolved_cols.get("packed_on", ""), "") if resolved_cols.get("packed_on") else now_iso()[:10],
            })
        else:
            unresolved_rows.append({
                "row_index": r_idx,
                "order_id": order_id,
                "raw_sku": raw_leaf,
            })

    if not dry_run and len(matched_rows) == 0:
        raise HTTPException(status_code=400, detail="Nothing to commit — no rows matched.")

    if not dry_run:
        # Record movements
        movements = []
        for m in matched_rows:
            movements.append({
                "style_id": m["style_id"],
                "style_code": m["style_code"],
                "color": m["color"],
                "size": m["size"],
                "movement_type": "dispatched",
                "quantity": 1,
                "reference_id": m["order_release_id"] or m["order_id"],
                "platform": platform_lc,
                "created_at": now_iso(),
                "created_by": u["email"],
            })
        if movements:
            await db.fg_stock_movements.insert_many(movements)

    return {
        "platform": platform_lc,
        "total_rows": len(rows),
        "matched_count": len(matched_rows),
        "unresolved_count": len(unresolved_rows),
        "dry_run": dry_run,
    }


@online_orders_router.post("/online-orders/monthly-report-import", dependencies=[Depends(bulk_import_rate_limiter)])
async def import_monthly_report(
    file: UploadFile = File(...),
    platform: str = Query("myntra", description="Platform identifier"),
    dry_run: bool = False,
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.strip().lower()

    cfg_doc = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": "monthly_report", "active": True}
    )
    if not cfg_doc:
        raise HTTPException(400, f"No active monthly report config found for platform '{platform_lc}'")

    content = await file.read()
    sheet_loc = SheetLocator(**cfg_doc.get("sheet_locator", {"type": "first_sheet"}))
    header_loc = HeaderLocator(**cfg_doc.get("header_locator", {"type": "fixed_row", "row": 0}))
    skip_rows = int(cfg_doc.get("skip_rows_after_header", 0) or 0)

    headers, rows = _parse_tabular_bytes(
        content=content,
        filename=file.filename or "",
        sheet_locator=sheet_loc,
        header_locator=header_loc,
        skip_rows_after_header=skip_rows,
    )

    if not rows:
        raise HTTPException(400, "No data rows found in monthly report file")

    col_map = cfg_doc.get("column_map") or {}
    resolved_cols = {}
    for canon_field, target_name in col_map.items():
        if target_name:
            actual = _resolve_column(target_name, headers)
            if actual:
                resolved_cols[canon_field] = actual

    from routes.sku_map import resolve_style, split_leaf_sku

    matched_count = 0
    unresolved_count = 0
    records = []

    for r in rows:
        raw_leaf = r.get(resolved_cols.get("leaf_sku", ""), "").strip() if resolved_cols.get("leaf_sku") else ""
        if not raw_leaf:
            continue

        cleaned_leaf = strip_known_prefixes(raw_leaf, cfg_doc.get("known_sku_prefixes_to_strip", []))
        group_id, size_token, _ = split_leaf_sku(cleaned_leaf)

        result = await resolve_style(
            source_type="online_channel",
            source_name=platform_lc,
            external_sku=cleaned_leaf,
            external_color=None,
            external_size=size_token or None,
            db=db,
        )

        if result["matched"]:
            matched_count += 1
            records.append({
                "platform": platform_lc,
                "order_id": r.get(resolved_cols.get("order_id", ""), ""),
                "order_release_id": r.get(resolved_cols.get("order_release_id", ""), ""),
                "style_id": result["style_id"],
                "style_code": result["style_code"],
                "color": result["color"],
                "size": result["size"],
                "order_status": r.get(resolved_cols.get("order_status", ""), ""),
                "packed_on": r.get(resolved_cols.get("packed_on", ""), ""),
                "delivered_on": r.get(resolved_cols.get("delivered_on", ""), ""),
                "cancelled_on": r.get(resolved_cols.get("cancelled_on", ""), ""),
                "rto_creation_date": r.get(resolved_cols.get("rto_creation_date", ""), ""),
                "return_creation_date": r.get(resolved_cols.get("return_creation_date", ""), ""),
                "final_amount": float(r.get(resolved_cols.get("final_amount", ""), 0.0) or 0.0) if resolved_cols.get("final_amount") else 0.0,
                "created_at": now_iso(),
            })
        else:
            unresolved_count += 1

    if not dry_run and records:
        await db.online_orders_monthly.insert_many(records)

    return {
        "platform": platform_lc,
        "total_rows": len(rows),
        "matched_count": matched_count,
        "unresolved_count": unresolved_count,
        "dry_run": dry_run,
    }


@online_orders_router.post("/online-orders/settlement-import", dependencies=[Depends(bulk_import_rate_limiter)])
async def import_settlement(
    file: UploadFile = File(...),
    platform: str = Query(..., description="Platform identifier"),
    dry_run: bool = False,
    request: Request = None,
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = get_db()
    platform_lc = platform.strip().lower()

    cfg_doc = await db.order_import_format_configs.find_one(
        {"platform": platform_lc, "role": "settlement", "active": True}
    )
    if not cfg_doc:
        raise HTTPException(400, f"No active settlement import config found for platform '{platform_lc}'")

    content = await file.read()
    sheet_loc = SheetLocator(**cfg_doc.get("sheet_locator", {"type": "first_sheet"}))
    header_loc = HeaderLocator(**cfg_doc.get("header_locator", {"type": "fixed_row", "row": 0}))
    skip_rows = int(cfg_doc.get("skip_rows_after_header", 0) or 0)

    headers, rows = _parse_tabular_bytes(
        content=content,
        filename=file.filename or "",
        sheet_locator=sheet_loc,
        header_locator=header_loc,
        skip_rows_after_header=skip_rows,
    )

    if not rows:
        raise HTTPException(400, "No data rows found in settlement file")

    col_map = cfg_doc.get("column_map") or {}
    resolved_cols = {}
    for canon_field, target_name in col_map.items():
        if target_name:
            actual = _resolve_column(target_name, headers)
            if actual:
                resolved_cols[canon_field] = actual

    from routes.sku_map import resolve_style, split_leaf_sku

    matched_count = 0
    unresolved_count = 0
    records = []

    for r in rows:
        raw_leaf = r.get(resolved_cols.get("leaf_sku", ""), "").strip() if resolved_cols.get("leaf_sku") else ""
        order_ref = r.get(resolved_cols.get("order_ref", ""), "").strip() if resolved_cols.get("order_ref") else ""
        if not raw_leaf and not order_ref:
            continue

        cleaned_leaf = strip_known_prefixes(raw_leaf, cfg_doc.get("known_sku_prefixes_to_strip", []))
        group_id, size_token, _ = split_leaf_sku(cleaned_leaf)

        result = await resolve_style(
            source_type="online_channel",
            source_name=platform_lc,
            external_sku=cleaned_leaf,
            external_color=None,
            external_size=size_token or None,
            db=db,
        )

        gross = float(r.get(resolved_cols.get("gross_amount", ""), 0.0) or 0.0) if resolved_cols.get("gross_amount") else 0.0
        comm = float(r.get(resolved_cols.get("commission", ""), 0.0) or 0.0) if resolved_cols.get("commission") else 0.0
        ship = float(r.get(resolved_cols.get("shipping_fee", ""), 0.0) or 0.0) if resolved_cols.get("shipping_fee") else 0.0
        rto = float(r.get(resolved_cols.get("rto_charge", ""), 0.0) or 0.0) if resolved_cols.get("rto_charge") else 0.0
        net = float(r.get(resolved_cols.get("net_payout", ""), 0.0) or 0.0) if resolved_cols.get("net_payout") else 0.0

        if result["matched"]:
            matched_count += 1
            records.append({
                "platform": platform_lc,
                "order_ref": order_ref,
                "leaf_sku": cleaned_leaf,
                "style_id": result["style_id"],
                "style_code": result["style_code"],
                "color": result["color"],
                "size": result["size"],
                "gross_amount": gross,
                "commission": comm,
                "shipping_fee": ship,
                "rto_charge": rto,
                "net_payout": net,
                "settlement_date": r.get(resolved_cols.get("settlement_date", ""), "") if resolved_cols.get("settlement_date") else now_iso()[:10],
                "payment_id": r.get(resolved_cols.get("payment_id", ""), "") if resolved_cols.get("payment_id") else "",
                "created_at": now_iso(),
            })
        else:
            unresolved_count += 1

    if not dry_run and len(records) == 0:
        raise HTTPException(status_code=400, detail="Nothing to commit — no rows matched.")

    if not dry_run and records:
        await db.online_settlements.insert_many(records)

    return {
        "platform": platform_lc,
        "total_rows": len(rows),
        "matched_count": matched_count,
        "unresolved_count": unresolved_count,
        "dry_run": dry_run,
    }


@online_orders_router.get("/online-orders/settlements")
async def list_settlements(
    request: Request,
    platform: Optional[str] = None,
    order_ref: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    await _get_user(request)
    db = get_db()
    q = {}
    if platform:
        q["platform"] = platform.lower()
    if order_ref:
        q["order_ref"] = {"$regex": re.escape(order_ref), "$options": "i"}
    if from_date or to_date:
        dq = {}
        if from_date: dq["$gte"] = from_date
        if to_date: dq["$lte"] = to_date + "T23:59:59"
        q["settlement_date"] = dq
    docs = await db.online_settlements.find(q).sort("settlement_date", -1).to_list(2000)
    return [stringify(d) for d in docs]


@online_orders_router.get("/online-orders/settlement-summary")
async def settlement_summary(
    request: Request,
    platform: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    await _get_user(request)
    db = get_db()
    q = {}
    if platform:
        q["platform"] = platform.lower()
    if from_date or to_date:
        dq = {}
        if from_date: dq["$gte"] = from_date
        if to_date: dq["$lte"] = to_date + "T23:59:59"
        q["settlement_date"] = dq

    docs = await db.online_settlements.find(q).to_list(10000)
    tot_gross = sum(float(d.get("gross_amount", 0) or 0) for d in docs)
    tot_comm = sum(float(d.get("commission", 0) or 0) for d in docs)
    tot_ship = sum(float(d.get("shipping_fee", 0) or 0) for d in docs)
    tot_rto = sum(float(d.get("rto_charge", 0) or 0) for d in docs)
    tot_net = sum(float(d.get("net_payout", 0) or 0) for d in docs)

    return {
        "count": len(docs),
        "total_gross": round(tot_gross, 2),
        "total_commission": round(tot_comm, 2),
        "total_shipping": round(tot_ship, 2),
        "total_rto": round(tot_rto, 2),
        "total_net_payout": round(tot_net, 2),
    }


@online_orders_router.get("/online-orders/reconciliation-summary")
async def reconciliation_summary(
    request: Request,
    platform: Optional[str] = None,
):
    await _get_user(request)
    db = get_db()
    pq = {"channel": platform.lower()} if platform else {}
    total_orders = await db.production_jobs.count_documents({**pq, "source_type": "online_channel"})

    sq = {"platform": platform.lower()} if platform else {}
    settled_orders = await db.online_settlements.count_documents(sq)

    return {
        "platform": platform or "all",
        "total_orders": total_orders,
        "settled_orders": settled_orders,
        "unsettled_orders": max(0, total_orders - settled_orders),
    }


async def _parse_and_resolve_order_row(
    raw_row: dict,
    cfg: dict,
    platform: str,
    picklist_batch_id: Optional[str] = None,
    db=None,
) -> dict:
    if db is None:
        db = get_db()
    col_map = cfg.get("column_map") or {}
    leaf_col = col_map.get("leaf_sku") or "leaf_sku"
    raw_leaf = (raw_row.get(leaf_col) or "").strip()

    prefixes = cfg.get("known_sku_prefixes_to_strip", [])
    replacements = cfg.get("known_sku_prefix_replacements", {})
    cleaned_leaf = raw_leaf
    for wrong, right in (replacements or {}).items():
        if cleaned_leaf.startswith(wrong):
            cleaned_leaf = right + cleaned_leaf[len(wrong):]
    cleaned_leaf = strip_known_prefixes(cleaned_leaf, prefixes)

    from routes.sku_map import resolve_style, split_leaf_sku
    group_id, size_token, _ = split_leaf_sku(cleaned_leaf)

    size_col = col_map.get("size")
    color_col = col_map.get("color")
    price_col = col_map.get("selling_price")
    order_id_col = col_map.get("order_id")

    size_val = (raw_row.get(size_col) if size_col else "") or size_token or ""
    color_val = (raw_row.get(color_col) if color_col else "") or ""
    order_id = (raw_row.get(order_id_col) if order_id_col else "") or picklist_batch_id or ""

    price_str = (raw_row.get(price_col) if price_col else "0") or "0"
    try:
        selling_price = float(re.sub(r"[^\d.]", "", str(price_str)) or 0.0)
    except Exception:
        selling_price = 0.0

    import server
    resolve_style_fn = getattr(server, "resolve_style", resolve_style)
    res = await resolve_style_fn(
        source_type="online_channel",
        source_name=platform,
        external_sku=cleaned_leaf,
        external_color=color_val or None,
        external_size=size_val or None,
        db=db,
    )

    gst_warning = None
    if res["matched"] and res.get("style_id"):
        try:
            style = await db.styles.find_one({"_id": ObjectId(res["style_id"])})
        except Exception:
            style = await db.styles.find_one({"_id": res["style_id"]})
        if style:
            current_gst = float(style.get("gst_pct", 5.0) or 5.0)
            suggested_gst = 18.0 if selling_price > 2500.0 else 5.0
            if current_gst != suggested_gst:
                gst_warning = f"Selling price ₹{selling_price} suggests {suggested_gst:.0f}% GST, but style is currently set to {current_gst:.0f}%"

    matched = bool(res.get("matched")) and (res.get("matched_exact") is not False) and (not res.get("unmapped_size")) and (not res.get("unmapped_color"))

    exc_reason = None
    if not matched:
        if res.get("unmapped_size") or res.get("size_matched_exact") is False:
            exc_reason = f"Unmapped color/size: size '{size_val}' not in size_map for SKU '{cleaned_leaf}'"
        elif res.get("unmapped_color") or res.get("color_matched_exact") is False:
            exc_reason = f"Unmapped color/size: color '{color_val}' not in color_map for SKU '{cleaned_leaf}'"
        else:
            exc_reason = f"SKU '{cleaned_leaf}' not found in Style Master or SKU Mappings"

    return {
        "raw_leaf_sku": raw_leaf,
        "leaf_sku": cleaned_leaf,
        "order_id": order_id,
        "color": res.get("color") or color_val,
        "size": res.get("size") or size_val,
        "selling_price": selling_price,
        "matched": matched,
        "matched_exact": res.get("matched_exact", res["matched"]),
        "style_id": res.get("style_id"),
        "style_code": res.get("style_code"),
        "match_via": res.get("match_via"),
        "gst_mismatch_warning": gst_warning,
        "exception_reason": exc_reason,
    }


# Backwards compatibility aliases
_seed_order_import_configs = _seed_order_import_format_configs
import_online_orders_configured = import_configured_online_orders
