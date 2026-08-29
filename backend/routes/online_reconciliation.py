"""Online Reconciliation, Settlement Import & Online Profitability Engine Routes."""

import re
import io
import csv
import logging
from collections import defaultdict
from datetime import datetime, timezone, date as _date, timedelta as _td
from typing import Optional, List, Dict, Any, Literal
from bson import ObjectId
from fastapi import APIRouter, HTTPException, Request, Depends, UploadFile, File, Query
from fastapi.responses import Response

from models.online_reconciliation import (
    DailyPaymentRow,
    DailyPaymentImportIn,
    SettlementImportIn,
    NonOrderDeductionRow,
    NonOrderDeductionIn,
    MonthlyOrderRow,
    StyleCostSnapshotIn,
    ReconciliationRunIn,
)
from auth import require_roles

log = logging.getLogger(__name__)

online_reconciliation_router = APIRouter(prefix="/api", tags=["Online Reconciliation & Profitability"])


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stringify(doc: dict) -> dict:
    if not doc:
        return {}
    d = dict(doc)
    if "_id" in d:
        d["id"] = str(d.pop("_id"))
    for k, v in list(d.items()):
        if isinstance(v, ObjectId):
            d[k] = str(v)
    return d


def oid(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(400, "Invalid object ID format")


async def _get_user(request: Request):
    import server
    if getattr(server, "get_current_user", None) is not None:
        return await server.get_current_user(request)
    from auth import get_current_user_factory
    db = getattr(request.app, "mongodb", None) or server.db
    fn = await get_current_user_factory(db)
    return await fn(request)


async def log_activity_db(db, action: str, category: str, details: str, email: str):
    try:
        await db.audit_logs.insert_one({
            "action": action,
            "category": category,
            "details": details,
            "by": email,
            "created_at": now_iso()
        })
    except Exception:
        pass


def _clean_key(k: Any) -> str:
    if k is None:
        return ""
    s = str(k).strip().lower()
    return re.sub(r'[\s_]+', '_', s)


def _get_float_val(row: dict, keys: list[str]) -> float:
    for k in keys:
        clean_k = _clean_key(k)
        for rk, rv in row.items():
            if _clean_key(rk) == clean_k or clean_k in _clean_key(rk):
                if rv is not None and rv != "":
                    try:
                        return float(str(rv).replace(",", "").replace("₹", "").strip())
                    except Exception:
                        pass
    return 0.0


def _get_str_val(row: dict, keys: list[str]) -> str:
    for k in keys:
        clean_k = _clean_key(k)
        for rk, rv in row.items():
            if _clean_key(rk) == clean_k or clean_k in _clean_key(rk):
                if rv is not None:
                    return str(rv).strip()
    return ""


# ═══════════════════════════════════════════════════════════════════════
# ══ ONLINE PROFITABILITY ENGINE ════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════

_PLATFORM_FEE_FIELDS = [
    "commission_amount_incl_gst_postpaid", "commission_amount_incl_gst_prepaid",
    "fixed_fee_postpaid",                  "fixed_fee_prepaid",
    "logistics_cost_forward_incl_tax_postpaid", "logistics_cost_forward_incl_tax_prepaid",
    "pick_and_pack_fees_postpaid",         "pick_and_pack_fees_prepaid",
    "tech_enablement_charges_postpaid",    "tech_enablement_charges_prepaid",
    "royalty_charges_postpaid",            "royalty_charges_prepaid",
]

_LOGISTICS_FEE_FIELDS = [
    "logistics_cost_forward_incl_tax_postpaid", "logistics_cost_forward_incl_tax_prepaid",
    "logistics_cost_reverse_incl_tax_postpaid", "logistics_cost_reverse_incl_tax_prepaid",
]

_SETTLED_AMOUNT_FIELDS = ["settled_amount_postpaid", "settled_amount_prepaid"]
_PENDING_AMOUNT_FIELDS = [
    "amount_pending_settlement_postpaid",
    "amount_pending_settlement_prepaid",
]

_CUSTOMER_PAID_FIELDS = [
    "customer_paid_amount_postpaid",
    "customer_paid_amount_prepaid",
    "customer_paid_amount",
]


async def _collection_exists(name: str, db=None) -> bool:
    if db is None:
        import server
        db = server.db
    try:
        names = await db.list_collection_names()
        return name in names
    except Exception:
        return False


async def _sum_settlement_fields(
    coll_name: str,
    fields: List[str],
    match: Dict[str, Any],
    db=None,
) -> float:
    if db is None:
        import server
        db = server.db
    if not await _collection_exists(coll_name, db=db):
        return 0.0
    coll = db[coll_name]
    sum_exprs = {
        "total": {"$sum": {
            "$add": [
                {"$ifNull": [f"${f}", 0]}
                for f in fields
            ]
        }}
    } if fields else {"total": {"$sum": 0}}
    try:
        pipeline = [
            {"$match": match},
            {"$group": {"_id": None, **sum_exprs}},
        ]
        agg = await coll.aggregate(pipeline).to_list(1)
        if not agg: return 0.0
        return float(agg[0].get("total") or 0.0)
    except Exception as e:
        log.warning(f"_sum_settlement_fields failed on {coll_name}: {e}")
        return 0.0


def _build_settlement_match(
    platform: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    style_id: Optional[str],
) -> Dict[str, Any]:
    q: Dict[str, Any] = {}
    if platform: q["platform"] = platform.lower()
    if style_id: q["style_id"] = style_id
    if date_from or date_to:
        rng: Dict[str, Any] = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to
        q["$or"] = [{"packing_date": rng}, {"delivery_date": rng}]
    return q


async def _compute_online_profitability(
    platform:  Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    style_id:  Optional[str] = None,
    db=None,
) -> Dict[str, Any]:
    import server
    if db is None:
        db = server.db
    from server import compute_style_costing_async

    notes: List[str] = []
    item_match: Dict[str, Any] = {}
    if platform: item_match["platform"] = platform.lower()
    if style_id:
        try: item_match["style_id"] = ObjectId(style_id)
        except Exception: item_match["style_id"] = style_id
    if date_from or date_to:
        rng: Dict[str, Any] = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to
        item_match["packed_on"] = rng

    sold_pipeline = [
        {"$match": {**item_match, "is_net_sold": True}},
        {"$group": {
            "_id":         "$style_id",
            "style_code":  {"$first": "$style_code"},
            "color":       {"$first": "$color"},
            "units_sold":  {"$sum": {"$ifNull": ["$qty", 1]}},
            "item_final_amount": {"$sum": {"$ifNull": ["$final_amount", 0]}},
            "order_release_ids": {"$addToSet": "$order_release_id"},
        }},
    ]
    sold_rows = await db.online_order_items.aggregate(sold_pipeline).to_list(5000)

    ret_pipeline = [
        {"$match": {**item_match, "was_returned_to_stock": True}},
        {"$group": {"_id": "$style_id", "returned_units": {"$sum": {"$ifNull": ["$qty", 1]}}}},
    ]
    ret_rows = await db.online_order_items.aggregate(ret_pipeline).to_list(5000)
    returned_by_style = {str(r["_id"]): r["returned_units"] for r in ret_rows if r.get("_id")}

    style_ids = [r["_id"] for r in sold_rows if r.get("_id")]
    style_cost_map: Dict[str, Dict[str, Any]] = {}
    if style_ids:
        style_docs = await db.styles.find({"_id": {"$in": style_ids}}).to_list(len(style_ids))
        s_id_strs = [str(s.get("_id") or s.get("id")) for s in style_docs if (s.get("_id") or s.get("id"))]
        s_code_strs = [s.get("code") for s in style_docs if s.get("code")]
        jobs_by_style_id = defaultdict(list)
        jobs_by_style_code = defaultdict(list)
        pj_col = getattr(db, "production_jobs", None)
        if pj_col is not None and (s_id_strs or s_code_strs):
            job_or = []
            if s_id_strs:
                job_or.append({"style_id": {"$in": s_id_strs}})
            if s_code_strs:
                job_or.append({"style_code": {"$in": s_code_strs}})
            try:
                all_jobs = await pj_col.find({"$or": job_or} if len(job_or) > 1 else job_or[0]).to_list(50000)
                for job in all_jobs:
                    jid = job.get("style_id")
                    if jid:
                        jobs_by_style_id[str(jid)].append(job)
                    jcode = job.get("style_code")
                    if jcode:
                        jobs_by_style_code[jcode].append(job)
            except Exception:
                pass

        from routes.styles import compute_style_costing_from_jobs
        for s in style_docs:
            try:
                sid = str(s.get("_id") or s.get("id") or "")
                scode = s.get("code", "")
                matched_jobs = []
                seen_job_ids = set()
                for job in jobs_by_style_id.get(sid, []):
                    job_id_key = str(job.get("_id", id(job)))
                    if job_id_key not in seen_job_ids:
                        seen_job_ids.add(job_id_key)
                        matched_jobs.append(job)
                for job in jobs_by_style_code.get(scode, []):
                    job_id_key = str(job.get("_id", id(job)))
                    if job_id_key not in seen_job_ids:
                        seen_job_ids.add(job_id_key)
                        matched_jobs.append(job)

                c = compute_style_costing_from_jobs(s, matched_jobs)
                if not s.get("bom") and not s.get("labor") and hasattr(server, "compute_style_costing_async"):
                    try:
                        c_async = await server.compute_style_costing_async(s, db)
                        if c_async and float(c_async.get("total_cost", 0) or 0) > 0:
                            c = c_async
                    except Exception:
                        pass
                style_cost_map[str(s["_id"])] = {
                    "style_code":        s.get("code") or s.get("style_code"),
                    "unit_cogs":         c.get("total_cost") or 0,
                    "materials_cost":    c.get("materials_cost"),
                    "labor_cost":        c.get("labor_cost"),
                    "labor_source":      c.get("labor_source", "estimated"),
                    "cost_is_estimated": c.get("labor_source", "estimated") == "estimated",
                    "is_assigned":       c.get("is_assigned", False),
                    "actual_labor_cost": c.get("actual_labor_cost"),
                    "planned_labor_cost": c.get("planned_labor_cost"),
                    "overhead_cost":     c.get("overhead_cost"),
                    "packing_cost":      c.get("packing_cost"),
                }
            except Exception as e:
                log.warning(f"compute_style_costing_from_jobs failed for style {s.get('_id')}: {e}")
                style_cost_map[str(s["_id"])] = {
                    "style_code":        s.get("code"),
                    "unit_cogs":         0,
                    "labor_source":      "estimated",
                    "cost_is_estimated": True,
                }

    net_units_sold = 0
    total_net_cogs = 0.0
    fallback_revenue = 0.0
    by_style_acc: Dict[str, Dict[str, Any]] = {}

    for r in sold_rows:
        sid_raw = r.get("_id")
        sid_str = str(sid_raw) if sid_raw else None
        units = int(r.get("units_sold") or 0)
        item_rev = float(r.get("item_final_amount") or 0.0)

        cost_info = style_cost_map.get(sid_str) if sid_str else None
        unit_cogs = float(cost_info["unit_cogs"]) if cost_info else 0.0
        line_cogs = unit_cogs * units

        net_units_sold += units
        total_net_cogs += line_cogs
        fallback_revenue += item_rev

        key = sid_str or f"code_{r.get('style_code')}"
        if key not in by_style_acc:
            by_style_acc[key] = {
                "style_id":          sid_str,
                "style_code":        (cost_info.get("style_code") if cost_info else None) or r.get("style_code") or "—",
                "color":             r.get("color") or "—",
                "units_sold":        0,
                "returned_units":    returned_by_style.get(sid_str, 0) if sid_str else 0,
                "unit_cogs":         round(unit_cogs, 2),
                "cogs":              0.0,
                "fallback_revenue":  0.0,
                "materials_cost":    cost_info.get("materials_cost", 0.0) if cost_info else 0.0,
                "labor_cost":        cost_info.get("labor_cost", 0.0) if cost_info else 0.0,
                "cost_is_estimated": cost_info.get("cost_is_estimated", True) if cost_info else True,
                "labor_source":      cost_info.get("labor_source", "estimated") if cost_info else "estimated",
                "is_assigned":       cost_info.get("is_assigned", False) if cost_info else False,
                "actual_labor_cost": cost_info.get("actual_labor_cost") if cost_info else None,
                "planned_labor_cost": cost_info.get("planned_labor_cost") if cost_info else None,
            }
        by_style_acc[key]["units_sold"] += units
        by_style_acc[key]["cogs"] += line_cogs
        by_style_acc[key]["fallback_revenue"] += item_rev

    smatch = _build_settlement_match(platform, date_from, date_to, style_id)
    fwd_exists = await _collection_exists("settlement_forward", db=db)
    rev_exists = await _collection_exists("settlement_reverse", db=db)
    fwd_un_exists = await _collection_exists("settlement_unsettled_forward", db=db)
    rev_un_exists = await _collection_exists("settlement_unsettled_reverse", db=db)
    phase_3_here = fwd_exists or rev_exists or fwd_un_exists or rev_un_exists

    if not phase_3_here:
        notes.append("Phase 3 settlement collections not yet imported; settlement figures are 0.")

    settled_fwd = await _sum_settlement_fields("settlement_forward", _SETTLED_AMOUNT_FIELDS, smatch, db=db) if fwd_exists else 0.0
    settled_rev = await _sum_settlement_fields("settlement_reverse", _SETTLED_AMOUNT_FIELDS, smatch, db=db) if rev_exists else 0.0
    pending_fwd = await _sum_settlement_fields("settlement_unsettled_forward", _PENDING_AMOUNT_FIELDS, smatch, db=db) if fwd_un_exists else 0.0
    pending_rev = await _sum_settlement_fields("settlement_unsettled_reverse", _PENDING_AMOUNT_FIELDS, smatch, db=db) if rev_un_exists else 0.0

    total_revenue_settled = settled_fwd - settled_rev
    total_revenue_pending = pending_fwd - pending_rev

    fees_fwd = await _sum_settlement_fields("settlement_forward", _PLATFORM_FEE_FIELDS, smatch, db=db) if fwd_exists else 0.0
    fees_rev = await _sum_settlement_fields("settlement_reverse", _PLATFORM_FEE_FIELDS, smatch, db=db) if rev_exists else 0.0
    fees_fwd_un = await _sum_settlement_fields("settlement_unsettled_forward", _PLATFORM_FEE_FIELDS, smatch, db=db) if fwd_un_exists else 0.0
    fees_rev_un = await _sum_settlement_fields("settlement_unsettled_reverse", _PLATFORM_FEE_FIELDS, smatch, db=db) if rev_un_exists else 0.0
    total_platform_fees = (fees_fwd + fees_fwd_un) - (fees_rev + fees_rev_un)

    cost_of_returns_logistics = fees_rev + fees_rev_un

    if phase_3_here and (settled_fwd + settled_rev) > 0:
        revenue_for_profit = total_revenue_settled
        rev_source_used = "settlements (Phase 3)"
    elif fallback_revenue > 0:
        revenue_for_profit = fallback_revenue - total_platform_fees
        rev_source_used = "item-level final_amount minus platform fees"
        notes.append("Using item-level final_amount minus platform fees as revenue fallback.")
    else:
        revenue_for_profit = 0.0
        rev_source_used = "none"
        notes.append("No revenue records available for this filter range.")

    gross_profit = revenue_for_profit - total_net_cogs
    gross_margin_pct = (
        round((gross_profit / revenue_for_profit) * 100.0, 2)
        if revenue_for_profit > 0 else 0.0
    )

    by_style_list: List[Dict[str, Any]] = []
    for key, row in by_style_acc.items():
        sid = row["style_id"]
        cogs_i = row["cogs"]
        u_sold = row["units_sold"]
        u_ret  = row["returned_units"]
        total_attempts = u_sold + u_ret
        ret_rate = round((u_ret / total_attempts) * 100.0, 2) if total_attempts > 0 else 0.0

        if phase_3_here:
            sm_style = _build_settlement_match(platform, date_from, date_to, sid)
            sf = await _sum_settlement_fields("settlement_forward", _SETTLED_AMOUNT_FIELDS, sm_style, db=db) if fwd_exists else 0.0
            sr = await _sum_settlement_fields("settlement_reverse", _SETTLED_AMOUNT_FIELDS, sm_style, db=db) if rev_exists else 0.0
            ff = await _sum_settlement_fields("settlement_forward", _PLATFORM_FEE_FIELDS, sm_style, db=db) if fwd_exists else 0.0
            fr = await _sum_settlement_fields("settlement_reverse", _PLATFORM_FEE_FIELDS, sm_style, db=db) if rev_exists else 0.0
            rev_i = sf - sr
            fee_i = ff - fr
            source_i = "settlements"
        else:
            rev_i = row["fallback_revenue"]
            fee_i = 0.0
            source_i = "fallback"

        if source_i == "fallback":
            eff_rev_i = rev_i - fee_i
        else:
            eff_rev_i = rev_i

        profit_i = eff_rev_i - cogs_i
        margin_i = round((profit_i / eff_rev_i) * 100.0, 2) if eff_rev_i > 0 else 0.0

        by_style_list.append({
            "style_id":          sid,
            "style_code":        row["style_code"],
            "color":             row["color"],
            "units_sold":        u_sold,
            "returned_units":    u_ret,
            "return_rate_pct":   ret_rate,
            "unit_cogs":         row["unit_cogs"],
            "materials_cost":    row.get("materials_cost"),
            "labor_cost":        row.get("labor_cost"),
            "cogs":              round(cogs_i, 2),
            "revenue_settled":   round(rev_i, 2),
            "platform_fees":     round(fee_i, 2),
            "profit":            round(profit_i, 2),
            "margin_pct":        margin_i,
            "is_estimated":      source_i == "fallback",
            "cost_is_estimated": row["cost_is_estimated"],
            "labor_source":      row.get("labor_source", "estimated"),
            "is_assigned":       row.get("is_assigned", False),
            "actual_labor_cost": row.get("actual_labor_cost"),
            "planned_labor_cost": row.get("planned_labor_cost"),
            "revenue_source":    source_i,
        })

    return {
        "period": {
            "date_from": date_from,
            "date_to":   date_to,
        },
        "platform":                   platform,
        "style_id":                   style_id,
        "net_units_sold":             net_units_sold,
        "total_net_cogs":             round(total_net_cogs, 2),
        "total_revenue_settled":      round(total_revenue_settled, 2),
        "total_revenue_pending":      round(total_revenue_pending, 2),
        "total_platform_fees":        round(total_platform_fees, 2),
        "cost_of_returns_logistics":  round(cost_of_returns_logistics, 2),
        "gross_profit":               round(gross_profit, 2),
        "gross_margin_pct":           gross_margin_pct,
        "revenue_source_used":        rev_source_used,
        "is_estimated":               any(r.get("is_estimated", False) for r in by_style_list) if by_style_list else (not phase_3_here),
        "cost_is_estimated":          any(r.get("cost_is_estimated", False) for r in by_style_list) if by_style_list else True,
        "phase_3_available":          phase_3_here,
        "by_style":                   by_style_list,
        "notes":                      notes,
        "computed_at":                now_iso(),
    }


async def _materialise_profitability_snapshot(
    platform:  Optional[str],
    date_from: Optional[str],
    date_to:   Optional[str],
    style_id:  Optional[str],
    result:    Dict[str, Any],
    db=None,
):
    if db is None:
        import server
        db = server.db
    filt = {
        "platform":  platform.lower() if platform else "(all)",
        "date_from": date_from or "",
        "date_to":   date_to   or "",
        "style_id":  style_id  or "",
    }
    doc = {
        **filt,
        "snapshot":   result,
        "computed_at": now_iso(),
    }
    await db.online_profitability_daily.update_one(
        filt,
        {"$set": doc},
        upsert=True,
    )


async def _materialise_profitability_range(
    platform:  Optional[str],
    date_from: Optional[str],
    date_to:   Optional[str],
    db=None,
) -> Dict[str, Any]:
    if db is None:
        import server
        db = server.db
    days_rebuilt = 0
    if date_from and date_to:
        try:
            d_from = _date.fromisoformat(date_from)
            d_to   = _date.fromisoformat(date_to)
        except ValueError:
            d_from = d_to = None
        if d_from and d_to and d_from <= d_to:
            d = d_from
            while d <= d_to:
                iso = d.isoformat()
                try:
                    day_result = await _compute_online_profitability(
                        platform=platform, date_from=iso, date_to=iso, style_id=None, db=db,
                    )
                    await _materialise_profitability_snapshot(
                        platform=platform, date_from=iso, date_to=iso,
                        style_id=None, result=day_result, db=db,
                    )
                    days_rebuilt += 1
                except Exception as e:
                    log.warning(f"Day rollup failed for {iso}/{platform}: {e}")
                d = _date.fromordinal(d.toordinal() + 1)

    agg_result = await _compute_online_profitability(
        platform=platform, date_from=date_from, date_to=date_to, style_id=None, db=db,
    )
    await _materialise_profitability_snapshot(
        platform=platform, date_from=date_from, date_to=date_to,
        style_id=None, result=agg_result, db=db,
    )
    return {
        "days_rebuilt":          days_rebuilt,
        "aggregate_computed_at": now_iso(),
    }


# ───────────── Online Profitability Endpoints ─────────────

@online_reconciliation_router.get("/reports/online-profitability")
async def online_profitability(
    request: Request,
    platform: Optional[str] = None,
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD (inclusive)"),
    date_to:   Optional[str] = Query(None, description="YYYY-MM-DD (inclusive)"),
    style_id:  Optional[str] = None,
    materialise: bool = Query(False, description="If true, also upsert into online_profitability_daily"),
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    result = await _compute_online_profitability(
        platform=platform, date_from=date_from, date_to=date_to, style_id=style_id, db=db,
    )
    if materialise:
        await _materialise_profitability_snapshot(
            platform=platform, date_from=date_from, date_to=date_to,
            style_id=style_id, result=result, db=db,
        )
        result["materialised"] = True
    return result


@online_reconciliation_router.post("/reports/online-profitability/rebuild")
async def rebuild_online_profitability(
    request:  Request,
    platform: Optional[str] = None,
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD (inclusive)"),
    date_to:   Optional[str] = Query(None, description="YYYY-MM-DD (inclusive)"),
):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    summary = await _materialise_profitability_range(
        platform=platform, date_from=date_from, date_to=date_to, db=db,
    )
    await log_activity_db(
        db,
        "REBUILD", "online_profitability",
        f"Rebuilt profitability rollup for platform={platform or '(all)'} {date_from or '-'}..{date_to or '-'}: {summary['days_rebuilt']} days.",
        u.get("email") or u.get("name", ""),
    )
    return {"ok": True, "platform": platform, "period": {"from": date_from, "to": date_to}, **summary}


@online_reconciliation_router.get("/reports/online-profitability-materialised")
async def list_materialised_profitability(
    request: Request,
    platform: Optional[str] = None,
    limit: int = 50,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    q: Dict[str, Any] = {}
    if platform: q["platform"] = platform.lower()
    docs = await db.online_profitability_daily.find(q).sort("computed_at", -1).limit(limit).to_list(limit)
    for d in docs: d.pop("_id", None)
    return docs


@online_reconciliation_router.get("/reports/online-profitability/trend")
async def online_profitability_trend(
    request:   Request,
    platform:  Optional[str] = None,
    date_from: str = Query(..., description="YYYY-MM-DD (inclusive)"),
    date_to:   str = Query(..., description="YYYY-MM-DD (inclusive)"),
    bucket:    Literal["day", "week"] = "day",
    style_id:  Optional[str] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    from server import compute_style_costing_async

    try:
        d_from = _date.fromisoformat(date_from)
        d_to   = _date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(400, "date_from / date_to must be YYYY-MM-DD")
    if d_from > d_to:
        raise HTTPException(400, "date_from must be <= date_to")

    buckets: List[Dict[str, Any]] = []
    if bucket == "day":
        d = d_from
        while d <= d_to:
            iso = d.isoformat()
            buckets.append({"key": iso, "from": iso, "to": iso})
            d = _date.fromordinal(d.toordinal() + 1)
    else:
        d = d_from
        while d <= d_to:
            wk_start = d - _td(days=d.weekday())
            wk_end   = wk_start + _td(days=6)
            if wk_end > d_to: wk_end = d_to
            if wk_start < d_from: wk_start = d_from
            buckets.append({
                "key":  f"W{wk_start.isocalendar().week:02d} {wk_start.strftime('%b %d')}",
                "from": wk_start.isoformat(),
                "to":   wk_end.isoformat(),
            })
            d = wk_end + _td(days=1)

    base_item: Dict[str, Any] = {}
    if platform: base_item["platform"] = platform.lower()
    if style_id:
        try: base_item["style_id"] = ObjectId(style_id)
        except Exception: base_item["style_id"] = style_id

    fwd_exists = await _collection_exists("settlement_forward", db=db)
    rev_exists = await _collection_exists("settlement_reverse", db=db)
    fwd_un_exists = await _collection_exists("settlement_unsettled_forward", db=db)
    rev_un_exists = await _collection_exists("settlement_unsettled_reverse", db=db)

    async def _sum_by(coll: str, fields: List[str], q: Dict[str, Any]) -> float:
        if not await _collection_exists(coll, db=db): return 0.0
        return await _sum_settlement_fields(coll, fields, q, db=db)

    async def _count_units(match: Dict[str, Any]) -> int:
        pipeline = [
            {"$match": match},
            {"$group": {"_id": None, "n": {"$sum": {"$ifNull": ["$qty", 1]}}}},
        ]
        agg = await db.online_order_items.aggregate(pipeline).to_list(1)
        return int(agg[0]["n"]) if agg else 0

    _FEE_GROUPS = {
        "commission":       ["commission_amount_incl_gst_postpaid", "commission_amount_incl_gst_prepaid"],
        "fixed_fee":        ["fixed_fee_postpaid", "fixed_fee_prepaid"],
        "logistics_fwd":    ["logistics_cost_forward_incl_tax_postpaid", "logistics_cost_forward_incl_tax_prepaid"],
        "logistics_rev":    ["logistics_cost_reverse_incl_tax_postpaid", "logistics_cost_reverse_incl_tax_prepaid"],
        "pick_and_pack":    ["pick_and_pack_fees_postpaid", "pick_and_pack_fees_prepaid"],
        "tech_enablement":  ["tech_enablement_charges_postpaid", "tech_enablement_charges_prepaid"],
        "royalty":          ["royalty_charges_postpaid", "royalty_charges_prepaid"],
    }

    result_rows: List[Dict[str, Any]] = []
    for b in buckets:
        packed_match = {**base_item, "packed_on": {"$gte": b["from"], "$lte": b["to"]}}
        units_packed   = await _count_units({**packed_match, "was_packed": True})
        units_returned = await _count_units({**packed_match, "was_returned_to_stock": True})
        net_pipeline = [
            {"$match": {**packed_match, "is_net_sold": True}},
            {"$group": {"_id": "$style_id", "n": {"$sum": {"$ifNull": ["$qty", 1]}}, "amt": {"$sum": {"$ifNull": ["$final_amount", 0]}}}},
        ]
        net_rows = await db.online_order_items.aggregate(net_pipeline).to_list(2000)
        net_units_sold = sum(int(r.get("n") or 0) for r in net_rows)
        item_final_amount = sum(float(r.get("amt") or 0) for r in net_rows)

        style_ids = [r["_id"] for r in net_rows if r.get("_id")]
        cogs = 0.0
        if style_ids:
            style_docs = await db.styles.find({"_id": {"$in": style_ids}}).to_list(len(style_ids))
            s_id_strs = [str(s.get("_id") or s.get("id")) for s in style_docs if (s.get("_id") or s.get("id"))]
            s_code_strs = [s.get("code") for s in style_docs if s.get("code")]
            jobs_by_style_id = defaultdict(list)
            jobs_by_style_code = defaultdict(list)
            pj_col = getattr(db, "production_jobs", None)
            if pj_col is not None and (s_id_strs or s_code_strs):
                job_or = []
                if s_id_strs:
                    job_or.append({"style_id": {"$in": s_id_strs}})
                if s_code_strs:
                    job_or.append({"style_code": {"$in": s_code_strs}})
                try:
                    all_jobs = await pj_col.find({"$or": job_or} if len(job_or) > 1 else job_or[0]).to_list(50000)
                    for job in all_jobs:
                        jid = job.get("style_id")
                        if jid:
                            jobs_by_style_id[str(jid)].append(job)
                        jcode = job.get("style_code")
                        if jcode:
                            jobs_by_style_code[jcode].append(job)
                except Exception:
                    pass

            from routes.styles import compute_style_costing_from_jobs
            unit_cost_map = {}
            for s in style_docs:
                try:
                    sid = str(s.get("_id") or s.get("id") or "")
                    scode = s.get("code", "")
                    matched_jobs = []
                    seen_job_ids = set()
                    for job in jobs_by_style_id.get(sid, []):
                        job_id_key = str(job.get("_id", id(job)))
                        if job_id_key not in seen_job_ids:
                            seen_job_ids.add(job_id_key)
                            matched_jobs.append(job)
                    for job in jobs_by_style_code.get(scode, []):
                        job_id_key = str(job.get("_id", id(job)))
                        if job_id_key not in seen_job_ids:
                            seen_job_ids.add(job_id_key)
                            matched_jobs.append(job)
                    c_res = compute_style_costing_from_jobs(s, matched_jobs)
                    if not s.get("bom") and not s.get("labor") and hasattr(server, "compute_style_costing_async"):
                        try:
                            c_async = await server.compute_style_costing_async(s, db)
                            if c_async and float(c_async.get("total_cost", 0) or 0) > 0:
                                c_res = c_async
                        except Exception:
                            pass
                    unit_cost_map[str(s["_id"])] = float(c_res.get("total_cost") or 0)
                except Exception:
                    unit_cost_map[str(s["_id"])] = 0.0
            for r in net_rows:
                sid = str(r.get("_id"))
                cogs += unit_cost_map.get(sid, 0.0) * int(r.get("n") or 0)

        smatch = _build_settlement_match(platform, b["from"], b["to"], style_id)
        rev_fwd = await _sum_by("settlement_forward", _SETTLED_AMOUNT_FIELDS, smatch) if fwd_exists else 0.0
        rev_rev = await _sum_by("settlement_reverse", _SETTLED_AMOUNT_FIELDS, smatch) if rev_exists else 0.0
        pen_fwd = await _sum_by("settlement_unsettled_forward", _PENDING_AMOUNT_FIELDS, smatch) if fwd_un_exists else 0.0
        pen_rev = await _sum_by("settlement_unsettled_reverse", _PENDING_AMOUNT_FIELDS, smatch) if rev_un_exists else 0.0

        fees_out: Dict[str, float] = {}
        for grp, fields in _FEE_GROUPS.items():
            f_fwd = await _sum_by("settlement_forward",           fields, smatch) if fwd_exists else 0.0
            f_rev = await _sum_by("settlement_reverse",           fields, smatch) if rev_exists else 0.0
            f_fu  = await _sum_by("settlement_unsettled_forward", fields, smatch) if fwd_un_exists else 0.0
            f_ru  = await _sum_by("settlement_unsettled_reverse", fields, smatch) if rev_un_exists else 0.0
            if grp == "logistics_rev":
                fees_out[grp] = round(f_rev + f_ru, 2)
            else:
                fees_out[grp] = round(f_fwd + f_fu - f_rev - f_ru, 2)

        revenue_settled = rev_fwd - rev_rev
        phase_3_here_b = (rev_fwd + rev_rev + pen_fwd + pen_rev) > 0
        revenue_effective = revenue_settled if phase_3_here_b else item_final_amount
        gross_profit = revenue_effective - cogs

        result_rows.append({
            "date":              b["key"],
            "from":              b["from"],
            "to":                b["to"],
            "units_packed":      units_packed,
            "units_returned":    units_returned,
            "net_units_sold":    net_units_sold,
            "cogs":              round(cogs, 2),
            "revenue_settled":   round(revenue_settled, 2),
            "revenue_pending":   round(pen_fwd - pen_rev, 2),
            "revenue_effective": round(revenue_effective, 2),
            "gross_profit":      round(gross_profit, 2),
            "fees":              fees_out,
            "phase_3_here":      phase_3_here_b,
        })

    return {
        "platform": platform,
        "bucket":   bucket,
        "period":   {"from": date_from, "to": date_to},
        "style_id": style_id,
        "rows":     result_rows,
    }


@online_reconciliation_router.get("/reports/online-profitability/export")
async def online_profitability_export(
    request:   Request,
    platform:  Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    style_id:  Optional[str] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    result = await _compute_online_profitability(
        platform=platform, date_from=date_from, date_to=date_to, style_id=style_id, db=db,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    label_font  = Font(bold=True)

    ws["A1"] = "Online Profitability Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    meta_rows = [
        ("Platform",   platform or "(all)"),
        ("From",       date_from or "(any)"),
        ("To",         date_to or "(any)"),
        ("Style filter", style_id or "(none)"),
        ("Phase 3 available", "Yes" if result.get("phase_3_available") else "No"),
        ("Revenue source", result.get("revenue_source_used", "")),
        ("Generated at", now_iso()),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)

    metrics = [
        ("Net Units Sold",             result.get("net_units_sold")),
        ("Total Net COGS (₹)",         result.get("total_net_cogs")),
        ("Revenue Settled (₹)",        result.get("total_revenue_settled")),
        ("Revenue Pending (₹)",        result.get("total_revenue_pending")),
        ("Platform Fees (₹)",          result.get("total_platform_fees")),
        ("Cost of Returns Logistics (₹)", result.get("cost_of_returns_logistics")),
        ("Gross Profit (₹)",           result.get("gross_profit")),
        ("Gross Margin (%)",           result.get("gross_margin_pct")),
    ]
    start_row = 3 + len(meta_rows) + 2
    ws.cell(row=start_row, column=1, value="Metric").font = header_font
    ws.cell(row=start_row, column=1).fill = header_fill
    ws.cell(row=start_row, column=2, value="Value").font = header_font
    ws.cell(row=start_row, column=2).fill = header_fill
    for i, (k, v) in enumerate(metrics, start=start_row + 1):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("By Style")
    headers = [
        "Style Code", "Color", "Units Sold", "Returned Units",
        "Unit COGS (₹)", "Total COGS (₹)",
        "Revenue Settled (₹)", "Platform Fees (₹)",
        "Profit (₹)", "Margin (%)", "Return Rate (%)", "Revenue Source",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    rows = sorted(result.get("by_style") or [], key=lambda r: r.get("profit", 0))
    for ri, r in enumerate(rows, start=2):
        ws2.cell(row=ri, column=1,  value=r.get("style_code"))
        ws2.cell(row=ri, column=2,  value=r.get("color"))
        ws2.cell(row=ri, column=3,  value=r.get("units_sold"))
        ws2.cell(row=ri, column=4,  value=r.get("returned_units"))
        ws2.cell(row=ri, column=5,  value=r.get("unit_cogs"))
        ws2.cell(row=ri, column=6,  value=r.get("cogs"))
        ws2.cell(row=ri, column=7,  value=r.get("revenue_settled"))
        ws2.cell(row=ri, column=8,  value=r.get("platform_fees"))
        ws2.cell(row=ri, column=9,  value=r.get("profit"))
        ws2.cell(row=ri, column=10, value=r.get("margin_pct"))
        ws2.cell(row=ri, column=11, value=r.get("return_rate_pct"))
        ws2.cell(row=ri, column=12, value=r.get("revenue_source"))
    for c in range(1, 13):
        col_letter = ws2.cell(row=1, column=c).column_letter
        ws2.column_dimensions[col_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="online_profitability_report.xlsx"'}
    )


# ───────────── Multi-Report Reconciliation Endpoints ─────────────

@online_reconciliation_router.post("/online-reconciliation/cost-snapshots")
async def create_cost_snapshot(payload: StyleCostSnapshotIn, request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    doc = {
        "style_id": payload.style_id,
        "style_code": payload.style_code.strip().upper(),
        "effective_date": payload.effective_date.strip(),
        "total_cost": float(payload.total_cost),
        "material_cost": float(payload.material_cost or 0),
        "labor_cost": float(payload.labor_cost or 0),
        "notes": payload.notes or "",
        "created_at": now_iso(),
        "created_by": u.get("email") or u.get("name", ""),
    }
    res = await db.style_cost_snapshots.insert_one(doc)
    doc["_id"] = res.inserted_id
    await log_activity_db(db, "CREATE", "style_cost_snapshots", f"Created cost snapshot for {payload.style_code} effective {payload.effective_date} (₹{payload.total_cost})", u.get("email") or u.get("name", ""))
    return stringify(doc)


@online_reconciliation_router.get("/online-reconciliation/cost-snapshots")
async def list_cost_snapshots(request: Request, style_code: Optional[str] = None):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    q = {}
    if style_code:
        q["style_code"] = style_code.strip().upper()
    docs = await db.style_cost_snapshots.find(q).sort("effective_date", -1).to_list(1000)
    return [stringify(d) for d in docs]


@online_reconciliation_router.post("/online-reconciliation/clear-test-data")
async def clear_reconciliation_test_data(request: Request):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    await db.online_daily_payments.delete_many({})
    await db.online_settlements_detailed.delete_many({})
    await db.online_non_order_deductions.delete_many({})
    await db.online_monthly_order_reports.delete_many({})
    await db.style_cost_snapshots.delete_many({})
    return {"ok": True}


@online_reconciliation_router.post("/online-reconciliation/import-daily-payments")
async def import_daily_payments(request: Request, file: UploadFile = File(...)):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    content = await file.read()
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))

    rows_to_insert = []
    for row in reader:
        pay_type_raw = _get_str_val(row, ["payment_type", "type", "pay_type"]).lower()
        pay_type = "prepaid" if "prepaid" in pay_type_raw else "postpaid"
        ord_type_raw = _get_str_val(row, ["order_type", "order_type_forward_reverse"]).lower()
        ord_type = "Reverse" if "reverse" in ord_type_raw else "Forward"

        r_doc = {
            "neft_ref": _get_str_val(row, ["neft_ref", "neft ref", "utr", "payment_ref"]),
            "settled_amount": _get_float_val(row, ["settled_amount", "settled amount", "amount", "net_amount"]),
            "commission": _get_float_val(row, ["commission", "commission_amount"]),
            "shipping_fee": _get_float_val(row, ["shipping_fee", "shipping fee", "logistics_fee"]),
            "tds": _get_float_val(row, ["tds", "tds_amount"]),
            "payment_type": pay_type,
            "order_type": ord_type,
            "order_release_id": _get_str_val(row, ["order_release_id", "release id"]),
            "seller_order_id": _get_str_val(row, ["seller_order_id", "order id", "seller order id"]),
            "order_line_id": _get_str_val(row, ["order_line_id", "line id"]),
            "return_id": _get_str_val(row, ["return_id"]),
            "payment_date": _get_str_val(row, ["payment_date", "date"]),
            "filename": file.filename,
            "imported_at": now_iso(),
            "imported_by": u.get("email") or u.get("name", ""),
        }
        rows_to_insert.append(r_doc)

    if rows_to_insert:
        await db.online_daily_payments.insert_many(rows_to_insert)
    await log_activity_db(db, "IMPORT", "online_daily_payments", f"Imported {len(rows_to_insert)} daily payment rows from '{file.filename}'", u.get("email") or u.get("name", ""))
    return {"ok": True, "count": len(rows_to_insert), "filename": file.filename}


@online_reconciliation_router.post("/online-reconciliation/import-settlements")
async def import_settlements(request: Request, file: UploadFile = File(...)):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    content = await file.read()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    settlements_inserted = 0
    non_order_inserted = 0

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        name_lower = sheetname.lower()
        rows_data = list(sheet.iter_rows(values_only=True))
        if not rows_data or len(rows_data) < 3:
            continue

        header_row = [str(c or "").strip() for c in rows_data[2]]
        is_non_order = "non_order_deduction" in name_lower
        is_reverse = "reverse" in name_lower
        is_unsettled = "unsettled" in name_lower
        settlement_status = "unsettled" if is_unsettled else "settled"
        direction = "reverse" if is_reverse else "forward"

        for row_idx in range(3, len(rows_data)):
            r_vals = rows_data[row_idx]
            if not r_vals or all(c is None or str(c).strip() == "" for c in r_vals):
                continue

            row_dict = {header_row[i]: r_vals[i] for i in range(min(len(header_row), len(r_vals)))}

            if is_non_order:
                ded_doc = {
                    "sheet_name": sheetname,
                    "seller_id": _get_str_val(row_dict, ["seller_id", "seller id"]),
                    "settlement_amount": _get_float_val(row_dict, ["settlement_amount", "amount"]),
                    "settlement_type": _get_str_val(row_dict, ["settlement_type", "type"]),
                    "utr": _get_str_val(row_dict, ["utr", "neft_ref", "reference"]),
                    "invoice_ref": _get_str_val(row_dict, ["invoice_ref", "invoice"]),
                    "settlement_date": _get_str_val(row_dict, ["settlement_date", "date"]),
                    "settlement_description": _get_str_val(row_dict, ["settlement_description", "description", "remarks"]),
                    "filename": file.filename,
                    "imported_at": now_iso(),
                    "imported_by": u.get("email") or u.get("name", ""),
                }
                await db.online_non_order_deductions.insert_one(ded_doc)
                non_order_inserted += 1
            else:
                s_doc = {
                    "sheet_name": sheetname,
                    "settlement_status": settlement_status,
                    "direction": direction,
                    "order_release_id": _get_str_val(row_dict, ["order_release_id", "release id"]),
                    "seller_order_id": _get_str_val(row_dict, ["seller_order_id", "order id", "seller order id"]),
                    "sku_id": _get_str_val(row_dict, ["sku_id", "sku id"]),
                    "style_id": _get_str_val(row_dict, ["style_id", "style id"]),
                    "seller_sku_code": _get_str_val(row_dict, ["seller_sku_code", "seller sku"]),
                    "settled_amount_postpaid": _get_float_val(row_dict, ["settled_amount_postpaid"]),
                    "settled_amount_prepaid": _get_float_val(row_dict, ["settled_amount_prepaid"]),
                    "amount_pending_settlement_postpaid": _get_float_val(row_dict, ["amount_pending_settlement_postpaid"]),
                    "amount_pending_settlement_prepaid": _get_float_val(row_dict, ["amount_pending_settlement_prepaid"]),
                    "commission": _get_float_val(row_dict, ["commission_amount_incl_gst", "commission"]),
                    "logistics_cost_forward": _get_float_val(row_dict, ["logistics_cost_forward_incl_tax", "logistics_cost_forward"]),
                    "logistics_cost_reverse": _get_float_val(row_dict, ["logistics_cost_reverse_incl_tax", "logistics_cost_reverse"]),
                    "reverse_additional_charges": _get_float_val(row_dict, ["reverse_additional_charges"]),
                    "fixed_fee": _get_float_val(row_dict, ["fixed_fee"]),
                    "pick_and_pack_fees": _get_float_val(row_dict, ["pick_and_pack_fees"]),
                    "tech_enablement_charges": _get_float_val(row_dict, ["tech_enablement_charges"]),
                    "tds": _get_float_val(row_dict, ["tds"]),
                    "tcs": _get_float_val(row_dict, ["tcs"]),
                    "gst": _get_float_val(row_dict, ["gst"]),
                    "return_date": _get_str_val(row_dict, ["return_date", "return date"]),
                    "return_type": _get_str_val(row_dict, ["return_type", "return type"]),
                    "neft_ref": _get_str_val(row_dict, ["neft_ref", "utr", "payment_ref"]),
                    "filename": file.filename,
                    "imported_at": now_iso(),
                    "imported_by": u.get("email") or u.get("name", ""),
                }
                await db.online_settlements_detailed.insert_one(s_doc)
                settlements_inserted += 1

    await log_activity_db(db, "IMPORT", "online_settlements_detailed", f"Imported {settlements_inserted} settlement rows and {non_order_inserted} non-order deductions from '{file.filename}'", u.get("email") or u.get("name", ""))
    return {
        "ok": True,
        "settlements_count": settlements_inserted,
        "non_order_deductions_count": non_order_inserted,
        "filename": file.filename,
    }


@online_reconciliation_router.post("/online-reconciliation/import-monthly-report")
async def import_monthly_reconciliation_report(request: Request, file: UploadFile = File(...)):
    u = await _get_user(request)
    require_roles("admin", "manager")(u)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    content = await file.read()
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))

    rows_to_insert = []
    for row in reader:
        seller_order_id = _get_str_val(row, ["seller order id", "seller_order_id", "order id"])
        if not seller_order_id:
            continue
        m_doc = {
            "seller_order_id": seller_order_id,
            "order_release_id": _get_str_val(row, ["order release id", "order_release_id"]),
            "sku_id": _get_str_val(row, ["sku id", "sku_id"]),
            "style_id": _get_str_val(row, ["style id", "style_id"]),
            "seller_sku_code": _get_str_val(row, ["seller sku code", "seller_sku_code"]),
            "size": _get_str_val(row, ["size"]),
            "order_status": _get_str_val(row, ["order status", "status"]),
            "packed_on": _get_str_val(row, ["packed on", "packed_on", "packed_date"]),
            "shipped_on": _get_str_val(row, ["shipped on", "shipped_on"]),
            "delivered_on": _get_str_val(row, ["delivered on", "delivered_on"]),
            "cancelled_on": _get_str_val(row, ["cancelled on", "cancelled_on"]),
            "rto_return_creation_date": _get_str_val(row, ["rto/return creation date", "return_date"]),
            "final_amount": _get_float_val(row, ["final amount", "final_amount"]),
            "seller_price": _get_float_val(row, ["seller price", "seller_price"]),
            "filename": file.filename,
            "imported_at": now_iso(),
            "imported_by": u.get("email") or u.get("name", ""),
        }
        rows_to_insert.append(m_doc)

    if rows_to_insert:
        await db.online_monthly_order_reports.insert_many(rows_to_insert)
    await log_activity_db(db, "IMPORT", "online_monthly_order_reports", f"Imported {len(rows_to_insert)} monthly order report rows from '{file.filename}'", u.get("email") or u.get("name", ""))
    return {"ok": True, "count": len(rows_to_insert), "filename": file.filename}


@online_reconciliation_router.post("/online-reconciliation/run")
@online_reconciliation_router.get("/online-reconciliation/summary")
async def run_online_reconciliation(
    request: Request,
    aged_pending_days: int = Query(30, description="Days threshold for aged pending classification"),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    await _get_user(request)
    db = getattr(request.app, "mongodb", None) or getattr(__import__("server"), "db")
    from server import compute_style_costing

    try:
        aged_days_val = int(aged_pending_days) if not hasattr(aged_pending_days, 'default') else int(aged_pending_days.default)
    except Exception:
        aged_days_val = 30

    m_query = {}
    if from_date or to_date:
        d_q = {}
        if from_date: d_q["$gte"] = from_date
        if to_date:   d_q["$lte"] = to_date
        m_query["$or"] = [{"packed_on": d_q}, {"delivered_on": d_q}, {"imported_at": d_q}]

    monthly_orders = await db.online_monthly_order_reports.find(m_query).to_list(10000)
    settlements = await db.online_settlements_detailed.find({}).to_list(10000)

    settle_by_order_id = defaultdict(list)
    settle_by_rel_sku = defaultdict(list)
    for st in settlements:
        soid = st.get("seller_order_id")
        if soid:
            settle_by_order_id[soid].append(st)
        rel_id = st.get("order_release_id")
        sku_id = st.get("sku_id") or st.get("style_id")
        if rel_id and sku_id:
            settle_by_rel_sku[f"{rel_id}_{sku_id}"].append(st)

    daily_payments = await db.online_daily_payments.find({}).to_list(10000)

    sku_maps = await db.marketplace_style_color_mapping.find({}).to_list(5000)
    sku_map_dict = {}
    for sm in sku_maps:
        k = sm.get("raw_sku") or f"{sm.get('marketplace_style_code')}_{sm.get('marketplace_color_code')}"
        sku_map_dict[k] = sm.get("erp_style_code")
        if sm.get("marketplace_style_code"):
            sku_map_dict[sm["marketplace_style_code"]] = sm.get("erp_style_code")

    db_sku_maps = await db.sku_map.find({}).to_list(5000)
    for sm in db_sku_maps:
        ext_sku = sm.get("external_sku")
        st_code = sm.get("style_code") or sm.get("erp_style_code")
        if ext_sku and st_code:
            sku_map_dict[ext_sku] = st_code
            sku_map_dict[ext_sku.strip().upper()] = st_code
            sku_map_dict[ext_sku.strip().lower()] = st_code

    styles = await db.styles.find({}).to_list(1000)
    style_dict = {s.get("code"): s for s in styles if s.get("code")}

    snapshots = await db.style_cost_snapshots.find({}).sort("effective_date", 1).to_list(5000)
    snapshots_by_style = defaultdict(list)
    for snap in snapshots:
        snapshots_by_style[snap["style_code"]].append(snap)

    today_dt = datetime.now(timezone.utc).date()

    settled_count = 0
    pending_count = 0
    aged_pending_count = 0
    unmatched_count = 0
    total_matched = 0

    reconciled_lines = []
    unreconciled_orders = []

    cogs_exact_snapshot_count = 0
    cogs_total_requiring = 0

    return_charges_by_style = defaultdict(float)
    for st in settlements:
        if st.get("direction") == "reverse" or "reverse" in str(st.get("sheet_name", "")).lower():
            sku_code = st.get("seller_sku_code") or st.get("style_id") or ""
            st_style = sku_map_dict.get(sku_code) or st.get("style_id") or sku_code or "UNKNOWN"
            chg = float(st.get("logistics_cost_reverse", 0) or 0) + float(st.get("reverse_additional_charges", 0) or 0)
            if chg > 0:
                return_charges_by_style[st_style] += chg

    for mo in monthly_orders:
        soid = mo.get("seller_order_id")
        rel_id = mo.get("order_release_id")
        sku_id = mo.get("sku_id") or mo.get("style_id")

        matches = settle_by_order_id.get(soid) or settle_by_rel_sku.get(f"{rel_id}_{sku_id}") or []

        if not matches:
            reasons = []
            if not soid:
                reasons.append("Missing seller_order_id")
            reasons.append("Absent from settlement files")

            unmatched_count += 1
            unreconciled_orders.append({
                "seller_order_id": soid or "—",
                "order_release_id": rel_id or "—",
                "seller_sku_code": mo.get("seller_sku_code") or "—",
                "order_status": mo.get("order_status") or "Unknown",
                "reasons": reasons,
                "packed_on": mo.get("packed_on") or "—",
            })
            continue

        total_matched += 1
        has_settled = any(st.get("settlement_status") == "settled" for st in matches)
        has_unsettled = any(st.get("settlement_status") == "unsettled" for st in matches)

        order_date_str = mo.get("packed_on") or mo.get("shipped_on") or mo.get("delivered_on") or str(mo.get("imported_at", ""))[:10]
        try:
            o_date = datetime.strptime(order_date_str[:10], "%Y-%m-%d").date()
            age_days = (today_dt - o_date).days
        except Exception:
            age_days = 0

        if has_settled:
            rec_status = "settled"
            settled_count += 1
        elif has_unsettled:
            if age_days > aged_days_val:
                rec_status = "aged_pending"
                aged_pending_count += 1
            else:
                rec_status = "pending"
                pending_count += 1
        else:
            rec_status = "unmatched"
            unmatched_count += 1

        settled_postpaid = sum(st.get("settled_amount_postpaid", 0) for st in matches)
        settled_prepaid = sum(st.get("settled_amount_prepaid", 0) for st in matches)
        settled_amount = settled_postpaid + settled_prepaid

        commission = sum(st.get("commission", 0) for st in matches)
        logistics_fwd = sum(st.get("logistics_cost_forward", 0) for st in matches)
        logistics_rev = sum(st.get("logistics_cost_reverse", 0) for st in matches)
        rev_add_charges = sum(st.get("reverse_additional_charges", 0) for st in matches)
        fixed_fee = sum(st.get("fixed_fee", 0) for st in matches)
        pick_pack = sum(st.get("pick_and_pack_fees", 0) for st in matches)
        tech_enablement = sum(st.get("tech_enablement_charges", 0) for st in matches)
        taxes = sum(st.get("tds", 0) + st.get("tcs", 0) + st.get("gst", 0) for st in matches)

        seller_sku = mo.get("seller_sku_code") or ""
        erp_style_code = sku_map_dict.get(seller_sku) or mo.get("style_id") or seller_sku.split("-")[0]
        is_cancelled_before_prod = bool(mo.get("cancelled_on")) and not bool(mo.get("packed_on"))

        unit_cogs = 0.0
        cost_estimated = False

        if not is_cancelled_before_prod:
            cogs_total_requiring += 1
            snaps = snapshots_by_style.get(erp_style_code) or []
            active_snap = None
            for s_item in reversed(snaps):
                if s_item.get("effective_date", "") <= order_date_str[:10]:
                    active_snap = s_item
                    break

            if active_snap:
                unit_cogs = float(active_snap.get("total_cost", 0))
                cogs_exact_snapshot_count += 1
                cost_estimated = False
            else:
                if snaps:
                    unit_cogs = float(snaps[0].get("total_cost", 0))
                else:
                    st_obj = style_dict.get(erp_style_code)
                    if st_obj:
                        c_res = compute_style_costing(st_obj)
                        unit_cogs = float(c_res.get("total_cost", 0))
                    else:
                        unit_cogs = 0.0
                cost_estimated = True
                unreconciled_orders.append({
                    "seller_order_id": soid or "—",
                    "order_release_id": rel_id or "—",
                    "seller_sku_code": seller_sku,
                    "order_status": mo.get("order_status") or "Unknown",
                    "reasons": ["No active costing snapshot on order date (cost estimated)"],
                    "packed_on": order_date_str,
                })

        total_line_costs = commission + fixed_fee + logistics_fwd + logistics_rev + rev_add_charges + pick_pack + tech_enablement + taxes + unit_cogs
        net_profit = settled_amount - total_line_costs

        reconciled_lines.append({
            "seller_order_id": soid,
            "order_release_id": rel_id,
            "style_code": erp_style_code,
            "seller_sku_code": seller_sku,
            "order_date": order_date_str,
            "status": rec_status,
            "settled_amount": round(settled_amount, 2),
            "commission": round(commission, 2),
            "fixed_fee": round(fixed_fee, 2),
            "logistics_cost": round(logistics_fwd + logistics_rev + rev_add_charges, 2),
            "pick_and_pack_fees": round(pick_pack, 2),
            "tech_enablement_charges": round(tech_enablement, 2),
            "taxes": round(taxes, 2),
            "actual_cogs": round(unit_cogs, 2),
            "cost_estimated": cost_estimated,
            "net_profit": round(net_profit, 2),
        })

    neft_daily = defaultdict(float)
    for dp in daily_payments:
        neft = dp.get("neft_ref") or "UNSPECIFIED"
        neft_daily[neft] += float(dp.get("settled_amount", 0) or 0)

    neft_settled = defaultdict(float)
    for st in settlements:
        if st.get("settlement_status") == "settled":
            neft = st.get("neft_ref") or "UNSPECIFIED"
            amt = float(st.get("settled_amount_postpaid", 0) or 0) + float(st.get("settled_amount_prepaid", 0) or 0)
            neft_settled[neft] += amt

    all_nefts = set(list(neft_daily.keys()) + list(neft_settled.keys()))
    neft_mismatches = []
    for neft in all_nefts:
        if neft == "UNSPECIFIED" and len(all_nefts) > 1:
            continue
        d_amt = neft_daily.get(neft, 0.0)
        s_amt = neft_settled.get(neft, 0.0)
        diff = round(d_amt - s_amt, 2)
        if abs(diff) > 0.01:
            neft_mismatches.append({
                "neft_ref": neft,
                "daily_payment_amount": round(d_amt, 2),
                "settlement_file_amount": round(s_amt, 2),
                "difference": diff,
            })

    non_order_docs = await db.online_non_order_deductions.find({}).sort("settlement_date", -1).to_list(1000)
    total_non_order_deductions = sum(float(d.get("settlement_amount", 0) or 0) for d in non_order_docs)

    total_monthly_lines = len(monthly_orders)
    join_rate_pct = round((total_matched / total_monthly_lines * 100), 2) if total_monthly_lines > 0 else 100.0
    cogs_resolution_rate_pct = round((cogs_exact_snapshot_count / cogs_total_requiring * 100), 2) if cogs_total_requiring > 0 else 100.0

    style_profitability = defaultdict(lambda: {"units": 0, "settled": 0.0, "cogs": 0.0, "fees": 0.0, "net_profit": 0.0, "cost_estimated_count": 0})
    for rl in reconciled_lines:
        sc = rl["style_code"]
        style_profitability[sc]["units"] += 1
        style_profitability[sc]["settled"] += rl["settled_amount"]
        style_profitability[sc]["cogs"] += rl["actual_cogs"]
        fees = rl["commission"] + rl["fixed_fee"] + rl["logistics_cost"] + rl["pick_and_pack_fees"] + rl["tech_enablement_charges"] + rl["taxes"]
        style_profitability[sc]["fees"] += fees
        style_profitability[sc]["net_profit"] += rl["net_profit"]
        if rl["cost_estimated"]:
            style_profitability[sc]["cost_estimated_count"] += 1

    by_style_list = [
        {
            "style_code": sc,
            "units": val["units"],
            "settled_amount": round(val["settled"], 2),
            "actual_cogs": round(val["cogs"], 2),
            "platform_fees": round(val["fees"], 2),
            "net_profit": round(val["net_profit"], 2),
            "cost_estimated_count": val["cost_estimated_count"],
        }
        for sc, val in style_profitability.items()
    ]

    return {
        "join_rate_pct": join_rate_pct,
        "cogs_resolution_rate_pct": cogs_resolution_rate_pct,
        "total_monthly_report_units": total_monthly_lines,
        "settled_count": settled_count,
        "pending_count": pending_count,
        "aged_pending_count": aged_pending_count,
        "unmatched_count": unmatched_count,
        "return_charges_by_style": {k: round(v, 2) for k, v in return_charges_by_style.items()},
        "total_non_order_deductions": round(total_non_order_deductions, 2),
        "non_order_deductions_ledger": [stringify(d) for d in non_order_docs[:100]],
        "neft_mismatches": neft_mismatches,
        "unreconciled_orders": unreconciled_orders[:200],
        "profitability_by_style": by_style_list,
    }


@online_reconciliation_router.get("/online-reconciliation/unreconciled-orders")
async def list_unreconciled_orders(request: Request):
    await _get_user(request)
    summary = await run_online_reconciliation(request)
    return summary.get("unreconciled_orders", [])
