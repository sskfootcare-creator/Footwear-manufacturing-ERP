"""Online Profitability & Multi-Report Reconciliation Tests.

Run with:
    cd backend && .venv/Scripts/pytest tests/test_online_reconciliation.py -v
"""
import os
import sys
import io
import pytest
import httpx
import openpyxl

# Ensure backend root is in sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@sskfootcare.com")
ADMIN_PASS = os.environ.get("ADMIN_PASSWORD", "Admin@123")


@pytest.fixture(scope="module")
def client():
    try:
        c = httpx.Client(base_url="http://localhost:8000/api", timeout=10)
        r = c.post("/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
        if r.status_code == 200:
            yield c
            return
    except Exception:
        pass

    from fastapi.testclient import TestClient
    from server import app
    with TestClient(app, base_url="http://testserver/api") as tc:
        r = tc.post("/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
        assert r.status_code == 200, f"Login failed: {r.text}"
        yield tc


def generate_sample_daily_payment_csv():
    content = (
        "NEFT_Ref,Settled_Amount,Commission,Shipping_Fee,TDS,Payment_Type,Order_Type,order_release_id,seller_order_id,order_line_id,return_id,Payment_Date\n"
        "NEFT123,450.00,45.00,30.00,4.50,prepaid,Forward,REL101,SO101,LINE101,,2026-07-01\n"
        "NEFT123,550.00,55.00,30.00,5.50,postpaid,Forward,REL102,SO102,LINE102,,2026-07-02\n"
    )
    return content.encode("utf-8")


def generate_sample_settled_excel():
    wb = openpyxl.Workbook()
    # Sheet 1: forward_settled
    ws1 = wb.active
    ws1.title = "forward_settled"
    ws1.append(["Title Row 1"])
    ws1.append(["Subtitle Row 2"])
    ws1.append(["order_release_id", "seller_order_id", "sku_id", "style_id", "Settled_Amount_Postpaid", "Settled_Amount_Prepaid", "Commission_Amount_incl_GST", "Logistics_Cost_Forward_incl_Tax", "Fixed_Fee", "Pick_and_Pack_Fees", "Tech_Enablement_Charges", "neft_ref"])
    ws1.append(["REL101", "SO101", "SKU101", "STYLE-A", 0, 450.00, 45.00, 30.00, 10.00, 5.00, 2.00, "NEFT123"])
    ws1.append(["REL102", "SO102", "SKU102", "STYLE-A", 550.00, 0, 55.00, 30.00, 10.00, 5.00, 2.00, "NEFT123"])

    # Sheet 2: reverse_settled
    ws2 = wb.create_sheet(title="reverse_settled")
    ws2.append(["Title Row 1"])
    ws2.append(["Subtitle Row 2"])
    ws2.append(["order_release_id", "seller_order_id", "sku_id", "style_id", "Logistics_Cost_Reverse_incl_Tax", "Reverse_additional_charges", "return_date", "return_type"])
    ws2.append(["REL103", "SO103", "SKU103", "STYLE-A", 40.00, 15.00, "2026-07-05", "damaged"])

    # Sheet 3: forward_non_order_deduction
    ws3 = wb.create_sheet(title="forward_non_order_deduction")
    ws3.append(["Title Row 1"])
    ws3.append(["Subtitle Row 2"])
    ws3.append(["Seller_ID", "Settlement_Amount", "Settlement_Type", "UTR", "Invoice_Ref", "Settlement_Date", "Settlement_Description"])
    ws3.append(["SELLER123", 1200.00, "Platform Penalty", "NEFT999", "INV999", "2026-07-04", "Cataloging Penalty"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_sample_unsettled_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "forward_unsettled"
    ws.append(["Title Row 1"])
    ws.append(["Subtitle Row 2"])
    ws.append(["order_release_id", "seller_order_id", "sku_id", "style_id", "Amount_pending_settlement_Postpaid", "Amount_pending_settlement_Prepaid"])
    ws.append(["REL104", "SO104", "SKU104", "STYLE-B", 350.00, 0])
    ws.append(["REL105", "SO105", "SKU105", "STYLE-B", 400.00, 0])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_sample_monthly_report_csv():
    content = (
        "seller order id,order release id,sku id,style id,seller sku code,size,order status,packed on,shipped on,delivered on,cancelled on,rto/return creation date,final amount,seller price\n"
        "SO101,REL101,SKU101,STYLE-A,SKU-STYLE-A,8,Delivered,2026-07-01,2026-07-01,2026-07-03,,,600,600\n"
        "SO102,REL102,SKU102,STYLE-A,SKU-STYLE-A,9,Delivered,2026-07-02,2026-07-02,2026-07-04,,,700,700\n"
        "SO103,REL103,SKU103,STYLE-A,SKU-STYLE-A,8,Returned,2026-07-02,2026-07-02,2026-07-04,,2026-07-05,600,600\n"
        "SO104,REL104,SKU104,STYLE-B,SKU-STYLE-B,7,Shipped,2026-07-15,2026-07-15,,,,500,500\n"
        "SO105,REL105,SKU105,STYLE-B,SKU-STYLE-B,8,Shipped,2026-05-10,2026-05-10,,,,500,500\n"
        "SO106,REL106,SKU106,STYLE-B,SKU-STYLE-B,9,Delivered,2026-07-10,2026-07-10,2026-07-12,,,500,500\n"
    )
    return content.encode("utf-8")


class TestOnlineReconciliation:

    def test_reconciliation_pipeline(self, client):
        # 0. Clean test collections via API endpoint for idempotent test runs
        r_clear = client.post("/online-reconciliation/clear-test-data")
        assert r_clear.status_code == 200

        # 1. Create Style Cost Snapshot
        snap_payload = {
            "style_code": "STYLE-A",
            "effective_date": "2026-06-01",
            "total_cost": 210.0,
            "material_cost": 150.0,
            "labor_cost": 60.0,
            "notes": "Effective June 2026 snapshot"
        }
        r_snap = client.post("/online-reconciliation/cost-snapshots", json=snap_payload)
        assert r_snap.status_code == 200, r_snap.text
        assert r_snap.json()["style_code"] == "STYLE-A"

        # 2. Import Daily Payments (prepaid.csv)
        files_dp = {"file": ("prepaid.csv", generate_sample_daily_payment_csv(), "text/csv")}
        r_dp = client.post("/online-reconciliation/import-daily-payments", files=files_dp)
        assert r_dp.status_code == 200, r_dp.text
        assert r_dp.json()["count"] == 2

        # 3. Import Settled Excel (settled.xlsx)
        files_st = {"file": ("settled.xlsx", generate_sample_settled_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r_st = client.post("/online-reconciliation/import-settlements", files=files_st)
        assert r_st.status_code == 200, r_st.text
        assert r_st.json()["settlements_count"] >= 3
        assert r_st.json()["non_order_deductions_count"] >= 1

        # 4. Import Unsettled Excel (unsettled.xlsx)
        files_unst = {"file": ("unsettled.xlsx", generate_sample_unsettled_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r_unst = client.post("/online-reconciliation/import-settlements", files=files_unst)
        assert r_unst.status_code == 200, r_unst.text
        assert r_unst.json()["settlements_count"] >= 2

        # 5. Import Monthly Order Report (monthly_order_report.csv)
        files_m = {"file": ("monthly_order_report.csv", generate_sample_monthly_report_csv(), "text/csv")}
        r_m = client.post("/online-reconciliation/import-monthly-report", files=files_m)
        assert r_m.status_code == 200, r_m.text
        assert r_m.json()["count"] == 6

        # 6. Run Reconciliation Engine & Check Summary
        r_run = client.post("/online-reconciliation/run")
        assert r_run.status_code == 200, r_run.text
        summary = r_run.json()

        # Verify Join Rate % and COGS Resolution Rate % explicitly
        assert "join_rate_pct" in summary
        assert "cogs_resolution_rate_pct" in summary
        assert summary["join_rate_pct"] > 0
        assert summary["cogs_resolution_rate_pct"] > 0

        # Verify Counts: settled, pending, aged_pending, unmatched
        assert summary["settled_count"] >= 3  # SO101, SO102, SO103 matched to settled sheets
        assert summary["pending_count"] >= 1  # SO104 (recent pending)
        assert summary["aged_pending_count"] >= 1  # SO105 (older than 30 days pending)
        assert summary["unmatched_count"] >= 1  # SO106 (absent from settlements)

        # Verify Return Charges sum by style: STYLE-A should be 40 + 15 = 55.00
        ret_charges = summary["return_charges_by_style"]
        assert "STYLE-A" in ret_charges
        assert ret_charges["STYLE-A"] == 55.00

        # Verify Non-Order Deductions Total
        assert summary["total_non_order_deductions"] == 1200.00

        # Verify Unreconciled Orders listing with explicit reasons
        unrec = summary["unreconciled_orders"]
        assert len(unrec) > 0
        so106_unrec = next((u for u in unrec if u["seller_order_id"] == "SO106"), None)
        assert so106_unrec is not None
        assert "Absent from settlement files" in so106_unrec["reasons"]

        # 7. Check unreconciled orders dedicated endpoint
        r_unrec_ep = client.get("/online-reconciliation/unreconciled-orders")
        assert r_unrec_ep.status_code == 200
        assert len(r_unrec_ep.json()) > 0
