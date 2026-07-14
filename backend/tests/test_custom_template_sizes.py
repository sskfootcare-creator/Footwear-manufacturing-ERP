import openpyxl
from openpyxl.styles import Font
from packing_list import _classify_header_simple, _expand_lines, build_from_template
import io

def test_classify_size_headers():
    assert _classify_header_simple("Style") == "style"
    assert _classify_header_simple("Colour") == "color"
    assert _classify_header_simple("39") == "size_col:39"
    assert _classify_header_simple("42") == "size_col:42"
    assert _classify_header_simple("8") == "size_col:8"
    assert _classify_header_simple("8.5") == "size_col:8.5"
    assert _classify_header_simple("Qty") == "quantity"

def test_expand_lines_matrix():
    # Setup a mock workbook with size columns
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers
    ws.cell(1, 1).value = "Style"
    ws.cell(1, 2).value = "Color"
    ws.cell(1, 3).value = "39"
    ws.cell(1, 4).value = "40"
    ws.cell(1, 5).value = "Qty"
    
    # Marker
    ws.cell(2, 1).value = "{{lines}}"
    # Give some style to the marker row to verify copy
    font = Font(name="Arial", size=12, bold=True)
    ws.cell(2, 1).font = font
    
    po = {
        "line_items": [
            {"style_code": "ST-01", "color": "Black", "size": "39", "quantity": 10, "unit_price": 100},
            {"style_code": "ST-01", "color": "Black", "size": "40", "quantity": 15, "unit_price": 100},
            {"style_code": "ST-02", "color": "Brown", "size": "39", "quantity": 20, "unit_price": 120},
        ]
    }
    
    _expand_lines(ws, po, {"pcs_per_box": 10})
    
    # Verify rows generated: 2 rows (ST-01 Black, ST-02 Brown)
    # Row 2 should be ST-01 Black
    assert ws.cell(2, 1).value == "ST-01"
    assert ws.cell(2, 2).value == "Black"
    assert ws.cell(2, 3).value == 10
    assert ws.cell(2, 4).value == 15
    assert ws.cell(2, 5).value == 25  # aggregated qty
    assert ws.cell(2, 1).font.bold is True
    assert ws.cell(2, 1).font.name == "Arial"
    
    # Row 3 should be ST-02 Brown
    assert ws.cell(3, 1).value == "ST-02"
    assert ws.cell(3, 2).value == "Brown"
    assert ws.cell(3, 3).value == 20
    assert ws.cell(3, 4).value == ""
    assert ws.cell(3, 5).value == 20
    # Style copied
    assert ws.cell(3, 1).font.bold is True
    assert ws.cell(3, 1).font.name == "Arial"

def test_expand_lines_with_cartons():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Style"
    ws.cell(1, 2).value = "Color"
    ws.cell(1, 3).value = "Size"
    ws.cell(1, 4).value = "Qty"
    ws.cell(1, 5).value = "Ctn No"
    ws.cell(2, 1).value = "{{lines}}"
    
    po = {
        "line_items": [
            {"style_code": "ST-01", "color": "Black", "size": "39", "quantity": 100, "unit_price": 100},
        ]
    }
    
    cartons = [
        {"box_number": 1, "qty": 30, "size": "39", "color": "Black", "style_code": "ST-01"},
        {"box_number": 2, "qty": 30, "size": "39", "color": "Black", "style_code": "ST-01"},
        {"box_number": 3, "qty": 40, "size": "39", "color": "Black", "style_code": "ST-01"},
    ]
    
    _expand_lines(ws, po, {"pcs_per_box": 30}, cartons=cartons)
    
    # We should have 2 groups/rows because of different carton quantities:
    # Row 2: ST-01 Black 39, carton count 2 (qty 30), range 1 - 2, total pairs 60
    # Row 3: ST-01 Black 39, carton count 1 (qty 40), range 3, total pairs 40
    assert ws.cell(2, 1).value == "ST-01"
    assert ws.cell(2, 4).value == 60
    assert ws.cell(2, 5).value == "1 - 2"
    
    assert ws.cell(3, 1).value == "ST-01"
    assert ws.cell(3, 4).value == 40
    assert ws.cell(3, 5).value == "3"
