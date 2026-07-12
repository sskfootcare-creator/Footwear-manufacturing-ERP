"""Iteration 11 tests: FREE PO extractor, Packing List, Packing Templates, Auto-archive, A4 PDF."""
import base64
import io
import os
import time

import openpyxl
import pdfplumber
import pytest
import requests

def _read_env():
    p = "/app/frontend/.env"
    if os.path.exists(p):
        for line in open(p):
            if line.startswith("REACT_APP_BACKEND_URL="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8000")

BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or _read_env()).rstrip("/")
ADMIN_EMAIL = "admin@sskfootcare.com"
ADMIN_PASS = "Admin@123"


@pytest.fixture(scope="session")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
    assert r.status_code == 200, f"Login failed: {r.text}"
    return s


# ----------------- helpers -----------------
def _build_test_xlsx_po() -> bytes:
    """Build a synthetic PO xlsx with header/line-items, deterministic content."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO"
    ws["A1"] = "Purchase Order No: 2226000123"
    ws["A2"] = "PO Date: 15.01.2026"
    ws["A3"] = "Delivery Date: 28.02.2026"
    ws["A4"] = "Buyer: TEST CLIENT PVT LTD"
    ws["A5"] = "Vendor: SSK FOOTCARE MANUFACTURING LLP"
    ws["A7"] = ""
    headers = ["Style", "Description", "Color", "Size", "HSN", "Quantity", "Rate", "Amount"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=9, column=i, value=h)
    rows = [
        ("ZFLWTEST1", "Mens Loafer", "BROWN", "8", "64029990", 10, 500.00, 5000.00),
        ("ZFLWTEST1", "Mens Loafer", "BROWN", "9", "64029990", 12, 500.00, 6000.00),
        ("ZFLWTEST2", "Mens Boot", "BLACK", "7", "64029990", 8, 750.00, 6000.00),
    ]
    for r, row in enumerate(rows, 10):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    ws["A14"] = "Grand Total: 17000.00"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _small_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "PACKING TEMPLATE"
    ws["A2"] = "PO: {{po_number}}"
    ws["A3"] = "Client: {{client_name}}"
    ws["A5"] = "Style"
    ws["B5"] = "Color"
    ws["C5"] = "Qty"
    ws["A6"] = "{{lines}}"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------- FREE PO Extractor -----------------
class TestPOExtractor:
    def test_extract_xlsx_deterministic(self, session):
        data = _build_test_xlsx_po()
        files = {"file": ("test_po.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        # multipart - use a fresh session w/o JSON content-type
        s = requests.Session()
        s.cookies.update(session.cookies)
        r = s.post(f"{BASE_URL}/api/pos/extract", files=files)
        assert r.status_code == 200, f"extract failed: {r.status_code} {r.text}"
        body = r.json()
        # Some endpoints wrap in {data: ...}
        d = body.get("data", body) if isinstance(body, dict) else body
        assert d.get("po_number") == "2226000123", f"po_number mismatch: {d.get('po_number')}"
        assert d.get("po_date") == "2026-01-15", f"po_date mismatch: {d.get('po_date')}"
        assert d.get("delivery_date") == "2026-02-28"
        assert "TEST CLIENT" in (d.get("client_name") or "").upper()
        assert "SSK FOOTCARE" in (d.get("vendor_name") or "").upper()
        items = d.get("line_items") or []
        assert len(items) == 3, f"expected 3 line items got {len(items)}"
        assert d.get("total_quantity") == 30
        assert d.get("subtotal") == 17000.0
        assert d.get("grand_total") >= 17000.0


# ----------------- Packing List Default -----------------
class TestPackingListDefault:
    @pytest.fixture(scope="class")
    def po_id(self, session):
        r = session.get(f"{BASE_URL}/api/pos")
        assert r.status_code == 200
        pos = r.json()
        assert len(pos) >= 1, "no PO available"
        return pos[0]["id"]

    def test_default_packing_list(self, session, po_id):
        r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id})
        assert r.status_code == 200, f"got {r.status_code}: {r.text[:300]}"
        assert "spreadsheet" in r.headers.get("content-type", "") or "xlsx" in r.headers.get("content-disposition", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws["A1"].value == "PACKING LIST", f"title got {ws['A1'].value}"
        assert "SSK FOOTCARE" in (ws["B2"].value or "")
        # client name at G3
        assert (ws["G3"].value or "").strip() != ""
        # PO number at C13
        assert (str(ws["C13"].value or "")).strip() != ""
        # header at row 16
        row16 = [ws.cell(16, c).value for c in range(1, 20)]
        joined = " ".join(str(v or "") for v in row16).upper()
        for kw in ["SITE CODE", "STYLE", "COLOUR", "CTN", "PCS/CTN", "TOTAL PCS", "NET WT", "GROSS WT"]:
            assert kw in joined, f"missing {kw} in header row: {joined}"
        # at least one line in row 17
        row17 = [ws.cell(17, c).value for c in range(1, 20)]
        assert any(v not in (None, "") for v in row17), "no line items"


# ----------------- Packing List for specific job -----------------
class TestPackingListJob:
    def test_packing_for_jobs(self, session):
        # find a production job
        r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        assert r.status_code == 200
        jobs = r.json()
        if not jobs:
            pytest.skip("no production jobs")
        job = jobs[0]
        po_id = job.get("po_id")
        jid = job.get("id")
        assert po_id and jid
        r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id, "job_ids": [jid]})
        assert r.status_code == 200, f"{r.status_code}: {r.text[:200]}"
        # verify packing_generated_at set
        r2 = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        match = next((j for j in r2.json() if j.get("id") == jid), None)
        assert match is not None
        assert match.get("packing_generated_at"), "packing_generated_at not set"


# ----------------- Packing Templates CRUD -----------------
class TestPackingTemplates:
    def test_crud(self, session):
        r = session.get(f"{BASE_URL}/api/packing-templates")
        assert r.status_code == 200
        initial = r.json()
        assert isinstance(initial, list)

        # POST a valid xlsx template
        tpl = _small_template_xlsx()
        b64 = base64.b64encode(tpl).decode()
        r = session.post(f"{BASE_URL}/api/packing-templates", json={
            "client_name": "TEST_CLIENT_X", "name": "TEST_TPL", "file_b64": b64,
        })
        assert r.status_code in (200, 201), f"{r.status_code}: {r.text[:200]}"
        created = r.json()
        tid = created.get("id") or created.get("_id")
        assert tid

        # GET to verify
        r = session.get(f"{BASE_URL}/api/packing-templates")
        assert any((t.get("id") == tid) for t in r.json())

        # POST non-xlsx → should 400
        bad = base64.b64encode(b"not an xlsx file").decode()
        r = session.post(f"{BASE_URL}/api/packing-templates", json={
            "client_name": "TEST_CLIENT_X", "name": "BAD", "file_b64": bad,
        })
        assert r.status_code == 400, f"non-xlsx should fail; got {r.status_code}"

        # DELETE
        r = session.delete(f"{BASE_URL}/api/packing-templates/{tid}")
        assert r.status_code in (200, 204)
        r = session.get(f"{BASE_URL}/api/packing-templates")
        assert not any(t.get("id") == tid for t in r.json())


# ----------------- Auto-archive -----------------
class TestAutoArchive:
    def test_auto_archive_flow(self, session):
        r = session.get(f"{BASE_URL}/api/production/jobs")
        jobs = r.json()
        if not jobs:
            pytest.skip("no production jobs")
        # pick one not yet archived
        target = jobs[0]
        jid = target["id"]
        po_id = target["po_id"]

        # move to dispatched
        r = session.patch(f"{BASE_URL}/api/production/jobs/{jid}", json={"stage": "dispatched"})
        assert r.status_code == 200, f"stage update failed: {r.status_code} {r.text[:200]}"

        # invoice
        r = session.post(f"{BASE_URL}/api/invoices/job", json={"po_id": po_id, "job_ids": [jid]})
        assert r.status_code == 200, f"invoice gen failed: {r.status_code} {r.text[:200]}"

        # packing
        r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id, "job_ids": [jid]})
        assert r.status_code == 200, f"packing failed: {r.status_code} {r.text[:200]}"

        time.sleep(0.5)

        # default list excludes
        r = session.get(f"{BASE_URL}/api/production/jobs")
        assert jid not in [j["id"] for j in r.json()], "archived job leaked into default list"

        # include_archived shows it with archived=True
        r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        match = next((j for j in r.json() if j["id"] == jid), None)
        assert match is not None, "job missing from include_archived list"
        assert match.get("archived") is True, f"archived flag not True: {match.get('archived')}"

        # archive endpoint includes it
        r = session.get(f"{BASE_URL}/api/production/archive")
        assert r.status_code == 200
        assert jid in [j["id"] for j in r.json()]


# ----------------- Production Card PDF A4 -----------------
class TestProductionCardPDF:
    def test_pdf_a4(self, session):
        r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        jobs = r.json()
        if not jobs:
            pytest.skip("no jobs")
        jid = jobs[0]["id"]
        r = session.post(f"{BASE_URL}/api/production/card.pdf", json={"job_ids": [jid]})
        assert r.status_code == 200
        assert r.content.startswith(b"%PDF")
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            page = pdf.pages[0]
            # A4 portrait = 595 x 842 pt
            assert round(page.width) == 595, f"page width {page.width}"
            assert round(page.height) == 842, f"page height {page.height}"
            txt = "\n".join((p.extract_text() or "") for p in pdf.pages)
            assert "SSK FOOTCARE MANUFACTURING LLP" in txt
            assert "PROCESS TALLY" in txt
            assert "KARIGAR" in txt
            assert "SIZE BREAKDOWN" in txt


# ----------------- Regression -----------------
class TestRegression:
    def test_smoke(self, session):
        endpoints = [
            "/api/pos", "/api/workers", "/api/reports/payroll",
            "/api/dashboard/overdue", "/api/reports/monthly-production",
            "/api/reports/karigar-output", "/api/reports/cost-variance",
            "/api/reports/stage-cycle-time", "/api/reports/defect-rate",
            "/api/settings/stage-durations",
        ]
        for ep in endpoints:
            r = session.get(f"{BASE_URL}{ep}")
            assert r.status_code == 200, f"{ep} -> {r.status_code}"
