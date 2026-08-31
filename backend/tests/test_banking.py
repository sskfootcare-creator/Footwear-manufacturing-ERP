"""Tests for Bank Accounts, Bank Statement Lines, and Payment/Expense/Settlement model integration."""

import pytest
from io import BytesIO
from datetime import datetime, timezone
from bson import ObjectId
from unittest.mock import MagicMock, AsyncMock
from fastapi import HTTPException

from models.banking import (
    BankAccountIn,
    BankAccountUpdate,
    BankStatementLineIn,
    BankStatementLineUpdate,
    MatchedTo,
    TransferConfirmIn,
    CashWithdrawalConfirmIn,
)
from models.vendors import PaymentIn
from models.expenses import ExpenseIn, ExpenseUpdate, RecurringExpenseIn
from models.online_reconciliation import SettlementImportIn, DailyPaymentRow, NonOrderDeductionRow
import routes.banking as banking_routes
from routes.banking import (
    list_bank_accounts,
    create_bank_account,
    get_bank_account,
    update_bank_account,
    delete_bank_account,
    list_statement_lines,
    create_statement_lines,
    match_statement_line,
    _is_cash_withdrawal_candidate,
    get_suggested_cash_withdrawals,
    confirm_cash_withdrawal,
    list_cash_ledger,
    get_cash_ledger_detail,
)


def test_bank_account_models_validation():
    """Verify BankAccountIn validation for online_channel and b2b_client accounts."""
    hdfc = BankAccountIn(
        name="HDFC - Online",
        bank_name="HDFC",
        account_number_last4="4321",
        account_type="online_channel",
        opening_balance=150000.50,
        opening_balance_date="2026-04-01",
        active=True,
    )
    assert hdfc.name == "HDFC - Online"
    assert hdfc.bank_name == "HDFC"
    assert hdfc.account_type == "online_channel"
    assert hdfc.opening_balance == 150000.50

    uco = BankAccountIn(
        name="UCO Bank - Offline",
        bank_name="UCO Bank",
        account_number_last4="8765",
        account_type="b2b_client",
        opening_balance=50000.0,
        opening_balance_date="2026-04-01",
        active=True,
    )
    assert uco.name == "UCO Bank - Offline"
    assert uco.account_type == "b2b_client"


def test_bank_statement_line_models_validation():
    """Verify BankStatementLineIn and MatchedTo structure validation."""
    line = BankStatementLineIn(
        bank_account_id=str(ObjectId()),
        date="2026-08-15",
        narration="NEFT-MYNTRA DESIGNS-SETTLEMENT-AUG15",
        reference_no="UTR987654321",
        debit_amount=0.0,
        credit_amount=45230.75,
        running_balance=195231.25,
        match_status="matched",
        matched_to=MatchedTo(type="settlement", ref_id="settle_123"),
        remarks="Initial remark",
    )
    assert line.credit_amount == 45230.75
    assert line.match_status == "matched"
    assert line.matched_to.type == "settlement"
    assert line.matched_to.ref_id == "settle_123"
    assert line.remarks == "Initial remark"

    # Default remarks should be empty string
    line_default = BankStatementLineIn(
        bank_account_id=str(ObjectId()),
        date="2026-08-15",
        narration="Test line",
    )
    assert line_default.remarks == ""

    # BankStatementLineUpdate remarks
    update_with_remarks = BankStatementLineUpdate(remarks="Updated remark note")
    assert update_with_remarks.remarks == "Updated remark note"
    update_default = BankStatementLineUpdate()
    assert update_default.remarks == ""


def test_payment_and_expense_models_backward_compatibility():
    """Verify PaymentIn and ExpenseIn accept bank_account_id alongside legacy fields."""
    # Payment without bank_account_id (legacy)
    p1 = PaymentIn(
        amount=10000.0,
        payment_date="2026-08-20",
        mode="NEFT",
        bank="HDFC Bank",
    )
    assert p1.bank == "HDFC Bank"
    assert p1.bank_account_id is None

    # Payment with bank_account_id
    acc_id = str(ObjectId())
    p2 = PaymentIn(
        amount=25000.0,
        payment_date="2026-08-20",
        mode="RTGS",
        bank="HDFC Bank",
        bank_account_id=acc_id,
    )
    assert p2.bank_account_id == acc_id
    assert p2.bank == "HDFC Bank"

    # ExpenseIn with bank_account_id
    e1 = ExpenseIn(
        category="Rent & Utilities",
        amount=45000.0,
        date="2026-08-01",
        payee="Landlord Realty",
        bank_account_id=acc_id,
    )
    assert e1.bank_account_id == acc_id

    # Online reconciliation records
    s1 = SettlementImportIn(
        seller_order_id="ORD123",
        settled_amount_postpaid=1200.0,
        bank_account_id=acc_id,
    )
    assert s1.bank_account_id == acc_id


@pytest.mark.anyio
async def test_bank_accounts_crud_routes(monkeypatch):
    """Test full CRUD lifecycle of bank accounts via API route handlers."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    # 1. Create HDFC - Online
    acc_oid_1 = ObjectId()
    created_doc_1 = {
        "_id": acc_oid_1,
        "name": "HDFC - Online",
        "bank_name": "HDFC",
        "account_number_last4": "4321",
        "account_type": "online_channel",
        "opening_balance": 100000.0,
        "opening_balance_date": "2026-04-01",
        "active": True,
        "created_at": "2026-08-30T10:00:00Z",
        "created_by": "admin@sskfootcare.com",
    }
    mock_db.bank_accounts.insert_one = AsyncMock(return_value=MagicMock(inserted_id=acc_oid_1))
    mock_db.bank_accounts.find_one = AsyncMock(return_value=created_doc_1)

    req = MagicMock()
    payload = BankAccountIn(
        name="HDFC - Online",
        bank_name="HDFC",
        account_number_last4="4321",
        account_type="online_channel",
        opening_balance=100000.0,
        opening_balance_date="2026-04-01",
        active=True,
    )
    res = await create_bank_account(payload, req)
    assert res["id"] == str(acc_oid_1)
    assert res["name"] == "HDFC - Online"
    assert res["account_type"] == "online_channel"

    # 2. List accounts
    acc_oid_2 = ObjectId()
    created_doc_2 = {
        "_id": acc_oid_2,
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "account_number_last4": "8765",
        "account_type": "b2b_client",
        "opening_balance": 50000.0,
        "opening_balance_date": "2026-04-01",
        "active": True,
    }
    
    mock_cursor = MagicMock()
    mock_cursor.sort = MagicMock(return_value=mock_cursor)
    mock_cursor.to_list = AsyncMock(return_value=[created_doc_1, created_doc_2])
    mock_db.bank_accounts.find = MagicMock(return_value=mock_cursor)

    accs = await list_bank_accounts(req, active=True)
    assert len(accs) == 2
    assert accs[0]["name"] == "HDFC - Online"
    assert accs[1]["name"] == "UCO Bank - Offline"

    # 3. Get single account
    mock_db.bank_accounts.find_one = AsyncMock(return_value=created_doc_1)
    single = await get_bank_account(str(acc_oid_1), req)
    assert single["id"] == str(acc_oid_1)
    assert single["bank_name"] == "HDFC"

    # 4. Update account
    updated_doc = dict(created_doc_1, opening_balance=125000.0)
    mock_db.bank_accounts.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    mock_db.bank_accounts.find_one = AsyncMock(return_value=updated_doc)

    patch_res = await update_bank_account(str(acc_oid_1), BankAccountUpdate(opening_balance=125000.0), req)
    assert patch_res["opening_balance"] == 125000.0

    # 5. Delete account (no statement lines -> hard delete)
    mock_db.bank_statement_lines.count_documents = AsyncMock(return_value=0)
    mock_db.bank_accounts.delete_one = AsyncMock(return_value=MagicMock(deleted_count=1))

    del_res = await delete_bank_account(str(acc_oid_1), req)
    assert del_res["ok"] is True
    assert "deleted" in del_res["message"]


@pytest.mark.anyio
async def test_bank_statement_lines_routes(monkeypatch):
    """Test importing and matching bank statement lines."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    req = MagicMock()
    acc_id = str(ObjectId())
    line_id = ObjectId()

    # 1. Insert statement line
    lines_in = [
        BankStatementLineIn(
            bank_account_id=acc_id,
            date="2026-08-25",
            narration="ACH CR-MYNTRA DESIGNS-AUG25",
            reference_no="ACH123456",
            credit_amount=78500.0,
            debit_amount=0.0,
            running_balance=250000.0,
            match_status="unmatched",
        )
    ]
    mock_db.bank_statement_lines.insert_many = AsyncMock(return_value=MagicMock(inserted_ids=[line_id]))
    create_res = await create_statement_lines(lines_in, req)
    assert create_res["ok"] is True
    assert create_res["inserted_count"] == 1

    # 2. Match statement line
    matched_doc = {
        "_id": line_id,
        "bank_account_id": acc_id,
        "date": "2026-08-25",
        "narration": "ACH CR-MYNTRA DESIGNS-AUG25",
        "reference_no": "ACH123456",
        "credit_amount": 78500.0,
        "debit_amount": 0.0,
        "running_balance": 250000.0,
        "match_status": "matched",
        "matched_to": {"type": "settlement", "ref_id": "settle_myntra_aug25"},
    }
    mock_db.bank_statement_lines.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    mock_db.bank_statement_lines.find_one = AsyncMock(return_value=matched_doc)

    match_payload = BankStatementLineUpdate(
        match_status="matched",
        matched_to=MatchedTo(type="settlement", ref_id="settle_myntra_aug25"),
        remarks="Verified against order batch #44",
    )
    match_res = await match_statement_line(str(line_id), match_payload, req)
    assert match_res["match_status"] == "matched"
    assert match_res["matched_to"]["type"] == "settlement"
    assert match_res["matched_to"]["ref_id"] == "settle_myntra_aug25"

    # 3. Update remark only
    remark_doc = {**matched_doc, "remarks": "Updated remark only"}
    mock_db.bank_statement_lines.find_one = AsyncMock(return_value=remark_doc)
    remark_payload = BankStatementLineUpdate(remarks="Updated remark only")
    remark_res = await match_statement_line(str(line_id), remark_payload, req)
    assert remark_res["remarks"] == "Updated remark only"


@pytest.mark.anyio
async def test_hdfc_statement_import_csv(monkeypatch):
    """Test importing HDFC bank statement CSV with custom column map layout."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    hdfc_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC - Online",
        "bank_name": "HDFC",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Date",
                "narration": "Narration",
                "reference": "Chq./Ref.No.",
                "debit_amount": "Withdrawal Amt.",
                "credit_amount": "Deposit Amt.",
                "balance": "Closing Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=hdfc_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    hdfc_csv_content = (
        "Date,Narration,Chq./Ref.No.,Withdrawal Amt.,Deposit Amt.,Closing Balance\n"
        "15/08/2026,ACH CR-MYNTRA DESIGNS-SETTLEMENT,ACH001,,45250.50,145250.50\n"
        "16/08/2026,UPI-SWIGGY-EXPENSE,UPI999,350.00,,144900.50\n"
        "17/08/2026,NEFT DR-LEATHER SUPPLIER VENDOR,NEFT777,25000.00,,119900.50\n"
    ).encode("utf-8")

    from fastapi import UploadFile
    from io import BytesIO
    file = UploadFile(filename="hdfc_aug_statement.csv", file=BytesIO(hdfc_csv_content))

    req = MagicMock()
    res = await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert res["ok"] is True
    assert res["inserted_count"] == 3
    assert len(inserted_docs) == 3

    # Row 1 check
    assert inserted_docs[0]["date"] == "2026-08-15"
    assert inserted_docs[0]["narration"] == "ACH CR-MYNTRA DESIGNS-SETTLEMENT"
    assert inserted_docs[0]["reference_no"] == "ACH001"
    assert inserted_docs[0]["credit_amount"] == 45250.50
    assert inserted_docs[0]["debit_amount"] == 0.0
    assert inserted_docs[0]["running_balance"] == 145250.50
    assert inserted_docs[0]["match_status"] == "unmatched"

    # Row 2 check
    assert inserted_docs[1]["date"] == "2026-08-16"
    assert inserted_docs[1]["debit_amount"] == 350.00
    assert inserted_docs[1]["running_balance"] == 144900.50


@pytest.mark.anyio
async def test_uco_statement_import_csv(monkeypatch):
    """Test importing UCO bank statement CSV with different column names and date formats."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Txn Date",
                "narration": "Description",
                "reference": "Ref No",
                "debit_amount": "Debit",
                "credit_amount": "Credit",
                "balance": "Balance",
            },
            "date_format": "%d-%m-%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    uco_csv_content = (
        "Txn Date,Description,Ref No,Debit,Credit,Balance\n"
        "20-08-2026,CHQ DEP - BAXTER RETAIL CLIENT,CHQ54321,,120000.00,520000.00\n"
        "21-08-2026,RTGS - RAW MATERIAL AGRA,RTGS1234,80000.00,,440000.00\n"
    ).encode("utf-8")

    from fastapi import UploadFile
    from io import BytesIO
    file = UploadFile(filename="uco_statement.csv", file=BytesIO(uco_csv_content))

    req = MagicMock()
    res = await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert res["ok"] is True
    assert res["inserted_count"] == 2
    assert len(inserted_docs) == 2

    # Row 1 check
    assert inserted_docs[0]["date"] == "2026-08-20"
    assert inserted_docs[0]["narration"] == "CHQ DEP - BAXTER RETAIL CLIENT"
    assert inserted_docs[0]["reference_no"] == "CHQ54321"
    assert inserted_docs[0]["credit_amount"] == 120000.00
    assert inserted_docs[0]["debit_amount"] == 0.0
    assert inserted_docs[0]["running_balance"] == 520000.00
    assert inserted_docs[0]["match_status"] == "unmatched"

    # Row 2 check
    assert inserted_docs[1]["date"] == "2026-08-21"
    assert inserted_docs[1]["debit_amount"] == 80000.00
    assert inserted_docs[1]["running_balance"] == 440000.00


@pytest.mark.anyio
async def test_hdfc_statement_import_xlsx(monkeypatch):
    """Test importing HDFC bank statement Excel workbook (.xlsx)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement"
    
    # Headers
    ws.append(["Date", "Narration", "Chq./Ref.No.", "Withdrawal Amt.", "Deposit Amt.", "Closing Balance"])
    # Data rows
    ws.append(["15/08/2026", "ACH CR-MYNTRA DESIGNS-SETTLEMENT", "ACH001", "", 45250.50, 145250.50])
    ws.append(["16/08/2026", "UPI-SWIGGY-EXPENSE", "UPI999", 350.00, "", 144900.50])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    hdfc_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC - Online",
        "bank_name": "HDFC",
        "statement_format": {
            "sheet_locator": {"type": "fixed_name", "name": "Statement"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Date",
                "narration": "Narration",
                "reference": "Chq./Ref.No.",
                "debit_amount": "Withdrawal Amt.",
                "credit_amount": "Deposit Amt.",
                "balance": "Closing Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=hdfc_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    from fastapi import UploadFile
    file = UploadFile(filename="hdfc_statement.xlsx", file=bio)

    req = MagicMock()
    res = await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert res["ok"] is True
    assert res["inserted_count"] == 2
    assert len(inserted_docs) == 2
    assert inserted_docs[0]["credit_amount"] == 45250.50
    assert inserted_docs[0]["running_balance"] == 145250.50


@pytest.mark.anyio
async def test_uco_statement_import_real_xls(monkeypatch):
    """Test importing the real legacy UCO Bank .xls file via xlrd reader."""
    import os
    from pathlib import Path
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixture_path = fixtures_dir / "uco_statement_real.xls"
    downloads_path = Path(r"C:\Users\Dell\Downloads\OpTransactionHistoryUX531-08-2026 11_06_17 .xls")

    if fixture_path.exists():
        with open(fixture_path, "rb") as f:
            uco_xls_content = f.read()
    elif downloads_path.exists():
        with open(downloads_path, "rb") as f:
            uco_xls_content = f.read()
    else:
        pytest.skip("Real UCO Bank .xls file not found")

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "scan_for_columns", "must_contain_any": ["Tran. Date", "Narration"]},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Tran. Date",
                "narration": "Narration",
                "reference": "Chq. No.",
                "debit_amount": "Withdrawl",
                "credit_amount": "Deposit",
                "balance": "Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    from fastapi import UploadFile
    from io import BytesIO
    file = UploadFile(filename="OpTransactionHistoryUX531-08-2026 11_06_17 .xls", file=BytesIO(uco_xls_content))

    req = MagicMock()
    res = await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert res["ok"] is True
    assert res["inserted_count"] == 10
    assert len(inserted_docs) == 10

    # Ensure none of the metadata rows (Account Details, IFSC, Balance Details, etc.) were parsed as transactions
    narrations = [d["narration"] for d in inserted_docs]
    assert not any("Account Details" in n or "IFSC" in n or "Balance Details" in n for n in narrations)

    # Detailed assertion of all 10 real transaction rows
    expected_rows = [
        {"date": "2026-08-13", "debit": 10000.00, "credit": 0.0, "balance": 25304.00, "narr": "IB-To:18600110093108-TRTRumesh"},
        {"date": "2026-08-13", "debit": 25000.00, "credit": 0.0, "balance": 304.00, "narr": "IB-To:18600110093108-TRTRJeevrajPU"},
        {"date": "2026-08-18", "debit": 10.00, "credit": 0.0, "balance": 294.00, "narr": "MPAY/IMPSP/TRTR/623032483866/SBIN/XX131411/IMPS"},
        {"date": "2026-08-21", "debit": 0.0, "credit": 450000.00, "balance": 450294.00, "narr": "RTGS/HDFCR52026082198849971/SSK FOOTCARE MANUFACTU"},
        {"date": "2026-08-21", "debit": 206.50, "credit": 0.0, "balance": 450087.50, "narr": "Charges for CARD-ISSUE"},
        {"date": "2026-08-21", "debit": 450000.00, "credit": 0.0, "balance": 87.50, "narr": "IB-To:18600510000946-TRTRMysteva"},
        {"date": "2026-08-26", "debit": 0.0, "credit": 550000.00, "balance": 550087.50, "narr": "RTGS/HDFCR52026082650776660/SSK FOOTCARE MANUFACTU"},
        {"date": "2026-08-26", "debit": 16005.90, "credit": 0.0, "balance": 534081.60, "narr": "MPAY/IMPSP/TRTR/623834195323/ICIC/XX556181/IMPS"},
        {"date": "2026-08-26", "debit": 221000.00, "credit": 0.0, "balance": 313081.60, "narr": "eRTGS/UCBAR52026082600471738/UMESH CHOTURAM SUWASI"},
        {"date": "2026-08-26", "debit": 60000.00, "credit": 0.0, "balance": 253081.60, "narr": "IB-To:18600110093108-TRTRPukhrajMaterial"},
    ]

    for idx, exp in enumerate(expected_rows):
        doc = inserted_docs[idx]
        assert doc["date"] == exp["date"], f"Row {idx} date mismatch"
        assert doc["debit_amount"] == exp["debit"], f"Row {idx} debit mismatch"
        assert doc["credit_amount"] == exp["credit"], f"Row {idx} credit mismatch"
        assert doc["running_balance"] == exp["balance"], f"Row {idx} balance mismatch"
        assert doc["narration"] == exp["narr"], f"Row {idx} narration mismatch"
        assert doc["match_status"] == "unmatched"



@pytest.mark.anyio
async def test_uco_statement_import_misnamed_extension_xls(monkeypatch):
    """Test importing legacy .xls file even when misnamed with .xlsx extension via magic bytes detection."""
    from pathlib import Path
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixture_path = fixtures_dir / "uco_statement_real.xls"
    if not fixture_path.exists():
        pytest.skip("Real UCO Bank .xls file not found")

    with open(fixture_path, "rb") as f:
        uco_xls_content = f.read()

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "scan_for_columns", "must_contain_any": ["Tran. Date", "Narration"]},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Tran. Date",
                "narration": "Narration",
                "reference": "Chq. No.",
                "debit_amount": "Withdrawl",
                "credit_amount": "Deposit",
                "balance": "Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    from fastapi import UploadFile
    from io import BytesIO
    # Note: misnamed as .xlsx but content has OLE2 d0 cf 11 e0 magic bytes
    file = UploadFile(filename="statement_misnamed.xlsx", file=BytesIO(uco_xls_content))

    req = MagicMock()
    res = await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert res["ok"] is True
    assert res["inserted_count"] == 10
    assert len(inserted_docs) == 10


@pytest.mark.anyio
async def test_statement_import_metadata_autofill_and_confirmation(monkeypatch):
    """Test extracting IFSC/Account No. from statement header and confirming before saving."""
    from pathlib import Path
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixture_path = fixtures_dir / "uco_statement_real.xls"
    if not fixture_path.exists():
        pytest.skip("Real UCO Bank .xls file not found")

    with open(fixture_path, "rb") as f:
        uco_xls_content = f.read()

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    # Account is missing IFSC and full account_number
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "account_number_last4": "6054",
        "account_number": None,
        "ifsc": None,
        "branch": None,
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "scan_for_columns"},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Tran. Date",
                "narration": "Narration",
                "reference": "Chq. No.",
                "debit_amount": "Withdrawl",
                "credit_amount": "Deposit",
                "balance": "Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)
    mock_db.bank_accounts.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    mock_db.bank_statement_lines.insert_many = AsyncMock(return_value=MagicMock(inserted_ids=[ObjectId() for _ in range(10)]))

    from fastapi import UploadFile
    from io import BytesIO

    # 1. Normal import without confirm_account_update:
    # Must surface suggestions and require confirmation, but NOT modify bank_account in DB
    file1 = UploadFile(filename="uco_statement.xls", file=BytesIO(uco_xls_content))
    req = MagicMock()
    res1 = await routes.banking.import_bank_statement(acc_id, req, file=file1, dry_run=False, confirm_account_update=False)

    assert res1["ok"] is True
    assert res1["inserted_count"] == 10
    assert res1.get("requires_account_confirmation") is True
    assert res1.get("suggested_account_update") == {
        "account_number": "18600210006054",
        "ifsc": "UCBA0001860",
        "branch": "KOPAR KHAIRANE",
    }
    # Confirm bank_accounts.update_one was NOT called (no silent overwrite)
    mock_db.bank_accounts.update_one.assert_not_called()

    # 2. Import with confirm_account_update=True:
    # Must update bank_account in DB with the confirmed fields
    file2 = UploadFile(filename="uco_statement.xls", file=BytesIO(uco_xls_content))
    res2 = await routes.banking.import_bank_statement(acc_id, req, file=file2, dry_run=False, confirm_account_update=True)

    assert res2["ok"] is True
    assert res2.get("applied_account_update") == {
        "account_number": "18600210006054",
        "ifsc": "UCBA0001860",
        "branch": "KOPAR KHAIRANE",
    }
    mock_db.bank_accounts.update_one.assert_called_once()
    call_args = mock_db.bank_accounts.update_one.call_args[0]
    assert call_args[0] == {"_id": ObjectId(acc_id)}
    set_fields = call_args[1]["$set"]
    assert set_fields["account_number"] == "18600210006054"
    assert set_fields["ifsc"] == "UCBA0001860"
    assert set_fields["branch"] == "KOPAR KHAIRANE"


@pytest.mark.anyio
async def test_statement_import_unsupported_format_error(monkeypatch):
    """Test that uploading an unsupported format (like PDF or binary) returns a clear 400 error."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "bank_name": "UCO Bank",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Date",
                "narration": "Narration",
            },
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)

    from fastapi import UploadFile, HTTPException
    from io import BytesIO
    # Binary PDF-like content
    fake_pdf = b"%PDF-1.4\x00\x01\x02\x03\xff\xfe\xfd"
    file = UploadFile(filename="bank_statement.pdf", file=BytesIO(fake_pdf))

    req = MagicMock()
    with pytest.raises(HTTPException) as exc_info:
        await routes.banking.import_bank_statement(acc_id, req, file=file, dry_run=False)

    assert exc_info.value.status_code == 400
    assert "Unsupported file format" in exc_info.value.detail


@pytest.mark.anyio
async def test_uco_statement_preview_and_commit_flow(monkeypatch):
    """
    Re-test exact bug report scenario:
    1. Select UCO account.
    2. Upload real .xls file.
    3. Click Preview Layout (dry_run=True) -> Confirm it succeeds and shows accurate transaction preview.
    4. Confirm & Commit Import (dry_run=False) -> Confirm all 10 bank_statement_lines are inserted matching real file data.
    """
    from pathlib import Path
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixture_path = fixtures_dir / "uco_statement_real.xls"
    if not fixture_path.exists():
        pytest.skip("Real UCO Bank .xls file not found")

    with open(fixture_path, "rb") as f:
        uco_xls_content = f.read()

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    uco_acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Current A/C",
        "bank_name": "UCO Bank",
        "account_number_last4": "6054",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {
                "type": "scan_for_columns",
                "must_contain_any": ["Tran. Date", "Withdrawl", "Deposit", "Balance", "Narration"]
            },
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Tran. Date",
                "narration": "Narration",
                "reference": "Chq. No.",
                "debit_amount": "Withdrawl",
                "credit_amount": "Deposit",
                "balance": "Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=uco_acc_doc)

    inserted_docs = []
    async def mock_insert_many(docs):
        inserted_docs.extend(docs)
        return MagicMock(inserted_ids=[ObjectId() for _ in docs])
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)

    from fastapi import UploadFile
    from io import BytesIO

    # ── STEP 1: PREVIEW LAYOUT (dry_run=True) ──
    file_preview = UploadFile(
        filename="OpTransactionHistoryUX531-08-2026 11_06_17 .xls",
        file=BytesIO(uco_xls_content)
    )
    req = MagicMock()
    preview_res = await routes.banking.import_bank_statement(
        acc_id, req, file=file_preview, dry_run=True
    )

    assert preview_res["ok"] is True
    assert preview_res["dry_run"] is True
    assert preview_res["parsed_count"] == 10
    assert preview_res["total_file_rows"] == 10
    assert len(preview_res["sample"]) > 0
    # Confirm DB insert was NOT called during preview
    mock_db.bank_statement_lines.insert_many.assert_not_called()
    assert len(inserted_docs) == 0

    # ── STEP 2: CONFIRM & COMMIT IMPORT (dry_run=False) ──
    file_commit = UploadFile(
        filename="OpTransactionHistoryUX531-08-2026 11_06_17 .xls",
        file=BytesIO(uco_xls_content)
    )
    commit_res = await routes.banking.import_bank_statement(
        acc_id, req, file=file_commit, dry_run=False
    )

    assert commit_res["ok"] is True
    assert commit_res["dry_run"] is False
    assert commit_res["inserted_count"] == 10
    mock_db.bank_statement_lines.insert_many.assert_called_once()
    assert len(inserted_docs) == 10

    # Confirm created statement lines match exact real file transactions
    assert inserted_docs[0]["date"] == "2026-08-13"
    assert inserted_docs[0]["debit_amount"] == 10000.00
    assert inserted_docs[0]["credit_amount"] == 0.0
    assert inserted_docs[0]["running_balance"] == 25304.00
    assert inserted_docs[0]["narration"] == "IB-To:18600110093108-TRTRumesh"

    assert inserted_docs[3]["date"] == "2026-08-21"
    assert inserted_docs[3]["debit_amount"] == 0.0
    assert inserted_docs[3]["credit_amount"] == 450000.00
    assert inserted_docs[3]["running_balance"] == 450294.00
    assert inserted_docs[3]["narration"] == "RTGS/HDFCR52026082198849971/SSK FOOTCARE MANUFACTU"

    assert inserted_docs[9]["date"] == "2026-08-26"
    assert inserted_docs[9]["debit_amount"] == 60000.00
    assert inserted_docs[9]["credit_amount"] == 0.0
    assert inserted_docs[9]["running_balance"] == 253081.60
    assert inserted_docs[9]["narration"] == "IB-To:18600110093108-TRTRPukhrajMaterial"


@pytest.mark.anyio
async def test_auto_reconcile_online_account_credits(monkeypatch):
    """Test auto-reconciliation on online account: confident match, ambiguous match, and no match."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC - Online",
        "account_type": "online_channel",
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    line1_id = ObjectId()
    line2_id = ObjectId()
    line3_id = ObjectId()

    unmatched_lines = [
        # Confident match candidate: amount 45250.0 on 2026-08-15 (matches 45250.5 on 2026-08-14)
        {
            "_id": line1_id,
            "bank_account_id": acc_id,
            "date": "2026-08-15",
            "credit_amount": 45250.0,
            "debit_amount": 0.0,
            "match_status": "unmatched",
            "narration": "ACH CR-MYNTRA DESIGNS",
        },
        # Ambiguous candidate: amount 10000.0 on 2026-08-16 (two settlements of 10000.0)
        {
            "_id": line2_id,
            "bank_account_id": acc_id,
            "date": "2026-08-16",
            "credit_amount": 10000.0,
            "debit_amount": 0.0,
            "match_status": "unmatched",
            "narration": "ACH CR-FLIPKART",
        },
        # No match candidate: amount 99999.0
        {
            "_id": line3_id,
            "bank_account_id": acc_id,
            "date": "2026-08-17",
            "credit_amount": 99999.0,
            "debit_amount": 0.0,
            "match_status": "unmatched",
            "narration": "UNKNOWN CREDIT",
        },
    ]

    mock_cursor = MagicMock()
    mock_cursor.sort = MagicMock(return_value=mock_cursor)
    mock_cursor.to_list = AsyncMock(return_value=unmatched_lines)
    mock_db.bank_statement_lines.find = MagicMock(return_value=mock_cursor)

    settle1_id = ObjectId()
    settle2_id = ObjectId()
    settle3_id = ObjectId()

    settlement_docs = [
        # Exactly 1 match for Line 1
        {
            "_id": settle1_id,
            "order_release_id": "ORD001",
            "net_payout": 45250.50,
            "settlement_date": "2026-08-14",
            "bank_account_id": None,
        },
        # 2 settlements matching Line 2 (ambiguous!)
        {
            "_id": settle2_id,
            "order_release_id": "ORD002",
            "net_payout": 10000.0,
            "settlement_date": "2026-08-16",
            "bank_account_id": None,
        },
        {
            "_id": settle3_id,
            "order_release_id": "ORD003",
            "net_payout": 10000.0,
            "settlement_date": "2026-08-16",
            "bank_account_id": None,
        },
    ]

    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=settlement_docs)))
    mock_db.payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))

    updated_statement_lines = {}
    async def mock_update_stmt(q, u):
        updated_statement_lines[str(q["_id"])] = u["$set"]
        return MagicMock(matched_count=1)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=mock_update_stmt)

    updated_settlements = {}
    async def mock_update_settle(q, u):
        updated_settlements[str(q["_id"])] = u["$set"]
        return MagicMock(matched_count=1)
    mock_db.online_settlements.update_one = AsyncMock(side_effect=mock_update_settle)

    req = MagicMock()
    res = await routes.banking.reconcile_bank_account(
        id=acc_id,
        request=req,
        date_window_days=3,
        amount_tolerance=1.0,
        min_confidence=0.70,
        dry_run=False,
    )

    assert res["ok"] is True
    assert res["total_unmatched_evaluated"] == 3
    assert res["auto_matched_count"] == 1
    assert res["ambiguous_count"] == 1
    assert res["no_match_count"] == 1

    # Confirmed match verification
    assert str(line1_id) in updated_statement_lines
    assert updated_statement_lines[str(line1_id)]["match_status"] == "matched"
    assert updated_statement_lines[str(line1_id)]["matched_to"]["type"] == "settlement"
    assert updated_statement_lines[str(line1_id)]["matched_to"]["ref_id"] == str(settle1_id)

    # Confirmed settlement bank_account_id set
    assert str(settle1_id) in updated_settlements
    assert updated_settlements[str(settle1_id)]["bank_account_id"] == acc_id

    # Line 2 and Line 3 must NOT have been updated (remained unmatched)
    assert str(line2_id) not in updated_statement_lines
    assert str(line3_id) not in updated_statement_lines


@pytest.mark.anyio
async def test_auto_reconcile_b2b_account_credits_and_debits(monkeypatch):
    """Test auto-reconciliation on offline account with client payment credit & expense/vendor debits."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "UCO Bank - Offline",
        "account_type": "b2b_client",
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    credit_line_id = ObjectId()
    debit_vendor_line_id = ObjectId()
    debit_expense_line_id = ObjectId()

    unmatched_lines = [
        # Credit line: client payment 120,000 on 2026-08-20
        {
            "_id": credit_line_id,
            "bank_account_id": acc_id,
            "date": "2026-08-20",
            "credit_amount": 120000.0,
            "debit_amount": 0.0,
            "match_status": "unmatched",
            "narration": "CHQ DEP - BAXTER RETAIL",
        },
        # Debit line 1: vendor payment 25,000 on 2026-08-22
        {
            "_id": debit_vendor_line_id,
            "bank_account_id": acc_id,
            "date": "2026-08-22",
            "credit_amount": 0.0,
            "debit_amount": 25000.0,
            "match_status": "unmatched",
            "narration": "NEFT - AGRA SOLE SUPPLIER",
        },
        # Debit line 2: office rent expense 45,000 on 2026-08-01
        {
            "_id": debit_expense_line_id,
            "bank_account_id": acc_id,
            "date": "2026-08-01",
            "credit_amount": 0.0,
            "debit_amount": 45000.0,
            "match_status": "unmatched",
            "narration": "NEFT - FACTORY RENT",
        },
    ]

    mock_cursor = MagicMock()
    mock_cursor.sort = MagicMock(return_value=mock_cursor)
    mock_cursor.to_list = AsyncMock(return_value=unmatched_lines)
    mock_db.bank_statement_lines.find = MagicMock(return_value=mock_cursor)

    client_pay_id = ObjectId()
    vendor_pay_id = ObjectId()
    expense_id = ObjectId()

    # Client payment
    client_pay_doc = {
        "_id": client_pay_id,
        "payment_date": "2026-08-19",
        "amount": 120000.0,
        "client_name": "Baxter Retail",
        "type": "client_payment",
        "vendor_id": None,
        "bank_account_id": None,
    }

    # Vendor payment
    vendor_pay_doc = {
        "_id": vendor_pay_id,
        "payment_date": "2026-08-22",
        "amount": 25000.0,
        "type": "vendor_payment",
        "vendor_id": "v_agra_1",
        "bank_account_id": None,
    }

    # Expense
    expense_doc = {
        "_id": expense_id,
        "date": "2026-08-01",
        "amount": 45000.0,
        "category": "Rent & Utilities",
        "payee": "Factory Landlord",
        "bank_account_id": None,
    }

    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    
    # Setup payments and expenses find responses
    def mock_payments_find(q):
        if q.get("type") == {"$ne": "vendor_payment"}:
            return [client_pay_doc]
        return [vendor_pay_doc]
    mock_db.payments.find = MagicMock(side_effect=lambda q: MagicMock(to_list=AsyncMock(return_value=mock_payments_find(q))))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[expense_doc])))

    updated_stmt = {}
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=lambda q, u: updated_stmt.update({str(q["_id"]): u["$set"]}))
    mock_db.payments.update_one = AsyncMock()
    mock_db.expenses.update_one = AsyncMock()

    req = MagicMock()
    res = await routes.banking.reconcile_bank_account(
        id=acc_id,
        request=req,
        date_window_days=3,
        amount_tolerance=1.0,
        min_confidence=0.70,
        dry_run=False,
    )

    assert res["ok"] is True
    assert res["total_unmatched_evaluated"] == 3
    assert res["auto_matched_count"] == 3
    assert res["ambiguous_count"] == 0
    assert res["no_match_count"] == 0

    # Verify all 3 matched
    assert updated_stmt[str(credit_line_id)]["matched_to"]["type"] == "payment"
    assert updated_stmt[str(credit_line_id)]["matched_to"]["ref_id"] == str(client_pay_id)

    assert updated_stmt[str(debit_vendor_line_id)]["matched_to"]["type"] == "vendor_payment"
    assert updated_stmt[str(debit_vendor_line_id)]["matched_to"]["ref_id"] == str(vendor_pay_id)

    assert updated_stmt[str(debit_expense_line_id)]["matched_to"]["type"] == "expense"
    assert updated_stmt[str(debit_expense_line_id)]["matched_to"]["ref_id"] == str(expense_id)


@pytest.mark.anyio
async def test_suggest_and_confirm_transfers_and_reconciliation_summary(monkeypatch):
    """
    Test scenario:
    1. Debit in HDFC (₹50,000 on Aug 15) and Credit in UCO (₹50,000 on Aug 16) surfaced as suggested transfer pair.
    2. Explicit confirmation marks both lines match_status='transfer' pointing to each other.
    3. Reconciliation summary strictly excludes the transfer pair from income/expenses.
    """
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    hdfc_acc_id = str(ObjectId())
    uco_acc_id = str(ObjectId())

    acc_docs = [
        {"_id": ObjectId(hdfc_acc_id), "name": "HDFC - Online", "bank_name": "HDFC", "account_type": "online_channel"},
        {"_id": ObjectId(uco_acc_id), "name": "UCO Bank - Offline", "bank_name": "UCO Bank", "account_type": "b2b_client"},
    ]
    mock_db.bank_accounts.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=acc_docs)))

    hdfc_debit_trf_id = ObjectId()
    uco_credit_trf_id = ObjectId()
    uco_income_line_id = ObjectId()
    hdfc_expense_line_id = ObjectId()

    # In-memory statement lines store
    statement_store = {
        str(hdfc_debit_trf_id): {
            "_id": hdfc_debit_trf_id,
            "bank_account_id": hdfc_acc_id,
            "date": "2026-08-15",
            "debit_amount": 50000.0,
            "credit_amount": 0.0,
            "match_status": "unmatched",
            "narration": "NEFT OUT - TRF TO UCO BANK",
        },
        str(uco_credit_trf_id): {
            "_id": uco_credit_trf_id,
            "bank_account_id": uco_acc_id,
            "date": "2026-08-16",
            "debit_amount": 0.0,
            "credit_amount": 50000.0,
            "match_status": "unmatched",
            "narration": "NEFT IN - TRF FROM HDFC ONLINE",
        },
        str(uco_income_line_id): {
            "_id": uco_income_line_id,
            "bank_account_id": uco_acc_id,
            "date": "2026-08-18",
            "debit_amount": 0.0,
            "credit_amount": 100000.0,
            "match_status": "matched",
            "narration": "CHQ DEP - CLIENT PAYMENT",
        },
        str(hdfc_expense_line_id): {
            "_id": hdfc_expense_line_id,
            "bank_account_id": hdfc_acc_id,
            "date": "2026-08-19",
            "debit_amount": 20000.0,
            "credit_amount": 0.0,
            "match_status": "matched",
            "narration": "RAW MATERIAL PAYMENT",
        },
    }

    def mock_stmt_find(q):
        status_filter = q.get("match_status")
        res = list(statement_store.values())
        if status_filter:
            res = [r for r in res if r.get("match_status") == status_filter]
        mock_cursor = MagicMock()
        mock_cursor.sort = MagicMock(return_value=mock_cursor)
        mock_cursor.to_list = AsyncMock(return_value=res)
        return mock_cursor

    mock_db.bank_statement_lines.find = MagicMock(side_effect=mock_stmt_find)

    async def mock_stmt_find_one(q):
        oid_key = str(q.get("_id"))
        return statement_store.get(oid_key)
    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=mock_stmt_find_one)

    async def mock_stmt_update_one(q, u):
        oid_key = str(q.get("_id"))
        if oid_key in statement_store:
            statement_store[oid_key].update(u.get("$set", {}))
        return MagicMock(matched_count=1)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=mock_stmt_update_one)

    req = MagicMock()

    # Step 1: Scan for suggested transfer pairs
    suggestions = await routes.banking.get_suggested_transfers(
        request=req,
        date_window_days=3,
        amount_tolerance=1.0,
    )
    assert suggestions["ok"] is True
    assert suggestions["total_suggestions"] == 1
    pair = suggestions["pairs"][0]
    assert pair["from_line"]["id"] == str(hdfc_debit_trf_id)
    assert pair["to_line"]["id"] == str(uco_credit_trf_id)
    assert pair["amount_diff"] == 0.0
    assert pair["day_diff"] == 1

    # Both lines must still be unmatched before confirmation
    assert statement_store[str(hdfc_debit_trf_id)]["match_status"] == "unmatched"
    assert statement_store[str(uco_credit_trf_id)]["match_status"] == "unmatched"

    # Step 2: Explicitly confirm the transfer pair
    from models.banking import TransferConfirmIn
    confirm_res = await routes.banking.confirm_transfer_pair(
        payload=TransferConfirmIn(
            from_line_id=str(hdfc_debit_trf_id),
            to_line_id=str(uco_credit_trf_id),
            notes="Liquidity rebalance to UCO account",
        ),
        request=req,
    )
    assert confirm_res["ok"] is True
    assert statement_store[str(hdfc_debit_trf_id)]["match_status"] == "transfer"
    assert statement_store[str(hdfc_debit_trf_id)]["matched_to"]["ref_id"] == str(uco_credit_trf_id)
    assert statement_store[str(uco_credit_trf_id)]["match_status"] == "transfer"
    assert statement_store[str(uco_credit_trf_id)]["matched_to"]["ref_id"] == str(hdfc_debit_trf_id)

    # Step 3: Verify reconciliation summary excludes transfer pair from income/expenses
    summary_res = await routes.banking.get_reconciliation_summary(request=req)
    assert summary_res["ok"] is True
    s = summary_res["summary"]

    # Total income is ONLY the ₹100,000 client payment (the ₹50,000 transfer credit is excluded!)
    assert s["total_income"] == 100000.0
    assert s["matched_income"] == 100000.0
    assert s["unmatched_income"] == 0.0

    # Total expenses is ONLY the ₹20,000 raw material expense (the ₹50,000 transfer debit is excluded!)
    assert s["total_expenses"] == 20000.0
    assert s["matched_expenses"] == 20000.0
    assert s["unmatched_expenses"] == 0.0

    # Net operating cashflow is 100,000 - 20,000 = 80,000
    assert s["net_operating_cashflow"] == 80000.0

    # Transfers tracked separately
    assert s["total_transfers_volume"] == 50000.0
    assert s["transfer_lines_count"] == 2


@pytest.mark.anyio
async def test_unmatched_erp_candidates_listing(monkeypatch):
    """Test get_unmatched_erp_candidates returns settlements, payments, and expenses."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    
    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC - Online",
        "account_type": "online_channel",
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    settlements = [
        {"_id": ObjectId(), "platform": "myntra", "seller_order_id": "ORD999", "net_payout": 3450.0, "settlement_date": "2026-08-20", "bank_account_id": None}
    ]
    expenses = [
        {"_id": ObjectId(), "category": "Office", "payee": "Broadband Provider", "amount": 1500.0, "date": "2026-08-21", "bank_account_id": None}
    ]
    vendor_payments = [
        {"_id": ObjectId(), "type": "vendor_payment", "vendor_name": "Sole Corp", "amount": 12000.0, "payment_date": "2026-08-22", "bank_account_id": None}
    ]

    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(sort=MagicMock(return_value=MagicMock(limit=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=settlements)))))))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(sort=MagicMock(return_value=MagicMock(limit=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=expenses)))))))
    mock_db.payments.find = MagicMock(return_value=MagicMock(sort=MagicMock(return_value=MagicMock(limit=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=vendor_payments)))))))

    req = MagicMock()
    res = await routes.banking.get_unmatched_erp_candidates(
        request=req,
        bank_account_id=acc_id,
        side="all",
    )

    assert res["ok"] is True
    assert res["total"] == 3
    types = [c["type"] for c in res["candidates"]]
    assert "settlement" in types
    assert "expense" in types
    assert "vendor_payment" in types


@pytest.mark.anyio
async def test_auto_reconcile_expense_with_pre_set_bank_account(monkeypatch):
    """Ensure auto-reconcile on Account A does NOT match an expense pre-assigned to Account B."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}

    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    account_a_id = str(ObjectId())
    account_b_id = str(ObjectId())

    acc_a_doc = {
        "_id": ObjectId(account_a_id),
        "name": "Account A - UCO Bank",
        "account_type": "b2b_client",
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_a_doc)

    # Statement line on Account A (debit of 15000.0 on 2026-08-10)
    line_id = ObjectId()
    unmatched_lines = [
        {
            "_id": line_id,
            "bank_account_id": account_a_id,
            "date": "2026-08-10",
            "credit_amount": 0.0,
            "debit_amount": 15000.0,
            "match_status": "unmatched",
            "narration": "OFFICE EXPENSE PAYMENT",
        }
    ]

    mock_cursor = MagicMock()
    mock_cursor.sort = MagicMock(return_value=mock_cursor)
    mock_cursor.to_list = AsyncMock(return_value=unmatched_lines)
    mock_db.bank_statement_lines.find = MagicMock(return_value=mock_cursor)

    # Expense is pre-assigned to Account B
    expense_doc = {
        "_id": ObjectId(),
        "date": "2026-08-10",
        "amount": 15000.0,
        "category": "Office & Administrative",
        "payee": "Stationery Mart",
        "bank_account_id": account_b_id,
    }

    # When querying expenses, simulate MongoDB filtering on bank_account_id
    def mock_expenses_find(query):
        acc_filter = query.get("bank_account_id", {})
        allowed_accounts = acc_filter.get("$in", []) if isinstance(acc_filter, dict) else [acc_filter]
        matching = [expense_doc] if expense_doc.get("bank_account_id") in allowed_accounts else []
        return MagicMock(to_list=AsyncMock(return_value=matching))

    mock_db.expenses.find = MagicMock(side_effect=mock_expenses_find)
    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))

    updated_stmt = {}
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=lambda q, u: updated_stmt.update({str(q["_id"]): u["$set"]}))
    mock_db.expenses.update_one = AsyncMock()

    req = MagicMock()
    # Run auto-reconcile on Account A
    res = await routes.banking.reconcile_bank_account(
        id=account_a_id,
        request=req,
        date_window_days=3,
        amount_tolerance=1.0,
        dry_run=False,
    )

    assert res["ok"] is True
    assert res["total_unmatched_evaluated"] == 1
    assert res["auto_matched_count"] == 0
    assert res["no_match_count"] == 1
    assert str(line_id) not in updated_stmt
    mock_db.expenses.update_one.assert_not_called()


def test_matched_to_cash_withdrawal_model():
    """Verify MatchedTo model accepts cash_withdrawal type."""
    m = MatchedTo(type="cash_withdrawal", ref_id="cash_123")
    assert m.type == "cash_withdrawal"
    assert m.ref_id == "cash_123"


def test_cash_withdrawal_pattern_detection():
    """Verify pattern detector detects ATM, CASH, SELF, CWDR, etc. in narrations."""
    assert _is_cash_withdrawal_candidate("ATM WDL / 9876 / KOTAK") is True
    assert _is_cash_withdrawal_candidate("CHQ PAID - SELF") is True
    assert _is_cash_withdrawal_candidate("SELF CHEQUE 504011") is True
    assert _is_cash_withdrawal_candidate("CASH WITHDRAWAL BRANCH AGRA") is True
    assert _is_cash_withdrawal_candidate("CWDR-4321-NEW DELHI") is True
    assert _is_cash_withdrawal_candidate("EAW CASH DISPENSE") is True

    # Negative patterns
    assert _is_cash_withdrawal_candidate("NEFT-MYNTRA DESIGNS-SETTLEMENT") is False
    assert _is_cash_withdrawal_candidate("RTGS TO RELIANCE RETAIL") is False
    assert _is_cash_withdrawal_candidate("MONTHLY WAREHOUSE RENT") is False
    assert _is_cash_withdrawal_candidate("") is False
    assert _is_cash_withdrawal_candidate(None) is False


@pytest.mark.anyio
async def test_cash_withdrawal_routes_and_cash_ledger(monkeypatch):
    """Verify suggested cash withdrawals, confirmation flow, cash_ledger collection, and summary."""
    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}

    import routes.banking
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = ObjectId()
    line_id = ObjectId()
    cash_ledger_id = ObjectId()

    # 1. Mock statement line for suggestion
    atm_line_doc = {
        "_id": line_id,
        "bank_account_id": str(acc_id),
        "date": "2026-08-28",
        "narration": "ATM CASH WDL - SELF CHQ",
        "reference_no": "ATM5566",
        "debit_amount": 25000.0,
        "credit_amount": 0.0,
        "running_balance": 175000.0,
        "match_status": "unmatched",
    }

    acc_doc = {
        "_id": acc_id,
        "name": "HDFC Primary Current",
        "bank_name": "HDFC Bank",
        "account_type": "b2b_client",
        "opening_balance": 200000.0,
    }

    mock_db.bank_statement_lines.find = MagicMock(return_value=MagicMock(
        sort=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[atm_line_doc])))
    ))
    mock_db.bank_accounts.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[acc_doc])))

    req = MagicMock()

    # 2. Get suggested cash withdrawals
    suggestions_res = await get_suggested_cash_withdrawals(req, bank_account_id=str(acc_id))
    assert suggestions_res["ok"] is True
    assert suggestions_res["total_suggestions"] == 1
    candidate = suggestions_res["candidates"][0]
    assert candidate["id"] == str(line_id)
    assert candidate["amount"] == 25000.0
    assert "cash withdrawal pattern" in candidate["suggestion_reason"]

    # 3. Confirm cash withdrawal
    mock_db.bank_statement_lines.find_one = AsyncMock(return_value=atm_line_doc)
    mock_db.cash_ledger.insert_one = AsyncMock(return_value=MagicMock(inserted_id=cash_ledger_id))
    mock_db.bank_statement_lines.update_one = AsyncMock(return_value=MagicMock(matched_count=1))

    confirm_payload = CashWithdrawalConfirmIn(
        statement_line_id=str(line_id),
        notes="Cash drawn for factory floor weekly expenses",
    )
    confirm_res = await confirm_cash_withdrawal(confirm_payload, req)
    assert confirm_res["ok"] is True
    assert confirm_res["cash_ledger_id"] == str(cash_ledger_id)
    assert confirm_res["amount"] == 25000.0
    assert confirm_res["remaining_balance"] == 25000.0

    # Verify cash_ledger insert
    inserted_cash_doc = mock_db.cash_ledger.insert_one.call_args[0][0]
    assert inserted_cash_doc["amount"] == 25000.0
    assert inserted_cash_doc["remaining_balance"] == 25000.0
    assert inserted_cash_doc["bank_account_id"] == str(acc_id)
    assert inserted_cash_doc["source_statement_line_id"] == str(line_id)
    assert inserted_cash_doc["notes"] == "Cash drawn for factory floor weekly expenses"

    # Verify statement line update to matched_to: cash_withdrawal
    updated_stmt_set = mock_db.bank_statement_lines.update_one.call_args[0][1]["$set"]
    assert updated_stmt_set["match_status"] == "matched"
    assert updated_stmt_set["matched_to"]["type"] == "cash_withdrawal"
    assert updated_stmt_set["matched_to"]["ref_id"] == str(cash_ledger_id)

    # 4. List cash ledger entries
    created_cash_doc = {
        "_id": cash_ledger_id,
        "bank_account_id": str(acc_id),
        "source_statement_line_id": str(line_id),
        "date": "2026-08-28",
        "amount": 25000.0,
        "remaining_balance": 25000.0,
        "notes": "Cash drawn for factory floor weekly expenses",
    }
    mock_db.cash_ledger.find = MagicMock(return_value=MagicMock(
        sort=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[created_cash_doc])))
    ))

    ledger_res = await list_cash_ledger(req, bank_account_id=str(acc_id))
    assert ledger_res["ok"] is True
    assert ledger_res["total_count"] == 1
    assert ledger_res["total_withdrawn"] == 25000.0
    assert ledger_res["total_remaining_balance"] == 25000.0
    assert len(ledger_res["items"]) == 1

    # 5. Reconciliation Summary with confirmed cash withdrawal
    matched_line_doc = {
        **atm_line_doc,
        "match_status": "matched",
        "matched_to": {"type": "cash_withdrawal", "ref_id": str(cash_ledger_id)},
    }
    mock_db.bank_statement_lines.find = MagicMock(return_value=MagicMock(
        to_list=AsyncMock(return_value=[matched_line_doc])
    ))
    mock_db.cash_ledger.find = MagicMock(return_value=MagicMock(
        to_list=AsyncMock(return_value=[created_cash_doc])
    ))

    summary_res = await routes.banking.get_reconciliation_summary(req, bank_account_id=str(acc_id))
    assert summary_res["ok"] is True
    summary_data = summary_res["summary"]
    # Reconciled debit (matched expense)
    assert summary_data["matched_expenses"] == 25000.0
    assert summary_data["unmatched_expenses"] == 0.0
    assert summary_data["total_cash_in_hand"] == 25000.0
    account_stats = summary_res["accounts"][0]
    assert account_stats["total_reconciled_debits"] == 25000.0


@pytest.mark.anyio
async def test_full_cash_withdrawal_to_wage_payments_audit_trail_flow(monkeypatch):
    """
    Full Flow Verification:
    1. Import statement line with cash withdrawal.
    2. Confirm statement line as cash_withdrawal -> creates cash_ledger entry.
    3. Record wage payments for 3 workers funded by this cash_ledger entry.
    4. View cash ledger detail & statement line detail:
       - Confirm all linked wage payments are visible with correct amounts & workers.
       - Confirm remaining unallocated balance is correctly shown (15000 - 13500 = 1500).
    """
    mock_db = MagicMock()
    acc_id = ObjectId()
    line_id = ObjectId()
    cash_ledger_id = ObjectId()

    # 1. Statement line doc
    atm_line_doc = {
        "_id": line_id,
        "bank_account_id": str(acc_id),
        "date": "2026-08-25",
        "narration": "ATM CASH WDL - SELF CHQ",
        "reference_no": "ATM5544",
        "debit_amount": 15000.0,
        "credit_amount": 0.0,
        "running_balance": 85000.0,
        "match_status": "unmatched",
    }

    # Store maps for mock_db
    statement_lines_store = {str(line_id): dict(atm_line_doc)}
    cash_ledger_store = {}
    wage_payments_store = {}

    # Setup mock_db methods
    async def mock_find_one_line(query):
        return statement_lines_store.get(str(query.get("_id")))

    async def mock_update_line(query, update):
        lid = str(query.get("_id"))
        if lid in statement_lines_store:
            statement_lines_store[lid].update(update.get("$set", {}))
        return MagicMock(matched_count=1)

    async def mock_insert_cash(doc):
        doc["_id"] = cash_ledger_id
        cash_ledger_store[str(cash_ledger_id)] = doc
        return MagicMock(inserted_id=cash_ledger_id)

    async def mock_find_one_cash(query):
        return cash_ledger_store.get(str(query.get("_id")))

    async def mock_update_cash(query, update):
        cid = str(query.get("_id"))
        if cid in cash_ledger_store:
            doc = cash_ledger_store[cid]
            if "$inc" in update:
                for k, v in update["$inc"].items():
                    doc[k] = round(doc.get(k, 0.0) + v, 2)
            if "$set" in update:
                doc.update(update["$set"])
        return MagicMock(matched_count=1)

    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=mock_find_one_line)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=mock_update_line)
    mock_db.cash_ledger.insert_one = AsyncMock(side_effect=mock_insert_cash)
    mock_db.cash_ledger.find_one = AsyncMock(side_effect=mock_find_one_cash)
    mock_db.cash_ledger.update_one = AsyncMock(side_effect=mock_update_cash)
    mock_db.cash_ledger.find = MagicMock(return_value=MagicMock(
        to_list=AsyncMock(side_effect=lambda limit: list(cash_ledger_store.values()))
    ))
    mock_db.wage_payments.find = MagicMock(return_value=MagicMock(
        to_list=AsyncMock(side_effect=lambda limit: list(wage_payments_store.values()))
    ))
    mock_db.bank_statement_lines.count_documents = AsyncMock(return_value=1)
    mock_db.bank_statement_lines.find = MagicMock(return_value=MagicMock(
        sort=MagicMock(return_value=MagicMock(
            skip=MagicMock(return_value=MagicMock(
                limit=MagicMock(return_value=MagicMock(
                    to_list=AsyncMock(side_effect=lambda limit: list(statement_lines_store.values()))
                ))
            ))
        ))
    ))

    req = MagicMock()
    req.state = MagicMock()
    req.state.user = {"email": "admin@sskfootcare.com", "role": "admin", "name": "Admin"}
    monkeypatch.setattr(banking_routes, "_get_user", AsyncMock(return_value=req.state.user))
    monkeypatch.setattr(banking_routes, "_get_db", lambda r: mock_db)

    # 2. Confirm Cash Withdrawal
    payload = CashWithdrawalConfirmIn(
        statement_line_id=str(line_id),
        notes="Cash withdrawal for factory floor worker payouts",
    )
    confirm_res = await confirm_cash_withdrawal(payload, req)
    assert confirm_res["ok"] is True
    assert confirm_res["amount"] == 15000.0
    assert confirm_res["remaining_balance"] == 15000.0

    # Verify statement line was marked matched
    assert statement_lines_store[str(line_id)]["match_status"] == "matched"
    assert statement_lines_store[str(line_id)]["matched_to"]["type"] == "cash_withdrawal"
    assert statement_lines_store[str(line_id)]["matched_to"]["ref_id"] == str(cash_ledger_id)

    # 3. Simulate 3 WagePaymentIn records drawing from this cash_ledger entry
    w1_id, w2_id, w3_id = str(ObjectId()), str(ObjectId()), str(ObjectId())
    wp1_doc = {
        "_id": ObjectId(),
        "worker_id": w1_id,
        "worker_name": "Ramesh Karigar",
        "amount": 5000.0,
        "period_from": "2026-08-01",
        "period_to": "2026-08-15",
        "paid_via": "cash",
        "cash_ledger_id": str(cash_ledger_id),
        "date": "2026-08-26",
        "notes": "Fortnightly wage payout",
    }
    wp2_doc = {
        "_id": ObjectId(),
        "worker_id": w2_id,
        "worker_name": "Suresh Karigar",
        "amount": 6000.0,
        "period_from": "2026-08-01",
        "period_to": "2026-08-15",
        "paid_via": "cash",
        "cash_ledger_id": str(cash_ledger_id),
        "date": "2026-08-26",
        "notes": "Piece-rate wages",
    }
    wp3_doc = {
        "_id": ObjectId(),
        "worker_id": w3_id,
        "worker_name": "Dinesh Karigar",
        "amount": 2500.0,
        "period_from": "2026-08-01",
        "period_to": "2026-08-15",
        "paid_via": "cash",
        "cash_ledger_id": str(cash_ledger_id),
        "date": "2026-08-26",
        "notes": "Stitching wages + bonus",
        "override_reason": "Festival advance included",
    }

    # Save to wage payments store
    wage_payments_store[str(wp1_doc["_id"])] = wp1_doc
    wage_payments_store[str(wp2_doc["_id"])] = wp2_doc
    wage_payments_store[str(wp3_doc["_id"])] = wp3_doc

    # Draw down cash_ledger remaining_balance (5000 + 6000 + 2500 = 13500)
    cash_ledger_store[str(cash_ledger_id)]["remaining_balance"] = 1500.0

    # 4. View Cash Ledger Detail via GET /banking/cash-ledger/{id}
    detail_res = await get_cash_ledger_detail(str(cash_ledger_id), req)
    assert detail_res["ok"] is True
    assert detail_res["withdrawal_amount"] == 15000.0
    assert detail_res["allocated_amount"] == 13500.0
    assert detail_res["remaining_balance"] == 1500.0
    assert detail_res["wage_payment_count"] == 3

    worker_names = [w["worker_name"] for w in detail_res["wage_payments"]]
    assert "Ramesh Karigar" in worker_names
    assert "Suresh Karigar" in worker_names
    assert "Dinesh Karigar" in worker_names

    # 5. View Statement Lines via GET /banking/statement-lines
    lines_res = await list_statement_lines(req)
    line_item = lines_res["items"][0]
    assert line_item["matched_to"]["type"] == "cash_withdrawal"
    assert "cash_ledger_info" in line_item

    cl_info = line_item["cash_ledger_info"]
    assert cl_info["withdrawal_amount"] == 15000.0
    assert cl_info["allocated_amount"] == 13500.0
    assert cl_info["remaining_balance"] == 1500.0
    assert cl_info["wage_payment_count"] == 3


def test_expense_paid_via_cash_models_validation():
    """Verify ExpenseIn and ExpenseUpdate accept paid_via and cash_ledger_id."""
    e_bank = ExpenseIn(
        category="Rent & Utilities",
        amount=25000.0,
        date="2026-08-01",
        payee="Landlord",
        paid_via="bank",
        bank_account_id="acc_1",
    )
    assert e_bank.paid_via == "bank"
    assert e_bank.bank_account_id == "acc_1"
    assert e_bank.cash_ledger_id is None

    e_cash = ExpenseIn(
        category="Office & Administrative",
        amount=1200.0,
        date="2026-08-02",
        payee="Stationery Mart",
        paid_via="cash",
        cash_ledger_id="cash_leg_99",
    )
    assert e_cash.paid_via == "cash"
    assert e_cash.cash_ledger_id == "cash_leg_99"

    up = ExpenseUpdate(paid_via="cash", cash_ledger_id="cash_leg_100")
    assert up.paid_via == "cash"
    assert up.cash_ledger_id == "cash_leg_100"


@pytest.mark.anyio
async def test_create_cash_expense_success_and_drawdown(monkeypatch):
    """Verify creating a cash expense decrements cash_ledger remaining_balance via atomic conditional update."""
    from routes.expenses import create_expense

    cash_id = ObjectId()
    cash_entry = {
        "_id": cash_id,
        "amount": 10000.0,
        "remaining_balance": 10000.0,
        "date": "2026-08-10",
        "notes": "ATM Withdrawal",
    }
    expenses_store = {}

    mock_db = MagicMock()
    mock_db.cash_ledger.find_one = AsyncMock(return_value=cash_entry)

    async def mock_update_cash(q, update):
        gte_val = q.get("remaining_balance", {}).get("$gte", 0.0)
        if cash_entry["remaining_balance"] >= gte_val:
            inc_val = update.get("$inc", {}).get("remaining_balance", 0.0)
            cash_entry["remaining_balance"] += inc_val
            return MagicMock(modified_count=1)
        return MagicMock(modified_count=0)
    mock_db.cash_ledger.update_one = AsyncMock(side_effect=mock_update_cash)

    async def mock_insert_exp(doc):
        eid = ObjectId()
        doc["_id"] = eid
        expenses_store[str(eid)] = doc
        return MagicMock(inserted_id=eid)
    mock_db.expenses.insert_one = AsyncMock(side_effect=mock_insert_exp)

    req = MagicMock()
    req.app.mongodb = mock_db
    monkeypatch.setattr("routes.expenses._get_user", AsyncMock(return_value={"email": "admin@ssk.com", "role": "admin"}))

    payload = ExpenseIn(
        category="Transport & Logistics",
        amount=2400.0,
        date="2026-08-11",
        payee="Agra Transport Co",
        notes="Carton shipping freight",
        paid_via="cash",
        cash_ledger_id=str(cash_id),
    )

    res = await create_expense(payload, req)
    assert res["amount"] == 2400.0
    assert res["paid_via"] == "cash"
    assert res["cash_ledger_id"] == str(cash_id)
    assert res["bank_account_id"] is None

    # Check cash ledger decremented from 10000 to 7600
    assert cash_entry["remaining_balance"] == 7600.0


@pytest.mark.anyio
async def test_create_cash_expense_overdraft_rejected(monkeypatch):
    """Verify attempting to pay more cash than available in the pool is rejected."""
    from routes.expenses import create_expense

    cash_id = ObjectId()
    cash_entry = {
        "_id": cash_id,
        "amount": 5000.0,
        "remaining_balance": 1500.0,  # Only 1500 remaining
        "date": "2026-08-10",
    }

    mock_db = MagicMock()
    async def mock_update_cash(q, update):
        gte_val = q.get("remaining_balance", {}).get("$gte", 0.0)
        if cash_entry["remaining_balance"] >= gte_val:
            inc_val = update.get("$inc", {}).get("remaining_balance", 0.0)
            cash_entry["remaining_balance"] += inc_val
            return MagicMock(modified_count=1)
        return MagicMock(modified_count=0)
    mock_db.cash_ledger.update_one = AsyncMock(side_effect=mock_update_cash)
    mock_db.cash_ledger.find_one = AsyncMock(return_value=cash_entry)

    req = MagicMock()
    req.app.mongodb = mock_db
    monkeypatch.setattr("routes.expenses._get_user", AsyncMock(return_value={"email": "admin@ssk.com", "role": "admin"}))

    payload = ExpenseIn(
        category="Rent & Utilities",
        amount=3000.0,  # Exceeds 1500
        date="2026-08-11",
        payee="Factory Landlord",
        paid_via="cash",
        cash_ledger_id=str(cash_id),
    )

    with pytest.raises(HTTPException) as exc:
        await create_expense(payload, req)
    assert exc.value.status_code == 400
    assert "Insufficient cash in ledger entry" in exc.value.detail


@pytest.mark.anyio
async def test_concurrent_cash_expense_requests_prevent_race_condition(monkeypatch):
    """Verify two concurrent cash-expense requests with balance only for one: exactly one succeeds, one rejected, balance >= 0."""
    import asyncio
    from routes.expenses import create_expense

    cash_id = ObjectId()
    cash_entry = {
        "_id": cash_id,
        "amount": 5000.0,
        "remaining_balance": 5000.0,
        "date": "2026-08-10",
        "notes": "Cash pool",
    }
    expenses_store = {}
    db_lock = asyncio.Lock()

    mock_db = MagicMock()

    async def mock_update_cash(q, update):
        # Simulate MongoDB atomic conditional update with a lock
        async with db_lock:
            gte_val = q.get("remaining_balance", {}).get("$gte", 0.0)
            if cash_entry["remaining_balance"] >= gte_val:
                inc_val = update.get("$inc", {}).get("remaining_balance", 0.0)
                cash_entry["remaining_balance"] += inc_val
                return MagicMock(modified_count=1)
            return MagicMock(modified_count=0)

    async def mock_find_one(q):
        async with db_lock:
            return dict(cash_entry)

    async def mock_insert_exp(doc):
        eid = ObjectId()
        doc["_id"] = eid
        expenses_store[str(eid)] = doc
        return MagicMock(inserted_id=eid)

    mock_db.cash_ledger.update_one = AsyncMock(side_effect=mock_update_cash)
    mock_db.cash_ledger.find_one = AsyncMock(side_effect=mock_find_one)
    mock_db.expenses.insert_one = AsyncMock(side_effect=mock_insert_exp)

    req = MagicMock()
    req.app.mongodb = mock_db
    monkeypatch.setattr("routes.expenses._get_user", AsyncMock(return_value={"email": "admin@ssk.com", "role": "admin"}))

    # Two concurrent requests for 4000.0 each when pool only has 5000.0
    payload1 = ExpenseIn(
        category="Raw Materials",
        amount=4000.0,
        date="2026-08-11",
        payee="Supplier A",
        paid_via="cash",
        cash_ledger_id=str(cash_id),
    )
    payload2 = ExpenseIn(
        category="Raw Materials",
        amount=4000.0,
        date="2026-08-11",
        payee="Supplier B",
        paid_via="cash",
        cash_ledger_id=str(cash_id),
    )

    results = await asyncio.gather(
        create_expense(payload1, req),
        create_expense(payload2, req),
        return_exceptions=True,
    )

    successes = [r for r in results if isinstance(r, dict) and r.get("paid_via") == "cash"]
    failures = [r for r in results if isinstance(r, HTTPException)]

    # Exactly one must succeed and exactly one must fail
    assert len(successes) == 1
    assert len(failures) == 1
    assert failures[0].status_code == 400
    assert "Insufficient cash in ledger entry" in failures[0].detail

    # Remaining balance must be 5000 - 4000 = 1000 (never negative)
    assert cash_entry["remaining_balance"] == 1000.0
    assert len(expenses_store) == 1


@pytest.mark.anyio
async def test_cash_withdrawal_combined_wages_and_expenses_audit_trail(monkeypatch):
    """Verify cash withdrawal audit trail combines both wage payments and general cash expenses."""
    cash_ledger_id = ObjectId()
    cash_doc = {
        "_id": cash_ledger_id,
        "bank_account_id": "acc_1",
        "amount": 20000.0,
        "remaining_balance": 8000.0,  # 20000 - (7000 wages + 5000 expenses) = 8000
        "date": "2026-08-01",
        "notes": "Cash withdrawal for factory floor",
    }

    wages = [
        {
            "_id": ObjectId(),
            "worker_id": "w_1",
            "worker_name": "Ramesh Karigar",
            "amount": 4000.0,
            "date": "2026-08-05",
            "paid_via": "cash",
            "cash_ledger_id": str(cash_ledger_id),
            "period_from": "2026-08-01",
            "period_to": "2026-08-15",
            "notes": "Cutting wages",
        },
        {
            "_id": ObjectId(),
            "worker_id": "w_2",
            "worker_name": "Suresh Karigar",
            "amount": 3000.0,
            "date": "2026-08-06",
            "paid_via": "cash",
            "cash_ledger_id": str(cash_ledger_id),
            "period_from": "2026-08-01",
            "period_to": "2026-08-15",
            "notes": "Lasting wages",
        },
    ]

    expenses = [
        {
            "_id": ObjectId(),
            "category": "Raw Materials",
            "amount": 3500.0,
            "date": "2026-08-07",
            "payee": "Local Thread Supplier",
            "paid_via": "cash",
            "cash_ledger_id": str(cash_ledger_id),
            "notes": "Cash buy for urgent thread stock",
        },
        {
            "_id": ObjectId(),
            "category": "Transport & Logistics",
            "amount": 1500.0,
            "date": "2026-08-08",
            "payee": "Tempo Driver Raju",
            "paid_via": "cash",
            "cash_ledger_id": str(cash_ledger_id),
            "notes": "Delivery freight charge",
        },
    ]

    mock_db = MagicMock()
    mock_db.cash_ledger.find_one = AsyncMock(return_value=cash_doc)
    mock_db.wage_payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=wages)))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=expenses)))

    req = MagicMock()
    monkeypatch.setattr(banking_routes, "_get_user", AsyncMock(return_value={"email": "admin@ssk.com", "role": "admin"}))
    monkeypatch.setattr(banking_routes, "_get_db", lambda r: mock_db)

    detail = await get_cash_ledger_detail(str(cash_ledger_id), req)
    assert detail["ok"] is True
    assert detail["withdrawal_amount"] == 20000.0
    assert detail["allocated_amount"] == 12000.0  # 7000 + 5000
    assert detail["remaining_balance"] == 8000.0
    assert detail["wage_payment_count"] == 2
    assert detail["expense_count"] == 2
    assert detail["disbursement_count"] == 4

    disb_types = [d["type"] for d in detail["disbursements"]]
    assert "wage_payment" in disb_types
    assert "expense" in disb_types

    disb_titles = [d["title"] for d in detail["disbursements"]]
    assert "Ramesh Karigar" in disb_titles
    assert "Local Thread Supplier" in disb_titles
    assert "Tempo Driver Raju" in disb_titles


@pytest.mark.anyio
async def test_statement_reimport_duplicate_skipping_and_reporting(monkeypatch):
    """Verify importing a statement, then re-importing the exact same file: 0 new, N skipped as duplicates."""
    from fastapi import UploadFile
    from io import BytesIO
    import routes.banking

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC Primary Current A/C",
        "bank_name": "HDFC",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Date",
                "narration": "Narration",
                "reference": "Chq./Ref.No.",
                "debit_amount": "Withdrawal Amt.",
                "credit_amount": "Deposit Amt.",
                "balance": "Closing Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    db_statement_lines = []

    async def mock_insert_many(docs):
        inserted_ids = []
        for d in docs:
            doc_copy = dict(d)
            oid_val = ObjectId()
            doc_copy["_id"] = oid_val
            db_statement_lines.append(doc_copy)
            inserted_ids.append(oid_val)
        return MagicMock(inserted_ids=inserted_ids)

    def mock_find(query):
        matched = [
            d for d in db_statement_lines
            if str(d.get("bank_account_id")) == str(query.get("bank_account_id"))
        ]
        if "date" in query and isinstance(query["date"], dict):
            gte = query["date"].get("$gte")
            lte = query["date"].get("$lte")
            if gte:
                matched = [d for d in matched if d.get("date", "") >= gte]
            if lte:
                matched = [d for d in matched if d.get("date", "") <= lte]
        return MagicMock(to_list=AsyncMock(return_value=matched))

    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)
    mock_db.bank_statement_lines.find = MagicMock(side_effect=mock_find)

    csv_content = (
        "Date,Narration,Chq./Ref.No.,Withdrawal Amt.,Deposit Amt.,Closing Balance\n"
        "15/08/2026,ACH CR-MYNTRA DESIGNS-SETTLEMENT,ACH001,,45250.50,145250.50\n"
        "16/08/2026,UPI-SWIGGY-EXPENSE,UPI999,350.00,,144900.50\n"
        "17/08/2026,NEFT DR-LEATHER SUPPLIER VENDOR,NEFT777,25000.00,,119900.50\n"
    ).encode("utf-8")

    req = MagicMock()

    # Pass 1: Initial import -> 3 new rows inserted, 0 skipped
    file1 = UploadFile(filename="hdfc_aug.csv", file=BytesIO(csv_content))
    res1 = await routes.banking.import_bank_statement(acc_id, req, file=file1, dry_run=False)
    assert res1["ok"] is True
    assert res1["inserted_count"] == 3
    assert res1["skipped_count"] == 0
    assert len(db_statement_lines) == 3

    # Pass 2: Re-import exact same file -> 0 new, 3 skipped as duplicates
    file2 = UploadFile(filename="hdfc_aug.csv", file=BytesIO(csv_content))
    res2 = await routes.banking.import_bank_statement(acc_id, req, file=file2, dry_run=False)
    assert res2["ok"] is True
    assert res2["inserted_count"] == 0
    assert res2["skipped_count"] == 3
    assert "0 new, 3 skipped as duplicates" in res2["message"]
    # Total rows in DB remains exactly 3, not 6
    assert len(db_statement_lines) == 3


@pytest.mark.anyio
async def test_statement_overlapping_import_partial_skipping(monkeypatch):
    """Verify importing a statement with partial overlap: overlapping rows skipped, genuinely new rows inserted."""
    from fastapi import UploadFile
    from io import BytesIO
    import routes.banking

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC Primary Current A/C",
        "bank_name": "HDFC",
        "statement_format": {
            "sheet_locator": {"type": "first_sheet"},
            "header_locator": {"type": "fixed_row", "row": 0},
            "skip_rows_after_header": 0,
            "column_map": {
                "date": "Date",
                "narration": "Narration",
                "reference": "Chq./Ref.No.",
                "debit_amount": "Withdrawal Amt.",
                "credit_amount": "Deposit Amt.",
                "balance": "Closing Balance",
            },
            "date_format": "%d/%m/%Y",
        }
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    db_statement_lines = []

    async def mock_insert_many(docs):
        inserted_ids = []
        for d in docs:
            doc_copy = dict(d)
            oid_val = ObjectId()
            doc_copy["_id"] = oid_val
            db_statement_lines.append(doc_copy)
            inserted_ids.append(oid_val)
        return MagicMock(inserted_ids=inserted_ids)

    def mock_find(query):
        matched = [
            d for d in db_statement_lines
            if str(d.get("bank_account_id")) == str(query.get("bank_account_id"))
        ]
        if "date" in query and isinstance(query["date"], dict):
            gte = query["date"].get("$gte")
            lte = query["date"].get("$lte")
            if gte:
                matched = [d for d in matched if d.get("date", "") >= gte]
            if lte:
                matched = [d for d in matched if d.get("date", "") <= lte]
        return MagicMock(to_list=AsyncMock(return_value=matched))

    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=mock_insert_many)
    mock_db.bank_statement_lines.find = MagicMock(side_effect=mock_find)

    # First export: Aug 15 to Aug 17 (3 rows)
    csv1 = (
        "Date,Narration,Chq./Ref.No.,Withdrawal Amt.,Deposit Amt.,Closing Balance\n"
        "15/08/2026,ACH CR-MYNTRA DESIGNS-SETTLEMENT,ACH001,,45250.50,145250.50\n"
        "16/08/2026,UPI-SWIGGY-EXPENSE,UPI999,350.00,,144900.50\n"
        "17/08/2026,NEFT DR-LEATHER SUPPLIER VENDOR,NEFT777,25000.00,,119900.50\n"
    ).encode("utf-8")

    # Second export: Aug 17 to Aug 19 (1 overlapping row on Aug 17, 2 new rows on Aug 18 and 19)
    csv2 = (
        "Date,Narration,Chq./Ref.No.,Withdrawal Amt.,Deposit Amt.,Closing Balance\n"
        "17/08/2026,NEFT DR-LEATHER SUPPLIER VENDOR,NEFT777,25000.00,,119900.50\n"
        "18/08/2026,ACH CR-AMAZON PAYOUT,ACH002,,62000.00,181900.50\n"
        "19/08/2026,CHQ DEP-OFFLINE BUYER,CHQ888,,15000.00,196900.50\n"
    ).encode("utf-8")

    req = MagicMock()

    # Import 1
    file1 = UploadFile(filename="export_part1.csv", file=BytesIO(csv1))
    res1 = await routes.banking.import_bank_statement(acc_id, req, file=file1, dry_run=False)
    assert res1["inserted_count"] == 3
    assert res1["skipped_count"] == 0

    # Dry run of Import 2 -> reports 2 new, 1 duplicate
    file2_preview = UploadFile(filename="export_part2.csv", file=BytesIO(csv2))
    preview_res = await routes.banking.import_bank_statement(acc_id, req, file=file2_preview, dry_run=True)
    assert preview_res["new_count"] == 2
    assert preview_res["skipped_count"] == 1
    assert "2 new, 1 skipped as duplicates" in preview_res["message"]

    # Import 2 commit -> inserts 2 new, skips 1 duplicate
    file2_commit = UploadFile(filename="export_part2.csv", file=BytesIO(csv2))
    res2 = await routes.banking.import_bank_statement(acc_id, req, file=file2_commit, dry_run=False)
    assert res2["inserted_count"] == 2
    assert res2["skipped_count"] == 1
    assert "2 new, 1 skipped as duplicates" in res2["message"]

    # Total distinct rows in DB is 3 + 2 = 5
    assert len(db_statement_lines) == 5


@pytest.mark.anyio
async def test_reclassify_cash_withdrawal_with_dependents_blocked(monkeypatch):
    """Verify attempting to reclassify/rematch a cash withdrawal line with dependent wage payments/expenses is blocked with a clear error."""
    import routes.banking
    from routes.banking import confirm_transfer_pair, match_statement_line
    from models.banking import TransferConfirmIn, BankStatementLineUpdate

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    from_line_id = ObjectId()
    to_line_id = ObjectId()
    cash_ledger_id = ObjectId()

    # from_line was confirmed as cash_withdrawal
    from_line_doc = {
        "_id": from_line_id,
        "bank_account_id": "acc_1",
        "date": "2026-08-10",
        "narration": "ATM CASH WITHDRAWAL",
        "debit_amount": 10000.0,
        "credit_amount": 0.0,
        "match_status": "matched",
        "matched_to": {"type": "cash_withdrawal", "ref_id": str(cash_ledger_id)},
    }
    to_line_doc = {
        "_id": to_line_id,
        "bank_account_id": "acc_2",
        "date": "2026-08-10",
        "narration": "TRANSFER DEPOSIT",
        "debit_amount": 0.0,
        "credit_amount": 10000.0,
        "match_status": "unmatched",
        "matched_to": None,
    }
    cash_ledger_doc = {
        "_id": cash_ledger_id,
        "source_statement_line_id": str(from_line_id),
        "amount": 10000.0,
        "remaining_balance": 3000.0,
    }
    # Has 2 dependent wage payments!
    dependent_wages = [
        {"_id": ObjectId(), "cash_ledger_id": str(cash_ledger_id), "amount": 4000.0, "worker_name": "Ramesh"},
        {"_id": ObjectId(), "cash_ledger_id": str(cash_ledger_id), "amount": 3000.0, "worker_name": "Suresh"},
    ]

    async def mock_find_one_line(q):
        qid = str(q.get("_id"))
        if qid == str(from_line_id):
            return dict(from_line_doc)
        if qid == str(to_line_id):
            return dict(to_line_doc)
        return None

    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=mock_find_one_line)
    mock_db.cash_ledger.find_one = AsyncMock(return_value=cash_ledger_doc)
    mock_db.wage_payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=dependent_wages)))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))

    req = MagicMock()

    # 1. Attempt to confirm as transfer pair -> must be blocked
    transfer_payload = TransferConfirmIn(
        from_line_id=str(from_line_id),
        to_line_id=str(to_line_id),
        notes="Reclassifying to transfer",
    )
    with pytest.raises(HTTPException) as exc_info:
        await confirm_transfer_pair(transfer_payload, req)
    assert exc_info.value.status_code == 400
    assert "active dependent records" in exc_info.value.detail
    assert "2 wage payment(s)" in exc_info.value.detail
    assert "reclassify as transfer" in exc_info.value.detail

    # 2. Attempt to rematch/unmatch directly -> must also be blocked
    rematch_payload = BankStatementLineUpdate(
        match_status="unmatched",
    )
    with pytest.raises(HTTPException) as exc_info2:
        await match_statement_line(str(from_line_id), rematch_payload, req)
    assert exc_info2.value.status_code == 400
    assert "active dependent records" in exc_info2.value.detail


@pytest.mark.anyio
async def test_rematch_line_without_dependents_succeeds(monkeypatch):
    """Verify a statement line with no dependents can be freely reclassified/rematched."""
    import routes.banking
    from routes.banking import confirm_transfer_pair, match_statement_line
    from models.banking import TransferConfirmIn, BankStatementLineUpdate

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    from_line_id = ObjectId()
    to_line_id = ObjectId()
    cash_ledger_id = ObjectId()

    # from_line was confirmed as cash_withdrawal, but NO wage payments or expenses were drawn from it
    from_line_doc = {
        "_id": from_line_id,
        "bank_account_id": "acc_1",
        "date": "2026-08-10",
        "narration": "ATM CASH WITHDRAWAL",
        "debit_amount": 5000.0,
        "credit_amount": 0.0,
        "match_status": "matched",
        "matched_to": {"type": "cash_withdrawal", "ref_id": str(cash_ledger_id)},
    }
    to_line_doc = {
        "_id": to_line_id,
        "bank_account_id": "acc_2",
        "date": "2026-08-10",
        "narration": "TRANSFER DEPOSIT",
        "debit_amount": 0.0,
        "credit_amount": 5000.0,
        "match_status": "unmatched",
        "matched_to": None,
    }
    cash_ledger_doc = {
        "_id": cash_ledger_id,
        "source_statement_line_id": str(from_line_id),
        "amount": 5000.0,
        "remaining_balance": 5000.0,
    }

    async def mock_find_one_line(q):
        qid = str(q.get("_id"))
        if qid == str(from_line_id):
            return dict(from_line_doc)
        if qid == str(to_line_id):
            return dict(to_line_doc)
        return None

    async def mock_find_one_cash(q):
        if str(q.get("_id")) == str(cash_ledger_id) or str(q.get("source_statement_line_id")) == str(from_line_id):
            return dict(cash_ledger_doc)
        return None

    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=mock_find_one_line)
    mock_db.bank_statement_lines.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    mock_db.cash_ledger.find_one = AsyncMock(side_effect=mock_find_one_cash)
    mock_db.cash_ledger.delete_one = AsyncMock(return_value=MagicMock(deleted_count=1))
    mock_db.wage_payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))

    req = MagicMock()

    # Confirm transfer pair -> succeeds because 0 dependents exist
    transfer_payload = TransferConfirmIn(
        from_line_id=str(from_line_id),
        to_line_id=str(to_line_id),
        notes="Reclassified to transfer",
    )
    res = await confirm_transfer_pair(transfer_payload, req)
    assert res["ok"] is True
    assert "Transfer pair successfully confirmed" in res["message"]
    # Cleaned up the empty unused cash ledger entry
    mock_db.cash_ledger.delete_one.assert_called_once_with({"_id": cash_ledger_id})


@pytest.mark.anyio
async def test_multi_concurrent_cash_expense_requests_high_contention(monkeypatch):
    """Verify N-way simultaneous concurrent cash-expense requests (5 coroutines for 1200 each against 3000 pool). Exactly 2 succeed, 3 rejected, balance is 600."""
    import asyncio
    from routes.expenses import create_expense

    cash_id = ObjectId()
    cash_entry = {
        "_id": cash_id,
        "amount": 3000.0,
        "remaining_balance": 3000.0,
        "date": "2026-08-10",
    }
    expenses_store = {}
    db_lock = asyncio.Lock()

    mock_db = MagicMock()

    async def mock_update_cash(q, update):
        # Simulate slight async context switch before lock to maximize interleaving
        await asyncio.sleep(0.001)
        async with db_lock:
            gte_val = q.get("remaining_balance", {}).get("$gte", 0.0)
            if cash_entry["remaining_balance"] >= gte_val:
                inc_val = update.get("$inc", {}).get("remaining_balance", 0.0)
                cash_entry["remaining_balance"] = round(cash_entry["remaining_balance"] + inc_val, 2)
                return MagicMock(modified_count=1)
            return MagicMock(modified_count=0)

    async def mock_find_one(q):
        async with db_lock:
            return dict(cash_entry)

    async def mock_insert_exp(doc):
        await asyncio.sleep(0.001)
        eid = ObjectId()
        doc["_id"] = eid
        expenses_store[str(eid)] = doc
        return MagicMock(inserted_id=eid)

    mock_db.cash_ledger.update_one = AsyncMock(side_effect=mock_update_cash)
    mock_db.cash_ledger.find_one = AsyncMock(side_effect=mock_find_one)
    mock_db.expenses.insert_one = AsyncMock(side_effect=mock_insert_exp)

    req = MagicMock()
    req.app.mongodb = mock_db
    monkeypatch.setattr("routes.expenses._get_user", AsyncMock(return_value={"email": "admin@ssk.com", "role": "admin"}))

    # 5 concurrent requests of 1200 each against 3000 pool
    payloads = [
        ExpenseIn(
            category="Raw Materials",
            amount=1200.0,
            date="2026-08-11",
            payee=f"Supplier {i}",
            paid_via="cash",
            cash_ledger_id=str(cash_id),
        )
        for i in range(5)
    ]

    results = await asyncio.gather(
        *(create_expense(p, req) for p in payloads),
        return_exceptions=True,
    )

    successes = [r for r in results if isinstance(r, dict) and r.get("paid_via") == "cash"]
    failures = [r for r in results if isinstance(r, HTTPException)]

    # Exactly 2 succeed (2400 <= 3000), and exactly 3 fail (600 < 1200)
    assert len(successes) == 2
    assert len(failures) == 3
    for f in failures:
        assert f.status_code == 400
        assert "Insufficient cash in ledger entry" in f.detail

    assert cash_entry["remaining_balance"] == 600.0
    assert len(expenses_store) == 2


@pytest.mark.anyio
async def test_bulk_confirm_matches_confidence_threshold_filtering(monkeypatch):
    """Verify bulk auto-reconcile confirms only matches >= min_confidence and leaves lower confidence matches for individual review."""
    import routes.banking
    from routes.banking import reconcile_bank_account

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    acc_doc = {
        "_id": ObjectId(acc_id),
        "name": "HDFC Current",
        "account_type": "b2b_client",
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)

    line1_id = ObjectId()
    line2_id = ObjectId()
    line3_id = ObjectId()

    # In-memory store of statement lines
    db_lines = {
        str(line1_id): {
            "_id": line1_id,
            "bank_account_id": acc_id,
            "date": "2026-08-10",
            "debit_amount": 5000.0,
            "credit_amount": 0.0,
            "narration": "NEFT DR RENT PAYMENT",
            "match_status": "unmatched",
        },
        str(line2_id): {
            "_id": line2_id,
            "bank_account_id": acc_id,
            "date": "2026-08-10",
            "debit_amount": 3000.0,
            "credit_amount": 0.0,
            "narration": "NEFT DR VENDOR XYZ",
            "match_status": "unmatched",
        },
        str(line3_id): {
            "_id": line3_id,
            "bank_account_id": acc_id,
            "date": "2026-08-10",
            "debit_amount": 1500.0,
            "credit_amount": 0.0,
            "narration": "UPI EXPENSE TEA",
            "match_status": "unmatched",
        },
    }

    # ERP records:
    # 1. Expense 1: Exact date ("2026-08-10") & exact amount (5000.0) -> Confidence >= 95%
    exp1_id = ObjectId()
    exp1 = {"_id": exp1_id, "amount": 5000.0, "date": "2026-08-10", "payee": "Landlord Rent", "bank_account_id": None}

    # 2. Vendor payment 2: 2 days offset ("2026-08-12") & slight amount diff (2999.50 vs 3000.0) -> Confidence ≈ 68%
    vp2_id = ObjectId()
    vp2 = {"_id": vp2_id, "type": "vendor_payment", "amount": 2999.50, "payment_date": "2026-08-12", "vendor_name": "Leather Corp", "bank_account_id": None}

    # 3. Expense 3: 1 day offset ("2026-08-11") & amount (1500.0) -> Confidence ≈ 85%
    exp3_id = ObjectId()
    exp3 = {"_id": exp3_id, "amount": 1500.0, "date": "2026-08-11", "payee": "Tea Pantry", "bank_account_id": None}

    def mock_find_lines(q):
        unmatched = [d for d in db_lines.values() if d["match_status"] == "unmatched"]
        return MagicMock(sort=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=unmatched))))

    async def mock_update_line(q, update):
        lid = str(q["_id"])
        if lid in db_lines:
            db_lines[lid].update(update.get("$set", {}))
        return MagicMock(matched_count=1)

    mock_db.bank_statement_lines.find = MagicMock(side_effect=mock_find_lines)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=mock_update_line)
    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[vp2])))
    mock_db.payments.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[exp1, exp3])))
    mock_db.expenses.update_one = AsyncMock(return_value=MagicMock(matched_count=1))

    req = MagicMock()

    # Pass 1: Run with default conservative threshold (95%+)
    res1 = await reconcile_bank_account(
        acc_id,
        req,
        date_window_days=3,
        amount_tolerance=1.0,
        min_confidence=0.95,
        dry_run=False,
    )
    assert res1["ok"] is True
    assert res1["total_unmatched_evaluated"] == 3
    # Exactly Line 1 (exact date+amount) has confidence >= 95% and is auto-matched
    assert res1["auto_matched_count"] == 1
    # Lines 2 and 3 remain for individual review
    assert res1["pending_review_count"] == 2
    assert db_lines[str(line1_id)]["match_status"] == "matched"
    assert db_lines[str(line2_id)]["match_status"] == "unmatched"
    assert db_lines[str(line3_id)]["match_status"] == "unmatched"

    # Pass 2: Run with 80%+ threshold
    res2 = await reconcile_bank_account(
        acc_id,
        req,
        date_window_days=3,
        amount_tolerance=1.0,
        min_confidence=0.80,
        dry_run=False,
    )
    assert res2["ok"] is True
    assert res2["total_unmatched_evaluated"] == 2
    # Line 3 (1-day offset, ~85% confidence) is now auto-matched
    assert res2["auto_matched_count"] == 1
    # Line 2 (2-day offset, ~75% confidence) still remains for review
    assert res2["pending_review_count"] == 1
    assert db_lines[str(line3_id)]["match_status"] == "matched"
    assert db_lines[str(line2_id)]["match_status"] == "unmatched"


@pytest.mark.anyio
async def test_period_lock_blocks_edits_and_admin_unlock_allows_edits(monkeypatch):
    """
    Verify:
    1. Lock a reconciliation period (2026-08-01 to 2026-08-31).
    2. Attempt to edit/unmatch/rematch a line within the period -> blocked with clear 400 error.
    3. Non-admin unlock attempt is rejected with 403.
    4. Admin unlock with audit reason succeeds.
    5. Edits/unmatching within the unlocked period are permitted again.
    """
    import routes.banking
    from routes.banking import (
        lock_reconciliation_period,
        unlock_reconciliation_period,
        match_statement_line,
        list_period_locks,
    )
    from models.banking import PeriodLockIn, PeriodUnlockIn, BankStatementLineUpdate

    mock_db = MagicMock()
    admin_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin User"}
    manager_user = {"role": "manager", "email": "manager@sskfootcare.com", "name": "Manager User"}
    current_user = [admin_user]

    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(side_effect=lambda r: current_user[0]))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc_id = str(ObjectId())
    line_id = ObjectId()
    line_doc = {
        "_id": line_id,
        "bank_account_id": acc_id,
        "date": "2026-08-15",
        "debit_amount": 5000.0,
        "credit_amount": 0.0,
        "narration": "OFFICE EXPENSE",
        "match_status": "matched",
        "matched_to": {"type": "expense", "ref_id": str(ObjectId())},
    }

    # In-memory collections
    locks_store = []
    lines_store = {str(line_id): dict(line_doc)}

    # Mock reconciliation_locks collection
    mock_locks_col = MagicMock()

    async def mock_lock_insert(doc):
        d = dict(doc)
        d["_id"] = ObjectId()
        locks_store.append(d)
        return MagicMock(inserted_id=d["_id"])

    async def mock_lock_find_one(q):
        for l in locks_store:
            if l.get("status") != q.get("status", l.get("status")):
                continue
            if "period_from" in q and isinstance(q["period_from"], dict) and "$lte" in q["period_from"]:
                target_date = q["period_from"]["$lte"]
                if not (l["period_from"] <= target_date <= l["period_to"]):
                    continue
            elif "period_from" in q and q["period_from"] != l["period_from"]:
                continue
            if "period_to" in q and not isinstance(q["period_to"], dict) and q["period_to"] != l["period_to"]:
                continue
            return l
        return None

    async def mock_lock_update(q, u):
        lid = q.get("_id")
        for l in locks_store:
            if l["_id"] == lid:
                if "$set" in u:
                    l.update(u["$set"])
                if "$push" in u:
                    for pk, pv in u["$push"].items():
                        l.setdefault(pk, []).append(pv)
                return MagicMock(matched_count=1)
        return MagicMock(matched_count=0)

    def mock_lock_find(q):
        return MagicMock(sort=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=locks_store))))

    mock_locks_col.insert_one = AsyncMock(side_effect=mock_lock_insert)
    mock_locks_col.find_one = AsyncMock(side_effect=mock_lock_find_one)
    mock_locks_col.update_one = AsyncMock(side_effect=mock_lock_update)
    mock_locks_col.find = MagicMock(side_effect=mock_lock_find)
    mock_db.reconciliation_locks = mock_locks_col

    # Mock statement lines collection
    async def mock_stmt_find_one(q):
        lid = str(q.get("_id"))
        return lines_store.get(lid)

    async def mock_stmt_update_one(q, u):
        lid = str(q.get("_id"))
        if lid in lines_store:
            lines_store[lid].update(u.get("$set", {}))
            return MagicMock(matched_count=1)
        return MagicMock(matched_count=0)

    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=mock_stmt_find_one)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=mock_stmt_update_one)
    mock_db.cash_ledger = MagicMock(find_one=AsyncMock(return_value=None))

    req = MagicMock()

    # Step 1: Lock the period 2026-08-01 to 2026-08-31
    lock_res = await lock_reconciliation_period(
        PeriodLockIn(
            bank_account_id=acc_id,
            period_from="2026-08-01",
            period_to="2026-08-31",
            reason="August 2026 finalized for GST filing",
        ),
        req,
    )
    assert lock_res["ok"] is True
    assert len(locks_store) == 1
    assert locks_store[0]["status"] == "locked"
    assert locks_store[0]["locked_by"] == "admin@sskfootcare.com"

    # Step 2: Attempt to unmatch/edit statement line on 2026-08-15 -> MUST be blocked
    with pytest.raises(HTTPException) as exc_info:
        await match_statement_line(
            id=str(line_id),
            payload=BankStatementLineUpdate(match_status="unmatched"),
            request=req,
        )
    assert exc_info.value.status_code == 400
    assert "Reconciliation period (2026-08-01 to 2026-08-31)" in exc_info.value.detail
    assert "is finalized and locked" in exc_info.value.detail
    assert lines_store[str(line_id)]["match_status"] == "matched"

    # Step 3: Non-admin attempt to unlock -> 403 Forbidden
    current_user[0] = manager_user
    with pytest.raises(HTTPException) as exc_info:
        await unlock_reconciliation_period(
            PeriodUnlockIn(
                bank_account_id=acc_id,
                period_from="2026-08-01",
                period_to="2026-08-31",
                reason="Manager adjustment",
            ),
            req,
        )
    assert exc_info.value.status_code == 403

    # Step 4: Admin unlocks the period with an audit reason
    current_user[0] = admin_user
    unlock_res = await unlock_reconciliation_period(
        PeriodUnlockIn(
            bank_account_id=acc_id,
            period_from="2026-08-01",
            period_to="2026-08-31",
            reason="Admin audit correction for invoice mismatch",
        ),
        req,
    )
    assert unlock_res["ok"] is True
    assert locks_store[0]["status"] == "unlocked"
    assert locks_store[0]["unlocked_by"] == "admin@sskfootcare.com"
    assert len(locks_store[0]["history"]) == 2

    # Step 5: Now editing / unmatching the statement line SUCCEEDS
    edit_res = await match_statement_line(
        id=str(line_id),
        payload=BankStatementLineUpdate(match_status="unmatched"),
        request=req,
    )
    assert edit_res["match_status"] == "unmatched"
    assert lines_store[str(line_id)]["match_status"] == "unmatched"


@pytest.mark.anyio
async def test_export_reconciliation_report_excel_generation_and_content(monkeypatch):
    """
    Verify the Month-End Bank Reconciliation Excel Export generates a professional,
    CA/Accountant-grade workbook with Executive Summary, Balance Proof, Grouped
    Categories (Revenue, Expenses, Vendor, Transfers, Cash Withdrawals + Karigar Wages),
    and Pending/Unmatched items.
    """
    import io
    import openpyxl
    import routes.banking
    from routes.banking import export_reconciliation_report

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "ca.auditor@sskfootcare.com", "name": "Chief Auditor"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)

    acc1_id = ObjectId()
    acc2_id = ObjectId()
    acc_doc = {
        "_id": acc1_id,
        "name": "HDFC Current Account",
        "bank_name": "HDFC Bank",
        "account_number_last4": "5678",
        "account_number": "50200012345678",
        "ifsc": "HDFC0001234",
        "branch": "Agra Civil Lines",
        "opening_balance": 100000.0,
    }
    acc2_doc = {
        "_id": acc2_id,
        "name": "UCO Bank - Factory",
        "bank_name": "UCO Bank",
        "account_number_last4": "9999",
        "opening_balance": 25000.0,
    }
    mock_db.bank_accounts.find_one = AsyncMock(return_value=acc_doc)
    mock_db.bank_accounts.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[acc_doc, acc2_doc])))

    # 1. Statement lines with mixed transaction types
    line_rev_id = ObjectId()
    line_exp_id = ObjectId()
    line_vp_id = ObjectId()
    line_trf_id = ObjectId()
    line_cash_id = ObjectId()
    line_unmatched_id = ObjectId()

    client_pay_id = ObjectId()
    expense_id = ObjectId()
    vendor_pay_id = ObjectId()
    cash_ledger_id = ObjectId()

    statement_lines = [
        # Revenue / Client receipt (Credit ₹75,000)
        {
            "_id": line_rev_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-05",
            "credit_amount": 75000.0,
            "debit_amount": 0.0,
            "narration": "NEFT CR - BAXTER RETAIL CLIENT PAYMENT",
            "reference_no": "UTR12345678",
            "match_status": "matched",
            "matched_to": {"type": "payment", "ref_id": str(client_pay_id)},
        },
        # Direct Operating Expense (Debit ₹15,000)
        {
            "_id": line_exp_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-10",
            "credit_amount": 0.0,
            "debit_amount": 15000.0,
            "narration": "NEFT DR - FACTORY ELECTRICITY UPPCL",
            "reference_no": "BILL9988",
            "match_status": "matched",
            "matched_to": {"type": "expense", "ref_id": str(expense_id)},
        },
        # Vendor Raw Material Payment (Debit ₹30,000)
        {
            "_id": line_vp_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-14",
            "credit_amount": 0.0,
            "debit_amount": 30000.0,
            "narration": "RTGS DR - AGRA LEATHER SUPPLIERS",
            "reference_no": "UTR887766",
            "match_status": "matched",
            "matched_to": {"type": "vendor_payment", "ref_id": str(vendor_pay_id)},
        },
        # Cash Withdrawal for Karigar Wages (Debit ₹20,000)
        {
            "_id": line_cash_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-18",
            "credit_amount": 0.0,
            "debit_amount": 20000.0,
            "narration": "ATM CASH WDL - SELF FOR KARIGAR PAYOUTS",
            "reference_no": "ATM5544",
            "match_status": "matched",
            "matched_to": {"type": "cash_withdrawal", "ref_id": str(cash_ledger_id)},
        },
        # Inter-Account Transfer to UCO Bank (Debit ₹10,000)
        {
            "_id": line_trf_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-22",
            "credit_amount": 0.0,
            "debit_amount": 10000.0,
            "narration": "TRF DR - LIQUIDITY TRANSFER TO UCO FACTORY",
            "reference_no": "TRF1122",
            "match_status": "transfer",
            "matched_to": {"type": "transfer", "ref_id": str(ObjectId())},
            "transfer_notes": "Treasury liquidity rebalance",
            "confirmed_by": "admin@sskfootcare.com",
        },
        # Unmatched Statement Debit (Debit ₹2,500 pending)
        {
            "_id": line_unmatched_id,
            "bank_account_id": str(acc1_id),
            "date": "2026-08-25",
            "credit_amount": 0.0,
            "debit_amount": 2500.0,
            "narration": "BANK CHARGES / ANNUAL SMS FEE",
            "reference_no": "CHG4433",
            "match_status": "unmatched",
            "remarks": "Bank fees to be booked under Bank Charges expense",
        },
    ]

    mock_db.bank_statement_lines.find = MagicMock(
        return_value=MagicMock(sort=MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=statement_lines))))
    )
    mock_db.bank_statement_lines.find_one = AsyncMock(return_value=None)

    # 2. Linked ERP Records
    client_pay_doc = {
        "_id": client_pay_id,
        "client_name": "Baxter Retail Pvt Ltd",
        "payment_no": "PAY-2026-088",
        "amount": 75000.0,
    }
    expense_doc = {
        "_id": expense_id,
        "payee": "UP Power Corporation",
        "category": "Rent & Utilities",
        "amount": 15000.0,
    }
    vendor_pay_doc = {
        "_id": vendor_pay_id,
        "vendor_name": "Agra Leather Tanneries",
        "payment_no": "VP-2026-044",
        "amount": 30000.0,
    }
    cash_ledger_doc = {
        "_id": cash_ledger_id,
        "source_statement_line_id": str(line_cash_id),
        "amount": 20000.0,
        "remaining_balance": 5000.0,
    }
    wage_payments = [
        {"_id": ObjectId(), "cash_ledger_id": str(cash_ledger_id), "worker_name": "Ramesh Kumar", "amount": 8000.0, "paid_via": "cash"},
        {"_id": ObjectId(), "cash_ledger_id": str(cash_ledger_id), "worker_name": "Suresh Chand", "amount": 7000.0, "paid_via": "cash"},
    ]

    mock_db.payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[client_pay_doc, vendor_pay_doc])))
    mock_db.expenses.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[expense_doc])))
    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.cash_ledger.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[cash_ledger_doc])))
    mock_db.wage_payments.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=wage_payments)))

    # 3. Active period lock document
    lock_doc = {
        "_id": ObjectId(),
        "status": "locked",
        "bank_account_id": str(acc1_id),
        "period_from": "2026-08-01",
        "period_to": "2026-08-31",
        "locked_at": "2026-08-31T23:59:59Z",
        "locked_by": "admin@sskfootcare.com",
    }
    mock_db.reconciliation_locks = MagicMock(find_one=AsyncMock(return_value=lock_doc))

    req = MagicMock()

    # Step 1: Call export endpoint
    response = await export_reconciliation_report(
        request=req,
        bank_account_id=str(acc1_id),
        from_date="2026-08-01",
        to_date="2026-08-31",
        format="excel",
    )

    assert response.status_code == 200
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=" in response.headers["Content-Disposition"]
    assert "HDFC_Current_Account" in response.headers["Content-Disposition"]

    # Step 2: Load and verify the Excel workbook with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(response.body))
    sheet_names = wb.sheetnames

    # All 7 accountant tabs must exist
    assert "Executive Summary" in sheet_names
    assert "1. Revenue & Receipts" in sheet_names
    assert "2. Operating Expenses" in sheet_names
    assert "3. Vendor Payments" in sheet_names
    assert "4. Cash & Karigar Wages" in sheet_names
    assert "5. Inter-Account Transfers" in sheet_names
    assert "6. Pending & Unmatched" in sheet_names

    # Inspect Tab 1: Executive Summary
    ws_summary = wb["Executive Summary"]
    assert ws_summary["A1"].value == "SSK FOOTCARE - BANK RECONCILIATION STATEMENT"
    assert "HDFC Current Account" in ws_summary["B4"].value
    assert "FINALIZED & LOCKED" in ws_summary["D5"].value

    # Check that amounts were correctly totaled on summary sheet
    summary_cells = [cell.value for row in ws_summary.iter_rows() for cell in row if cell.value is not None]
    assert any("Bank Statement Opening Balance" in str(v) for v in summary_cells)
    assert any("Matched Inflows" in str(v) for v in summary_cells)
    assert any("Matched Outflows" in str(v) for v in summary_cells)
    assert any("Reconciliation Variance" in str(v) for v in summary_cells)

    # Inspect Tab 2: Revenue
    ws_rev = wb["1. Revenue & Receipts"]
    assert ws_rev["D4"].value == "Baxter Retail Pvt Ltd"
    assert ws_rev["F4"].value == 75000.0

    # Inspect Tab 3: Operating Expenses
    ws_exp = wb["2. Operating Expenses"]
    assert ws_exp["D4"].value == "UP Power Corporation"
    assert ws_exp["E4"].value == "Rent & Utilities"
    assert ws_exp["F4"].value == 15000.0

    # Inspect Tab 4: Vendor Payments
    ws_vp = wb["3. Vendor Payments"]
    assert ws_vp["D4"].value == "Agra Leather Tanneries"
    assert ws_vp["F4"].value == 30000.0

    # Inspect Tab 5: Cash Withdrawals & Karigar Wages
    ws_cash = wb["4. Cash & Karigar Wages"]
    assert ws_cash["C4"].value == 20000.0  # Total withdrawal
    assert ws_cash["D4"].value == 15000.0  # Disbursed to karigars (8k + 7k)
    assert ws_cash["E4"].value == 5000.0   # Remaining unallocated cash
    assert "Ramesh Kumar" in ws_cash["G4"].value
    assert "Suresh Chand" in ws_cash["G4"].value

    # Inspect Tab 6: Transfers
    ws_trf = wb["5. Inter-Account Transfers"]
    assert ws_trf["E4"].value == 10000.0

    # Inspect Tab 7: Pending & Unmatched
    ws_unmatched = wb["6. Pending & Unmatched"]
    assert ws_unmatched["C4"].value == "BANK CHARGES / ANNUAL SMS FEE"
    assert ws_unmatched["E4"].value == 2500.0
    assert ws_unmatched["H4"].value == "Bank fees to be booked under Bank Charges expense"


@pytest.mark.anyio
async def test_total_cash_in_hand_live_aggregation_across_withdrawals_and_disbursements(monkeypatch):
    """
    Verify that Total Cash in Hand accurately sums all remaining_balance values across
    active cash_ledger pools, updates live when cash is withdrawn, and updates as
    funds are drawn down for wages/expenses.
    """
    import server
    import routes.banking
    from routes.banking import get_reconciliation_summary

    mock_db = MagicMock()
    mock_user = {"role": "admin", "email": "admin@sskfootcare.com", "name": "Admin"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=mock_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)
    monkeypatch.setattr(server, "db", mock_db)

    # 1. Initially two cash pools exist (e.g. ₹20,000 withdrawn, ₹15,000 remaining; ₹10,000 withdrawn, ₹10,000 remaining)
    pool1_id = ObjectId()
    pool2_id = ObjectId()
    cash_ledger_docs = [
        {"_id": pool1_id, "amount": 20000.0, "remaining_balance": 15000.0},
        {"_id": pool2_id, "amount": 10000.0, "remaining_balance": 10000.0},
    ]

    mock_db.cash_ledger.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=cash_ledger_docs)))
    mock_db.bank_accounts.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.bank_statement_lines.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.online_settlements.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.jobs.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.production_jobs.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.pos.find = MagicMock(return_value=MagicMock(to_list=AsyncMock(return_value=[])))
    mock_db.materials.count_documents = AsyncMock(return_value=5)
    mock_db.styles.count_documents = AsyncMock(return_value=8)
    mock_db.pos.count_documents = AsyncMock(return_value=0)
    mock_db.jobs.count_documents = AsyncMock(return_value=0)

    req = MagicMock()

    # Step 1: Banking summary returns Total Cash in Hand = 15,000 + 10,000 = 25,000
    res_summary = await get_reconciliation_summary(req)
    assert res_summary["summary"]["total_cash_in_hand"] == 25000.0
    assert res_summary["summary"]["total_cash_withdrawn"] == 30000.0

    # Step 2: Dashboard stats live computation also reflects Total Cash in Hand = 25,000
    dash_stats = await server._compute_dashboard_stats_live()
    assert dash_stats["total_cash_in_hand"] == 25000.0

    # Step 3: Draw down cash: Pool 1 spends ₹5,000 (remaining: ₹10,000), Pool 2 spends ₹8,000 (remaining: ₹2,000)
    cash_ledger_docs[0]["remaining_balance"] = 10000.0
    cash_ledger_docs[1]["remaining_balance"] = 2000.0

    # Step 4: Verify both banking summary and dashboard stats immediately reflect the updated ₹12,000 total cash in hand
    res_summary_updated = await get_reconciliation_summary(req)
    assert res_summary_updated["summary"]["total_cash_in_hand"] == 12000.0

    dash_stats_updated = await server._compute_dashboard_stats_live()
    assert dash_stats_updated["total_cash_in_hand"] == 12000.0


@pytest.mark.anyio
async def test_full_end_to_end_bank_reconciliation_workflow(monkeypatch):
    """
    Comprehensive End-to-End Walkthrough of the full banking & reconciliation journey:
    1. Create Bank Account with opening balance.
    2. Import CSV Statement with deduplication safeguards.
    3. Auto-reconcile & match client revenue and expenses.
    4. Confirm Cash Withdrawal, creating an active cash_ledger pool.
    5. Disburse Karigar Wage Payment & Cash Expense atomically from the cash pool.
    6. Verify Total Cash in Hand updates in real time.
    7. Verify Statement Line Reclassification Guard prevents orphaning active dependents.
    8. Finalize and Lock Reconciliation Period (blocking subsequent edits).
    9. Export 7-tab Accountant / CA Audit Excel Workbook and verify balance proof.
    10. Unlock period with admin audit reason.
    """
    import io
    import openpyxl
    from fastapi import UploadFile
    import server
    import routes.banking
    from routes.banking import (
        create_bank_account,
        import_bank_statement,
        match_statement_line,
        confirm_cash_withdrawal,
        confirm_transfer_pair,
        lock_reconciliation_period,
        unlock_reconciliation_period,
        export_reconciliation_report,
        get_reconciliation_summary,
    )
    import routes.workers
    import routes.expenses
    from routes.workers import create_wage_payment
    from routes.expenses import create_expense
    from models.workers import WagePaymentIn
    from models.expenses import ExpenseIn
    from models.banking import CashWithdrawalConfirmIn, PeriodLockIn, PeriodUnlockIn

    # In-memory realistic collections
    accounts_store = {}
    lines_store = {}
    cash_ledger_store = {}
    wage_payments_store = {}
    expenses_store = {}
    payments_store = {}
    locks_store = []

    mock_db = MagicMock()

    class MockCursor:
        def __init__(self, docs):
            self.docs = list(docs) if docs else []

        def sort(self, *args, **kwargs):
            return self

        def skip(self, *args, **kwargs):
            return self

        def limit(self, *args, **kwargs):
            return self

        async def to_list(self, length=None):
            return list(self.docs)

    # Generic mongo find/update mocks backed by our in-memory stores
    def _mock_find_accounts(query=None):
        docs = list(accounts_store.values())
        return MockCursor(docs)

    def _mock_find_lines(query=None):
        docs = list(lines_store.values())
        if query:
            if "bank_account_id" in query:
                docs = [d for d in docs if str(d.get("bank_account_id")) == str(query["bank_account_id"])]
        return MockCursor(docs)

    def _mock_find_cash(query=None):
        docs = list(cash_ledger_store.values())
        if query and "bank_account_id" in query:
            docs = [d for d in docs if str(d.get("bank_account_id")) == str(query["bank_account_id"])]
        return MockCursor(docs)

    def _mock_find_wage_payments(query=None):
        docs = list(wage_payments_store.values())
        if query:
            if "cash_ledger_id" in query:
                docs = [d for d in docs if str(d.get("cash_ledger_id")) == str(query["cash_ledger_id"])]
            if "worker_id" in query:
                docs = [d for d in docs if str(d.get("worker_id")) == str(query["worker_id"])]
        return MockCursor(docs)

    def _mock_find_expenses(query=None):
        docs = list(expenses_store.values())
        if query and "cash_ledger_id" in query:
            docs = [d for d in docs if str(d.get("cash_ledger_id")) == str(query["cash_ledger_id"])]
        return MockCursor(docs)

    mock_db.bank_accounts.find = MagicMock(side_effect=_mock_find_accounts)
    mock_db.bank_statement_lines.find = MagicMock(side_effect=_mock_find_lines)
    mock_db.cash_ledger.find = MagicMock(side_effect=_mock_find_cash)
    mock_db.wage_payments.find = MagicMock(side_effect=_mock_find_wage_payments)
    mock_db.expenses.find = MagicMock(side_effect=_mock_find_expenses)
    mock_db.payments.find = MagicMock(return_value=MockCursor(list(payments_store.values())))
    mock_db.online_settlements.find = MagicMock(return_value=MockCursor([]))
    mock_db.jobs.find = MagicMock(return_value=MockCursor([]))
    mock_db.pos.find = MagicMock(return_value=MockCursor([]))
    mock_db.materials.count_documents = AsyncMock(return_value=10)
    mock_db.styles.count_documents = AsyncMock(return_value=12)
    mock_db.pos.count_documents = AsyncMock(return_value=0)
    mock_db.jobs.count_documents = AsyncMock(return_value=0)
    mock_db.production_jobs.find = MagicMock(return_value=MockCursor([]))
    mock_db.styles.find = MagicMock(return_value=MockCursor([]))

    async def _mock_find_one_account(query):
        return accounts_store.get(str(query.get("_id")))

    async def _mock_find_one_line(query):
        return lines_store.get(str(query.get("_id")))

    async def _mock_find_one_cash(query):
        return cash_ledger_store.get(str(query.get("_id")))

    async def _mock_insert_account(doc):
        aid = ObjectId()
        doc["_id"] = aid
        accounts_store[str(aid)] = doc
        return MagicMock(inserted_id=aid)

    async def _mock_insert_lines(docs):
        inserted_ids = []
        for d in docs:
            lid = ObjectId()
            d["_id"] = lid
            lines_store[str(lid)] = d
            inserted_ids.append(lid)
        return MagicMock(inserted_ids=inserted_ids)

    async def _mock_insert_cash(doc):
        cid = ObjectId()
        doc["_id"] = cid
        cash_ledger_store[str(cid)] = doc
        return MagicMock(inserted_id=cid)

    async def _mock_insert_wage_payment(doc):
        wpid = ObjectId()
        doc["_id"] = wpid
        wage_payments_store[str(wpid)] = doc
        return MagicMock(inserted_id=wpid)

    async def _mock_insert_expense(doc):
        eid = ObjectId()
        doc["_id"] = eid
        expenses_store[str(eid)] = doc
        return MagicMock(inserted_id=eid)

    async def _mock_update_line(query, update):
        lid_str = str(query.get("_id"))
        if lid_str in lines_store:
            if "$set" in update:
                lines_store[lid_str].update(update["$set"])
            return MagicMock(matched_count=1, modified_count=1)
        return MagicMock(matched_count=0, modified_count=0)

    async def _mock_update_cash(query, update):
        cid_str = str(query.get("_id"))
        if cid_str in cash_ledger_store:
            entry = cash_ledger_store[cid_str]
            # Conditional $gte check
            if "$gte" in query.get("remaining_balance", {}):
                req_bal = query["remaining_balance"]["$gte"]
                if entry.get("remaining_balance", 0.0) < req_bal:
                    return MagicMock(matched_count=0, modified_count=0)
            if "$inc" in update:
                inc_val = update["$inc"].get("remaining_balance", 0.0)
                entry["remaining_balance"] = round(entry.get("remaining_balance", 0.0) + inc_val, 2)
            if "$set" in update:
                entry.update(update["$set"])
            return MagicMock(matched_count=1, modified_count=1)
        return MagicMock(matched_count=0, modified_count=0)

    async def _mock_find_one_lock(query):
        for l in locks_store:
            if l.get("status") == "locked":
                # Check date overlap
                if query.get("period_to", {}).get("$gte") and l.get("period_to") < query["period_to"]["$gte"]:
                    continue
                if query.get("period_from", {}).get("$lte") and l.get("period_from") > query["period_from"]["$lte"]:
                    continue
                return l
        return None

    async def _mock_update_lock(query, update, upsert=False):
        for l in locks_store:
            matches_id = ("_id" in query and (str(l.get("_id")) == str(query["_id"]) or l.get("_id") == query["_id"]))
            matches_key = (bool(query.get("bank_account_id")) and l.get("bank_account_id") == query.get("bank_account_id") and l.get("period_from") == query.get("period_from"))
            if matches_id or matches_key:
                if "$set" in update:
                    l.update(update["$set"])
                if "$push" in update:
                    l.setdefault("history", []).append(update["$push"]["history"])
                return MagicMock(matched_count=1, modified_count=1)
        if upsert:
            new_lock = dict(query)
            if "$set" in update:
                new_lock.update(update["$set"])
            if "$setOnInsert" in update:
                new_lock.update(update["$setOnInsert"])
            new_lock["_id"] = ObjectId()
            locks_store.append(new_lock)
            return MagicMock(matched_count=0, upserted_id=new_lock["_id"])
        return MagicMock(matched_count=0, modified_count=0)

    async def _mock_find_one_wage_payment(query):
        return wage_payments_store.get(str(query.get("_id")))

    async def _mock_find_one_expense(query):
        return expenses_store.get(str(query.get("_id")))

    async def _mock_insert_lock(doc):
        lid = ObjectId()
        doc["_id"] = lid
        locks_store.append(doc)
        return MagicMock(inserted_id=lid)

    mock_db.bank_accounts.find_one = AsyncMock(side_effect=_mock_find_one_account)
    mock_db.bank_accounts.insert_one = AsyncMock(side_effect=_mock_insert_account)
    mock_db.bank_statement_lines.find_one = AsyncMock(side_effect=_mock_find_one_line)
    mock_db.bank_statement_lines.insert_many = AsyncMock(side_effect=_mock_insert_lines)
    mock_db.bank_statement_lines.update_one = AsyncMock(side_effect=_mock_update_line)
    mock_db.cash_ledger.find_one = AsyncMock(side_effect=_mock_find_one_cash)
    mock_db.cash_ledger.insert_one = AsyncMock(side_effect=_mock_insert_cash)
    mock_db.cash_ledger.update_one = AsyncMock(side_effect=_mock_update_cash)
    mock_db.wage_payments.insert_one = AsyncMock(side_effect=_mock_insert_wage_payment)
    mock_db.wage_payments.find_one = AsyncMock(side_effect=_mock_find_one_wage_payment)
    mock_db.expenses.insert_one = AsyncMock(side_effect=_mock_insert_expense)
    mock_db.expenses.find_one = AsyncMock(side_effect=_mock_find_one_expense)
    mock_db.audit_logs.insert_one = AsyncMock(return_value=MagicMock())
    mock_db.reconciliation_locks.find_one = AsyncMock(side_effect=_mock_find_one_lock)
    mock_db.reconciliation_locks.insert_one = AsyncMock(side_effect=_mock_insert_lock)
    mock_db.reconciliation_locks.update_one = AsyncMock(side_effect=_mock_update_lock)

    # Mock workers
    worker_id = ObjectId()
    worker_doc = {
        "_id": worker_id,
        "name": "Mukesh Karigar",
        "phone": "9876543210",
        "active": True,
        "rate_per_pair": 20.0,
    }
    mock_db.workers.find_one = AsyncMock(return_value=worker_doc)
    mock_db.workers.find = MagicMock(return_value=MockCursor([worker_doc]))
    mock_db.advances.find = MagicMock(return_value=MockCursor([]))

    admin_user = {"role": "admin", "email": "founder@sskfootcare.com", "name": "SSK Founder"}
    monkeypatch.setattr(routes.banking, "_get_user", AsyncMock(return_value=admin_user))
    monkeypatch.setattr(routes.banking, "_get_db", lambda r: mock_db)
    monkeypatch.setattr(routes.workers, "_get_user", AsyncMock(return_value=admin_user))
    monkeypatch.setattr(routes.expenses, "_get_user", AsyncMock(return_value=admin_user))
    monkeypatch.setattr(server, "db", mock_db)

    req = MagicMock()
    req.app.mongodb = mock_db
    req.state.user = admin_user

    # -------------------------------------------------------------------------
    # STEP 1: Create Bank Account
    # -------------------------------------------------------------------------
    acc_res = await create_bank_account(
        BankAccountIn(
            name="HDFC Factory Primary",
            bank_name="HDFC Bank",
            account_number_last4="4321",
            account_number="50200098764321",
            ifsc="HDFC0004321",
            branch="Sanjay Place Agra",
            account_type="b2b_client",
            opening_balance=50000.0,
        ),
        req,
    )
    acc_id = acc_res["id"]
    assert acc_res["name"] == "HDFC Factory Primary"
    assert acc_res["opening_balance"] == 50000.0

    # -------------------------------------------------------------------------
    # STEP 2: Import CSV Statement & Verify Deduplication Safeguards (Stage 4)
    # -------------------------------------------------------------------------
    csv_content = (
        "Date,Narration,Chq/Ref Number,Withdrawal Amt.,Deposit Amt.,Closing Balance\n"
        "2026-08-05,NEFT CR - BAXTER RETAIL CLIENT PAYMENT,UTR001,,40000.00,90000.00\n"
        "2026-08-10,NEFT DR - FACTORY SHED RENT AUGUST,RENT001,10000.00,,80000.00\n"
        "2026-08-15,ATM CASH WDL - SELF FOR FACTORY WAGES,ATM991,15000.00,,65000.00\n"
        "2026-08-20,NEFT DR - TRF TO UCO BANK FACTORY,TRF992,5000.00,,60000.00\n"
    ).encode("utf-8")

    file1 = UploadFile(filename="statement_aug_2026.csv", file=io.BytesIO(csv_content))
    import_res1 = await import_bank_statement(id=acc_id, file=file1, dry_run=False, confirm_account_update=False, request=req)
    assert import_res1["inserted_count"] == 4
    assert import_res1["skipped_count"] == 0
    assert len(lines_store) == 4

    # Re-import identical statement file: 0 inserted, 4 skipped
    file2 = UploadFile(filename="statement_aug_2026.csv", file=io.BytesIO(csv_content))
    import_res2 = await import_bank_statement(id=acc_id, file=file2, dry_run=False, confirm_account_update=False, request=req)
    assert import_res2["inserted_count"] == 0
    assert import_res2["skipped_count"] == 4
    assert len(lines_store) == 4

    # Locate individual imported statement lines
    all_lines = list(lines_store.values())
    line_rev = next(l for l in all_lines if l["credit_amount"] == 40000.0)
    line_rent = next(l for l in all_lines if l["debit_amount"] == 10000.0)
    line_cash = next(l for l in all_lines if l["debit_amount"] == 15000.0)
    line_trf = next(l for l in all_lines if l["debit_amount"] == 5000.0)

    # -------------------------------------------------------------------------
    # STEP 3: Match Client Revenue & Factory Rent
    # -------------------------------------------------------------------------
    client_pay_id = ObjectId()
    payments_store[str(client_pay_id)] = {
        "_id": client_pay_id,
        "client_name": "Baxter Retail Pvt Ltd",
        "payment_no": "PAY-2026-AUG-01",
        "amount": 40000.0,
    }
    await match_statement_line(
        id=str(line_rev["_id"]),
        payload=BankStatementLineUpdate(
            match_status="matched",
            matched_to={"type": "payment", "ref_id": str(client_pay_id)},
        ),
        request=req,
    )

    rent_exp_id = ObjectId()
    expenses_store[str(rent_exp_id)] = {
        "_id": rent_exp_id,
        "payee": "Agra Industrial Landlords",
        "category": "Rent & Utilities",
        "amount": 10000.0,
    }
    await match_statement_line(
        id=str(line_rent["_id"]),
        payload=BankStatementLineUpdate(
            match_status="matched",
            matched_to={"type": "expense", "ref_id": str(rent_exp_id)},
        ),
        request=req,
    )

    # -------------------------------------------------------------------------
    # STEP 4: Confirm Cash Withdrawal & Create Cash Pool (Stage 1 & 2 Foundation)
    # -------------------------------------------------------------------------
    cash_confirm_res = await confirm_cash_withdrawal(
        payload=CashWithdrawalConfirmIn(statement_line_id=str(line_cash["_id"])),
        request=req,
    )
    assert cash_confirm_res["ok"] is True
    cash_ledger_id = cash_confirm_res["cash_ledger_id"]
    assert cash_ledger_store[cash_ledger_id]["amount"] == 15000.0
    assert cash_ledger_store[cash_ledger_id]["remaining_balance"] == 15000.0

    # -------------------------------------------------------------------------
    # STEP 5: Disburse Karigar Wage Payment & Cash Expense from Cash Pool (Stages 1-3)
    # -------------------------------------------------------------------------
    # 5a. Pay Karigar Wage: ₹8,000
    wage_res = await create_wage_payment(
        wid=str(worker_id),
        payload=WagePaymentIn(
            worker_id=str(worker_id),
            amount=8000.0,
            date="2026-08-16",
            period_from="2026-08-01",
            period_to="2026-08-15",
            paid_via="cash",
            cash_ledger_id=cash_ledger_id,
            allow_overpayment=True,
            override_reason="Bi-weekly approved piece-rate payout",
        ),
        request=req,
    )
    assert wage_res["amount"] == 8000.0
    assert wage_res["paid_via"] == "cash"
    assert cash_ledger_store[cash_ledger_id]["remaining_balance"] == 7000.0

    # 5b. Pay Cash Expense: ₹2,000 for Factory Tea & Refreshments
    exp_res = await create_expense(
        ExpenseIn(
            category="Office & Administrative",
            payee="Factory Canteen Vendor",
            amount=2000.0,
            date="2026-08-17",
            paid_via="cash",
            cash_ledger_id=cash_ledger_id,
            notes="August bi-weekly tea and refreshments for karigars",
        ),
        req,
    )
    assert exp_res["amount"] == 2000.0
    assert exp_res["paid_via"] == "cash"
    assert cash_ledger_store[cash_ledger_id]["remaining_balance"] == 5000.0

    # -------------------------------------------------------------------------
    # STEP 6: Verify Live "Total Cash in Hand" Metric
    # -------------------------------------------------------------------------
    summary_res = await get_reconciliation_summary(req)
    assert summary_res["summary"]["total_cash_in_hand"] == 5000.0
    assert summary_res["summary"]["total_cash_withdrawn"] == 15000.0

    # -------------------------------------------------------------------------
    # STEP 7: Reclassification Safeguard (Stage 5)
    # -------------------------------------------------------------------------
    # Attempting to reclassify the withdrawal line to unmatched/transfer while dependents exist must be BLOCKED
    with pytest.raises(HTTPException) as exc_reclassify:
        await match_statement_line(
            id=str(line_cash["_id"]),
            payload=BankStatementLineUpdate(match_status="unmatched"),
            request=req,
        )
    assert exc_reclassify.value.status_code == 400
    assert "active dependent records" in exc_reclassify.value.detail

    # -------------------------------------------------------------------------
    # STEP 8: Lock Reconciliation Period (Stage 8)
    # -------------------------------------------------------------------------
    lock_res = await lock_reconciliation_period(
        PeriodLockIn(
            bank_account_id=acc_id,
            period_from="2026-08-01",
            period_to="2026-08-31",
            notes="August 2026 finalized and submitted for GST & CA audit",
        ),
        req,
    )
    assert lock_res["ok"] is True
    assert lock_res["lock"]["status"] == "locked"

    # Attempting any edits on August statement lines while locked must be BLOCKED
    with pytest.raises(HTTPException) as exc_locked:
        await match_statement_line(
            id=str(line_rev["_id"]),
            payload=BankStatementLineUpdate(match_status="unmatched"),
            request=req,
        )
    assert exc_locked.value.status_code == 400
    assert "finalized and locked" in exc_locked.value.detail

    # -------------------------------------------------------------------------
    # STEP 9: Export CA / Accountant Reconciliation Workbook (Stage 9)
    # -------------------------------------------------------------------------
    export_response = await export_reconciliation_report(
        request=req,
        bank_account_id=acc_id,
        from_date="2026-08-01",
        to_date="2026-08-31",
        format="excel",
    )
    assert export_response.status_code == 200
    assert export_response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(export_response.body))
    assert "Executive Summary" in wb.sheetnames
    assert "1. Revenue & Receipts" in wb.sheetnames
    assert "2. Operating Expenses" in wb.sheetnames
    assert "4. Cash & Karigar Wages" in wb.sheetnames

    # Check cash sheet shows withdrawal + karigar wage + cash expense breakdown
    ws_cash = wb["4. Cash & Karigar Wages"]
    assert ws_cash["C4"].value == 15000.0  # Total withdrawal
    assert ws_cash["D4"].value == 10000.0  # Disbursed (8k wage + 2k expense)
    assert ws_cash["E4"].value == 5000.0   # Remaining Cash in Hand
    assert "Mukesh Karigar" in ws_cash["G4"].value
    assert "Office & Administrative" in ws_cash["G4"].value

    # -------------------------------------------------------------------------
    # STEP 10: Admin Unlock
    # -------------------------------------------------------------------------
    unlock_res = await unlock_reconciliation_period(
        PeriodUnlockIn(
            bank_account_id=acc_id,
            period_from="2026-08-01",
            period_to="2026-08-31",
            reason="Admin unlocking to adjust misallocated transfer line",
        ),
        req,
    )
    assert unlock_res["ok"] is True
    assert locks_store[0]["status"] == "unlocked"















