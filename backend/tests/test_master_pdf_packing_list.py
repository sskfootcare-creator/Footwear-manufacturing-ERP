"""Automated tests for Master Visual Reference Packing List generation system.

Verifies:
1. Reference data fixture matching the master reference PDF (PO 2220007313).
2. Size matrix breakdown (36-42).
3. Excel output (.xlsx) and PDF output (.pdf) generation.
4. Grand Total dynamic formulas and calculation accuracy.
5. Order Summary calculation and #DIV/0! error protection when Order Qty is 0.
6. Input validation endpoint (/api/packing-list/validate).
7. Interactive preview endpoint (/api/packing-list/preview).
"""
import io
import pytest
from httpx import AsyncClient, ASGITransport
from bson import ObjectId
import openpyxl

import server
from server import app, oid, now_iso
from packing_list import build_default_packing_list, build_packing_list_pdf, VENDOR, DEFAULT_SIZES


class MockFindCursor:
    def __init__(self, docs):
        self._docs = list(docs)

    def sort(self, key, direction=1):
        if isinstance(key, list):
            key = key[0][0]
        reverse = (direction == -1)
        self._docs.sort(key=lambda d: d.get(key, "") or "", reverse=reverse)
        return self

    async def to_list(self, limit=None):
        if limit:
            return self._docs[:limit]
        return self._docs


def match_filter(doc, filter_dict):
    if not filter_dict:
        return True
    if "_id" in filter_dict:
        if str(doc.get("_id")) != str(filter_dict["_id"]):
            return False
    for k, v in filter_dict.items():
        if k in ("_id", "$or"):
            continue
        val = doc.get(k)
        if isinstance(v, dict) and "$in" in v:
            target_list = v["$in"]
            if isinstance(val, list):
                if not any(item in target_list for item in val):
                    return False
            else:
                if val not in target_list:
                    return False
        else:
            if val != v:
                return False
    return True


class MockCollection:
    def __init__(self):
        self.docs = []

    async def find_one(self, filter_dict):
        for d in self.docs:
            if match_filter(d, filter_dict):
                return d
        return None

    def find(self, filter_dict=None, projection=None):
        res = [d for d in self.docs if match_filter(d, filter_dict or {})]
        return MockFindCursor(res)

    async def insert_one(self, doc):
        doc_copy = dict(doc)
        if "_id" not in doc_copy:
            doc_copy["_id"] = ObjectId()
        self.docs.append(doc_copy)
        class InsertResult:
            inserted_id = doc_copy["_id"]
        return InsertResult()

    async def count_documents(self, filter_dict):
        res = [d for d in self.docs if match_filter(d, filter_dict)]
        return len(res)


class MockDB:
    def __init__(self):
        self.pos = MockCollection()
        self.packing_lists = MockCollection()
        self.packing_cartons = MockCollection()
        self.production_jobs = MockCollection()


@pytest.fixture
def mock_db(monkeypatch):
    mdb = MockDB()
    monkeypatch.setattr(server, "db", mdb)
    return mdb


@pytest.fixture
async def async_client():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as client:
        yield client


@pytest.fixture
def mock_user_override(monkeypatch):
    async def mock_get_current_user(request=None):
        return {
            "id": "admin_test_id",
            "email": "admin@sskfootcare.com",
            "role": "admin",
            "roles": ["admin"],
        }
    monkeypatch.setattr(server, "get_current_user", mock_get_current_user)


@pytest.fixture
def reference_po_data():
    """Master reference PDF data fixture (PO 2220007313)."""
    line_items = []
    # Style ZFLWWWFLTM154 - CREAM (4 pairs each of sizes 37, 38, 39, 40, 41 = 20 pairs total)
    for sz in ["37", "38", "39", "40", "41"]:
        line_items.append({
            "style_code": "ZFLWWWFLTM154",
            "description": "CREAM SLIPPER",
            "color": "CREAM",
            "size": sz,
            "quantity": 4,
            "unit_price": 100.0,
            "amount": 400.0,
        })
    # Style ZFLWWWFLTM154 - TAUPE (4 pairs each of sizes 37, 38, 39, 40, 41 = 20 pairs total)
    for sz in ["37", "38", "39", "40", "41"]:
        line_items.append({
            "style_code": "ZFLWWWFLTM154",
            "description": "TAUPE SLIPPER",
            "color": "TAUPE",
            "size": sz,
            "quantity": 4,
            "unit_price": 100.0,
            "amount": 400.0,
        })

    return {
        "po_number": "2220007313",
        "po_date": "12.02.2026",
        "client_name": "ZECODE-BANGLORE-2220 ZECODE-BANGLORE-2220",
        "client_address": "PLOT NO. 2J/2K, 3RD PHASE KIADB OBEDENAHALLI INDUSTRIAL AREA BANGLORE, KARNATAKA DODDABALLAPUR 561 BENGALURU KARNATAKA 561203",
        "client_gstin": "29AAACS6995D2ZX",
        "site_code": "ZC_BLR-WH",
        "carton_dim": "60x50x30 CMS",
        "total_quantity": 40,
        "line_items": line_items,
        "created_at": now_iso(),
    }


# ---------------------------------------------------------------------------
# Test Cases
# ---------------------------------------------------------------------------

def test_excel_packing_list_generation(reference_po_data):
    """Test generating Excel Packing List matching reference PDF."""
    xlsx_bytes = build_default_packing_list(reference_po_data)
    assert len(xlsx_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    ws = wb.active
    assert ws.title == "Packing list"

    # Title check
    assert ws["A1"].value == "PACKING LIST"

    # Vendor & PO Info check
    assert VENDOR["name"] in str(ws["B2"].value)
    assert ws["B8"].value == "2220007313"
    assert ws["B9"].value == "12.02.2026"
    assert ws["O9"].value == "60x50x30 CMS"

    # Headers check
    headers = [ws.cell(row=10, column=c).value for c in range(1, 18)]
    assert "SITE CODE" in headers
    assert "Style" in headers
    assert "Colour" in headers
    assert "CTN .NO" in headers
    assert "36" in headers and "42" in headers
    assert "GROSS WEIGHT" in headers

    # Grand Total formulas check
    gt_row = 34
    assert "=SUM(E11:E33)" in ws[f"E{gt_row}"].value
    assert "=SUM(N11:N33)" in ws[f"N{gt_row}"].value
    assert "=SUM(O11:O33)" in ws[f"O{gt_row}"].value

    # Order Summary formulas & #DIV/0! protection check
    summary_start = 37
    # Size 36 order qty formula check
    assert ws[f"F{summary_start+4}"].value == "=IF(F38=0, 0, F40/F38)"


def test_pdf_packing_list_generation(reference_po_data):
    """Test generating vector PDF Packing List matching reference PDF."""
    pdf_bytes = build_packing_list_pdf(reference_po_data)
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b"%PDF")


@pytest.mark.anyio
async def test_preview_packing_list_endpoint(mock_db, mock_user_override, async_client, reference_po_data):
    """Test POST /api/packing-list/preview endpoint."""
    res_po = await mock_db.pos.insert_one(reference_po_data)
    pid = str(res_po.inserted_id)

    response = await async_client.post("/api/packing-list/preview", json={"po_id": pid})
    assert response.status_code == 200
    data = response.json()

    assert data["vendor"]["name"] == VENDOR["name"]
    assert data["po"]["po_number"] == "2220007313"
    assert data["po"]["total_pcs"] == 40
    assert data["po"]["total_cartons"] == 2
    assert len(data["rows"]) == 2
    assert data["grand_total"]["total_pcs"] == 40
    assert data["order_summary"]["total_order_qty"] == 40
    assert data["order_summary"]["total_pack_qty"] == 40
    assert data["order_summary"]["total_excess_short"] == 0
    assert data["order_summary"]["total_excess_short_pct"] == "0.00%"


@pytest.mark.anyio
async def test_validate_packing_list_endpoint(mock_db, mock_user_override, async_client, reference_po_data):
    """Test POST /api/packing-list/validate endpoint."""
    res_po = await mock_db.pos.insert_one(reference_po_data)
    pid = str(res_po.inserted_id)

    response = await async_client.post("/api/packing-list/validate", json={"po_id": pid})
    assert response.status_code == 200
    data = response.json()
    assert data["valid"] is True
    assert len(data["errors"]) == 0

    # Test invalid PO ID
    bad_res = await async_client.post("/api/packing-list/validate", json={"po_id": str(ObjectId())})
    assert bad_res.status_code == 200
    assert bad_res.json()["valid"] is False


@pytest.mark.anyio
async def test_po_packing_list_export_endpoints(mock_db, mock_user_override, async_client, reference_po_data):
    """Test GET /api/pos/{pid}/packing-list.pdf and GET /api/pos/{pid}/packing-list.xlsx."""
    res_po = await mock_db.pos.insert_one(reference_po_data)
    pid = str(res_po.inserted_id)

    # PDF Export
    pdf_res = await async_client.get(f"/api/pos/{pid}/packing-list.pdf")
    assert pdf_res.status_code == 200
    assert pdf_res.headers["content-type"] == "application/pdf"
    assert pdf_res.content.startswith(b"%PDF")

    # XLSX Export
    xlsx_res = await async_client.get(f"/api/pos/{pid}/packing-list.xlsx")
    assert xlsx_res.status_code == 200
    assert "spreadsheetml" in xlsx_res.headers["content-type"]
    assert len(xlsx_res.content) > 0
