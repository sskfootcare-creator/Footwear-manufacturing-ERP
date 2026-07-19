"""Iteration 12 tests:
- FREE PO extractor on real SHEIN PDF (/tmp/newpo.pdf, 126 lines)
- Packing list with full manual / shipping fields + persistence + re-download
- Merged packing list endpoint (same client, sectioned mode, cross-client 400)
- Auto-pick template via alias matching + fallback to default
- Regression: extractor on /tmp/test_po.xlsx + smoke endpoints
"""
import base64
import io
import os
import re

import openpyxl
import pytest
import requests


def _read_env():
    p = "/app/frontend/.env"
    if os.path.exists(p):
        for line in open(p):
            if line.startswith("REACT_APP_BACKEND_URL="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8000")


BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or _read_env()).rstrip("/")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL",    "admin@sskfootcare.com")
ADMIN_PASS  = os.environ.get("ADMIN_PASSWORD", "Admin@123")

SHEIN_PDF = "/tmp/newpo.pdf"
TEST_XLSX = "/tmp/test_po.xlsx"


@pytest.fixture(scope="session")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
    assert r.status_code == 200, f"Login failed: {r.text}"
    return s


@pytest.fixture(scope="session")
def upload_session(session):
    """Session without forced JSON content-type for multipart uploads."""
    s = requests.Session()
    s.cookies.update(session.cookies)
    return s


# -------------------------------------------------------------------
# 1. SHEIN PDF extractor
# -------------------------------------------------------------------
class TestSheinPDFExtract:
    @pytest.fixture(scope="class")
    def extracted(self, upload_session):
        if not os.path.exists(SHEIN_PDF):
            pytest.skip("SHEIN PDF not present at /tmp/newpo.pdf")
        with open(SHEIN_PDF, "rb") as fh:
            files = {"file": ("newpo.pdf", fh.read(), "application/pdf")}
        r = upload_session.post(f"{BASE_URL}/api/pos/extract", files=files)
        assert r.status_code == 200, f"extract failed: {r.status_code} {r.text[:400]}"
        body = r.json()
        return body.get("data", body) if isinstance(body, dict) else body

    def test_top_fields(self, extracted):
        d = extracted
        assert d.get("po_number") == "5155396467", f"po_number={d.get('po_number')}"
        assert "NEXTGEN" in (d.get("client_name") or "").upper(), f"client_name={d.get('client_name')}"
        assert "SSK FOOTCARE" in (d.get("vendor_name") or "").upper(), f"vendor_name={d.get('vendor_name')}"
        assert d.get("po_date") == "2026-03-24", f"po_date={d.get('po_date')}"
        assert d.get("delivery_date") == "2026-06-20", f"delivery_date={d.get('delivery_date')}"
        # grand_total may be float; allow tolerance
        gt = float(d.get("grand_total") or 0)
        assert abs(gt - 512925) < 1.0, f"grand_total={gt}"
        assert d.get("total_quantity") == 2100, f"total_quantity={d.get('total_quantity')}"
        items = d.get("line_items") or []
        assert len(items) == 126, f"line_items_count={len(items)}"

    def test_first_line_item(self, extracted):
        items = extracted.get("line_items") or []
        assert items, "no line items extracted"
        # Locate the RRL-SH-10002 / BLACK / size 3 row
        match = next(
            (
                li for li in items
                if (li.get("style_code") or "").upper() == "RRL-SH-10002"
                and (li.get("color") or "").upper() == "BLACK"
                and str(li.get("size") or "").strip() == "3"
            ),
            None,
        )
        assert match is not None, f"sample line not found; got styles={[i.get('style_code') for i in items[:5]]}"
        desc = (match.get("description") or "").upper().replace(" ", "")
        assert "SHEIN" in desc and "WOMEN" in desc, f"description split issue: {match.get('description')}"
        assert match.get("hsn_code") == "64039990", f"hsn={match.get('hsn_code')}"
        assert int(match.get("quantity") or 0) == 10, f"qty={match.get('quantity')}"
        assert float(match.get("unit_price") or 0) == 235.0, f"unit_price={match.get('unit_price')}"
        assert float(match.get("amount") or 0) == 2350.0, f"amount={match.get('amount')}"
        mrp_raw = str(match.get("mrp") or "")
        assert mrp_raw.startswith("399"), f"mrp={mrp_raw}"


# -------------------------------------------------------------------
# 2. Regression: synthetic xlsx extractor
# -------------------------------------------------------------------
class TestRegressionXlsxExtract:
    def test_regression(self, upload_session):
        if not os.path.exists(TEST_XLSX):
            pytest.skip("test xlsx missing")
        with open(TEST_XLSX, "rb") as fh:
            files = {"file": ("test_po.xlsx", fh.read(),
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = upload_session.post(f"{BASE_URL}/api/pos/extract", files=files)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:200]}"
        body = r.json()
        d = body.get("data", body) if isinstance(body, dict) else body
        assert str(d.get("po_number") or "") == "2220008835", f"po_number={d.get('po_number')}"
        assert "SIYARAM" in (d.get("client_name") or "").upper(), f"client={d.get('client_name')}"
        assert len(d.get("line_items") or []) == 3, f"line_items={len(d.get('line_items') or [])}"


# -------------------------------------------------------------------
# 3. Packing list with manual fields + persistence + re-download
# -------------------------------------------------------------------
class TestPackingListFull:
    @pytest.fixture(scope="class")
    def po_id(self, session):
        r = session.get(f"{BASE_URL}/api/pos")
        assert r.status_code == 200
        pos = r.json()
        assert pos, "no POs available"
        return pos[0]["id"]

    def test_generate_with_all_options(self, session, po_id):
        opts = {
            "po_id": po_id,
            "carton_dim": "60x50x30 CMS",
            "pcs_per_box": 24,
            "net_wt_per_carton": 11.5,
            "gross_wt_per_carton": 13.5,
            "dispatch_date": "2026-02-15",
            "transporter": "XPRESS LOGISTICS",
            "vehicle_no": "MH04AB1234",
            "driver_name": "RAJU PATIL",
            "driver_phone": "9876543210",
            "site_code": "ST-9981",
            "destination": "NHAVA SHEVA",
            "port": "JNPT",
            "notes": "Handle with care - leather goods. Stack max 4 high.",
        }
        r = session.post(f"{BASE_URL}/api/packing-lists/job", json=opts)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:300]}"
        pl_id = r.headers.get("X-Packing-List-Id")
        assert pl_id, "missing X-Packing-List-Id header"

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active

        # Row 15 should contain shipping pairs
        row15 = " ".join(str(ws.cell(15, c).value or "") for c in range(1, 20)).upper()
        for kw in ["DISPATCH", "TRANSPORTER", "VEHICLE", "DRIVER"]:
            assert kw in row15, f"row15 missing {kw}; row15={row15}"
        # Concrete values populated
        full_text = "\n".join(
            " ".join(str(ws.cell(r_, c).value or "") for c in range(1, 25))
            for r_ in range(1, ws.max_row + 1)
        )
        assert "XPRESS LOGISTICS" in full_text.upper(), "transporter value missing"
        assert "MH04AB1234" in full_text.upper(), "vehicle_no missing"
        assert "RAJU PATIL" in full_text.upper(), "driver_name missing"
        assert "ST-9981" in full_text.upper(), "site_code value missing in body"
        assert "JNPT" in full_text.upper() or "NHAVA" in full_text.upper(), "port/destination missing"
        assert "HANDLE WITH CARE" in full_text.upper(), "notes block missing"

        # Persist class state
        TestPackingListFull._created_id = pl_id
        TestPackingListFull._orig_len = len(r.content)

    def test_list_packing_lists(self, session):
        r = session.get(f"{BASE_URL}/api/packing-lists")
        assert r.status_code == 200, r.text[:200]
        items = r.json()
        assert isinstance(items, list) and items, "list empty"
        ids = [it.get("id") or it.get("_id") for it in items]
        assert TestPackingListFull._created_id in ids, "newly created PL not in list"
        # file_b64 must be stripped from listing payload
        for it in items:
            assert "file_b64" not in it, "file_b64 leaked to client in listing"

    def test_redownload_file(self, session):
        pl_id = TestPackingListFull._created_id
        r = session.get(f"{BASE_URL}/api/packing-lists/{pl_id}/file")
        assert r.status_code == 200, f"{r.status_code}: {r.text[:200]}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "officedocument" in ct, f"bad content-type: {ct}"
        # bytes round-trip (allow minor zip variance — should be exact match)
        assert len(r.content) == TestPackingListFull._orig_len, (
            f"length mismatch: orig={TestPackingListFull._orig_len} dl={len(r.content)}"
        )
        # Verify it loads
        openpyxl.load_workbook(io.BytesIO(r.content))


# -------------------------------------------------------------------
# 4. Merged packing list endpoint
# -------------------------------------------------------------------
class TestMergedPackingList:
    def test_merge_same_client(self, session):
        # Move 2 production jobs to dispatched, attempt merged generation
        r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        assert r.status_code == 200
        all_jobs = r.json()
        
        # group by client via PO mapping
        r_po = session.get(f"{BASE_URL}/api/pos")
        po_to_client = {p["id"]: p.get("client_name", "") for p in r_po.json()}
        
        if len(all_jobs) < 2:
            import uuid
            style_payload = {
                "name": "Merge Style",
                "category": "Footwear",
                "base_size": "8",
                "bom": [],
                "labor": [],
            }
            r_style = session.post(f"{BASE_URL}/api/styles", json=style_payload, timeout=15)
            assert r_style.status_code == 200, r_style.text
            style_code = r_style.json()["code"]
            
            po_payload = {
                "po_number": f"PO-MERGE-{uuid.uuid4().hex[:8]}",
                "client_name": "Merge Client",
                "po_date": "2026-07-13",
                "delivery_date": "2026-08-13",
                "payment_terms": "30 Days Credit",
                "line_items": [
                    {
                        "style_code": style_code,
                        "color": "Black",
                        "size": "8",
                        "quantity": 50,
                        "unit_price": 400.0,
                        "amount": 20000.0
                    },
                    {
                        "style_code": style_code,
                        "color": "Black",
                        "size": "9",
                        "quantity": 40,
                        "unit_price": 400.0,
                        "amount": 16000.0
                    }
                ]
            }
            r_po = session.post(f"{BASE_URL}/api/pos", json=po_payload, timeout=15)
            assert r_po.status_code == 200, r_po.text
            po_id = r_po.json()["id"]
            po_to_client[po_id] = "Merge Client"
            
            r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
            assert r.status_code == 200
            all_jobs = r.json()
        # group by client via PO mapping
        r = session.get(f"{BASE_URL}/api/pos")
        po_to_client = {p["id"]: p.get("client_name", "") for p in r.json()}
        # bucket jobs by client
        buckets = {}
        for j in all_jobs:
            c = po_to_client.get(j.get("po_id"), "")
            buckets.setdefault(c, []).append(j)
        same_client = next((v for v in buckets.values() if len(v) >= 2), None)
        if not same_client:
            import uuid
            style_payload = {
                "name": "Merge Style",
                "category": "Footwear",
                "base_size": "8",
                "bom": [],
                "labor": [],
            }
            r_style = session.post(f"{BASE_URL}/api/styles", json=style_payload, timeout=15)
            assert r_style.status_code == 200, r_style.text
            style_code = r_style.json()["code"]
            
            po_payload = {
                "po_number": f"PO-MERGE-{uuid.uuid4().hex[:8]}",
                "client_name": "Merge Client",
                "po_date": "2026-07-13",
                "delivery_date": "2026-08-13",
                "payment_terms": "30 Days Credit",
                "line_items": [
                    {
                        "style_code": style_code,
                        "color": "Black",
                        "size": "8",
                        "quantity": 50,
                        "unit_price": 400.0,
                        "amount": 20000.0
                    },
                    {
                        "style_code": style_code,
                        "color": "Black",
                        "size": "9",
                        "quantity": 40,
                        "unit_price": 400.0,
                        "amount": 16000.0
                    }
                ]
            }
            r_po = session.post(f"{BASE_URL}/api/pos", json=po_payload, timeout=15)
            assert r_po.status_code == 200, r_po.text
            po_id = r_po.json()["id"]
            
            r_jobs = session.get(f"{BASE_URL}/api/production/jobs?source_type=all", timeout=15)
            assert r_jobs.status_code == 200
            same_client = [j for j in r_jobs.json() if j.get("po_id") == po_id]
            po_to_client[po_id] = "Merge Client"
        j1, j2 = same_client[:2]
        # PATCH both to dispatched (idempotent)
        for j in (j1, j2):
            session.patch(f"{BASE_URL}/api/production/jobs/{j['id']}", json={"stage": "dispatched"})

        body = {
            "job_ids": [j1["id"], j2["id"]],
            "sectioned": True,
            "site_code": "MERGED-001",
            "notes": "Merged packing test",
        }
        r = session.post(f"{BASE_URL}/api/packing-lists/merged", json=body)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:300]}"
        assert r.headers.get("X-Packing-List-Id"), "missing X-Packing-List-Id"
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Validate the file loads & contains data
        assert ws.max_row > 10, f"too few rows: {ws.max_row}"
        TestMergedPackingList._same_client_jobs = (j1, j2, po_to_client)

    def test_merge_cross_client_rejected(self, session):
        j1, j2, po_to_client = TestMergedPackingList._same_client_jobs
        # find a job whose client is DIFFERENT from j1's
        r = session.get(f"{BASE_URL}/api/production/jobs?include_archived=true")
        c1 = po_to_client.get(j1["po_id"], "")
        diff = next(
            (j for j in r.json() if po_to_client.get(j.get("po_id"), "") and po_to_client.get(j.get("po_id")) != c1),
            None,
        )
        if not diff:
            pytest.skip("no cross-client job to test rejection")
        r = session.post(f"{BASE_URL}/api/packing-lists/merged",
                         json={"job_ids": [j1["id"], diff["id"]]})
        assert r.status_code == 400, f"expected 400 for cross-client; got {r.status_code}: {r.text[:200]}"


# -------------------------------------------------------------------
# 5. Auto-pick template via alias
# -------------------------------------------------------------------
def _make_distinctive_xlsx_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "ZZ_TEST_TEMPLATE_MARKER_42"   # unique marker we can detect
    ws["A2"] = "PO: {{po_number}}"
    ws["A3"] = "Client: {{client_name}}"
    ws["A5"] = "Style"
    ws["B5"] = "Color"
    ws["C5"] = "Qty"
    ws["A6"] = "{{lines}}"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestAutoPickTemplate:
    def test_alias_match_then_fallback(self, session):
        # Find any PO + its client
        r = session.get(f"{BASE_URL}/api/pos")
        pos = r.json()
        assert pos, "no POs"
        target_po = pos[0]
        po_id = target_po["id"]
        client = target_po.get("client_name", "")
        assert client, "PO has no client_name"
        # Use first word of client name as alias
        alias_kw = client.split()[0]

        # 1. Generate without any template -> should fall back to default
        r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id})
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        default_a1 = wb.active["A1"].value or ""
        assert "ZZ_TEST_TEMPLATE_MARKER_42" not in default_a1, "default generation should NOT match marker"

        # 2. Upload an alias-matching template
        tpl_bytes = _make_distinctive_xlsx_template()
        b64 = base64.b64encode(tpl_bytes).decode()
        r = session.post(f"{BASE_URL}/api/packing-templates", json={
            "client_name": "ZZ_TEST_AUTO_PICK",
            "name": "AUTO_PICK_TEMPLATE",
            "aliases": [alias_kw],
            "file_b64": b64,
        })
        assert r.status_code in (200, 201), f"{r.status_code}: {r.text[:200]}"
        tid = r.json().get("id")
        assert tid

        try:
            # 3. Re-generate without explicit template_id; auto-pick should use uploaded template
            r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id})
            assert r.status_code == 200
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            a1 = wb.active["A1"].value or ""
            assert "ZZ_TEST_TEMPLATE_MARKER_42" in a1, (
                f"auto-pick failed; A1='{a1}', alias={alias_kw}, client={client}"
            )
        finally:
            # 4. Cleanup -> delete template, then verify fallback again
            r = session.delete(f"{BASE_URL}/api/packing-templates/{tid}")
            assert r.status_code in (200, 204)

        r = session.post(f"{BASE_URL}/api/packing-lists/job", json={"po_id": po_id})
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        a1 = wb.active["A1"].value or ""
        assert "ZZ_TEST_TEMPLATE_MARKER_42" not in a1, "marker still present after deletion (no fallback)"


# -------------------------------------------------------------------
# 6. Regression smoke endpoints
# -------------------------------------------------------------------
class TestSmoke:
    def test_smoke_endpoints(self, session):
        for ep in [
            "/api/pos", "/api/workers", "/api/dashboard/overdue",
            "/api/reports/payroll", "/api/reports/monthly-production",
            "/api/production/archive", "/api/packing-templates",
        ]:
            r = session.get(f"{BASE_URL}{ep}")
            assert r.status_code == 200, f"{ep} -> {r.status_code} {r.text[:200]}"
