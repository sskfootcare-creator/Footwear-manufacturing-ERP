"""Verification tests for the rebuilt bulk_create_sku_map endpoint.

Test scenarios:
  1. XLSX upload - real-template-style data with 2 styles, 2 colors, multiple
     sizes, one row with a Dropbox image_url.  Confirms:
       - correct grouping into sku_map documents
       - size_map per group
       - image_url applied and normalized
  2. CSV upload - same columns, same semantics.
  3. Re-upload the SAME file - expect UPDATED not DUPLICATED (size_map merged).
  4. One intentionally bad row (nonexistent style_code) - reported in errors
     with the row number; all other valid rows still process successfully.
  5. Case-insensitive color grouping - "Gold" and "GOLD" collapse into one doc.
  6. Conflicting image_url within a group - warns but does not fail.
  7. Missing required fields - each missing field reported correctly.
"""
import io
import csv
import pytest
import openpyxl
from bson import ObjectId
from httpx import AsyncClient, ASGITransport

import server
from server import app


# -- Shared mock infrastructure -----------------------------------------------

class MockFindCursor:
    def __init__(self, docs):
        self._docs = list(docs)

    def sort(self, *a, **kw):
        return self

    async def to_list(self, limit=None):
        return self._docs[:limit] if limit else list(self._docs)


def _match(doc, f):
    if not f:
        return True
    for k, v in f.items():
        if k == "_id":
            if str(doc.get("_id")) != str(v):
                return False
            continue
        val = doc.get(k)
        if isinstance(v, dict):
            if "$in" in v and val not in v["$in"]:
                return False
            if "$ne" in v and val == v["$ne"]:
                return False
            if "$regex" in v:
                import re as _re
                flags = _re.IGNORECASE if v.get("$options") == "i" else 0
                if not _re.search(v["$regex"], str(val or ""), flags):
                    return False
        else:
            if val != v:
                return False
    return True


class MockCollection:
    def __init__(self):
        self.docs: list = []

    async def find_one(self, f=None, *a, **kw):
        matched = [d for d in self.docs if _match(d, f or {})]
        return matched[0] if matched else None

    def find(self, f=None, *a, **kw):
        return MockFindCursor([d for d in self.docs if _match(d, f or {})])

    async def insert_one(self, doc):
        if "_id" not in doc:
            doc["_id"] = ObjectId()
        self.docs.append(doc)
        class R:
            inserted_id = doc["_id"]
        return R()

    async def update_one(self, f, upd, upsert=False):
        for doc in self.docs:
            if _match(doc, f):
                if "$set" in upd:
                    doc.update(upd["$set"])
                return
        if upsert:
            new = {}
            if "$set" in upd:
                new.update(upd["$set"])
            new.update({k: v for k, v in f.items() if not k.startswith("$")})
            if "_id" not in new:
                new["_id"] = ObjectId()
            self.docs.append(new)

    async def delete_one(self, f):
        before = len(self.docs)
        self.docs = [d for d in self.docs if not _match(d, f)]
        class R:
            deleted_count = before - len(self.docs)
        return R()

    async def find_one_and_update(self, f, upd, upsert=False, return_document=True):
        for doc in self.docs:
            if _match(doc, f):
                if "$inc" in upd:
                    for k, v in upd["$inc"].items():
                        doc[k] = doc.get(k, 0) + v
                return doc
        if upsert:
            new = {"_id": f.get("_id", "seq")}
            if "$inc" in upd:
                for k, v in upd["$inc"].items():
                    new[k] = v
            self.docs.append(new)
            return new
        return None


class MockDB:
    def __init__(self):
        self.styles                    = MockCollection()
        self.sku_map                   = MockCollection()
        self.activity_log              = MockCollection()
        self.counters                  = MockCollection()
        self.production_jobs           = MockCollection()
        self.unmatched_production_jobs = MockCollection()


@pytest.fixture
def mock_db(monkeypatch):
    mdb = MockDB()
    monkeypatch.setattr(server, "db", mdb)
    return mdb


@pytest.fixture
async def async_client():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as client:
        yield client


@pytest.fixture
def mock_admin(monkeypatch):
    async def _user(request=None):
        return {"id": "u1", "email": "admin@test.com", "role": "admin", "roles": ["admin"]}
    monkeypatch.setattr(server, "get_current_user", _user)


# -- Helpers to build test files -----------------------------------------------

COLUMNS = [
    "style_code", "color", "size", "external_sku",
    "source_type", "source_name", "external_style_name", "image_url",
]

def _make_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNS)
    w.writeheader()
    for r in rows:
        w.writerow({c: r.get(c, "") for c in COLUMNS})
    return buf.getvalue().encode()


STYLE_A = "MYNT-HEEL-001"
STYLE_B = "MYNT-FLAT-002"
CLIENT  = "Myntra"

BASE_ROWS = [
    {"style_code": STYLE_A, "color": "Gold",   "size": "5",  "external_sku": "MYN-HA-5",  "source_type": "online_channel", "source_name": CLIENT, "external_style_name": "Strappy Heel", "image_url": "https://www.dropbox.com/s/abc/heel_gold.jpg?dl=0"},
    {"style_code": STYLE_A, "color": "Gold",   "size": "6",  "external_sku": "MYN-HA-6",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_A, "color": "Gold",   "size": "7",  "external_sku": "MYN-HA-7",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_A, "color": "Silver", "size": "5",  "external_sku": "MYN-HS-5",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_A, "color": "Silver", "size": "6",  "external_sku": "MYN-HS-6",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_B, "color": "Black",  "size": "4",  "external_sku": "MYN-FB-4",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_B, "color": "Black",  "size": "5",  "external_sku": "MYN-FB-5",  "source_type": "online_channel", "source_name": CLIENT},
    {"style_code": STYLE_B, "color": "Black",  "size": "6",  "external_sku": "MYN-FB-6",  "source_type": "online_channel", "source_name": CLIENT},
]


async def _seed_styles(mock_db):
    await mock_db.styles.insert_one({"_id": ObjectId(), "code": STYLE_A, "name": "Strappy Heel"})
    await mock_db.styles.insert_one({"_id": ObjectId(), "code": STYLE_B, "name": "Ballet Flat"})


# -- Tests --------------------------------------------------------------------

@pytest.mark.anyio
async def test_bulk_xlsx_correct_grouping_and_size_map(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    r = await async_client.post(
        "/api/sku-map/bulk",
        files={"file": ("sku_map.xlsx", _make_xlsx(BASE_ROWS), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created"] == 3
    assert data["updated"] == 0
    assert data["errors"] == []

    docs = mock_db.sku_map.docs
    assert len(docs) == 3

    gold = next(d for d in docs if d["style_code"] == STYLE_A and d["color_key"] == "gold")
    assert gold["size_map"] == {"5": "MYN-HA-5", "6": "MYN-HA-6", "7": "MYN-HA-7"}
    assert "dl.dropboxusercontent.com" in gold["image_url"]
    assert "dl=0" not in gold["image_url"]

    silver = next(d for d in docs if d["style_code"] == STYLE_A and d["color_key"] == "silver")
    assert silver["size_map"] == {"5": "MYN-HS-5", "6": "MYN-HS-6"}
    assert silver["image_url"] == ""

    black = next(d for d in docs if d["style_code"] == STYLE_B and d["color_key"] == "black")
    assert black["size_map"] == {"4": "MYN-FB-4", "5": "MYN-FB-5", "6": "MYN-FB-6"}


@pytest.mark.anyio
async def test_bulk_csv_same_semantics(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    r = await async_client.post(
        "/api/sku-map/bulk",
        files={"file": ("sku_map.csv", _make_csv(BASE_ROWS), "text/csv")},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created"] == 3
    assert data["errors"] == []
    assert len(mock_db.sku_map.docs) == 3


@pytest.mark.anyio
async def test_bulk_reupload_updates_not_duplicates(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    xlsx = _make_xlsx(BASE_ROWS)
    r1 = await async_client.post("/api/sku-map/bulk", files={"file": ("f.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r1.json()["created"] == 3

    r2 = await async_client.post("/api/sku-map/bulk", files={"file": ("f.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r2.status_code == 200
    data = r2.json()
    assert data["created"] == 0
    assert data["updated"] == 3
    assert data["errors"] == []
    assert len(mock_db.sku_map.docs) == 3


@pytest.mark.anyio
async def test_bulk_reupload_new_size_merged(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    initial = [
        {"style_code": STYLE_A, "color": "Gold", "size": "5", "external_sku": "MYN-HA-5", "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": STYLE_A, "color": "Gold", "size": "6", "external_sku": "MYN-HA-6", "source_type": "online_channel", "source_name": CLIENT},
    ]
    r1 = await async_client.post("/api/sku-map/bulk", files={"file": ("i.xlsx", _make_xlsx(initial), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r1.json()["created"] == 1

    extended = initial + [{"style_code": STYLE_A, "color": "Gold", "size": "7", "external_sku": "MYN-HA-7", "source_type": "online_channel", "source_name": CLIENT}]
    r2 = await async_client.post("/api/sku-map/bulk", files={"file": ("e.xlsx", _make_xlsx(extended), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r2.json()["updated"] == 1
    assert r2.json()["created"] == 0
    assert mock_db.sku_map.docs[0]["size_map"] == {"5": "MYN-HA-5", "6": "MYN-HA-6", "7": "MYN-HA-7"}


@pytest.mark.anyio
async def test_bulk_bad_row_reported_others_processed(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    rows = [
        {"style_code": STYLE_A,               "color": "Gold",  "size": "5", "external_sku": "MYN-HA-5", "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": "NONEXISTENT-STYLE-XYZ","color": "Gold",  "size": "5", "external_sku": "BAD-001",  "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": STYLE_B,               "color": "Black", "size": "4", "external_sku": "MYN-FB-4", "source_type": "online_channel", "source_name": CLIENT},
    ]
    r = await async_client.post("/api/sku-map/bulk", files={"file": ("mixed.xlsx", _make_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()

    assert len(data["errors"]) == 1
    err = data["errors"][0]
    assert err["row"] == 3
    assert "NONEXISTENT-STYLE-XYZ" in err["reason"]
    assert "Style Master" in err["reason"]
    assert data["created"] == 2
    assert len(mock_db.sku_map.docs) == 2


@pytest.mark.anyio
async def test_bulk_case_insensitive_color_grouping(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    rows = [
        {"style_code": STYLE_A, "color": "Gold",  "size": "5", "external_sku": "MYN-HA-5", "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": STYLE_A, "color": "GOLD",  "size": "6", "external_sku": "MYN-HA-6", "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": STYLE_A, "color": " gold ", "size": "7", "external_sku": "MYN-HA-7", "source_type": "online_channel", "source_name": CLIENT},
    ]
    r = await async_client.post("/api/sku-map/bulk", files={"file": ("color.xlsx", _make_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created"] == 1
    assert data["errors"] == []
    assert len(mock_db.sku_map.docs) == 1
    doc = mock_db.sku_map.docs[0]
    assert doc["size_map"] == {"5": "MYN-HA-5", "6": "MYN-HA-6", "7": "MYN-HA-7"}
    assert doc["color"] == "Gold"
    assert doc["color_key"] == "gold"


@pytest.mark.anyio
async def test_bulk_conflicting_image_url_warns_uses_first(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    url1 = "https://www.dropbox.com/s/aaa/first.jpg?dl=0"
    url2 = "https://www.dropbox.com/s/bbb/second.jpg?dl=0"
    rows = [
        {"style_code": STYLE_A, "color": "Gold", "size": "5", "external_sku": "MYN-HA-5", "source_type": "online_channel", "source_name": CLIENT, "image_url": url1},
        {"style_code": STYLE_A, "color": "Gold", "size": "6", "external_sku": "MYN-HA-6", "source_type": "online_channel", "source_name": CLIENT, "image_url": url2},
    ]
    r = await async_client.post("/api/sku-map/bulk", files={"file": ("img.xlsx", _make_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created"] == 1
    assert data["errors"] == []
    assert len(data["warnings"]) == 1
    assert "image_url" in data["warnings"][0]["reason"]
    doc = mock_db.sku_map.docs[0]
    assert "dl.dropboxusercontent.com" in doc["image_url"]
    assert "aaa" in doc["image_url"]


@pytest.mark.anyio
async def test_bulk_missing_required_fields_reported(mock_db, mock_admin, async_client):
    await _seed_styles(mock_db)
    rows = [
        {"style_code": STYLE_A, "color": "", "size": "", "external_sku": "X", "source_type": "online_channel", "source_name": CLIENT},
        {"style_code": STYLE_A, "color": "Gold", "size": "5", "external_sku": "MYN-HA-5", "source_type": "online_channel", "source_name": CLIENT},
    ]
    r = await async_client.post("/api/sku-map/bulk", files={"file": ("miss.xlsx", _make_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert len(data["errors"]) == 1
    assert data["errors"][0]["row"] == 2
    assert "color" in data["errors"][0]["reason"]
    assert "size"  in data["errors"][0]["reason"]
    assert data["created"] == 1


@pytest.mark.anyio
async def test_bulk_empty_file_returns_400(mock_db, mock_admin, async_client):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    buf = io.BytesIO()
    wb.save(buf)
    r = await async_client.post(
        "/api/sku-map/bulk",
        files={"file": ("empty.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert "empty" in r.json()["detail"].lower()


@pytest.mark.anyio
async def test_download_sku_map_template_xlsx(mock_db, mock_admin, async_client):
    r = await async_client.get("/api/sku-map/template?format=xlsx")
    assert r.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in r.headers["content-type"]
    assert len(r.content) > 100
    # Verify openpyxl can read it
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    assert "style_code" in headers
    assert "external_sku" in headers


@pytest.mark.anyio
async def test_download_sku_map_template_csv(mock_db, mock_admin, async_client):
    r = await async_client.get("/api/sku-map/template?format=csv")
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    text = r.content.decode("utf-8-sig")
    assert "style_code,color,size,external_sku" in text
