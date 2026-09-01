"""Packing-list Excel and PDF generator.

Master visual reference implementation matching the SSK Footcare Master Packing List format.

Features:
1. Exact visual layout matching the master reference PDF (A4 Landscape print layout).
2. Vendor block, Destination Hub block, PO meta, main table with footwear size matrix (36-42).
3. Dynamic cartons, partial cartons, net/gross weights, and carton dimensions.
4. Grand Total row with dynamic Excel formulas (=SUM(...)).
5. Order Summary section (Order Qty vs Pack Qty vs Excss/Short & Excss/Short % protected against #DIV/0!).
6. Authorised Signatory block.
7. Dual export support: Excel (.xlsx) and PDF (.pdf).
8. Custom template substitution (build_from_template) retained for backward compatibility.
"""
from __future__ import annotations
import io
import re
from copy import copy
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ---------------------------------------------------------------------------
# Default Colors & Borders for Excel
# ---------------------------------------------------------------------------
_HEADER_FILL = "D9D9D9"     # Light grey banner/header fill
_SUMMARY_FILL = "E2E8F0"    # Light slate fill
_TEXT_COLOR = "000000"

_thin_black = Side(style="thin", color="000000")
_box_black = Border(left=_thin_black, right=_thin_black, top=_thin_black, bottom=_thin_black)

VENDOR = {
    "name": "SSK FOOTCARE MANUFACTURING LLP",
    "address": "CHEMBUR MUMBAI - 400071 H 43 , NARAYAN NIWAS, OPP JETVAN\nGARDEN MUMBAI MUMBAI 400071 MAHARASHTRA",
    "gstin": "27AFKFS4410F1Z2",
}

DEFAULT_SIZES = ["36", "37", "38", "39", "40", "41", "42"]


def _set_cell(
    ws, coord: str, val, *,
    bold=False, fill=None, color=None, size=9, align="left", border=True, num_format=None
):
    cell = ws[coord]
    cell.value = val
    font_kwargs = {"name": "Calibri", "size": size, "bold": bold}
    if color:
        font_kwargs["color"] = color
    cell.font = Font(**font_kwargs)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = _box_black
    if num_format:
        cell.number_format = num_format


def _apply_border_range(ws, min_col, min_row, max_col, max_row):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _box_black


def build_default_packing_list(po: dict, options: dict | None = None) -> bytes:
    """Generate a packing list xlsx matching the master reference PDF exactly.

    `po` is the full PO document with `line_items`.
    `options` carries `{carton_dim, pcs_per_box, net_wt_per_carton, gross_wt_per_carton}`.
    """
    options = options or {}
    pcs_per_box = int(options.get("pcs_per_box") or 20)
    net_wt_unit = float(options.get("net_wt_per_carton") or 10.8)
    gross_wt_unit = float(options.get("gross_wt_per_carton") or 12.0)
    carton_dim = options.get("carton_dim") or po.get("carton_dim") or "60x50x30 CMS"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing list"
    ws.views.sheetView[0].showGridLines = True

    # 1. Title Banner (Row 1)
    ws.merge_cells("A1:Q1")
    _set_cell(ws, "A1", "PACKING LIST", bold=True, size=14, align="center", fill=_HEADER_FILL)
    ws.row_dimensions[1].height = 24

    # 2. Vendor Block (A2:F6) & Destination Hub Block (G2:Q6)
    _set_cell(ws, "A2", "VENDOR NAME :", bold=True, size=9)
    ws.merge_cells("B2:F2"); _set_cell(ws, "B2", VENDOR["name"], bold=True, size=9)
    ws.merge_cells("B3:F3"); _set_cell(ws, "B3", VENDOR["address"].split("\n")[0], size=8)
    ws.merge_cells("B4:F4"); _set_cell(ws, "B4", VENDOR["address"].split("\n")[1] if "\n" in VENDOR["address"] else "", size=8)
    ws.merge_cells("B5:F5"); _set_cell(ws, "B5", "", size=8)
    ws.merge_cells("A6:F6"); _set_cell(ws, "A6", f"GSTIN:- {VENDOR['gstin']}", bold=True, size=9)

    client_name = po.get("client_name") or "ZECODE-BANGLORE-2220 ZECODE-BANGLORE-2220"
    client_addr = po.get("client_address") or po.get("shipping_address") or "PLOT NO. 2J/2K, 3RD PHASE KIADB OBEDENAHALLI INDUSTRIAL AREA BANGLORE, KARNATAKA DODDABALLAPUR 561 BENGALURU KARNATAKA 561203"
    client_gstin = po.get("client_gstin") or "29AAACS6995D2ZX"

    ws.merge_cells("G2:Q2"); _set_cell(ws, "G2", "DESTINATION HUB", bold=True, size=9, align="center", fill=_HEADER_FILL)
    ws.merge_cells("G3:Q3"); _set_cell(ws, "G3", client_name, bold=True, size=9, align="center")
    ws.merge_cells("G4:Q4"); _set_cell(ws, "G4", client_addr, size=8, align="center")
    ws.merge_cells("G5:Q5"); _set_cell(ws, "G5", f"GSTIN:- {client_gstin}", bold=True, size=9, align="center")
    ws.merge_cells("G6:Q6"); _set_cell(ws, "G6", "EACHES", size=8, align="center")

    _apply_border_range(ws, 1, 2, 6, 6)
    _apply_border_range(ws, 7, 2, 17, 6)

    # 3. PO Information Rows (Row 8 & Row 9)
    total_qty = int(po.get("total_quantity") or sum(int(li.get("quantity") or 0) for li in po.get("line_items", [])))
    total_cartons = max(1, (total_qty + pcs_per_box - 1) // pcs_per_box)

    _set_cell(ws, "A8", "PO NO", bold=True, size=9)
    ws.merge_cells("B8:D8"); _set_cell(ws, "B8", po.get("po_number", ""), bold=True, size=9)
    _set_cell(ws, "E8", total_qty, bold=True, size=9, align="center")
    _set_cell(ws, "F8", "PCS", bold=True, size=9, align="center")
    _set_cell(ws, "G8", "BOX", bold=True, size=9, align="center")
    _set_cell(ws, "H8", total_cartons, bold=True, size=9, align="center")
    ws.merge_cells("I8:Q8"); _set_cell(ws, "I8", "", size=9)

    _set_cell(ws, "A9", "PO DATE", bold=True, size=9)
    ws.merge_cells("B9:D9"); _set_cell(ws, "B9", po.get("po_date", ""), bold=True, size=9)
    ws.merge_cells("E9:K9"); _set_cell(ws, "E9", "", size=9)
    ws.merge_cells("L9:N9"); _set_cell(ws, "L9", "CARTON DIMENTION", bold=True, size=9, align="center")
    ws.merge_cells("O9:Q9"); _set_cell(ws, "O9", carton_dim, bold=True, size=9, align="center")

    _apply_border_range(ws, 1, 8, 17, 9)

    # 4. Main Packing Table Header (Row 10)
    headers = [
        "SITE CODE", "Style", "Colour", "CTN .NO",
        "36", "37", "38", "39", "40", "41", "42",
        "PCS/CTN", "Per Carton", "TTL CTN", "Total PCS", "NET WEIGHT", "GROSS WEIGHT"
    ]
    for col_idx, h in enumerate(headers, start=1):
        col_let = get_column_letter(col_idx)
        _set_cell(ws, f"{col_let}10", h, bold=True, size=9, align="center", fill=_HEADER_FILL)
    ws.row_dimensions[10].height = 22

    # 5. Group Line Items by (Style, Colour)
    agg: dict[tuple[str, str], dict] = {}
    for li in po.get("line_items", []):
        style = str(li.get("style_code") or "").strip()
        color = str(li.get("color") or "").strip()
        key = (style, color)
        slot = agg.setdefault(key, {
            "style": style, "color": color,
            "by_size": {s: 0 for s in DEFAULT_SIZES}, "total": 0,
        })
        sz = str(li.get("size") or "").strip()
        qty = int(li.get("quantity") or 0)
        if sz in slot["by_size"]:
            slot["by_size"][sz] += qty
        slot["total"] += qty

    site_code = options.get("site_code") or po.get("site_code") or "ZC_BLR-WH"
    start_row = 11
    ctn_seq = 1
    current_row = start_row

    for (style, color), rec in agg.items():
        cartons_here = max(1, (rec["total"] + pcs_per_box - 1) // pcs_per_box)
        ctn_range = f"{ctn_seq}-{ctn_seq + cartons_here - 1}" if cartons_here > 1 else str(ctn_seq)
        row_net_wt = round(cartons_here * net_wt_unit, 3)
        row_gross_wt = round(cartons_here * gross_wt_unit, 3)

        _set_cell(ws, f"A{current_row}", site_code, size=9, align="center")
        _set_cell(ws, f"B{current_row}", style, size=9, bold=True, align="left")
        _set_cell(ws, f"C{current_row}", color, size=9, align="left")
        _set_cell(ws, f"D{current_row}", ctn_range, size=9, align="center")

        for idx, sz in enumerate(DEFAULT_SIZES):
            col_let = get_column_letter(5 + idx)
            sz_qty = rec["by_size"].get(sz, 0)
            _set_cell(ws, f"{col_let}{current_row}", sz_qty if sz_qty > 0 else "", size=9, align="center")

        _set_cell(ws, f"L{current_row}", rec["total"], size=9, align="center")
        _set_cell(ws, f"M{current_row}", pcs_per_box, size=9, align="center")
        _set_cell(ws, f"N{current_row}", cartons_here, size=9, align="center")
        _set_cell(ws, f"O{current_row}", rec["total"], size=9, align="center", bold=True)
        _set_cell(ws, f"P{current_row}", row_net_wt, size=9, align="right", num_format="0.000")
        _set_cell(ws, f"Q{current_row}", row_gross_wt, size=9, align="right", num_format="0.000")

        ws.row_dimensions[current_row].height = 18
        current_row += 1
        ctn_seq += cartons_here

    # Pad table to at least 23 data rows to match master reference PDF structure
    min_data_rows = 23
    end_data_row = max(start_row + min_data_rows - 1, current_row - 1)
    for r in range(current_row, end_data_row + 1):
        _set_cell(ws, f"A{r}", "", size=9)
        _set_cell(ws, f"B{r}", "", size=9)
        _set_cell(ws, f"C{r}", "", size=9)
        _set_cell(ws, f"D{r}", "", size=9)
        for idx in range(7):
            _set_cell(ws, f"{get_column_letter(5 + idx)}{r}", "", size=9)
        _set_cell(ws, f"L{r}", "", size=9)
        _set_cell(ws, f"M{r}", "", size=9)
        _set_cell(ws, f"N{r}", 0, size=9, align="center")
        _set_cell(ws, f"O{r}", 0, size=9, align="center")
        _set_cell(ws, f"P{r}", 0.000, size=9, align="right", num_format="0.000")
        _set_cell(ws, f"Q{r}", 0.000, size=9, align="right", num_format="0.000")
        ws.row_dimensions[r].height = 16

    gt_row = end_data_row + 1

    # 6. Grand Total Row
    ws.merge_cells(f"A{gt_row}:D{gt_row}")
    _set_cell(ws, f"A{gt_row}", "GRAND TOTAL", bold=True, size=9, align="center", fill=_HEADER_FILL)

    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(5 + idx)
        _set_cell(ws, f"{col_let}{gt_row}", f"=SUM({col_let}{start_row}:{col_let}{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)

    _set_cell(ws, f"L{gt_row}", "", fill=_HEADER_FILL)
    _set_cell(ws, f"M{gt_row}", "", fill=_HEADER_FILL)
    _set_cell(ws, f"N{gt_row}", f"=SUM(N{start_row}:N{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"O{gt_row}", f"=SUM(O{start_row}:O{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"P{gt_row}", f"=SUM(P{start_row}:P{end_data_row})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")
    _set_cell(ws, f"Q{gt_row}", f"=SUM(Q{start_row}:Q{end_data_row})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")
    ws.row_dimensions[gt_row].height = 20

    # 7. Order Summary & Signature Section
    summary_start_row = gt_row + 3

    # Order Summary Box (Left)
    ws.merge_cells(f"B{summary_start_row}:D{summary_start_row+3}")
    _set_cell(ws, f"B{summary_start_row}", "ORDER SUMMERY", bold=True, size=10, align="center")

    _set_cell(ws, f"E{summary_start_row}", "Size", bold=True, size=9, align="center", fill=_HEADER_FILL)
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)  # F..L
        _set_cell(ws, f"{col_let}{summary_start_row}", sz, bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"M{summary_start_row}", "TOTAL", bold=True, size=9, align="center", fill=_HEADER_FILL)

    # Calculate size order quantities for row 1
    size_order_map = {s: 0 for s in DEFAULT_SIZES}
    for li in po.get("line_items", []):
        sz = str(li.get("size") or "").strip()
        if sz in size_order_map:
            size_order_map[sz] += int(li.get("quantity") or 0)

    # Order Qty Row
    _set_cell(ws, f"E{summary_start_row+1}", "Order Qty", bold=True, size=9, align="center")
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+1}", size_order_map[sz], size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+1}", f"=SUM(F{summary_start_row+1}:L{summary_start_row+1})", bold=True, size=9, align="center")

    # Pack Qty Row
    _set_cell(ws, f"E{summary_start_row+2}", "Pack Qty", bold=True, size=9, align="center")
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)
        gt_size_col = get_column_letter(5 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+2}", f"={gt_size_col}{gt_row}", size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+2}", f"=SUM(F{summary_start_row+2}:L{summary_start_row+2})", bold=True, size=9, align="center")

    # Excess/Short Row
    _set_cell(ws, f"E{summary_start_row+3}", "Excss/Short", bold=True, size=9, align="center")
    for idx in range(len(DEFAULT_SIZES)):
        col_let = get_column_letter(6 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+3}", f"={col_let}{summary_start_row+2}-{col_let}{summary_start_row+1}", size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+3}", f"=SUM(F{summary_start_row+3}:L{summary_start_row+3})", bold=True, size=9, align="center")

    # Excess/Short % Row
    _set_cell(ws, f"E{summary_start_row+4}", "Excss/Short %", bold=True, size=9, align="center")
    for idx in range(len(DEFAULT_SIZES)):
        col_let = get_column_letter(6 + idx)
        # Protect against #DIV/0! when Order Qty is 0
        formula_str = f"=IF({col_let}{summary_start_row+1}=0, 0, {col_let}{summary_start_row+3}/{col_let}{summary_start_row+1})"
        _set_cell(ws, f"{col_let}{summary_start_row+4}", formula_str, size=9, align="center", num_format="0.00%")
    tot_pct_formula = f"=IF(M{summary_start_row+1}=0, 0, M{summary_start_row+3}/M{summary_start_row+1})"
    _set_cell(ws, f"M{summary_start_row+4}", tot_pct_formula, bold=True, size=9, align="center", num_format="0.00%")

    _apply_border_range(ws, 2, summary_start_row, 13, summary_start_row + 4)

    # Authorised Signatory Box (Right)
    sig_start_col = 14  # N
    sig_start_row = summary_start_row + 3
    ws.merge_cells(f"N{sig_start_row}:Q{sig_start_row+4}")
    _set_cell(ws, f"N{sig_start_row}", "AUTHORISED SIGNATORY", bold=True, size=9, align="center")
    cell_sig = ws[f"N{sig_start_row}"]
    cell_sig.alignment = Alignment(horizontal="center", vertical="top")

    # Shipping & Notes Block
    shipping_row = summary_start_row + 6
    disp_text = f"DISPATCH: {options.get('dispatch_date', '')} | TRANSPORTER: {options.get('transporter', '')} | VEHICLE: {options.get('vehicle_no', '')} | DRIVER: {options.get('driver_name', '')} {options.get('driver_phone', '')}"
    dest_text = f"DESTINATION: {options.get('destination', '')} | PORT: {options.get('port', '')}"
    notes_text = f"NOTES: {options.get('notes', '')}"
    if any([options.get("transporter"), options.get("vehicle_no"), options.get("dispatch_date"), options.get("notes"), options.get("destination")]):
        ws.merge_cells(f"A{shipping_row}:Q{shipping_row}")
        _set_cell(ws, f"A{shipping_row}", disp_text, size=8, bold=True)
        ws.merge_cells(f"A{shipping_row+1}:Q{shipping_row+1}")
        _set_cell(ws, f"A{shipping_row+1}", f"{dest_text} | {notes_text}", size=8)

    # 8. Column Widths & Print Settings
    col_widths = {
        "A": 14, "B": 22, "C": 14, "D": 12,
        "E": 6,  "F": 6,  "G": 6,  "H": 6,  "I": 6,  "J": 6,  "K": 6,
        "L": 10, "M": 11, "N": 10, "O": 11, "P": 13, "Q": 14,
    }
    for col_let, width in col_widths.items():
        ws.column_dimensions[col_let].width = width

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_dispatch_packing_list(cartons: list[dict], po: dict, invoice_no: str, options: dict | None = None) -> bytes:
    """Generate packing list xlsx matching the master reference PDF from actual carton records."""
    options = options or {}
    net_wt_unit = float(options.get("net_wt_per_carton") or 10.8)
    gross_wt_unit = float(options.get("gross_wt_per_carton") or 12.0)
    carton_dim = options.get("carton_dim") or po.get("carton_dim") or "60x50x30 CMS"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing list"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:Q1")
    _set_cell(ws, "A1", "PACKING LIST", bold=True, size=14, align="center", fill=_HEADER_FILL)
    ws.row_dimensions[1].height = 24

    # Vendor & Destination Blocks
    _set_cell(ws, "A2", "VENDOR NAME :", bold=True, size=9)
    ws.merge_cells("B2:F2"); _set_cell(ws, "B2", VENDOR["name"], bold=True, size=9)
    ws.merge_cells("B3:F3"); _set_cell(ws, "B3", VENDOR["address"].split("\n")[0], size=8)
    ws.merge_cells("B4:F4"); _set_cell(ws, "B4", VENDOR["address"].split("\n")[1] if "\n" in VENDOR["address"] else "", size=8)
    ws.merge_cells("B5:F5"); _set_cell(ws, "B5", "", size=8)
    ws.merge_cells("A6:F6"); _set_cell(ws, "A6", f"GSTIN:- {VENDOR['gstin']}", bold=True, size=9)

    client_name = po.get("client_name") or "ZECODE-BANGLORE-2220 ZECODE-BANGLORE-2220"
    client_addr = po.get("client_address") or po.get("shipping_address") or "PLOT NO. 2J/2K, 3RD PHASE KIADB OBEDENAHALLI INDUSTRIAL AREA BANGLORE, KARNATAKA DODDABALLAPUR 561 BENGALURU KARNATAKA 561203"
    client_gstin = po.get("client_gstin") or "29AAACS6995D2ZX"

    ws.merge_cells("G2:Q2"); _set_cell(ws, "G2", "DESTINATION HUB", bold=True, size=9, align="center", fill=_HEADER_FILL)
    ws.merge_cells("G3:Q3"); _set_cell(ws, "G3", client_name, bold=True, size=9, align="center")
    ws.merge_cells("G4:Q4"); _set_cell(ws, "G4", client_addr, size=8, align="center")
    ws.merge_cells("G5:Q5"); _set_cell(ws, "G5", f"GSTIN:- {client_gstin}", bold=True, size=9, align="center")
    ws.merge_cells("G6:Q6"); _set_cell(ws, "G6", "EACHES", size=8, align="center")

    _apply_border_range(ws, 1, 2, 6, 6)
    _apply_border_range(ws, 7, 2, 17, 6)

    total_qty = sum(c.get("qty", 0) for c in cartons)
    total_cartons = len(cartons)

    _set_cell(ws, "A8", "PO NO", bold=True, size=9)
    ws.merge_cells("B8:D8"); _set_cell(ws, "B8", po.get("po_number", ""), bold=True, size=9)
    _set_cell(ws, "E8", total_qty, bold=True, size=9, align="center")
    _set_cell(ws, "F8", "PCS", bold=True, size=9, align="center")
    _set_cell(ws, "G8", "BOX", bold=True, size=9, align="center")
    _set_cell(ws, "H8", total_cartons, bold=True, size=9, align="center")
    ws.merge_cells("I8:Q8"); _set_cell(ws, "I8", f"INVOICE NO: {invoice_no}", bold=True, size=9, align="center")

    _set_cell(ws, "A9", "PO DATE", bold=True, size=9)
    ws.merge_cells("B9:D9"); _set_cell(ws, "B9", po.get("po_date", ""), bold=True, size=9)
    ws.merge_cells("E9:K9"); _set_cell(ws, "E9", "", size=9)
    ws.merge_cells("L9:N9"); _set_cell(ws, "L9", "CARTON DIMENTION", bold=True, size=9, align="center")
    ws.merge_cells("O9:Q9"); _set_cell(ws, "O9", carton_dim, bold=True, size=9, align="center")

    _apply_border_range(ws, 1, 8, 17, 9)

    # Main Packing Table Header
    headers = [
        "SITE CODE", "Style", "Colour", "CTN .NO",
        "36", "37", "38", "39", "40", "41", "42",
        "PCS/CTN", "Per Carton", "TTL CTN", "Total PCS", "NET WEIGHT", "GROSS WEIGHT"
    ]
    for col_idx, h in enumerate(headers, start=1):
        col_let = get_column_letter(col_idx)
        _set_cell(ws, f"{col_let}10", h, bold=True, size=9, align="center", fill=_HEADER_FILL)
    ws.row_dimensions[10].height = 22

    # Group cartons by (style_code, color, size, qty)
    sorted_cartons = sorted(cartons, key=lambda c: c.get("box_number") or 0)
    groups = []
    curr = None
    for c in sorted_cartons:
        key = (c.get("style_code"), c.get("color"), c.get("size"), c.get("qty"))
        box_num = c.get("box_number")
        if curr and curr["key"] == key:
            curr["cartons"].append(c)
            curr["box_numbers"].append(box_num)
        else:
            if curr:
                groups.append(curr)
            curr = {"key": key, "cartons": [c], "box_numbers": [box_num]}
    if curr:
        groups.append(curr)

    site_code = options.get("site_code") or po.get("site_code") or "ZC_BLR-WH"
    start_row = 11
    current_row = start_row

    for g in groups:
        style_code, color, size_val, pcs_per_box = g["key"]
        num_boxes = len(g["cartons"])
        boxes = [b for b in g["box_numbers"] if b is not None]
        if boxes:
            min_b, max_b = min(boxes), max(boxes)
            ctn_range = f"{min_b}-{max_b}" if min_b != max_b else str(min_b)
        else:
            ctn_range = "1"

        row_total_pcs = num_boxes * (pcs_per_box or 20)
        row_net_wt = round(num_boxes * net_wt_unit, 3)
        row_gross_wt = round(num_boxes * gross_wt_unit, 3)

        _set_cell(ws, f"A{current_row}", site_code, size=9, align="center")
        _set_cell(ws, f"B{current_row}", style_code or "", size=9, bold=True, align="left")
        _set_cell(ws, f"C{current_row}", color or "", size=9, align="left")
        _set_cell(ws, f"D{current_row}", ctn_range, size=9, align="center")

        sz_str = str(size_val or "").strip()
        for idx, sz in enumerate(DEFAULT_SIZES):
            col_let = get_column_letter(5 + idx)
            val = row_total_pcs if sz == sz_str else ""
            _set_cell(ws, f"{col_let}{current_row}", val, size=9, align="center")

        _set_cell(ws, f"L{current_row}", row_total_pcs, size=9, align="center")
        _set_cell(ws, f"M{current_row}", pcs_per_box or 20, size=9, align="center")
        _set_cell(ws, f"N{current_row}", num_boxes, size=9, align="center")
        _set_cell(ws, f"O{current_row}", row_total_pcs, size=9, align="center", bold=True)
        _set_cell(ws, f"P{current_row}", row_net_wt, size=9, align="right", num_format="0.000")
        _set_cell(ws, f"Q{current_row}", row_gross_wt, size=9, align="right", num_format="0.000")

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Pad data rows
    min_data_rows = 23
    end_data_row = max(start_row + min_data_rows - 1, current_row - 1)
    for r in range(current_row, end_data_row + 1):
        _set_cell(ws, f"A{r}", "", size=9)
        _set_cell(ws, f"B{r}", "", size=9)
        _set_cell(ws, f"C{r}", "", size=9)
        _set_cell(ws, f"D{r}", "", size=9)
        for idx in range(7):
            _set_cell(ws, f"{get_column_letter(5 + idx)}{r}", "", size=9)
        _set_cell(ws, f"L{r}", "", size=9)
        _set_cell(ws, f"M{r}", "", size=9)
        _set_cell(ws, f"N{r}", 0, size=9, align="center")
        _set_cell(ws, f"O{r}", 0, size=9, align="center")
        _set_cell(ws, f"P{r}", 0.000, size=9, align="right", num_format="0.000")
        _set_cell(ws, f"Q{r}", 0.000, size=9, align="right", num_format="0.000")
        ws.row_dimensions[r].height = 16

    gt_row = end_data_row + 1

    # Grand Total Row
    ws.merge_cells(f"A{gt_row}:D{gt_row}")
    _set_cell(ws, f"A{gt_row}", "GRAND TOTAL", bold=True, size=9, align="center", fill=_HEADER_FILL)

    for idx in range(7):
        col_let = get_column_letter(5 + idx)
        _set_cell(ws, f"{col_let}{gt_row}", f"=SUM({col_let}{start_row}:{col_let}{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)

    _set_cell(ws, f"L{gt_row}", "", fill=_HEADER_FILL)
    _set_cell(ws, f"M{gt_row}", "", fill=_HEADER_FILL)
    _set_cell(ws, f"N{gt_row}", f"=SUM(N{start_row}:N{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"O{gt_row}", f"=SUM(O{start_row}:O{end_data_row})", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"P{gt_row}", f"=SUM(P{start_row}:P{end_data_row})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")
    _set_cell(ws, f"Q{gt_row}", f"=SUM(Q{start_row}:Q{end_data_row})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")
    ws.row_dimensions[gt_row].height = 20

    # Order Summary & Signature Section
    summary_start_row = gt_row + 3

    ws.merge_cells(f"B{summary_start_row}:D{summary_start_row+3}")
    _set_cell(ws, f"B{summary_start_row}", "ORDER SUMMERY", bold=True, size=10, align="center")

    _set_cell(ws, f"E{summary_start_row}", "Size", bold=True, size=9, align="center", fill=_HEADER_FILL)
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row}", sz, bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"M{summary_start_row}", "TOTAL", bold=True, size=9, align="center", fill=_HEADER_FILL)

    size_order_map = {s: 0 for s in DEFAULT_SIZES}
    for li in po.get("line_items", []):
        sz = str(li.get("size") or "").strip()
        if sz in size_order_map:
            size_order_map[sz] += int(li.get("quantity") or 0)

    # Order Qty Row
    _set_cell(ws, f"E{summary_start_row+1}", "Order Qty", bold=True, size=9, align="center")
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+1}", size_order_map[sz], size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+1}", f"=SUM(F{summary_start_row+1}:L{summary_start_row+1})", bold=True, size=9, align="center")

    # Pack Qty Row
    _set_cell(ws, f"E{summary_start_row+2}", "Pack Qty", bold=True, size=9, align="center")
    for idx, sz in enumerate(DEFAULT_SIZES):
        col_let = get_column_letter(6 + idx)
        gt_size_col = get_column_letter(5 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+2}", f"={gt_size_col}{gt_row}", size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+2}", f"=SUM(F{summary_start_row+2}:L{summary_start_row+2})", bold=True, size=9, align="center")

    # Excess/Short Row
    _set_cell(ws, f"E{summary_start_row+3}", "Excss/Short", bold=True, size=9, align="center")
    for idx in range(len(DEFAULT_SIZES)):
        col_let = get_column_letter(6 + idx)
        _set_cell(ws, f"{col_let}{summary_start_row+3}", f"={col_let}{summary_start_row+2}-{col_let}{summary_start_row+1}", size=9, align="center")
    _set_cell(ws, f"M{summary_start_row+3}", f"=SUM(F{summary_start_row+3}:L{summary_start_row+3})", bold=True, size=9, align="center")

    # Excess/Short % Row
    _set_cell(ws, f"E{summary_start_row+4}", "Excss/Short %", bold=True, size=9, align="center")
    for idx in range(len(DEFAULT_SIZES)):
        col_let = get_column_letter(6 + idx)
        formula_str = f"=IF({col_let}{summary_start_row+1}=0, 0, {col_let}{summary_start_row+3}/{col_let}{summary_start_row+1})"
        _set_cell(ws, f"{col_let}{summary_start_row+4}", formula_str, size=9, align="center", num_format="0.00%")
    tot_pct_formula = f"=IF(M{summary_start_row+1}=0, 0, M{summary_start_row+3}/M{summary_start_row+1})"
    _set_cell(ws, f"M{summary_start_row+4}", tot_pct_formula, bold=True, size=9, align="center", num_format="0.00%")

    _apply_border_range(ws, 2, summary_start_row, 13, summary_start_row + 4)

    # Authorised Signatory Box
    sig_start_row = summary_start_row + 3
    ws.merge_cells(f"N{sig_start_row}:Q{sig_start_row+4}")
    _set_cell(ws, f"N{sig_start_row}", "AUTHORISED SIGNATORY", bold=True, size=9, align="center")
    cell_sig = ws[f"N{sig_start_row}"]
    cell_sig.alignment = Alignment(horizontal="center", vertical="top")

    col_widths = {
        "A": 14, "B": 22, "C": 14, "D": 12,
        "E": 6,  "F": 6,  "G": 6,  "H": 6,  "I": 6,  "J": 6,  "K": 6,
        "L": 10, "M": 11, "N": 10, "O": 11, "P": 13, "Q": 14,
    }
    for col_let, width in col_widths.items():
        ws.column_dimensions[col_let].width = width

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_packing_list_pdf(po: dict, options: dict | None = None, cartons: list[dict] | None = None, invoice_no: str = "") -> bytes:
    """Generate a vector PDF packing list matching the visual layout of the master reference PDF using ReportLab."""
    options = options or {}
    pcs_per_box = int(options.get("pcs_per_box") or 20)
    net_wt_unit = float(options.get("net_wt_per_carton") or 10.8)
    gross_wt_unit = float(options.get("gross_wt_per_carton") or 12.0)
    carton_dim = options.get("carton_dim") or po.get("carton_dim") or "60x50x30 CMS"
    site_code = options.get("site_code") or po.get("site_code") or "ZC_BLR-WH"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.25 * inch,
        rightMargin=0.25 * inch,
        topMargin=0.25 * inch,
        bottomMargin=0.25 * inch,
    )

    styles_sheet = getSampleStyleSheet()
    normal_style = styles_sheet["Normal"]
    normal_style.fontSize = 7
    normal_style.leading = 9

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=normal_style,
        fontName="Helvetica-Bold",
        fontSize=11,
        alignment=1, # Center
    )

    bold_style = ParagraphStyle(
        "BoldStyle",
        parent=normal_style,
        fontName="Helvetica-Bold",
    )

    elements = []

    # 1. Title Header
    header_table = Table(
        [[Paragraph("PACKING LIST", title_style)]],
        colWidths=[11.2 * inch],
        rowHeights=[22]
    )
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D9D9D9')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)

    # 2. Vendor & Destination Info
    vendor_addr = VENDOR['address'].replace('\n', ' ')
    vendor_text = f"<b>VENDOR NAME :</b> {VENDOR['name']}<br/>{vendor_addr}<br/><b>GSTIN:-</b> {VENDOR['gstin']}"
    client_name = po.get("client_name") or "ZECODE-BANGLORE-2220 ZECODE-BANGLORE-2220"
    client_addr = po.get("client_address") or po.get("shipping_address") or "PLOT NO. 2J/2K, 3RD PHASE KIADB OBEDENAHALLI INDUSTRIAL AREA BANGLORE, KARNATAKA DODDABALLAPUR 561 BENGALURU KARNATAKA 561203"
    client_gstin = po.get("client_gstin") or "29AAACS6995D2ZX"
    dest_text = f"<b>DESTINATION HUB</b><br/><b>{client_name}</b><br/>{client_addr}<br/><b>GSTIN:-</b> {client_gstin}<br/>EACHES"

    info_table = Table(
        [[Paragraph(vendor_text, normal_style), Paragraph(dest_text, normal_style)]],
        colWidths=[5.6 * inch, 5.6 * inch]
    )
    info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (0, -1), 0.5, colors.black),
        ('BOX', (1, 0), (1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)

    # 3. PO Meta Row
    po_num = po.get("po_number", "")
    po_dt = po.get("po_date", "")
    total_qty = sum(c.get("qty", 0) for c in cartons) if cartons else int(po.get("total_quantity") or sum(int(li.get("quantity") or 0) for li in po.get("line_items", [])))
    total_cartons = len(cartons) if cartons else max(1, (total_qty + pcs_per_box - 1) // pcs_per_box)

    meta_row1 = [
        Paragraph(f"<b>PO NO</b>", normal_style),
        Paragraph(f"<b>{po_num}</b>", normal_style),
        Paragraph(f"<b>{total_qty}</b>", normal_style),
        Paragraph("<b>PCS</b>", normal_style),
        Paragraph("<b>BOX</b>", normal_style),
        Paragraph(f"<b>{total_cartons}</b>", normal_style),
        Paragraph(f"<b>INVOICE NO: {invoice_no}</b>" if invoice_no else "", normal_style),
        Paragraph("<b>CARTON DIMENTION</b>", normal_style),
        Paragraph(f"<b>{carton_dim}</b>", normal_style)
    ]
    meta_row2 = [
        Paragraph("<b>PO DATE</b>", normal_style),
        Paragraph(f"<b>{po_dt}</b>", normal_style),
        "", "", "", "", "", "", ""
    ]

    meta_table = Table(
        [meta_row1, meta_row2],
        colWidths=[1.0*inch, 1.8*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 2.0*inch, 2.0*inch, 2.0*inch]
    )
    meta_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('SPAN', (7, 0), (7, 1)),
        ('SPAN', (8, 0), (8, 1)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(meta_table)

    # 4. Main Packing Table Data Preparation
    table_data = [[
        "SITE CODE", "Style", "Colour", "CTN .NO",
        "36", "37", "38", "39", "40", "41", "42",
        "PCS/CTN", "Per Carton", "TTL CTN", "Total PCS", "NET WEIGHT", "GROSS WEIGHT"
    ]]

    size_totals = {s: 0 for s in DEFAULT_SIZES}
    grand_ttl_ctn = 0
    grand_total_pcs = 0
    grand_net_wt = 0.0
    grand_gross_wt = 0.0

    if cartons:
        sorted_cartons = sorted(cartons, key=lambda c: c.get("box_number") or 0)
        groups = []
        curr = None
        for c in sorted_cartons:
            key = (c.get("style_code"), c.get("color"), c.get("size"), c.get("qty"))
            box_num = c.get("box_number")
            if curr and curr["key"] == key:
                curr["cartons"].append(c)
                curr["box_numbers"].append(box_num)
            else:
                if curr: groups.append(curr)
                curr = {"key": key, "cartons": [c], "box_numbers": [box_num]}
        if curr: groups.append(curr)

        for g in groups:
            st, col, sz_val, box_qty = g["key"]
            n_boxes = len(g["cartons"])
            boxes = [b for b in g["box_numbers"] if b is not None]
            c_range = f"{min(boxes)}-{max(boxes)}" if boxes and min(boxes) != max(boxes) else (str(boxes[0]) if boxes else "1")
            r_pcs = n_boxes * (box_qty or 20)
            r_net = n_boxes * net_wt_unit
            r_gross = n_boxes * gross_wt_unit

            sz_str = str(sz_val or "").strip()
            row = [site_code, st or "", col or "", c_range]
            for sz in DEFAULT_SIZES:
                if sz == sz_str:
                    row.append(str(r_pcs))
                    size_totals[sz] += r_pcs
                else:
                    row.append("")
            row.extend([str(r_pcs), str(box_qty or 20), str(n_boxes), str(r_pcs), f"{r_net:.3f}", f"{r_gross:.3f}"])
            table_data.append(row)

            grand_ttl_ctn += n_boxes
            grand_total_pcs += r_pcs
            grand_net_wt += r_net
            grand_gross_wt += r_gross
    else:
        agg: dict[tuple[str, str], dict] = {}
        for li in po.get("line_items", []):
            st = str(li.get("style_code") or "").strip()
            co = str(li.get("color") or "").strip()
            slot = agg.setdefault((st, co), {"style": st, "color": co, "by_size": {s: 0 for s in DEFAULT_SIZES}, "total": 0})
            sz = str(li.get("size") or "").strip()
            q = int(li.get("quantity") or 0)
            if sz in slot["by_size"]: slot["by_size"][sz] += q
            slot["total"] += q

        ctn_seq = 1
        for (st, co), rec in agg.items():
            n_boxes = max(1, (rec["total"] + pcs_per_box - 1) // pcs_per_box)
            c_range = f"{ctn_seq}-{ctn_seq + n_boxes - 1}" if n_boxes > 1 else str(ctn_seq)
            r_net = n_boxes * net_wt_unit
            r_gross = n_boxes * gross_wt_unit

            row = [site_code, st, co, c_range]
            for sz in DEFAULT_SIZES:
                cnt = rec["by_size"].get(sz, 0)
                if cnt > 0:
                    row.append(str(cnt))
                    size_totals[sz] += cnt
                else:
                    row.append("")
            row.extend([str(rec["total"]), str(pcs_per_box), str(n_boxes), str(rec["total"]), f"{r_net:.3f}", f"{r_gross:.3f}"])
            table_data.append(row)

            grand_ttl_ctn += n_boxes
            grand_total_pcs += rec["total"]
            grand_net_wt += r_net
            grand_gross_wt += r_gross
            ctn_seq += n_boxes

    # Pad table rows
    while len(table_data) < 20:
        table_data.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "0", "0", "0.000", "0.000"])

    # Grand Total Row
    gt_row = ["GRAND TOTAL", "", "", ""]
    for sz in DEFAULT_SIZES:
        gt_row.append(str(size_totals[sz]))
    gt_row.extend(["", "", str(grand_ttl_ctn), str(grand_total_pcs), f"{grand_net_wt:.3f}", f"{grand_gross_wt:.3f}"])
    table_data.append(gt_row)

    col_w = [0.8*inch, 1.4*inch, 0.9*inch, 0.7*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.7*inch, 0.9*inch, 0.9*inch]
    main_table = Table(table_data, colWidths=col_w)
    main_table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('SPAN', (0, len(table_data)-1), (3, len(table_data)-1)),
        ('BACKGROUND', (0, len(table_data)-1), (-1, len(table_data)-1), colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0, len(table_data)-1), (-1, len(table_data)-1), 'Helvetica-Bold'),
    ]
    main_table.setStyle(TableStyle(main_table_style))
    elements.append(main_table)
    elements.append(Spacer(1, 10))

    # 5. Order Summary & Signature Table
    order_qty_map = {s: 0 for s in DEFAULT_SIZES}
    for li in po.get("line_items", []):
        sz = str(li.get("size") or "").strip()
        if sz in order_qty_map:
            order_qty_map[sz] += int(li.get("quantity") or 0)

    total_order_qty = sum(order_qty_map.values())
    total_pack_qty = grand_total_pcs
    total_diff = total_pack_qty - total_order_qty
    total_diff_pct = (total_diff / total_order_qty * 100) if total_order_qty > 0 else 0.0

    summary_rows = [
        ["ORDER SUMMERY", "Size", "36", "37", "38", "39", "40", "41", "42", "TOTAL", "", "AUTHORISED SIGNATORY"],
        ["", "Order Qty"] + [str(order_qty_map[s]) for s in DEFAULT_SIZES] + [str(total_order_qty), "", ""],
        ["", "Pack Qty"] + [str(size_totals[s]) for s in DEFAULT_SIZES] + [str(total_pack_qty), "", ""],
        ["", "Excss/Short"] + [str(size_totals[s] - order_qty_map[s]) for s in DEFAULT_SIZES] + [str(total_diff), "", ""],
        ["", "Excss/Short %"] + [f"{((size_totals[s] - order_qty_map[s])/order_qty_map[s]*100 if order_qty_map[s]>0 else 0.0):.2f}%" for s in DEFAULT_SIZES] + [f"{total_diff_pct:.2f}%", "", ""]
    ]

    summary_w = [1.2*inch, 0.9*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.7*inch, 2.0*inch, 3.6*inch]
    summary_table = Table(summary_rows, colWidths=summary_w, rowHeights=[14, 14, 14, 14, 14])
    summary_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (0, 4)), # ORDER SUMMERY left box
        ('SPAN', (11, 0), (11, 4)), # AUTHORISED SIGNATORY right box
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('VALIGN', (11, 0), (11, 4), 'TOP'),
        ('BACKGROUND', (0, 0), (0, 4), colors.HexColor('#D9D9D9')),
        ('BACKGROUND', (1, 0), (9, 0), colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('GRID', (0, 0), (9, 4), 0.5, colors.black),
        ('BOX', (11, 0), (11, 4), 0.5, colors.black),
    ]))
    elements.append(summary_table)

    doc.build(elements)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Custom template substitution ----------------------------------------------
# ---------------------------------------------------------------------------
_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z_]+)\s*\}\}")


def build_from_template(template_bytes: bytes, po: dict, options: dict | None = None, cartons: list[dict] | None = None) -> bytes:
    """Fill an uploaded template file with placeholders."""
    options = options or {}
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    today = datetime.now().strftime("%d.%m.%Y")
    total_qty = int(po.get("total_quantity") or sum(int(li.get("quantity") or 0) for li in po.get("line_items", [])))
    pcs_per_box = int(options.get("pcs_per_box") or 20)

    if cartons is not None:
        total_cartons = len(cartons)
    else:
        total_cartons = max(1, (total_qty + pcs_per_box - 1) // pcs_per_box)

    scalars = {
        "po_number": po.get("po_number", ""),
        "po_date": po.get("po_date", ""),
        "client_name": po.get("client_name", ""),
        "client_address": po.get("client_address", "") or po.get("shipping_address", ""),
        "client_gstin": po.get("client_gstin", ""),
        "vendor_name": VENDOR["name"],
        "vendor_address": VENDOR["address"],
        "vendor_gstin": VENDOR["gstin"],
        "carton_dim": options.get("carton_dim") or "60x50x30 CMS",
        "total_pcs": str(total_qty),
        "total_cartons": str(total_cartons),
        "date": today,
    }

    for ws in wb.worksheets:
        _expand_lines(ws, po, options, cartons=cartons)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    def _sub(m):
                        k = m.group(1).lower()
                        return str(scalars.get(k, m.group(0)))
                    cell.value = _PLACEHOLDER_RE.sub(_sub, cell.value)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _classify_header_simple(header: str) -> str:
    """Classify a table header string into known packing-list fields or size columns."""
    h = (header or "").strip().lower()
    if h in ["style", "style code", "style_code", "style no", "style_no", "model"]:
        return "style"
    if h in ["colour", "color", "shade"]:
        return "color"
    if h in ["size", "sizes"]:
        return "size"
    if h in ["qty", "quantity", "pairs", "total qty", "total pairs", "total"]:
        return "quantity"
    if h in ["ctn no", "ctn", "carton", "carton no", "box no", "box", "carton range"]:
        return "ctn_no"
    if h in ["ean", "ean code", "ean_code", "barcode", "upc", "sku barcode"]:
        return "ean"
    if h in ["mrp", "rate", "price", "unit price"]:
        return "price"
    # Check if header is a numeric size like "39", "42", "8", "8.5"
    try:
        float(str(header).strip())
        return f"size_col:{str(header).strip()}"
    except (ValueError, AttributeError):
        pass
    return "other"


def _expand_lines(ws, po: dict, options: dict | None = None, cartons: list[dict] | None = None):
    """Expand {{lines}} marker in a worksheet into tabular line item or carton rows."""
    options = options or {}
    marker_row = None
    marker_col = None
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        for c_idx, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str) and "{{lines}}" in cell.value:
                marker_row = r_idx
                marker_col = c_idx
                break
        if marker_row:
            break

    if not marker_row:
        return

    # Find header row (row above marker_row)
    header_row_idx = marker_row - 1 if marker_row > 1 else 1
    col_mapping = {}
    for c_idx, cell in enumerate(ws[header_row_idx], start=1):
        val = str(cell.value or "").strip()
        cls = _classify_header_simple(val)
        col_mapping[c_idx] = cls

    # Sample style from marker cell
    sample_cell = ws.cell(row=marker_row, column=marker_col)
    s_font = copy(sample_cell.font) if sample_cell.font else None
    s_align = copy(sample_cell.alignment) if sample_cell.alignment else None
    s_border = copy(sample_cell.border) if sample_cell.border else None

    # Clear marker cell
    sample_cell.value = ""

    if cartons:
        grouped_cartons = []
        for c in sorted(cartons, key=lambda x: x.get("box_number") or 0):
            sc = c.get("style_code") or ""
            col = c.get("color") or ""
            sz = str(c.get("size") or "")
            ean = c.get("ean_code") or ""
            q = c.get("qty") or 0
            box_num = c.get("box_number") or 1
            if grouped_cartons and grouped_cartons[-1]["key"] == (sc, col, sz, ean, q) and grouped_cartons[-1]["end_box"] + 1 == box_num:
                grouped_cartons[-1]["count"] += 1
                grouped_cartons[-1]["end_box"] = box_num
                grouped_cartons[-1]["total_pairs"] += q
            else:
                grouped_cartons.append({
                    "key": (sc, col, sz, ean, q),
                    "style_code": sc, "color": col, "size": sz, "ean_code": ean, "qty_per_box": q,
                    "count": 1, "start_box": box_num, "end_box": box_num, "total_pairs": q
                })

        for i, grp in enumerate(grouped_cartons):
            cur_row = marker_row + i
            if grp["start_box"] == grp["end_box"]:
                box_str = str(grp["start_box"])
            else:
                box_str = f"{grp['start_box']} - {grp['end_box']}"

            for c_idx, cls in col_mapping.items():
                cell = ws.cell(row=cur_row, column=c_idx)
                if s_font: cell.font = copy(s_font)
                if s_align: cell.alignment = copy(s_align)
                if s_border: cell.border = copy(s_border)

                if cls == "style":
                    cell.value = grp["style_code"]
                elif cls == "color":
                    cell.value = grp["color"]
                elif cls == "size":
                    cell.value = grp["size"]
                elif cls == "ean":
                    cell.value = grp.get("ean_code") or ""
                elif cls == "quantity":
                    cell.value = grp["total_pairs"]
                elif cls == "ctn_no":
                    cell.value = box_str
                elif cls.startswith("size_col:"):
                    sz = cls.split(":", 1)[1]
                    cell.value = grp["total_pairs"] if sz == grp["size"] else ""
    else:
        line_groups = {}
        for li in po.get("line_items", []):
            sc = li.get("style_code") or ""
            col = li.get("color") or ""
            key = (sc, col)
            if key not in line_groups:
                line_groups[key] = {
                    "style_code": sc, "color": col, "sizes": {}, "total_qty": 0
                }
            sz = str(li.get("size") or "")
            qty = int(li.get("quantity") or 0)
            line_groups[key]["sizes"][sz] = line_groups[key]["sizes"].get(sz, 0) + qty
            line_groups[key]["total_qty"] += qty

        for i, (key, grp) in enumerate(line_groups.items()):
            cur_row = marker_row + i
            for c_idx, cls in col_mapping.items():
                cell = ws.cell(row=cur_row, column=c_idx)
                if s_font: cell.font = copy(s_font)
                if s_align: cell.alignment = copy(s_align)
                if s_border: cell.border = copy(s_border)

                if cls == "style":
                    cell.value = grp["style_code"]
                elif cls == "color":
                    cell.value = grp["color"]
                elif cls == "quantity":
                    cell.value = grp["total_qty"]
                elif cls.startswith("size_col:"):
                    sz = cls.split(":", 1)[1]
                    cell.value = grp["sizes"].get(sz, "")


def build_carton_list_xlsx(cartons: list[dict], po: dict, invoice_no: str = "", options: dict | None = None) -> bytes:
    """Generate a simplified carton list xlsx spreadsheet."""
    options = options or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carton list"

    ws.merge_cells("A1:G1")
    _set_cell(ws, "A1", f"CARTON LIST - PO: {po.get('po_number', '')} {f'INVOICE: {invoice_no}' if invoice_no else ''}", bold=True, size=12, align="center", fill=_HEADER_FILL)

    headers = ["Carton No", "Style", "Colour", "Size", "Quantity (Pairs)", "Net Wt (Kg)", "Gross Wt (Kg)"]
    for i, h in enumerate(headers, start=1):
        col_let = get_column_letter(i)
        _set_cell(ws, f"{col_let}2", h, bold=True, size=9, align="center", fill=_HEADER_FILL)

    net_wt_unit = float(options.get("net_wt_per_carton") or 10.8)
    gross_wt_unit = float(options.get("gross_wt_per_carton") or 12.0)

    row_idx = 3
    if cartons:
        for c in sorted(cartons, key=lambda x: x.get("box_number") or 0):
            box_num = c.get("box_number") or row_idx - 2
            _set_cell(ws, f"A{row_idx}", box_num, size=9, align="center")
            _set_cell(ws, f"B{row_idx}", c.get("style_code") or "", size=9, align="left")
            _set_cell(ws, f"C{row_idx}", c.get("color") or "", size=9, align="left")
            _set_cell(ws, f"D{row_idx}", c.get("size") or "", size=9, align="center")
            _set_cell(ws, f"E{row_idx}", c.get("qty") or 20, size=9, align="center")
            _set_cell(ws, f"F{row_idx}", net_wt_unit, size=9, align="right", num_format="0.000")
            _set_cell(ws, f"G{row_idx}", gross_wt_unit, size=9, align="right", num_format="0.000")
            row_idx += 1
    else:
        for idx, li in enumerate(po.get("line_items", []), start=1):
            _set_cell(ws, f"A{row_idx}", idx, size=9, align="center")
            _set_cell(ws, f"B{row_idx}", li.get("style_code") or "", size=9, align="left")
            _set_cell(ws, f"C{row_idx}", li.get("color") or "", size=9, align="left")
            _set_cell(ws, f"D{row_idx}", li.get("size") or "", size=9, align="center")
            _set_cell(ws, f"E{row_idx}", li.get("quantity") or 0, size=9, align="center")
            _set_cell(ws, f"F{row_idx}", net_wt_unit, size=9, align="right", num_format="0.000")
            _set_cell(ws, f"G{row_idx}", gross_wt_unit, size=9, align="right", num_format="0.000")
            row_idx += 1

    gt_row = max(4, row_idx)
    ws.merge_cells(f"A{gt_row}:D{gt_row}")
    _set_cell(ws, f"A{gt_row}", "TOTAL", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"E{gt_row}", f"=SUM(E3:E{gt_row-1})", bold=True, size=9, align="center", fill=_HEADER_FILL)
    _set_cell(ws, f"F{gt_row}", f"=SUM(F3:F{gt_row-1})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")
    _set_cell(ws, f"G{gt_row}", f"=SUM(G3:G{gt_row-1})", bold=True, size=9, align="right", fill=_HEADER_FILL, num_format="0.000")

    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
