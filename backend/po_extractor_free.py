"""Free / offline PO extraction using pdfplumber + openpyxl.

This module is the *primary* extractor — it never calls an LLM service, never
needs an internet round-trip, and never expires. If it cannot extract a usable
PO it raises ``ExtractionFailed`` so the caller can decide whether to fall back
to an LLM (the LLM path is kept in ``po_extractor.py`` as an opt-in backup).

The strategy is deterministic:
  1.  PDF -> pdfplumber: pull text + tables.  Excel -> openpyxl cells.
  2.  Header/meta fields are detected via labelled regex (PO No, Date, GST etc).
  3.  Line-item table is detected by finding header tokens such as
      "Article", "Style", "Size", "Quantity", "Rate", "Amount". Whichever row
      contains the most of these is treated as the column header.
  4.  Each subsequent row that has a numeric quantity & price becomes a line
      item. Size variants on the same row are flattened.
  5.  Tax + grand-total are inferred from the bottom block (CGST/SGST/IGST).
"""
import io
import re
from datetime import datetime
from typing import Optional

import openpyxl
import pdfplumber


class ExtractionFailed(Exception):
    pass


# ---------- common helpers ----------
_HSN_CODES_FOOTWEAR = "64029990"


def _to_number(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₹", "").replace("Rs", "").replace("INR", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else 0.0


def _to_int(v) -> int:
    return int(round(_to_number(v)))


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _normalise_date(s: str) -> str:
    """Convert DD.MM.YYYY / DD/MM/YYYY / DD-MM-YYYY to YYYY-MM-DD. Returns empty on failure."""
    if not s:
        return ""
    s = str(s).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try regex: any 1-2 digits / 1-2 digits / 2-4 digits
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        try:
            return datetime(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _find_first(pattern: str, text: str, group: int = 1, flags=re.I) -> str:
    m = re.search(pattern, text, flags=flags)
    return _norm(m.group(group)) if m else ""


# ---------- PDF extraction ----------
def extract_po_from_pdf_local(file_bytes: bytes) -> dict:
    try:
        return _extract_pdf(file_bytes)
    except ExtractionFailed:
        raise
    except Exception as e:
        raise ExtractionFailed(f"PDF parse error: {e}") from e


def _extract_pdf(file_bytes: bytes) -> dict:
    full_text_parts = []
    all_tables = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            full_text_parts.append(t)
            for tab in (page.extract_tables() or []):
                if tab:
                    all_tables.append(tab)

    full_text = "\n".join(full_text_parts)
    if not full_text.strip():
        raise ExtractionFailed("PDF has no extractable text (scanned image?). Try Excel or LLM extractor.")

    data = _parse_meta(full_text)
    line_items = _parse_line_items_from_tables(all_tables, data.get("po_number", ""))
    # If the table parser found nothing meaningful, try a text-based fallback
    if not line_items:
        line_items = _parse_line_items_from_text(full_text)
    data["line_items"] = line_items

    # Totals: prefer explicit grand-total found in text; else compute
    _finalise_totals(data, full_text)
    return data


def _parse_meta(text: str) -> dict:
    """Return all header / metadata fields from free text."""
    # PO Number — try the strictest form first (digits-only or alphanum after explicit label)
    po_no = ""
    for pat in [
        r"(?:P\.?\s*O\.?|Purchase\s*Order)\s*(?:No\.?|#|Number)[\s:\-|]+([A-Z0-9][A-Z0-9\-_/]{3,})",
        r"\bOrder\s*(?:No\.?|#|Number)[\s:\-|]+([A-Z0-9][A-Z0-9\-_/]{3,})",
        r"\bP\.?O\.?\s*#[\s:\-|]*([A-Z0-9][A-Z0-9\-_/]{3,})",
    ]:
        m = re.search(pat, text, flags=re.I)
        if m:
            po_no = _norm(m.group(1))
            break

    po_date = _normalise_date(_find_first(r"(?:PO\s*Date|Order\s*Date|Date)[\s:\-|]+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text))
    delivery_date = _normalise_date(_find_first(r"(?:Delivery|Ship(?:ment)?|Due)\s*Date[\s:\-|]+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text))

    # Client / vendor — anchor on full word labels (avoid matching "To" inside "Total" or "Article")
    # Separator can be any combination of colon, dash, pipe, whitespace.
    client_name = ""
    for pat in [
        r"\b(?:Bill\s*To|Buyer|Customer|Consignee|Destination)(?:\s*Name)?\s*[\s:\-|]+\s*([A-Z][A-Za-z0-9 .,&'\-]{3,80})",
    ]:
        m = re.search(pat, text, flags=re.I)
        if m:
            client_name = _norm(m.group(1))
            break

    vendor_name = ""
    for pat in [
        r"\b(?:Vendor|Supplier|Sold\s*By|Seller)(?:\s*Name)?\s*[\s:\-|]+\s*([A-Z][A-Za-z0-9 .,&'\-]{3,80})",
    ]:
        m = re.search(pat, text, flags=re.I)
        if m:
            vendor_name = _norm(m.group(1))
            break

    return {
        "po_number": po_no or "",
        "po_date": po_date or "",
        "delivery_date": delivery_date or "",
        "client_name": client_name or "",
        "vendor_name": vendor_name or "",
        "client_address": "",
        "vendor_address": "",
        "billing_address": "",
        "shipping_address": "",
        "payment_terms": _find_first(r"Payment\s*Terms?\s*[:\-|]?\s*([^\n|]+)", text),
        "currency": "INR",
        "notes": "",
    }


# ---------- line-item table detection ----------
_HEADER_TOKENS = {
    "style": ["style", "article", "model", "item code", "sku"],
    "description": ["description", "particulars", "item name", "product"],
    "color": ["color", "colour", "shade"],
    "size": ["size", "uk size"],
    "hsn": ["hsn", "sac", "hsn code"],
    "quantity": ["quantity", "qty", "pcs", "pairs"],
    "unit_price": ["rate", "unit price", "price", "mrp", "unit rate"],
    "amount": ["amount", "total", "value", "net amount"],
}


def _classify_header(cell: str) -> str | None:
    s = _norm(cell).lower()
    for key, candidates in _HEADER_TOKENS.items():
        for c in candidates:
            if c == s or (len(s) <= 18 and c in s):
                return key
    return None


def _parse_line_items_from_tables(tables: list, po_no: str) -> list[dict]:
    items: list[dict] = []
    for tbl in tables:
        # Find header row (one with most recognised tokens)
        best_idx = -1
        best_score = 0
        best_map = {}
        for i, row in enumerate(tbl[:5]):  # header usually in first 5 rows
            classes = [_classify_header(c) for c in row]
            score = sum(1 for c in classes if c)
            if score > best_score:
                best_score, best_idx = score, i
                best_map = {j: c for j, c in enumerate(classes) if c}
        if best_score < 2 or best_idx < 0:
            continue

        for row in tbl[best_idx + 1:]:
            if not row or all((c is None or str(c).strip() == "") for c in row):
                continue
            rec = {"style_code": "", "description": "", "color": "", "size": "",
                   "hsn_code": "", "quantity": 0, "unit_price": 0.0, "amount": 0.0, "mrp": ""}
            for j, key in best_map.items():
                if j >= len(row):
                    continue
                val = row[j]
                if val is None:
                    continue
                sval = _norm(val)
                if key == "style":
                    rec["style_code"] = sval
                elif key == "description":
                    rec["description"] = sval
                elif key == "color":
                    rec["color"] = sval
                elif key == "size":
                    rec["size"] = sval
                elif key == "hsn":
                    rec["hsn_code"] = sval
                elif key == "quantity":
                    rec["quantity"] = _to_int(sval)
                elif key == "unit_price":
                    rec["unit_price"] = _to_number(sval)
                elif key == "amount":
                    rec["amount"] = _to_number(sval)
            # Filter rows: must have a non-empty style + positive qty + positive price
            if rec["quantity"] > 0 and rec["unit_price"] > 0 and (rec["style_code"] or rec["description"]):
                if not rec["hsn_code"]:
                    rec["hsn_code"] = _HSN_CODES_FOOTWEAR
                if not rec["amount"]:
                    rec["amount"] = round(rec["quantity"] * rec["unit_price"], 2)
                items.append(rec)
    return items


# ---------- text fallback for line items ----------
def _parse_line_items_from_text(text: str) -> list[dict]:
    """Lightweight fallback: look for repeating rows like 'STYLECODE  DESC  COLOR  SIZE  QTY  PRICE  AMOUNT'."""
    items = []
    for raw in text.splitlines():
        line = _norm(raw)
        # Heuristic: at least 4 tokens, with 2 large numbers near the end
        m = re.match(
            r"^([A-Z][A-Z0-9_\-]{2,})\s+(.+?)\s+(\d+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$",
            line,
        )
        if m:
            code, desc, qty, price, amount = m.groups()
            items.append({
                "style_code": code, "description": desc, "color": "", "size": "",
                "hsn_code": _HSN_CODES_FOOTWEAR,
                "quantity": int(qty),
                "unit_price": _to_number(price), "amount": _to_number(amount), "mrp": "",
            })
    return items


def _finalise_totals(data: dict, full_text: str) -> None:
    items = data.get("line_items", [])
    subtotal = sum(li.get("amount", 0) for li in items)
    total_qty = sum(li.get("quantity", 0) for li in items)

    # Try to detect explicit tax info
    cgst_amt = _to_number(_find_first(r"CGST[^\d-]*([0-9.,]+)", full_text))
    sgst_amt = _to_number(_find_first(r"SGST[^\d-]*([0-9.,]+)", full_text))
    igst_amt = _to_number(_find_first(r"IGST[^\d-]*([0-9.,]+)", full_text))
    grand    = _to_number(_find_first(r"(?:Grand\s*Total|Total\s*Amount|Net\s*Payable)[^\d-]*([0-9.,]+)", full_text))

    data["subtotal"] = round(subtotal, 2)
    data["total_quantity"] = total_qty
    data["cgst_rate"] = 0
    data["sgst_rate"] = 0
    data["igst_rate"] = 0
    data["cgst_amount"] = round(cgst_amt, 2)
    data["sgst_amount"] = round(sgst_amt, 2)
    data["igst_amount"] = round(igst_amt, 2)
    data["total_tax"] = round(cgst_amt + sgst_amt + igst_amt, 2)
    data["grand_total"] = round(grand if grand else subtotal + cgst_amt + sgst_amt + igst_amt, 2)


# ---------- Excel extraction ----------
def extract_po_from_xlsx_local(file_bytes: bytes) -> dict:
    try:
        return _extract_xlsx(file_bytes)
    except ExtractionFailed:
        raise
    except Exception as e:
        raise ExtractionFailed(f"Excel parse error: {e}") from e


def _extract_xlsx(file_bytes: bytes) -> dict:
    bio = io.BytesIO(file_bytes)
    try:
        wb = openpyxl.load_workbook(bio, data_only=True)
    except Exception as e:
        raise ExtractionFailed(f"Cannot open xlsx: {e}") from e

    # We build a (row, col) -> value map for the first sheet (and merge a flat textual rep for regex)
    ws = wb[wb.sheetnames[0]]
    grid = []
    flat_lines = []
    for row in ws.iter_rows(values_only=True):
        grid.append([("" if c is None else c) for c in row])
        flat_lines.append(" | ".join(("" if c is None else str(c)) for c in row))
    flat = "\n".join(flat_lines)

    data = _parse_meta(flat)
    items = _parse_xlsx_line_items(grid)
    if not items:
        items = _parse_line_items_from_text(flat)
    data["line_items"] = items
    _finalise_totals(data, flat)
    return data


def _parse_xlsx_line_items(grid: list[list]) -> list[dict]:
    """Detect the header row in the sheet and read rows below."""
    if not grid:
        return []
    best_row = -1
    best_map: dict[int, str] = {}
    best_score = 0
    # Header expected within first 30 rows
    for r_idx in range(min(len(grid), 30)):
        row = grid[r_idx]
        classes = [_classify_header(c) for c in row]
        score = sum(1 for c in classes if c)
        if score > best_score:
            best_score, best_row = score, r_idx
            best_map = {j: c for j, c in enumerate(classes) if c}

    if best_score < 2:
        return []

    items = []
    for r in grid[best_row + 1:]:
        if not any(str(c).strip() for c in r if c is not None):
            continue
        rec = {"style_code": "", "description": "", "color": "", "size": "",
               "hsn_code": "", "quantity": 0, "unit_price": 0.0, "amount": 0.0, "mrp": ""}
        for j, key in best_map.items():
            if j >= len(r):
                continue
            val = r[j]
            if key == "style":
                rec["style_code"] = _norm(val)
            elif key == "description":
                rec["description"] = _norm(val)
            elif key == "color":
                rec["color"] = _norm(val)
            elif key == "size":
                rec["size"] = _norm(val)
            elif key == "hsn":
                rec["hsn_code"] = _norm(val)
            elif key == "quantity":
                rec["quantity"] = _to_int(val)
            elif key == "unit_price":
                rec["unit_price"] = _to_number(val)
            elif key == "amount":
                rec["amount"] = _to_number(val)
        if rec["quantity"] > 0 and rec["unit_price"] > 0 and (rec["style_code"] or rec["description"]):
            if not rec["hsn_code"]:
                rec["hsn_code"] = _HSN_CODES_FOOTWEAR
            if not rec["amount"]:
                rec["amount"] = round(rec["quantity"] * rec["unit_price"], 2)
            items.append(rec)
    return items
