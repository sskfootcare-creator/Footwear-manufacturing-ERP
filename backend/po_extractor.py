"""PO extraction — free local parsing first, optional LLM fallback.

The local extractor (pdfplumber + openpyxl, in `po_extractor_free.py`) handles
nearly every footwear PO format the user uploads. If it fails because the PDF
is purely image-based or the layout is too exotic, we fall back to the
Google Gemini LLM call (only if `GEMINI_API_KEY` is configured AND the
fallback is enabled).
"""
import json
import os
import re
import tempfile
from pathlib import Path

from po_extractor_free import (
    ExtractionFailed,
    extract_po_from_pdf_local,
    extract_po_from_xlsx_local,
    STATE_CODE_MAP,
)


_USE_LLM_FALLBACK = os.environ.get("PO_EXTRACTOR_LLM_FALLBACK", "true").lower() == "true"


EXTRACTION_PROMPT = """You are an expert at extracting structured data from Purchase Order documents for footwear manufacturing.

Extract the following fields from the attached PO document and return ONLY valid JSON (no markdown, no commentary):

{
  "po_number": "string",
  "po_date": "YYYY-MM-DD",
  "client_name": "string",
  "client_address": "string",
  "client_gstin": "string",
  "client_state": "string",
  "client_state_code": "string",
  "vendor_name": "string",
  "vendor_address": "string",
  "billing_address": "string",
  "shipping_address": "string",
  "delivery_date": "YYYY-MM-DD",
  "payment_terms": "string",
  "currency": "INR",
  "line_items": [
    {
      "item_code": "string",
      "style_code": "string",
      "description": "string",
      "color": "string",
      "size": "string",
      "hsn_code": "string",
      "quantity": 0,
      "unit_price": 0.0,
      "amount": 0.0
    }
  ],
  "subtotal": 0.0,
  "cgst_rate": 0.0,
  "cgst_amount": 0.0,
  "sgst_rate": 0.0,
  "sgst_amount": 0.0,
  "igst_rate": 0.0,
  "igst_amount": 0.0,
  "total_tax": 0.0,
  "grand_total": 0.0,
  "total_quantity": 0,
  "notes": "string"
}

Rules:
- Convert dates from DD.MM.YYYY or DD/MM/YYYY to YYYY-MM-DD.
- For interstate transactions use IGST, intrastate use CGST+SGST.
- Each size variant should be its own line_item.
- All numeric values must be numbers, not strings.
- If a field is missing, use null or empty string.
- Return ONLY the JSON, no other text."""


def _clean_json(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


def _validate(data: dict) -> bool:
    """Treat the parse as successful when at least one usable line item came out."""
    if not isinstance(data, dict):
        return False
    items = data.get("line_items") or []
    return any((li.get("quantity") or 0) > 0 for li in items)


def _is_low_confidence(data: dict) -> bool:
    """Detect if locally extracted PO data is incomplete or low confidence.
    
    A result is considered low confidence if:
    1. It has no valid line items with quantity > 0.
    2. All line items are missing 'color' or all are missing 'size' (e.g., neither
       recognized local format where variant splitting silently failed).
    3. More than 50% of items are missing color or size attributes.
    4. All line items are missing an item identifier (style_code, item_code, or description).
    """
    if not isinstance(data, dict):
        return True
    items = data.get("line_items") or []
    valid_items = [li for li in items if (li.get("quantity") or 0) > 0]
    if not valid_items:
        return True

    total = len(valid_items)
    empty_color = sum(1 for li in valid_items if not str(li.get("color") or "").strip())
    empty_size = sum(1 for li in valid_items if not str(li.get("size") or "").strip())
    empty_identifier = sum(
        1 for li in valid_items
        if not str(li.get("style_code") or "").strip()
        and not str(li.get("item_code") or "").strip()
        and not str(li.get("description") or "").strip()
    )

    # If any essential footwear variant attribute is completely missing across all items
    if empty_color == total or empty_size == total or empty_identifier == total:
        return True

    # If more than half of items have missing color or missing size
    if empty_color > total / 2 or empty_size > total / 2:
        return True

    return False


# ---------------------- PUBLIC API (async) ----------------------
async def extract_po_from_pdf(file_bytes: bytes) -> dict:
    # 1) try the free local pipeline
    data = None
    try:
        data = extract_po_from_pdf_local(file_bytes)
        if _validate(data) and not _is_low_confidence(data):
            data["extraction_method"] = "local"
            data["confidence"] = "high"
            return data
    except ExtractionFailed:
        data = None

    # 2) optional LLM fallback (triggered if local failed OR result was low confidence)
    if _USE_LLM_FALLBACK and os.environ.get("GEMINI_API_KEY"):
        try:
            llm_data = await _llm_extract_pdf(file_bytes)
            if isinstance(llm_data, dict):
                llm_data["extraction_method"] = "gemini_fallback"
                llm_data["confidence"] = "high"
                llm_data["confidence_warning"] = (
                    "Local parser had low confidence (unrecognized format / missing color/size); "
                    "extracted via Gemini AI fallback."
                )
            return llm_data
        except Exception as e:
            # Bubble the local-parser result if we have one even though it's weak
            if data and _validate(data):
                data["extraction_method"] = "local_low_confidence"
                data["confidence"] = "low"
                data["confidence_warning"] = f"AI fallback failed ({e}); using weak local parse result."
                return data
            raise RuntimeError(f"PO extraction failed (local + LLM): {e}") from e

    # 3) nothing worked (or fallback disabled/unavailable)
    if data and _validate(data):
        data["extraction_method"] = "local_low_confidence"
        data["confidence"] = "low"
        data["confidence_warning"] = "Low confidence extraction: color/size could not be determined from this format."
        return data
    raise RuntimeError(
        "Could not extract PO from this PDF locally and the LLM fallback is disabled. "
        "Try uploading an Excel version of the PO, or set GEMINI_API_KEY and "
        "PO_EXTRACTOR_LLM_FALLBACK=true to enable AI fallback."
    )


async def extract_po_from_xlsx(file_bytes: bytes) -> dict:
    data = None
    try:
        data = extract_po_from_xlsx_local(file_bytes)
        if _validate(data) and not _is_low_confidence(data):
            data["extraction_method"] = "local"
            data["confidence"] = "high"
            return data
    except ExtractionFailed:
        data = None

    if _USE_LLM_FALLBACK and os.environ.get("GEMINI_API_KEY"):
        try:
            llm_data = await _llm_extract_xlsx(file_bytes)
            if isinstance(llm_data, dict):
                llm_data["extraction_method"] = "gemini_fallback"
                llm_data["confidence"] = "high"
                llm_data["confidence_warning"] = (
                    "Local parser had low confidence (unrecognized format / missing color/size); "
                    "extracted via Gemini AI fallback."
                )
            return llm_data
        except Exception as e:
            if data and _validate(data):
                data["extraction_method"] = "local_low_confidence"
                data["confidence"] = "low"
                data["confidence_warning"] = f"AI fallback failed ({e}); using weak local parse result."
                return data
            raise RuntimeError(f"PO extraction failed (local + LLM): {e}") from e

    if data and _validate(data):
        data["extraction_method"] = "local_low_confidence"
        data["confidence"] = "low"
        data["confidence_warning"] = "Low confidence extraction: color/size could not be determined from this format."
        return data
    raise RuntimeError("Could not extract PO from this Excel file (no recognisable line items).")


# ---------------------- LLM FALLBACK (lazy import) ----------------------
async def _llm_extract_pdf(file_bytes: bytes) -> dict:
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY is not set. Please set the GEMINI_API_KEY environment variable.")
    from google import genai
    client = genai.Client(api_key=api_key)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[
            genai.types.Part.from_bytes(data=file_bytes, mime_type="application/pdf"),
            EXTRACTION_PROMPT,
        ],
    )
    return _post_process_extracted_data(json.loads(_clean_json(response.text)))


async def _llm_extract_xlsx(file_bytes: bytes) -> dict:
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY is not set. Please set the GEMINI_API_KEY environment variable.")
    from google import genai
    import openpyxl
    client = genai.Client(api_key=api_key)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        parts = []
        for sn in wb.sheetnames:
            parts.append(f"### Sheet: {sn}")
            for row in wb[sn].iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    parts.append(" | ".join(cells))
        text_content = "\n".join(parts)
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"{EXTRACTION_PROMPT}\n\n--- Document Content ---\n{text_content}",
        )
        return _post_process_extracted_data(json.loads(_clean_json(response.text)))
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _post_process_extracted_data(data: dict) -> dict:
    if not isinstance(data, dict):
        return data
    cg = str(data.get("client_gstin") or "").strip().upper()
    if cg:
        cg = "".join(c for c in cg if c.isalnum())
        data["client_gstin"] = cg
        if len(cg) >= 2 and cg[:2].isdigit():
            if not data.get("client_state_code"):
                data["client_state_code"] = cg[:2]
            if not data.get("client_state"):
                data["client_state"] = STATE_CODE_MAP.get(cg[:2], "")
    return data

